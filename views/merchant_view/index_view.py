import os, random, datetime, threading, shortuuid
from openpyxl import Workbook
from flask import render_template, current_app, request, abort
from .base import FormViewBase, TableViewBase
from models.pay_table import MerchantTable, CollectionOrderTable, MerchantBankCardTable, BankTable, MerchantBillStatementTable, MerchantTunnleTable, TunnelTable, WithdrawTable, RechargeMoneyTable
from modules.google_module.google_verify import GooleVerifyCls
from common_utils.utils_funcs import getDayDateSilce, getMonthDateSilce, PagingCLS
from constants import BANK_CODE, PAY_METHOD, WITHDRAW_STATUS, ASSETS_FOLDER, EXPORT_FOLDER, BILL_STATEMEN_TYPES, CallbackState, MERCHANT_ROLES
from models.site_table import ExportStatu, ExportDataModel
from models.behalfPay import behalfPayOrderTable
from modules.view_helpres.view_func import MerchantUpdateAmout_func



class MerchantIndexView(FormViewBase):
    add_url_rules = [['/', 'merchant_index']]
    title = '世界付'

    def view_get(self):
        self.context['title'] = self.title
        time_start, time_end = getDayDateSilce()
        if self.is_merchant:
            merchant_id = self.current_user_dict.get('merchant_id')
        elif self.is_submerchant:
            merchant_id = self.current_user_dict.get('upper_data').get('merchant_id')
        else:
            return abort(404)
        orderDayCount = CollectionOrderTable.count({'order_time': {'$gte':time_start, '$lte': time_end}, 'merchant_id': merchant_id})
        self.context['orderDayCount'] = orderDayCount or 0
        time_start, time_end = getMonthDateSilce()
        orderMonthCount = CollectionOrderTable.count({'order_time': {'$gte':time_start, '$lte': time_end}, 'merchant_id': merchant_id})
        self.context['orderMonthCount'] = orderMonthCount or 0
        return render_template('merchant/index.html', **self.context)

    def post_other_way(self):
        if self.action == 'getWithdrawHtml':
            bankcark_html = ''
            for mc in MerchantBankCardTable.find_many({'merchant_uuid': self.current_user_dict.get('uuid')}):
                bank_data = BankTable.find_one({'uuid': mc.get('bank_uid')}) or {}
                bankcark_html += f'''
                    <option value="{mc.get('uuid')}">{bank_data.get('shortName')} - {mc.get('account')} - {mc.get('account_username')}</option>                
                '''

            html = f'''
                <div class="formBox">
                    <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">       
                        <div class="list-group-item">
                            <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行卡：</span>
                            <select class="form-control" id="bancardk_uid" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                                <option value="">选择银行卡</option>
                                {bankcark_html}
                            </select>
                        </div>
                        <div class="list-group-item">
                            <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">结算金额：</span>
                            <input type="number" class="form-control" id="jsamount" value="" onchange="monitorRechargeMoney()" placeholder="结算金额" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                        </div>                   
                        <div class="list-group-item">
                            <span style="width: 120px; text-align: right; display: inline-block; position: relative;">可用余额：</span>
                            <span style="display: inline-block;width: calc(100% - 180px);text-align: left;font-size: 16px;vertical-align: -1px;"><span id="balance_amount" data-amount="{ round(self.current_user_dict.get('balance_amount') or 0, 2) }">{ self.format_money(round(self.current_user_dict.get('balance_amount') or 0, 2)) }</span> VND</span>
                        </div>                                    
                        <div class="list-group-item">
                            <span style="width: 120px; text-align: right; display: inline-block; position: relative;">结算后余额：</span>
                            <span style="display: inline-block;width: calc(100% - 180px);text-align: left;font-size: 16px;vertical-align: -1px;"><span id="end_balance_amount">{ self.format_money(round(self.current_user_dict.get('balance_amount') or 0, 2)) }</span> VND</span>
                        </div>                                                 
                    </div>

                    <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                    <div style="position: relative; text-align: center">
                        <span class="btn btn-primary" onclick="post_jsamount()">确定</span>&emsp;
                        <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                    </div>                                                                                 
                </div>
            '''
            return self.xtjson.json_result(message=html)
        if self.action == 'jsamount':
            bancardk_uid = self.request_data.get('bancardk_uid')
            jsamount = self.request_data.get('jsamount')
            if not bancardk_uid or not jsamount:
                return self.xtjson.json_params_error()
            try:
                jsamount = int(jsamount)
            except:
                return self.xtjson.json_params_error('金额输入错误！')

            issued_money_rate = self.current_user_dict.get('issued_money_rate') or 0
            balance_amount = self.current_user_dict.get('balance_amount') or 0
            if jsamount > balance_amount:
                return self.xtjson.json_params_error('结算超额！')

            repay_amount = round(issued_money_rate * jsamount, 2)
            actual_amount = jsamount + repay_amount
            _state, _balance_amount = MerchantUpdateAmout_func(actual_amount, self.current_user_dict.get('uuid'), is_add=False)
            if not _state:
                return self.xtjson.json_params_error('结算余额！')

            bancardk_data = MerchantBankCardTable.find_one({'uuid': bancardk_uid}) or {}
            if not bancardk_data:
                return self.xtjson.json_params_error()

            banc_data = BankTable.find_one({'uuid': bancardk_data.get('bank_uid')}) or {}
            if not banc_data:
                return self.xtjson.json_params_error()

            _inter_data = {
                'merchant_uuid': self.current_user_dict.get('uuid'),
                'amount': jsamount,
                'repay_amount': repay_amount,
                'actual_amount': actual_amount,
                'statu': WITHDRAW_STATUS.review,
                'bankcard_uuid': bancardk_uid,
                'agentadmin_uuid': self.current_user_dict.get('agentadmin_uuid'),
                'payee_bank': banc_data.get('code'),
                'payee_bankcard': bancardk_data.get('account'),
                'payee_username': bancardk_data.get('account_username'),
            }
            WithdrawTable.insert_one(_inter_data)

            _mbs_data = {
                'merchant_uuid': self.current_user_dict.get('uuid'),
                'amount': int('-'+str(jsamount)),
                'balance_amount': _balance_amount,
                'note': '',
                'repay_amount': repay_amount,
                'bill_type': BILL_STATEMEN_TYPES.SETTLEMENT,
                'agentadmin_uuid': self.current_user_dict.get('agentadmin_uuid')
            }
            MerchantBillStatementTable.insert_one(_mbs_data)
            return self.xtjson.json_result()



class MerchantUserView(FormViewBase):
    add_url_rules = [['/user', 'merchant_user']]
    title = '商户信息'
    MCLS = MerchantTable

    def view_get(self):
        self.context['title'] = self.title

        if self.is_merchant:
            merchant_uuid = self.current_user_dict.get('uuid')
        elif self.is_submerchant:
            merchant_uuid = self.current_user_dict.get('upper_data').get('uuid')
        else:
            return abort(404)

        tunnle_datas = []
        for dt in MerchantTunnleTable.find_many({'merchant_uuid': merchant_uuid}):
            tunnle_data = TunnelTable.find_one({'uuid': dt.get('tunnle_id')})
            if tunnle_data.get('code') == PAY_METHOD.VNBANK:
                continue
            if tunnle_data.get('code') == PAY_METHOD.VNDIRECT:
                continue
            dt['tunnle_data'] = tunnle_data
            _rate = dt.get('rate') or 0
            _rate = round(_rate * 100, 5)
            dt['rate'] = _rate
            tunnle_datas.append(dt)

        logoUrl = self.current_user_dict.get('logoUrl') or ''
        logoName = logoUrl.rsplit('/', 1)[-1]

        self.context['logoUrl'] = logoUrl
        self.context['logoName'] = logoName
        self.context['PAY_METHOD'] = PAY_METHOD
        self.context['tunnle_datas'] = tunnle_datas
        return render_template('merchant/merchant_user.html', **self.context)

    def resetPasswordHtml(self):
        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">       
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">新密码：</span>
                        <input type="text" class="form-control" id="password" placeholder="新密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">确认密码：</span>
                        <input type="text" class="form-control" id="confirmPassword" placeholder="确认密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                    
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">Google验证码：</span>
                        <input type="text" class="form-control" id="verify_code" placeholder="Google验证码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                    
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="resetPassword_func()">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>
        '''
        return html

    def post_other_way(self):
        if self.action == 'resetPasswordHtml':
            html = self.resetPasswordHtml()
            return self.xtjson.json_result(message=html)
        if self.action == 'uploadLogo':
            fileobj = request.files['upload']
            fname, fext = os.path.splitext(fileobj.filename)
            relatively_path = f'/assets/upload/merchant/{self.current_user_dict.get("uuid")}/images/'
            import_folder = self.project_static_folder + relatively_path
            if not os.path.exists(import_folder):
                os.makedirs(import_folder)

            new_filename = shortuuid.uuid()
            fileobj.save(import_folder + new_filename + fext)
            filePath = f'/assets/upload/merchant/{self.current_user_dict.get("uuid")}/images/' + new_filename + fext
            MerchantTable.update_one({'uuid': self.current_user_dict.get('uuid')}, {'$set': {'logoUrl': filePath}})
            return self.xtjson.json_result(data={'filePath': filePath, 'logoNmae': new_filename + fext})
        if self.action == 'delLogo':
            MerchantTable.update_one({'uuid': self.current_user_dict.get('uuid')}, {'$set': {'logoUrl': ''}})
            return self.xtjson.json_result()

    def post_data_other_way(self):
        if self.action == 'resetPassword':
            password = self.request_data.get('password')
            verify_code = self.request_data.get('verify_code')

            if not password or not password.strip():
                return self.xtjson.json_params_error('请输入新密码！')

            if not verify_code or not verify_code.strip():
                return self.xtjson.json_params_error('请输入GOOGLE验证码！')

            if len(password.strip()) < 6:
                return self.xtjson.json_params_error('登录密码应为6~18位之间！')

            if len(password.strip()) > 18:
                return self.xtjson.json_params_error('登录密码应为6~18位之间！')

            if not password.strip().isalnum():
                return self.xtjson.json_params_error('密码应是字母和数字组成！')

            googleObj = GooleVerifyCls(pwd=self.data_uuid, s_label='pay2wold', account=self.data_dict.get('account'))
            if not googleObj.check_goole_code(verify_code):
                return self.xtjson.json_params_error('GOOGLE验证码输入错误!')

            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': {'password': self.MCLS.encry_password(password.strip())}})
            return self.xtjson.json_result()
        if self.action == 'updateOnlineStatu':
            if self.data_dict.get('statu'):
                self.data_from['statu'] = False
            else:
                self.data_from['statu'] = True
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            return self.xtjson.json_result()



class MerchantBankCardView(TableViewBase):
    add_url_rules = [['/merchantBankCard', 'merchantBankCard']]
    title = '商户银行卡'
    MCLS = MerchantBankCardTable
    template = 'merchant/bankCard.html'
    per_page = 30

    def get_filter_dict(self):
        if self.is_merchant:
            merchant_uuid = self.current_user_dict.get('uuid')
        elif self.is_submerchant:
            merchant_uuid = self.current_user_dict.get('upper_data').get('uuid')
        else:
            return abort(404)
        return {'merchant_uuid': merchant_uuid}

    def dealwith_main_context(self):
        datas = []
        all_datas = self.context.get('all_datas') or []
        for da in all_datas:
            _dd = BankTable.find_one({'uuid': da.get('bank_uid')})
            da['bank_data'] = _dd
            datas.append(da)

        self.context['all_datas'] = all_datas

        bank_datas = BankTable.find_many({})
        self.context['bank_datas'] = bank_datas

    def bankcard_html(self, data_dict={}):
        _action = 'add_bankcard'
        if data_dict:
            _action = 'edit_bankcard'

        bank_datas = BankTable.find_many({})

        bank_html = ''
        for bank_data in bank_datas:
            if not data_dict:
                bank_html += f'<span class="ant-tag" style="margin-bottom: 13px;margin-right: 10px;padding: 15px 15px;line-height: 0px;" data-bankuid="{ bank_data.get("uuid") }" onclick="seleBankCode_func($(this))">{ bank_data.get("shortName") or "" }</span>'
            else:
                bank_html += f'<span class="ant-tag { "ant-tag-red" if bank_data.get("uuid") == data_dict.get("bank_uid") else "" }" style="margin-bottom: 13px;margin-right: 10px;padding: 15px 15px;line-height: 0px;" data-bankuid="{ bank_data.get("uuid") }" onclick="seleBankCode_func($(this))">{ bank_data.get("shortName") or "" }</span>'

        bank_htmll = f'''    
                <div class="list-group-item" style="display: flex; align-items: center;justify-content: center;">
                    <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行类型：</span>
                    <input type="hidden" value="{ data_dict.get("bank_uid") or "" }" id="bank_uid">
                    <div class="form-control" style="display: inline-block; width: calc(100% - 180px); height: 300px; overflow-y: scroll; text-align: left;">
                        { bank_html }                        
                    </div>
                </div>                         
        '''

        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">    
                    { bank_htmll }                   
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行卡账户：</span>
                        <input type="text" class="form-control" id="account" value="{ data_dict.get('account') or '' }" placeholder="银行卡账户" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">持卡人：</span>
                        <input type="text" class="form-control" id="account_username" value="{ data_dict.get('account_username') or '' }" placeholder="持卡人" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                                        
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_bankcard_data('{_action}', '{ data_dict.get('uuid') if data_dict else '' }')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>
        '''
        return html

    def post_other_way(self):
        if self.action == 'add_bankcard_html':
            html = self.bankcard_html()
            return self.xtjson.json_result(message=html)
        if self.action == 'add_bankcard':
            bank_uid = self.request_data.get('bank_uid')
            account = self.request_data.get('account')
            account_username = self.request_data.get('account_username')
            if not bank_uid or not account or not account_username or not bank_uid.strip() or not account.strip() or not account_username.strip():
                return self.xtjson.json_params_error('缺少数据！')

            self.data_from.update({
                'bank_uid': bank_uid,
                'account': account,
                'account_username': account_username,
                'merchant_uuid': self.current_user_dict.get('uuid'),
            })
            self.MCLS.insert_one(self.data_from)
            return self.xtjson.json_result()

    def post_data_other_way(self):
        if self.action == 'edit_bankcard_html':
            html = self.bankcard_html(self.data_dict)
            return self.xtjson.json_result(message=html)
        if self.action == 'edit_bankcard':
            bank_uid = self.request_data.get('bank_uid')
            account = self.request_data.get('account')
            account_username = self.request_data.get('account_username')
            if not bank_uid or not account or not account_username or not bank_uid.strip() or not account.strip() or not account_username.strip():
                return self.xtjson.json_params_error('缺少数据！')

            self.data_from['bank_uid'] = bank_uid
            self.data_from['account'] = account
            self.data_from['account_username'] = account_username
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            return self.xtjson.json_result()



class MerchantOrderListView(TableViewBase):
    add_url_rules = [['/merchantOrder', 'MerchantOrderListView']]
    title = '代收订单'
    MCLS = CollectionOrderTable
    template = 'merchant/merchantOrder.html'
    per_page = 30
    sort = [['create_time', -1]]

    def exportData(self, datas, log_uuid, export_folder, filename):
        ''' 导出数据 '''
        export_data = ExportDataModel.find_one({'uuid': log_uuid})
        if not export_data:
            return
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1
            header = ['订单号', '商户名称', '商户订单号', '通道类型', '手续费', '订单金额', '支付状态', '回调状态', '订单时间', '回调时间']
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])
            mdaict = {}
            for data in datas:
                row += 1
                _md = mdaict.get(data.get('merchant_id'))
                if not _md:
                    _md = MerchantTable.find_one({'merchant_id': data.get('merchant_id')})
                    mdaict[data.get('merchant_id')] = _md

                wa.cell(row=row, column=1, value=str(data.get('order_id') or ''))
                wa.cell(row=row, column=2, value=str(_md.get('merchant_name') or ''))
                wa.cell(row=row, column=3, value=str(data.get('merchant_order_id') or ''))
                wa.cell(row=row, column=4, value=str(PAY_METHOD.name_dict.get(data.get('pay_method')) or ''))
                wa.cell(row=row, column=5, value=str(self.format_money(data.get('repay_amount')) or '0'))
                wa.cell(row=row, column=6, value=str(self.format_money(data.get('order_amount')) or '0'))
                wa.cell(row=row, column=7, value='已支付' if data.get('pay_statu') else '未支付')
                callback_statu_text = ''
                if data.get('callback_statu') == CallbackState.SUCCESS:
                    callback_statu_text = '已回调'
                if data.get('callback_statu') == CallbackState.NOT_CALLEDBACK:
                    callback_statu_text = '未回调'
                if data.get('callback_statu') == CallbackState.FAILED:
                    callback_statu_text = '回调失败'
                wa.cell(row=row, column=8, value=callback_statu_text)
                wa.cell(row=row, column=9, value=data.get('order_time').strftime('%Y-%m-%d %H:%M:%S')) if data.get('order_time') else ''
                wa.cell(row=row, column=10, value=data.get('callback_time').strftime('%Y-%m-%d %H:%M:%S')) if data.get('callback_time') else ''

                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def totalData(self):
        if self.is_merchant:
            merchant_id = self.current_user_dict.get('merchant_id')
        elif self.is_submerchant:
            merchant_id = self.current_user_dict.get('upper_data').get('merchant_id')
        else:
            return abort(404)
        filter_dict = {'merchant_id': merchant_id}
        fields = self.MCLS.fields()
        statu, res = self.search_func(fields)
        if not statu:
            return res
        filter_dict.update(res[0])

        # 订单金额统计
        _order_amount_total_ll = CollectionOrderTable.collection().aggregate([
            {"$match": filter_dict},
            {"$group": {"_id": None, "order_amount": {"$sum": '$order_amount'}}},
        ])
        order_amount_total_ll = list(_order_amount_total_ll)
        order_amount_tota = 0
        if order_amount_total_ll:
            order_amount_tota = order_amount_total_ll[0].get('order_amount')

        # 订单手续费统计
        _repay_amount_total_ll = CollectionOrderTable.collection().aggregate([
            {"$match": filter_dict},
            {"$group": {"_id": None, "repay_amount": {"$sum": '$repay_amount'}}},
        ])
        repay_amount_total_ll = list(_repay_amount_total_ll)
        repay_amount_total = 0
        if repay_amount_total_ll:
            repay_amount_total = repay_amount_total_ll[0].get('repay_amount')

        # 实付金额统计
        # fff1 = {}
        # fff1.update(filter_dict)
        # fff1.update({
        #     'pay_statu': True
        # })
        # _actual_amount_total_ll = CollectionOrderTable.collection().aggregate([
        #     {"$match": fff1},
        #     {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
        # ])
        # actual_amount_total_ll = list(_actual_amount_total_ll)
        # actual_amount_tota = 0
        # if actual_amount_total_ll:
        #     actual_amount_tota = actual_amount_total_ll[0].get('actual_amount')
        # actual_amount_count = CollectionOrderTable.count(fff1)

        # 回调金额统计
        fff2 = {}
        fff2.update(filter_dict)
        dd_time = fff2.get('order_time') or {}
        if dd_time:
            fff2.pop('order_time')
        fff2.update({
            'callback_statu': CallbackState.SUCCESS,
        })
        if dd_time:
            fff2['callback_time'] = dd_time
        _callback_amount_total_ll = CollectionOrderTable.collection().aggregate([
            {"$match": fff2},
            {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
        ])
        callback_amount_total_ll = list(_callback_amount_total_ll)
        callback_amount_tota = 0
        if callback_amount_total_ll:
            callback_amount_tota = callback_amount_total_ll[0].get('actual_amount')
        callback_amount_count = CollectionOrderTable.count(fff2)

        # 订单手续费统计
        _hd_repay_amount_total_ll = CollectionOrderTable.collection().aggregate([
            {"$match": fff2},
            {"$group": {"_id": None, "repay_amount": {"$sum": '$repay_amount'}}},
        ])
        hd_repay_amount_total_ll = list(_hd_repay_amount_total_ll)
        hd_repay_amount_total = 0
        if hd_repay_amount_total_ll:
            hd_repay_amount_total = hd_repay_amount_total_ll[0].get('repay_amount')


        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto; text-align: left; overflow-x: hidden; font-size: 14px; line-height: 25px; color: #333333;">
                    <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">订单金额：</span>{ self.format_money(order_amount_tota) }</p>
                    <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">订单手续费金额：</span>{ self.format_money(repay_amount_total) }</p>
                    <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">回调金额：</span>{ self.format_money(callback_amount_tota) }</p>
                    <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">回调笔数：</span>{ callback_amount_count }</p>
                    <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">回调手续费：</span>{ self.format_money(hd_repay_amount_total) }</p>
                </div>
            </div>
        '''
        '''
                    # <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">实付金额：</span>{ self.format_money(actual_amount_tota) }</p>
                    # <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">实付笔数：</span>{ actual_amount_count }</p>        
        '''
        return self.xtjson.json_result(message=html)

    def get_filter_dict(self):
        if self.is_merchant:
            merchant_id = self.current_user_dict.get('merchant_id')
        elif self.is_submerchant:
            merchant_id = self.current_user_dict.get('upper_data').get('merchant_id')
        else:
            return abort(404)
        return {'merchant_id': merchant_id}

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        datas = []
        for da in all_datas:
            merchant_data = MerchantTable.find_one({'merchant_id': da.get('merchant_id')})
            da['merchant_data'] = merchant_data
            datas.append(da)
        self.context['all_datas'] = datas
        self.context['PAY_METHOD'] = PAY_METHOD
        self.context['CallbackState'] = CallbackState

    def post_other_way(self):
        if self.action == 'export_order':
            if self.is_merchant:
                merchant_id = self.current_user_dict.get('merchant_id')
            elif self.is_submerchant:
                merchant_id = self.current_user_dict.get('upper_data').get('merchant_id')
            else:
                return abort(404)
            filter_dict = {'merchant_id': merchant_id}
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if not statu:
                return res
            filter_dict.update(res[0])
            sort_query = 'create_time'
            print(filter_dict)
            if 'order_time' in filter_dict:
                sort_query = 'order_time'
            elif 'callback_time' in filter_dict:
                sort_query = 'callback_time'
            datas = CollectionOrderTable.find_many(filter_dict, sort=[[sort_query, -1]])
            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            _out_data_dict = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas),
                'out_count': 0,
                'note': '代收订单-' + datetime.datetime.now().strftime('%Y%m%d%H%M%S'),
                'merchant_uuid': self.current_user_dict.get('uuid'),
                'agentadmin_uuid': self.current_user_dict.get('agentadmin_uuid'),
            }
            uuid = ExportDataModel.insert_one(_out_data_dict)
            threading.Thread(target=self.exportData, args=(datas, uuid, export_folder, filename)).start()
            return self.xtjson.json_result(message='数据导出中，请稍后到"文件下载"中查看数据！')
        if self.action == 'totalData':
            return self.totalData()



class MerchPayBehalfOrderView(TableViewBase):
    add_url_rules = [['/payBehalfOrder', 'MerchPayBehalfOrder']]
    title = '代付订单'
    MCLS = behalfPayOrderTable
    template = 'merchant/payBehalfOrder.html'
    per_page = 30
    sort = [['create_time', -1]]

    def exportData(self, datas, log_uuid, export_folder, filename):
        ''' 导出数据 '''
        export_data = ExportDataModel.find_one({'uuid': log_uuid})
        if not export_data:
            return
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1
            header = ['订单号', '商户名称', '商户订单号', '通道类型', '手续费', '订单金额', '支付状态', '回调状态', '订单时间', '回调时间']
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])
            mdaict = {}
            for data in datas:
                row += 1
                _md = mdaict.get(data.get('merchant_id'))
                if not _md:
                    _md = MerchantTable.find_one({'merchant_id': data.get('merchant_id')})
                    mdaict[data.get('merchant_id')] = _md

                wa.cell(row=row, column=1, value=str(data.get('order_id') or ''))
                wa.cell(row=row, column=2, value=str(_md.get('merchant_name') or ''))
                wa.cell(row=row, column=3, value=str(data.get('merchant_order_id') or ''))
                wa.cell(row=row, column=4, value=str(PAY_METHOD.name_dict.get(data.get('pay_method')) or ''))
                wa.cell(row=row, column=5, value=str(self.format_money(data.get('repay_amount')) or '0'))
                wa.cell(row=row, column=6, value=str(self.format_money(data.get('order_amount')) or '0'))
                wa.cell(row=row, column=7, value='已支付' if data.get('pay_statu') else '未支付')
                wa.cell(row=row, column=8, value='已通知' if data.get('callback_statu') else '未通知')
                wa.cell(row=row, column=9, value=data.get('order_time').strftime('%Y-%m-%d %H:%M:%S')) if data.get('order_time') else ''
                wa.cell(row=row, column=10, value=data.get('callback_time').strftime('%Y-%m-%d %H:%M:%S')) if data.get('callback_time') else ''

                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def get_filter_dict(self):
        if self.is_merchant:
            merchant_id = self.current_user_dict.get('merchant_id')
        elif self.is_submerchant:
            merchant_id = self.current_user_dict.get('upper_data').get('merchant_id')
        else:
            return abort(404)
        return {'merchant_id': merchant_id}

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        datas = []
        for da in all_datas:
            merchant_data = MerchantTable.find_one({'merchant_id': da.get('merchant_id')}) or {}
            da['merchant_data'] = merchant_data
            datas.append(da)
        self.context['all_datas'] = datas
        self.context['CallbackState'] = CallbackState

    def post_other_way(self):
        if self.action == 'export_order':
            if self.is_merchant:
                merchant_id = self.current_user_dict.get('merchant_id')
            elif self.is_submerchant:
                merchant_id = self.current_user_dict.get('upper_data').get('merchant_id')
            else:
                return abort(404)
            filter_dict = {'merchant_id': merchant_id}
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if not statu:
                return res
            filter_dict.update(res[0])
            sort_query = 'create_time'
            if 'order_time' in filter_dict:
                sort_query = 'order_time'
            elif 'callback_time' in filter_dict:
                sort_query = 'callback_time'
            datas = behalfPayOrderTable.find_many(filter_dict, sort=[[sort_query, -1]])
            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            _out_data_dict = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas),
                'out_count': 0,
                'note': '代付订单-' + datetime.datetime.now().strftime('%Y%m%d%H%M%S'),
                'merchant_uuid': self.current_user_dict.get('uuid'),
                'agentadmin_uuid': self.current_user_dict.get('agentadmin_uuid'),
            }
            uuid = ExportDataModel.insert_one(_out_data_dict)
            threading.Thread(target=self.exportData, args=(datas, uuid, export_folder, filename)).start()
            return self.xtjson.json_result(message='数据导出中，请稍后到"文件下载"中查看数据！')
        if self.action == 'get_total_info':
            if self.is_merchant:
                merchant_id = self.current_user_dict.get('merchant_id')
            elif self.is_submerchant:
                merchant_id = self.current_user_dict.get('upper_data').get('merchant_id')
            else:
                return abort(404)
            filter_dict = {'merchant_id': merchant_id}
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if not statu:
                return res
            filter_dict.update(res[0])

            # 总订单金额
            _order_amount_total = behalfPayOrderTable.collection().aggregate([
                {"$match": filter_dict},
                {"$group": {"_id": None, "order_amount": {"$sum": '$order_amount'}}},
            ])
            order_amount_total_ll = list(_order_amount_total)
            order_amount_tota = 0
            if order_amount_total_ll:
                order_amount_tota = order_amount_total_ll[0].get('order_amount')

            # 总支付金额
            _actual_amount_total = behalfPayOrderTable.collection().aggregate([
                {"$match": filter_dict},
                {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
            ])
            actual_amount_total_ll = list(_actual_amount_total)
            actual_amount_tota = 0
            if actual_amount_total_ll:
                actual_amount_tota = actual_amount_total_ll[0].get('actual_amount')

            # 总支付金额
            _repay_amount_total = behalfPayOrderTable.collection().aggregate([
                {"$match": filter_dict},
                {"$group": {"_id": None, "repay_amount": {"$sum": '$repay_amount'}}},
            ])
            repay_amount_total_ll = list(_repay_amount_total)
            repay_amount_tota = 0
            if repay_amount_total_ll:
                repay_amount_tota = repay_amount_total_ll[0].get('repay_amount')

            _data = {
                'order_amount_tota': self.format_money(order_amount_tota),
                'actual_amount_tota': self.format_money(actual_amount_tota),
                'repay_amount_tota': self.format_money(repay_amount_tota),
            }
            return self.xtjson.json_result(data=_data)



class MerchantBillStatementView(TableViewBase):
    add_url_rules = [['/billStatement', 'merchantBillStatement']]
    title = '资金流水'
    MCLS = MerchantBillStatementTable
    template = 'merchant/billStatement.html'
    per_page = 30

    def exportData(self, datas, log_uuid, export_folder, filename):
        ''' 导出数据 '''
        export_data = ExportDataModel.find_one({'uuid': log_uuid})
        if not export_data:
            return
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1

            header = ['商户名称', '账单类型', '订单号', '商户订单号', '交易金额', '手续费', '当前余额', '时间']
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])
            mdaict = {}
            for data in datas:
                row += 1
                _md = mdaict.get(data.get('merchant_uuid'))
                if not _md:
                    _md = MerchantTable.find_one({'uuid': data.get('merchant_uuid')})
                    mdaict[data.get('merchant_uuid')] = _md

                wa.cell(row=row, column=1, value=str(_md.get('merchant_name') or ''))
                wa.cell(row=row, column=2, value=BILL_STATEMEN_TYPES.name_dict.get(data.get('bill_type')) )
                wa.cell(row=row, column=3, value=str(data.get('order_id') or ''))
                wa.cell(row=row, column=4, value=str(data.get('merchant_order_id') or ''))
                wa.cell(row=row, column=5, value=str(round(data.get('amount') or 0, 2)))
                wa.cell(row=row, column=6, value=str(round(data.get('repay_amount') or 0, 2)))
                wa.cell(row=row, column=7, value=str(round(data.get('balance_amount') or 0, 2)))
                wa.cell(row=row, column=8, value=str(self.format_time_func(data.get('create_time'))))

                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def get_filter_dict(self):
        if self.is_merchant:
            merchant_uuid = self.current_user_dict.get('uuid')
        elif self.is_submerchant:
            merchant_uuid = self.current_user_dict.get('upper_data').get('uuid')
        else:
            return abort(404)
        return {'merchant_uuid': merchant_uuid}

    def round_func(self, data):
        try:
            return round(data, 2)
        except:
            return data

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        datas = []
        for da in all_datas:
            merchant_data = MerchantTable.find_one({'uuid': da.get('merchant_uuid')})
            da['merchant_data'] = merchant_data
            datas.append(da)
        self.context['all_datas'] = datas
        self.context['round_func'] = self.round_func
        self.context['BILL_STATEMEN_TYPES'] = BILL_STATEMEN_TYPES

    def post_other_way(self):
        if self.action == 'export_order':
            if self.is_merchant:
                merchant_uuid = self.current_user_dict.get('uuid')
            elif self.is_submerchant:
                merchant_uuid = self.current_user_dict.get('upper_data').get('uuid')
            else:
                return abort(404)
            filter_dict = {'merchant_uuid': merchant_uuid}
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if not statu:
                return res
            filter_dict.update(res[0])
            datas = self.MCLS.find_many(filter_dict)
            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            _out_data_dict = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas),
                'out_count': 0,
                'note': '资金流水-' + datetime.datetime.now().strftime('%Y%m%d%H%M%S'),
                'merchant_uuid': self.current_user_dict.get('uuid'),
                'agentadmin_uuid': self.current_user_dict.get('agentadmin_uuid'),
            }
            uuid = ExportDataModel.insert_one(_out_data_dict)
            threading.Thread(target=self.exportData, args=(datas, uuid, export_folder, filename)).start()
            return self.xtjson.json_result(message='数据导出中，请稍后到"文件下载"中查看数据！')



class WithdrawApplyView(TableViewBase):
    add_url_rules = [['/withdrawApply', 'WithdrawApply']]
    title = '结算申请'
    MCLS = WithdrawTable
    template = 'merchant/WithdrawApply.html'
    per_page = 30

    def get_context(self):
        return {'WITHDRAW_STATUS': WITHDRAW_STATUS}

    def get_filter_dict(self):
        if self.is_merchant:
            merchant_uuid = self.current_user_dict.get('uuid')
        elif self.is_submerchant:
            merchant_uuid = self.current_user_dict.get('upper_data').get('uuid')
        else:
            return abort(404)
        return {'merchant_uuid': merchant_uuid}

    def dealwith_main_context(self):
        datas = []
        all_datas = self.context.get('all_datas') or []
        for da in all_datas:
            _dd = MerchantBankCardTable.find_one({'uuid': da.get('bankcard_uuid')})
            da['bankcard_account'] = _dd.get('account') or ''
            datas.append(da)
            bank_data = BankTable.find_one({'code': da.get('payee_bank')})
            da['bank_data'] = bank_data
            datas.append(da)

        self.context['all_datas'] = all_datas

    def post_other_way(self):
        # 统计
        if self.action == 'get_total_info':
            filter_dict = self.get_filter_dict()
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if statu:
                filter_dict.update(res[0])

            # 下发成功总笔数
            fff1 = {}
            fff1.update(filter_dict)
            fff1['statu'] = WITHDRAW_STATUS.success
            number_count = self.MCLS.count(fff1) or 0

            # 下发总金额
            amount_total_ll = self.MCLS.collection().aggregate([
                {"$match": fff1},
                {"$group": {"_id": None, "amount": {"$sum": '$amount'}}},
            ])
            amount_total_l = list(amount_total_ll)
            amount_total = 0
            if amount_total_l:
                amount_total = amount_total_l[0].get('amount')

            _data = {
                'number_count': number_count,
                'amount_total': self.format_money(amount_total),
            }
            return self.xtjson.json_result(data=_data)



class ExportDownload(FormViewBase):
    title = '数据下载'
    MCLS = ExportDataModel
    add_url_rules = [['/exportDownload', 'exportDownload']]
    template = 'merchant/exportDownload.html'
    per_page = 20
    sort = [['create_time', -1]]

    def format_datetime(self, data):
        if isinstance(data, datetime.datetime):
            return data.strftime('%Y-%m-%d %H:%M:%S')
        return data

    def view_get(self):
        self.context['title'] = self.title
        self.context['format_datetime'] = self.format_datetime
        page = request.args.get('page', 1, int)
        skip = (page - 1) * self.per_page
        self.context['title'] = self.title
        filter_dict, context_res = {}, {}
        fields = self.MCLS.fields()
        self.context['FIELDS'] = fields
        statu, res = self.search_func(fields)
        if not statu:
            return res
        filter_dict.update(res[0])
        context_res.update(res[1])
        self.context.update(self.get_context())
        filter_dict.update(self.get_filter_dict())

        if self.is_merchant:
            merchant_uuid = self.current_user_dict.get('uuid')
        elif self.is_submerchant:
            merchant_uuid = self.current_user_dict.get('upper_data').get('uuid')
        else:
            return abort(404)
        filter_dict['merchant_uuid'] = merchant_uuid
        total = self.MCLS.count(filter_dict)
        all_datas = self.MCLS.find_many(filter_dict, limit=self.per_page, skip=skip, sort=self.sort)

        pagination = PagingCLS.pagination(page, self.per_page, total)
        self.context['total'] = total
        self.context['all_datas'] = all_datas
        self.context['pagination'] = pagination
        self.context['search_res'] = context_res
        return render_template(self.template, **self.context)

    def post_other_way(self):
        if self.action == 'del_all':
            self.MCLS.delete_many({})
            return self.xtjson.json_result()

    def post_data_other_way(self):
        if self.action == 'del':
            filePath = current_app.static_folder + '/' + current_app.config.get('PROJECT_NAME') + self.data_dict.get('path')
            os.remove(filePath)
            self.MCLS.delete_one({'uuid': self.data_uuid})
            return self.xtjson.json_result(message='数据删除成功！')



class MerchantReportView(FormViewBase):
    add_url_rules = [['/merchantReport', 'merchantReport']]
    title = '商户报表'
    template = 'merchant/merchantReport.html'

    def get_filter_dict(self):
        return {'uuid': self.current_user_dict.get('uuid')}

    def is_xhr(self):
        X_Requested_With = request.headers.get('X-Requested-With')

        if not X_Requested_With or X_Requested_With.lower() != 'xmlhttprequest':
            return
        return True

    def search_merchant_func(self):
        dataDate = request.args.get('dataDate')
        search_res = {}

        if dataDate and dataDate.strip():
            start_time, end_time = PagingCLS.by_silce(dataDate)
        else:
            crrdate = datetime.datetime.now()
            start_time, end_time = datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 0, 0,
                                                     0), datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 23,
                                                                           59, 59)
            dataDate = start_time.strftime('%Y-%m-%d %H:%M:%S') + '|' + end_time.strftime('%Y-%m-%d %H:%M:%S')

        search_res['dataDate'] = dataDate

        print(self.get_filter_dict())

        merchant_data = MerchantTable.find_one(self.get_filter_dict())

        mch_report_data = {
            'merchant_id': merchant_data.get('merchant_id') or '',
            'merchant_name': merchant_data.get('merchant_name') or '',
        }

        # 代收金额
        fff1 = {
            'callback_time': {'$gte': start_time, '$lte': end_time},
            'pay_statu': True,
            'callback_statu': CallbackState.SUCCESS,
            'merchant_id': merchant_data.get('merchant_id'),
        }
        ds_aomunt_total_ll = CollectionOrderTable.collection().aggregate([
            {"$match": fff1},
            {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
        ])
        ds_aomunt_total = 0
        ds_aomunt_total_l = list(ds_aomunt_total_ll)
        if ds_aomunt_total_l:
            ds_aomunt_total = ds_aomunt_total_l[0].get('actual_amount')
        mch_report_data['ds_aomunt_total'] = self.format_money(ds_aomunt_total)

        # 代收手续费
        fff2 = {
            'callback_time': {'$gte': start_time, '$lte': end_time},
            'pay_statu': True,
            'merchant_id': merchant_data.get('merchant_id'),
            'callback_statu': CallbackState.SUCCESS,
        }
        dssuf_aomunt_total_ll = CollectionOrderTable.collection().aggregate([
            {"$match": fff2},
            {"$group": {"_id": None, "repay_amount": {"$sum": '$repay_amount'}}},
        ])
        dssxf_aomunt_total = 0
        dssuf_aomunt_total_l = list(dssuf_aomunt_total_ll)
        if dssuf_aomunt_total_l:
            dssxf_aomunt_total = dssuf_aomunt_total_l[0].get('repay_amount')
        mch_report_data['dssxf_aomunt_total'] = self.format_money(dssxf_aomunt_total)

        # 代付金额
        fff3 = {
            'order_time': {'$gte': start_time, '$lte': end_time},
            'pay_statu': True,
            'merchant_id': merchant_data.get('merchant_id'),
        }
        df_aomunt_total_ll = behalfPayOrderTable.collection().aggregate([
            {"$match": fff3},
            {"$group": {"_id": None, "order_amount": {"$sum": '$order_amount'}}},
        ])
        df_aomunt_total = 0
        df_aomunt_total_l = list(df_aomunt_total_ll)
        if df_aomunt_total_l:
            df_aomunt_total = df_aomunt_total_l[0].get('order_amount')
        mch_report_data['df_aomunt_total'] = self.format_money(df_aomunt_total)

        # 代付手续费
        fff4 = {
            'order_time': {'$gte': start_time, '$lte': end_time},
            'pay_statu': True,
            'merchant_id': merchant_data.get('merchant_id'),
        }
        dfsxf_aomunt_total_ll = behalfPayOrderTable.collection().aggregate([
            {"$match": fff4},
            {"$group": {"_id": None, "repay_amount": {"$sum": '$repay_amount'}}},
        ])
        dfsxf_aomunt_total = 0
        dfsxf_aomunt_total_l = list(dfsxf_aomunt_total_ll)
        if dfsxf_aomunt_total_l:
            dfsxf_aomunt_total = dfsxf_aomunt_total_l[0].get('repay_amount')
        mch_report_data['dfsxf_aomunt_total'] = self.format_money(dfsxf_aomunt_total)

        # 手动下发
        fff5 = {
            'create_time': {'$gte': start_time, '$lte': end_time},
            'statu': WITHDRAW_STATUS.success,
            'merchant_uuid': merchant_data.get('uuid'),
        }
        sdxf_aomunt_total_ll = WithdrawTable.collection().aggregate([
            {"$match": fff5},
            {"$group": {"_id": None, "amount": {"$sum": '$amount'}}},
        ])
        sdxf_aomunt_total = 0
        sdxf_aomunt_total_l = list(sdxf_aomunt_total_ll)
        if sdxf_aomunt_total_l:
            sdxf_aomunt_total = sdxf_aomunt_total_l[0].get('amount')
        mch_report_data['sdxf_aomunt_total'] = self.format_money(sdxf_aomunt_total)

        # 手动下发手续费
        fff6 = {
            'create_time': {'$gte': start_time, '$lte': end_time},
            'statu': WITHDRAW_STATUS.success,
            'merchant_uuid': merchant_data.get('uuid'),
        }
        sdxfsxf_aomunt_total_ll = WithdrawTable.collection().aggregate([
            {"$match": fff6},
            {"$group": {"_id": None, "repay_amount": {"$sum": '$repay_amount'}}},
        ])
        sdxfsxf_aomunt_total = 0
        sdxfsxf_aomunt_total_l = list(sdxfsxf_aomunt_total_ll)
        if sdxfsxf_aomunt_total_l:
            sdxfsxf_aomunt_total = sdxfsxf_aomunt_total_l[0].get('repay_amount')
        mch_report_data['sdxfsxf_aomunt_total'] = self.format_money(sdxfsxf_aomunt_total)

        # 内部费用金额 内部收费
        fff7 = {
            'create_time': {'$gte': start_time, '$lte': end_time},
            'merchant_id': merchant_data.get('merchant_id'),
        }
        recharges = RechargeMoneyTable.find_many(fff7)
        internal_charge_amount = 0
        internal_fee_amount = 0
        for recharge in recharges:
            internal_charge_amount += recharge.get('amount')
            internal_fee_amount += recharge.get('repay_amount')

        mch_report_data['internal_charge_amount'] = self.format_money(internal_charge_amount)
        mch_report_data['internal_fee_amount'] = self.format_money(internal_fee_amount)

        end_mb_data = MerchantBillStatementTable.find_one(
            {'merchant_uuid': merchant_data.get('uuid'), 'create_time': {'$gte': start_time, '$lte': end_time}},
            sort=[['create_time', -1]]) or {}
        if not end_mb_data:
            end_mb_data = MerchantBillStatementTable.find_one(
                {'merchant_uuid': merchant_data.get('uuid'), 'create_time': {'$lte': start_time}},
                sort=[['create_time', -1]]) or {}
        if end_mb_data:
            mch_report_data['end_balance_amount'] = self.format_money(round(end_mb_data.get('balance_amount', 2)) or 0)
        else:
            mch_report_data['end_balance_amount'] = 0

        start_mb_data = MerchantBillStatementTable.find_one(
            {'merchant_uuid': merchant_data.get('uuid'), 'create_time': {'$lt': start_time}},
            sort=[['create_time', -1]]) or {}
        if start_mb_data:
            mch_report_data['start_balance_amount'] = self.format_money(
                round(start_mb_data.get('balance_amount'), 2) or 0)
        else:
            mch_report_data['start_balance_amount'] = 0


        end_balance_amount = float(str(mch_report_data.get('end_balance_amount')).replace(',', ''))
        start_balance_amount = float(str(mch_report_data.get('start_balance_amount')).replace(',', ''))
        wc_vv = start_balance_amount + float(mch_report_data.get('ds_aomunt_total').replace(',', '')) - float(
            mch_report_data.get('dssxf_aomunt_total').replace(',', '')) - float(
            mch_report_data.get('df_aomunt_total').replace(',', '')) - float(
            mch_report_data.get('dfsxf_aomunt_total').replace(',', '')) - float(
            mch_report_data.get('sdxf_aomunt_total').replace(',', '')) + float(
            mch_report_data.get('sdxfsxf_aomunt_total').replace(',', '')) + float(
            mch_report_data.get('internal_charge_amount').replace(',', '')) - float(
            mch_report_data.get('internal_fee_amount').replace(',', ''))

        wc_vvv = round(end_balance_amount - (wc_vv or 0), 2)
        if float(wc_vvv) == 0:
            wc_vvv = 0
        mch_report_data['wc_vv'] = self.format_money(wc_vvv or 0)

        return self.xtjson.json_result(data={'mch_report_data': mch_report_data, 'search_res': search_res})

    def view_get(self):
        if self.is_xhr():
            return self.search_merchant_func()
        search_res = {}
        crrdate = datetime.datetime.now()
        start_time, end_time = datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 0, 0, 0), datetime.datetime(
            crrdate.year, crrdate.month, crrdate.day, 23, 59, 59)
        dataDate = start_time.strftime('%Y-%m-%d %H:%M:%S') + '|' + end_time.strftime('%Y-%m-%d %H:%M:%S')
        search_res['dataDate'] = dataDate
        self.context['title'] = self.title
        self.context['search_res'] = search_res
        self.context['BILL_STATEMEN_TYPES'] = BILL_STATEMEN_TYPES
        return render_template(self.template, **self.context)



class SubAccountListView(TableViewBase):
    add_url_rules = [['/subAccountList', 'SubAccountListView']]
    title = '子账户列表'
    MCLS = MerchantTable
    template = 'merchant/account_list.html'
    per_page = 30

    def submerchant_html(self, data_dict={}):
        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">账户名称：</span>
                        <input type="text" class="form-control" id="account_name" value="{ data_dict.get('account_name') or '' }" placeholder="商户名称" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">账户：</span>
                        <input type="text" class="form-control" id="account" value="{ data_dict.get('account') or '' }" placeholder="账户" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span { '' if data_dict else 'class="loglable"' } style="width: 120px; text-align: right; display: inline-block; position: relative;">密码：</span>
                        <input type="text" class="form-control" id="password" value="" placeholder="密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span { '' if data_dict else 'class="loglable"' } style="width: 120px; text-align: right; display: inline-block; position: relative;">确认密码：</span>
                        <input type="text" class="form-control" id="confirm_password" value="{ data_dict.get('confirm_password') or '' }" placeholder="密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">备注：</span>
                        <input type="text" class="form-control" id="note" value="{ data_dict.get('note') or '' }" placeholder="备注" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                                    
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_sub_data('{ 'edit_submerchant_data' if data_dict else 'add_submerchant_data' }','{ self.data_uuid }')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>
        '''
        return self.xtjson.json_result(message=html)

    def get_filter_dict(self):
        return {'upper_mid': self.current_user_dict.get('uuid'), 'role_code': MERCHANT_ROLES.SUBMERCHANT}

    def post_other_way(self):
        if self.action == 'add_submerchant_html':
            return self.submerchant_html()
        if self.action == 'add_submerchant_data':
            account = self.request_data.get('account')
            account_name = self.request_data.get('account_name')
            password = self.request_data.get('password')
            note = self.request_data.get('note') or ''
            if not account or not account_name or not password:
                return self.xtjson.json_params_error('缺少数据！')
            if self.MCLS.find_one({'account':account.strip(), 'upper_mid': self.current_user_dict.get('uuid')}):
                return self.xtjson.json_params_error('该账户已存在！')
            if self.MCLS.find_one({'account_name':account_name.strip(), 'upper_mid': self.current_user_dict.get('uuid')}):
                return self.xtjson.json_params_error('该账户名称已存在！')
            if not password.isalnum():
                return self.xtjson.json_params_error('密码格式错误！')
            if len(password.strip()) < 5 or len(password.strip()) > 15:
                return self.xtjson.json_params_error('密码长度在5~15位之间！')
            _data = {
                'account': account,
                'account_name': account_name,
                'password': self.MCLS.encry_password(password),
                'note': note.strip(),
                'role_code': MERCHANT_ROLES.SUBMERCHANT,
                'upper_mid': self.current_user_dict.get('uuid'),
                'statu': True
            }
            self.MCLS.insert_one(_data)
            return self.xtjson.json_result()

    def post_data_other_way(self):
        if self.action == 'edit_submerchant_html':
            return self.submerchant_html(self.data_dict)
        if self.action == 'getGoogleQrcode':
            if not self.data_uuid:
                return self.xtjson.json_params_error()

            user_dict = self.MCLS.find_one({'uuid': self.data_uuid})
            if not user_dict:
                return self.xtjson.json_params_error()

            google_cls = GooleVerifyCls(pwd=self.data_uuid, account=user_dict.get('account'), s_label='pay2wold')
            generate_qrcode = google_cls.secret_generate_qrcode()
            return self.xtjson.json_result(data={'generate_qrcode': generate_qrcode})
        if self.action == 'del':
            self.MCLS.delete_one({'uuid': self.data_uuid})
            return self.xtjson.json_result()
        if self.action == 'edit_submerchant_data':
            account = self.request_data.get('account')
            account_name = self.request_data.get('account_name')
            password = self.request_data.get('password')
            note = self.request_data.get('note') or ''
            if not account or not account_name:
                return self.xtjson.json_params_error('缺少数据！')
            if password:
                if not password.isalnum():
                    return self.xtjson.json_params_error('密码格式错误！')
                if len(password.strip()) < 5 or len(password.strip()) > 15:
                    return self.xtjson.json_params_error('密码长度在5~15位之间！')
            _dd = self.MCLS.find_one({'account': account, 'upper_mid': self.current_user_dict.get('uuid')})
            if _dd and _dd.get('uuid') != self.data_uuid:
                return self.xtjson.json_params_error('该账户已存在！')
            _dd = self.MCLS.find_one({'account_name': account_name, 'upper_mid': self.current_user_dict.get('uuid')})
            if _dd and _dd.get('uuid') != self.data_uuid:
                return self.xtjson.json_params_error('该账户名已存在！')
            _data = {
                'account': account,
                'account_name': account_name,
                'note': note.strip(),
            }
            if password:
                _data['password'] = self.MCLS.encry_password(password)
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': _data})
            return self.xtjson.json_result()
        if self.action == 'update_statu':
            if self.data_dict.get('statu'):
                self.data_from['statu'] = False
            else:
                self.data_from['statu'] = True
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            return self.xtjson.json_result()
