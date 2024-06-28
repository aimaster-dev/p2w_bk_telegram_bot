'''
代付
'''
import os
import random
import time
import shortuuid
import json
import requests
import datetime
import threading
from flask import request, render_template, current_app, abort, url_for, session
from .cms_base import CmsTableViewBase, CmsFormViewBase
from models.behalfPay import behalfPayOrderTable, behalfPayCallbackLogTable, behalfPayScriptTable, behalfPayTaskTable, behalfPayOrderProcessTable, CnBankCardTable
from models.pay_table import MerchantTable,  BankCardTable,  BankTable, MerchantBillStatementTable, WithdrawTable, MerchantBankCardTable, WithdrawalCardTable, WithdrawalOrderLogTable
from models.cms_user import CmsUserTable

from constants import CallbackState, CallbankType, ROlE_ALL, taskStatus, ASSETS_FOLDER, EXPORT_FOLDER, ExportStatu,TUNNLE_METHOD, PAY_METHOD, WITHDRAW_STATUS, CallbackState, CallbankType, BILL_STATEMEN_TYPES, ROlE_ALL, ExportStatu, ASSETS_FOLDER, EXPORT_FOLDER, LOCATION_TYPE, ORDER_STATUS, ACCEPTED_STATUS, BEHAVIOR_TYPE, SELECET_COLS_SESSION_KEY, PayStatus
from models.pay_table import MerchantTable, BankCardTable, BankTable
from common_utils.utils_funcs import PagingCLS, encry_md5, is_valid_url, update_language
from modules.view_helpres.view_func import behalfPayCallbackOrderFunc, payBehalf_goback, getBehalfPayOrderId, payBehalf_deduct, getBankPayQrcode,getWithdrawalBankPayQrcode
from models.cms_user import CmsUserTable
from models.cms_table import SiteConfigTable
from models.site_table import ExportDataModel
from openpyxl import Workbook
from common_utils.lqredis import mc



class behalfPayOrderView(CmsTableViewBase):
    add_url_rules = [['/behalfPay', 'behalfPay']]
    per_page = 30
    MCLS = behalfPayOrderTable
    template = 'cms/behalfPay/behalfPay_list.html'
    title = '代付订单'

    def exportData(self, datas, log_uuid, export_folder, filename):
        ''' 导出数据 '''
        export_data = ExportDataModel.find_one({'uuid': log_uuid})
        if not export_data:
            return
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1
            header = ['订单号', '商户名称', '商户订单号', '收款银行', '收款账号', '收款人', '订单金额', '手续费',  '订单时间', '支付状态', '回调状态', '回调类型', '回调时间', '出款人员']
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])

            merchant_data_dict = {}
            uuuds = {}
            for data in datas:
                row += 1

                merchant_id = data.get('merchant_id') or ''
                merchant_data = merchant_data_dict.get(merchant_id)
                if not merchant_data:
                    merchant_data = MerchantTable.find_one({'merchant_id': merchant_id}) or {}
                    merchant_data_dict[merchant_id] = merchant_data

                out_money_userid = data.get('out_money_userid') or ''
                uuud_data = uuuds.get(out_money_userid)
                if not uuud_data:
                    uuud_data = CmsUserTable.find_one({'uuid': out_money_userid}) or {}
                    uuuds[out_money_userid] = uuud_data

                wa.cell(row=row, column=1, value=str(data.get('order_id') or ''))
                wa.cell(row=row, column=2, value=str(merchant_data.get('merchant_name') or ''))
                wa.cell(row=row, column=3, value=str(data.get('merchant_order_id') or ''))
                wa.cell(row=row, column=4, value=str(data.get('receive_bank_code') or ''))
                wa.cell(row=row, column=5, value=str(data.get('receive_account') or ''))
                wa.cell(row=row, column=6, value=str(data.get('receive_owner') or ''))
                wa.cell(row=row, column=7, value=self.format_money(str(data.get('order_amount') or '0')))
                wa.cell(row=row, column=8, value=self.format_money(str(data.get('repay_amount') or '0')))
                wa.cell(row=row, column=9, value=data.get('order_time').strftime('%Y-%m-%d %H:%M:%S'))
                statu_text = '未支付'
                if data.get('pay_statu'):
                    statu_text = '已支付'
                if data.get('reject_pay'):
                    statu_text = '拒绝支付'
                wa.cell(row=row, column=10, value=statu_text)
                if data.get('callback_statu') == CallbackState.SUCCESS:
                    call_text = '回调成功'
                elif data.get('callback_statu') == CallbackState.FAILED:
                    call_text = '回调失败'
                else:
                    call_text = '未回调'
                wa.cell(row=row, column=11, value=call_text)
                wa.cell(row=row, column=12, value=CallbankType.name_dict.get(data.get('callbank_type')) or '' if data.get('callback_statu') == CallbackState.SUCCESS else '')
                wa.cell(row=row, column=13, value=data.get('callback_time').strftime('%Y-%m-%d %H:%M:%S') if data.get('callback_statu') == CallbackState.SUCCESS else '')
                wa.cell(row=row, column=14, value=uuud_data.get('account') or '')

                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def fff_filter_dict(self):
        dd = {}
        fields = self.MCLS.fields()
        statu, res = self.search_func(fields)
        if statu:
            dd.update(res[0])
            self.search_dict.update(res[1])

        sort_type = request.args.get('sort_type')
        if sort_type and sort_type.strip():
            if len(sort_type.rsplit('_', 1)) == 2:
                filed, vv = sort_type.rsplit('_', 1)
                if vv == '0':
                    self.sort = [[filed, -1]]
                if vv == '1':
                    self.sort = [[filed, 1]]
            self.search_dict['sort_type'] = sort_type
        _order_time, start_time, end_time = self.get_reqorder()
        order_time = {'$gte': start_time, '$lte': end_time}

        if 'pay_time' not in dd and not request.args.get('order_time'):
            dd['order_time'] = order_time
            self.search_dict['order_time'] = _order_time
        if request.args.get('order_time'):
            dd['order_time'] = order_time
            self.search_dict['order_time'] = _order_time

        if request.args.get('pay_status'):
            pay_status = request.args.get('pay_status')
            if pay_status == '0':
                dd['pay_statu'] = False
                dd['reject_pay'] = {'$ne': True}
            elif pay_status == '1':
                dd['pay_statu'] = True
            elif pay_status == '2':
                dd['reject_pay'] = True
            self.search_dict['pay_status'] = pay_status

        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            dd['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            dd['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')

        searchType = request.args.get('searchType')
        searchValue = request.args.get('searchValue')
        if searchType and searchValue and searchValue.strip() and searchType.strip():
            if searchType == 'bankcard_account':
                bankcard_data = BankCardTable.find_one({'account': searchValue.strip()})
                if bankcard_data:
                    dd['bankcard_id'] = bankcard_data.get('uuid')
                else:
                    dd['bankcard_id'] = searchValue.strip()
            elif searchType == 'merchant_name':
                merchant_data = MerchantTable.find_one({'merchant_name': searchValue}) or {}
                dd['merchant_id'] = merchant_data.get('merchant_id')
                self.search_dict['merchant_name'] = searchValue
            elif searchType == 'agentadmin_account':
                agentadmin_data = CmsUserTable.find_one({'account': searchValue}) or {}
                dd['agentadmin_uuid'] = agentadmin_data.get('uuid')
                self.search_dict['agentadmin_account'] = searchValue
            elif searchType == 'out_money_userid':
                processor_data = CmsUserTable.find_one({'account': searchValue}) or {}
                dd['out_money_userid'] = processor_data.get('uuid')
                self.search_dict['out_money_userid'] = searchValue
            else:
                dd_fields = self.MCLS.fields()
                if hasattr(self.MCLS, 'field_search'):
                    field_search = getattr(self.MCLS, 'field_search')()
                    for db_field in field_search:
                        if db_field == searchType:
                            col_value = searchValue.strip().replace(',','')
                            field_cls = dd_fields.get(db_field)
                            if not field_cls:
                                break
                            statu, res = field_cls.search_validate(col_value)
                            if statu:
                                dd[db_field] = res
                            break
        if searchType:
            self.search_dict['searchType'] = searchType
        if searchValue:
            self.search_dict['searchValue'] = searchValue
        if self.current_admin_dict.get('role_code') in [ROlE_ALL.OUT_MONEY_USER, ROlE_ALL.SYS_OUT_MONEY_USER]:
            dd['out_money_userid'] = self.current_admin_dict.get('uuid')

        sys_payorder = request.args.get('sys_payorder')
        if sys_payorder == '1':
            dd['sys_payorder'] = True
            self.search_dict['sys_payorder'] = '1'
        elif sys_payorder == '0':
            dd['sys_payorder'] = False
            self.search_dict['sys_payorder'] = '0'
        else:
            self.search_dict['sys_payorder'] = sys_payorder or ''

        # is_search= False
        # if 'sort_type' in str(request.url):
        #     is_search = True
        self.search_dict['is_search'] = True
        return dd

    def getProxy(self, is_proxy_str):
        ip, port, username, password = is_proxy_str.replace(' ','').split(':')
        proxies = {
            "http": f"http://{username}:{password}@{ip}:{port}",
            "https": f"http://{username}:{password}@{ip}:{port}",
        }
        return proxies

    def detailsInfo_html(self):
        datas = behalfPayOrderProcessTable.find_many({'order_id': self.data_dict.get('order_id')}) or []
        if not datas:
            return self.xtjson.json_params_error('该订单没有流程信息！')
        html = '''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto; text-align: left; overflow-x: hidden; box-sizing: border-box;">               
                    <ul class="layui-timeline">        
        '''
        html += f'''
            <li class="layui-timeline-item">
                <i class="layui-icon layui-timeline-axis"></i>
                <div class="layui-timeline-content layui-text">
                    <div class="layui-timeline-title">{self.format_time_func(datas[0].get('create_time'))}，创建订单</div>
                </div>
            </li>
        '''
        if len(datas) == 1:
            html += '''
                        </ul>        
                    </div>
                </div>        
            '''
            return self.xtjson.json_result(message=html)

        for data in datas[1:]:
            html += f'''
                <li class="layui-timeline-item">
                    <i class="layui-icon layui-timeline-axis"></i>
                    <div class="layui-timeline-content layui-text">
                        <div class="layui-timeline-title">{ self.format_time_func(data.get('create_time')) }，{data.get('text') or ''}</div>
                    </div>
                </li>            
            '''

        if self.data_dict.get('note'):
            html += f'''
                <li class="layui-timeline-item">
                    <i class="layui-icon layui-timeline-axis"></i>
                    <div class="layui-timeline-content layui-text">
                        <div class="layui-timeline-title">备注：{self.data_dict.get('note') or ''}</div>
                    </div>
                </li>            
            '''


        html += '''
                    </ul>        
                </div>
            </div>        
        '''
        return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))

    def refreshOrderList(self):
        sort_type = request.args.get('sort_type')
        if sort_type and sort_type.strip():
            if len(sort_type.rsplit('_', 1)) == 2:
                filed, vv = sort_type.rsplit('_', 1)
                if vv == '0':
                    self.sort = [[filed, -1]]
                if vv == '1':
                    self.sort = [[filed, 1]]

        filter_dict = self.fff_filter_dict()
        page = request.args.get('page', 1, int)
        skip = (page - 1) * self.per_page
        all_datas = self.MCLS.find_many(filter_dict, limit=self.per_page, skip=skip, sort=self.sort)
        datas = []
        oumudd = {}
        for data in all_datas:
            merchant_data = MerchantTable.find_one({'merchant_id': data.get('merchant_id')}) or {}
            bank_data = BankTable.find_one({'code': data.get('receive_bank_code')}) or {}
            out_money_userid = data.get('out_money_userid')
            if out_money_userid:
                out_money_user_data = oumudd.get(out_money_userid) or ''
                if not out_money_user_data:
                    out_money_user_data = CmsUserTable.find_one({'uuid': out_money_userid}) or {}
                    oumudd[out_money_userid] = out_money_user_data
            else:
                out_money_user_data = {}
            _dd = {
                'data_uuid': data.get('uuid'),
                'order_id': data.get('order_id'),
                'merchant_name': merchant_data.get('merchant_name'),
                'merchant_order_id': data.get('merchant_order_id'),
                'shortName': bank_data.get('shortName'),
                'receive_account': data.get('receive_account'),
                'receive_owner': data.get('receive_owner'),
                'callback_statu': data.get('callback_statu'),
                'order_amount': self.format_money( data.get('order_amount') or 0),
                'repay_amount': self.format_money( data.get('repay_amount') or 0),
                'actual_amount': self.format_money( data.get('actual_amount') or 0),
                'order_time': self.format_time_func(data.get('order_time'), '%H:%M:%S'),
                'pay_time': self.format_time_func(data.get('pay_time'), '%H:%M:%S') or '',
                'callback_time': self.format_time_func(data.get('callback_time'), '%H:%M:%S') or '',
                'pay_statu': data.get('pay_statu') or False,
                'reject_pay': data.get('reject_pay') or False,
                'check_aname_state': data.get('check_aname_state'),
                'callbank_type': CallbankType.name_dict.get(data.get('callbank_type')) or '',
            }
            if self.current_admin_dict.get('role_code') in [ROlE_ALL.SUPERADMIN, ROlE_ALL.ADMINISTRATOR, ROlE_ALL.ADMINISTRATOR]:
                _dd['out_money_userid'] = out_money_user_data.get('account') or ''
                _dd['is_out_money_userid'] = True
            else:
                _dd['is_out_money_userid'] = False
            datas.append(_dd)
        return self.xtjson.json_result(data={'datas': datas})

    def get_context(self):
        return {'CallbackState': CallbackState}

    def get_filter_dict(self):
        ddd = self.fff_filter_dict()
        return ddd

    def get_reqorder(self):
        order_time = request.args.get('order_time')
        if order_time and order_time.strip():
            start_time, end_time = PagingCLS.by_silce(order_time)
        else:
            crrdate = datetime.datetime.now()
            start_time, end_time = datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 0, 0,0), datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 23,59, 59)
            order_time = start_time.strftime('%Y-%m-%d %H:%M:%S') + '|' + end_time.strftime('%Y-%m-%d %H:%M:%S')
        return order_time, start_time, end_time
    
    def get_testOrder_html(self):
        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">       
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">商户Id：</span>
                        <input type="text" class="form-control" id="test_mchId" placeholder="商户Id" value="1665526" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">商户订单号：</span>
                        <input type="text" class="form-control" id="test_mchOrderId" value="R{ str(int(time.time()*1000)) }" placeholder="商户订单号" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行Code：</span>
                        <input type="text" class="form-control" id="test_bankCode" value="" placeholder="商户订单号" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">收款银行卡号：</span>
                        <input type="text" class="form-control" id="test_bankAccount" value="" placeholder="收款银行卡号" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">收款人姓名：</span>
                        <input type="text" class="form-control" id="test_bankOwner" value="" placeholder="收款人姓名" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">金额：</span>
                        <input type="number" class="form-control" id="test_amount" value="" placeholder="金额" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">回调地址：</span>
                        <input type="text" class="form-control" id="test_notifyUrl" value="https://www.google.com" placeholder="回调地址" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                           
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_createTestOrder()">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>           
        '''
        return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))

    def createOrder_func(self):
        mchId = self.request_data.get('mchId')
        mchOrderId = self.request_data.get('mchOrderId')
        bankCode = self.request_data.get('bankCode')
        bankAccount = self.request_data.get('bankAccount')
        bankOwner = self.request_data.get('bankOwner')
        amount = self.request_data.get('amount')
        notifyUrl = self.request_data.get('notifyUrl')
        if not mchId or not mchOrderId or not bankCode or not bankAccount or not bankOwner or not amount or not notifyUrl:
            return self.xtjson.json_params_error('缺少数据！')

        _state, msg = is_valid_url(notifyUrl)
        if not _state:
            return self.xtjson.json_params_error('notifyUrl：参数错误！',code=403)

        if not bankCode or not bankCode.strip() or not bankCode.strip().isalpha():
            return self.xtjson.json_params_error('bankCode: 参数错误！',code=403)

        if not BankTable.find_one({'code': bankCode.strip()}):
            return self.xtjson.json_params_error('bankCode: 参数错误！',code=403)

        if behalfPayOrderTable.find_one({'merchant_order_id': mchOrderId.strip()}):
            return self.xtjson.json_params_error('mchOrderId：不可重复！', code=403)

        _site_data = SiteConfigTable.find_one({}) or {}
        if not _site_data.get('maintain_switch'):
            return self.xtjson.json_params_error('当前业务维护中，暂不可用！', code=407)

        maintain_bankcodes = _site_data.get('maintain_bankcodes') or ''
        maintain_bankcodes = maintain_bankcodes.split(',')
        if bankCode in maintain_bankcodes:
            return self.xtjson.json_params_error(f'{bankCode}：该银行维护中，暂不可用！', code=407)

        if not str(amount).isdigit():
            return self.xtjson.json_params_error('totalAmount: 参数错误！',code=403)

        try:
            amount = int(amount)
        except:
            return self.xtjson.json_params_error('totalAmount: 参数错误！',code=403)

        merchant_data = MerchantTable.find_one({'merchant_id': mchId})
        if not merchant_data:
            return self.xtjson.json_params_error('mchId：参数错误！', code=403)

        rate = merchant_data.get('payment_rate') or 0
        repay_amount = round(amount * rate, 2)

        balance_amount = merchant_data.get('balance_amount') or 0
        if balance_amount < (repay_amount+amount):
            return self.xtjson.json_params_error('商户余额不足！', code=408)

        ouuid = shortuuid.uuid()
        with mc.lock('insert_y_order'):
            _md = mchId[-3:]
            order_id = getBehalfPayOrderId(_md)

            _order_daat = {
                'uuid': ouuid,
                'merchant_id': mchId,
                'merchant_order_id': mchOrderId,
                'order_id': order_id,
                'order_amount': amount,
                'actual_amount': 0,
                'repay_amount': repay_amount,
                'order_time': datetime.datetime.now(),
                'pay_statu': False,
                'reject_pay': False,
                'is_task': False,
                'callback_statu': CallbackState.NOT_CALLEDBACK,
                'callback_url': notifyUrl,
                'bank_memo': order_id,
                'out_money_userid': '',
                'ip': '',

                'receive_bank_code': bankCode or '',
                'receive_account': bankAccount or '',
                'receive_owner': bankOwner or '',
                'agentadmin_uuid': merchant_data.get('agentadmin_uuid'),
            }
            behalfPayOrderTable.insert_one(_order_daat)

        behalfPayOrderProcessTable.insert_one({'order_id': order_id, 'text': '创建订单！'})
        payBehalf_deduct(order_uuid=ouuid, merchant_data=merchant_data)
        return self.xtjson.json_result()

    def createTestOrder(self):
        bankcc = [
            {
                "bankCode": "NAB",
                "bankAccount": "622226319100001",
                "bankOwner": "NGUYEN VU PHUONG",
            },
            {
                "bankCode": "VAB",
                "bankAccount": "00195363",
                "bankOwner": "NGO TAN PHUONG TAY",
            },
            {
                "bankCode": "SEAB",
                "bankAccount": "000009731380",
                "bankOwner": "DINH THI QUYNH NHU",
            },
            {
                "bankCode": "VAB",
                "bankAccount": "00340337",
                "bankOwner": "NGUYEN HONG SON",
            },
        ]
        for i in range(5):
            time.sleep(0.3)
            bank_data = random.choice(bankcc)
            bankCode = bank_data.get('bankCode')
            bankAccount = bank_data.get('bankAccount')
            bankOwner = bank_data.get('bankOwner')

            merchant_data = MerchantTable.find_one({})
            if not merchant_data:
                return self.xtjson.json_params_error('无商户数据！')
            mchId = MerchantTable.find_one({}).get('merchant_id')
            mchOrderId = 'Y'+str(int(time.time()*1000))
            amount = random.choice(list(range(10011, 10099)))
            notifyUrl = 'https://'+current_app.config.get('MAIN_DOMAIN')

            _site_data = SiteConfigTable.find_one({}) or {}
            if not _site_data.get('maintain_switch'):
                return self.xtjson.json_params_error('当前业务维护中，暂不可用！')

            # merchant_data = MerchantTable.find_one({'merchant_id': mchId})
            # if not merchant_data:
            #     return self.xtjson.json_params_error('商户不存在！')

            rate = merchant_data.get('payment_rate') or 0
            repay_amount = round(amount * rate, 2)

            balance_amount = merchant_data.get('balance_amount') or 0
            if balance_amount < (repay_amount + amount):
                return self.xtjson.json_params_error('商户余额不足！', code=408)

            ouuid = shortuuid.uuid()
            with mc.lock('insert_y_order'):
                _md = mchId[-3:]
                order_id = getBehalfPayOrderId(_md)
                _order_daat = {
                    'uuid': ouuid,
                    'merchant_id': mchId,
                    'merchant_order_id': mchOrderId,
                    'order_id': order_id,
                    'order_amount': amount,
                    'actual_amount': 0,
                    'repay_amount': repay_amount,
                    'order_time': datetime.datetime.now(),
                    'pay_statu': False,
                    'reject_pay': False,
                    'is_task': False,
                    'callback_statu': CallbackState.NOT_CALLEDBACK,
                    'callback_url': notifyUrl,
                    'bank_memo': order_id,
                    'out_money_userid': '',
                    'ip': '',

                    'receive_bank_code': bankCode or '',
                    'receive_account': bankAccount or '',
                    'receive_owner': bankOwner or '',
                    "out_money_userid": self.current_admin_dict.get("uuid") if self.current_admin_user.role_code == ROlE_ALL.SYS_OUT_MONEY_USER or self.current_admin_user.role_code == ROlE_ALL.OUT_MONEY_USER else '',
                    'agentadmin_uuid': merchant_data.get('agentadmin_uuid'),
                }
                # print('_order_daat:', _order_daat)
                # print('*'*50)
                behalfPayOrderTable.insert_one(_order_daat)

            behalfPayOrderProcessTable.insert_one({'order_id': order_id, 'text': '创建订单！'})
            payBehalf_deduct(order_uuid=ouuid, merchant_data=merchant_data)
        return self.xtjson.json_result()

    def getSelectedCardInfo(self, selectedcard, order_time):
        selectedcard["withdrawlogs"] = WithdrawalOrderLogTable.find_many({"former_card_uuid": selectedcard.get("uuid"), "payer_uuid": self.current_admin_dict["uuid"], "accepted_status": {"$in": ACCEPTED_STATUS.name_arr} , "order_time": order_time}, limit = 10, sort=[("order_time", -1)])

        result = list(WithdrawalOrderLogTable.collection().aggregate([
            {'$match': 
                {"former_card_uuid": selectedcard.get("uuid"), "order_time": order_time ,"accepted_status": {"$in": [ACCEPTED_STATUS.NOT_PROCESSED, ACCEPTED_STATUS.ACCEPTED]} }
            },
            { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'request_money'}"}}}
        ]))
        selectedcard["request_money"] = result[0].get("transfer_total") if result else 0

        result = list(WithdrawalOrderLogTable.collection().aggregate([
            {'$match': 
                {"former_card_uuid": selectedcard.get("uuid"),"behavior": BEHAVIOR_TYPE.TRANSFER_IN , "order_time": order_time ,"accepted_status": {"$in": [ACCEPTED_STATUS.NOT_PROCESSED, ACCEPTED_STATUS.ACCEPTED]} }
            },
            { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
        ]))
        selectedcard["transfer_in_money"] = result[0].get("transfer_total") if result else 0

        result = list(WithdrawalOrderLogTable.collection().aggregate([
            {'$match': 
                {"former_card_uuid": selectedcard.get("uuid"),"behavior": BEHAVIOR_TYPE.TRANSFER_OUT , "order_time": order_time ,"accepted_status": {"$in": [ACCEPTED_STATUS.NOT_PROCESSED, ACCEPTED_STATUS.ACCEPTED]} }
            },
            { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
        ]))
        selectedcard["transfer_out_money"] = result[0].get("transfer_total") if result else 0

        result = list(WithdrawalOrderLogTable.collection().aggregate([
            {'$match': 
                {"former_card_uuid": selectedcard.get("uuid"),"behavior": BEHAVIOR_TYPE.OTHER_TRANSFER_IN , "order_time": order_time ,"accepted_status": {"$in": [ACCEPTED_STATUS.NOT_PROCESSED, ACCEPTED_STATUS.ACCEPTED]} }
            },
            { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
        ]))
        selectedcard["other_transfer_in_money"] = result[0].get("transfer_total") if result else 0

        result = list(WithdrawalOrderLogTable.collection().aggregate([
            {'$match': 
                {"former_card_uuid": selectedcard.get("uuid"),"behavior": BEHAVIOR_TYPE.OTHER_TRANSFER_OUT , "order_time": order_time ,"accepted_status": {"$in": [ACCEPTED_STATUS.NOT_PROCESSED, ACCEPTED_STATUS.ACCEPTED]} }
            },
            { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
        ]))
        selectedcard["other_transfer_out_money"] = result[0].get("transfer_total") if result else 0

        result = list(WithdrawalOrderLogTable.collection().aggregate([
            {'$match': 
                {"former_card_uuid": selectedcard.get("uuid"),"behavior": BEHAVIOR_TYPE.ISSUED , "order_time": order_time ,"accepted_status": {"$in": [ACCEPTED_STATUS.NOT_PROCESSED, ACCEPTED_STATUS.ACCEPTED]} }
            },
            { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
        ]))
        selectedcard["issued_money"] = result[0].get("transfer_total") if result else 0
        
        selectedcard["plus_money"] = selectedcard["transfer_in_money"]+selectedcard.get("other_transfer_in_money")
        selectedcard["minus_money"] = selectedcard.get("transfer_out_money")+selectedcard.get("other_transfer_out_money")+selectedcard.get("issued_money")

        selectedcard["error"] = selectedcard.get("start_money") + selectedcard["transfer_in_money"]+selectedcard.get("other_transfer_in_money") - selectedcard.get("transfer_out_money")-selectedcard.get("other_transfer_out_money")-selectedcard.get("issued_money")-selectedcard.get("balance_amount")

    
    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        datas = []
        oumudd = {}
        merchant_datas = {}
        withdrawalcards = []
        selectedcard = {}
        for da in all_datas:
            merchant_data = merchant_datas.get(da.get('merchant_id')) or {}
            if not merchant_data:
                merchant_data = MerchantTable.find_one({'merchant_id': da.get('merchant_id')}) or {}
                merchant_datas[da.get('merchant_id')] = merchant_data
            da['merchant_data'] = merchant_data

            bank_data = BankTable.find_one({'code': da.get('receive_bank_code')}) or {}
            da['bank_data'] = bank_data
            out_money_userid = da.get('out_money_userid')
            if out_money_userid:
                out_money_user_data = oumudd.get(out_money_userid) or ''
                if not out_money_user_data:
                    out_money_user_data = CmsUserTable.find_one({'uuid': out_money_userid}) or {}
                    oumudd[out_money_userid] = out_money_user_data
            else:
                out_money_user_data = {}
            da['out_money_user_data'] = out_money_user_data or {}

            datas.append(da)

        system_paybehalf = True
        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            system_paybehalf = self.current_admin_dict.get('system_paybehalf') or False
        if self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            agentadmin_data = CmsUserTable.find_one({'uuid': self.current_admin_dict.get('agentadmin_uuid'), 'role_code': ROlE_ALL.AGENTADMIN}) or {}
            system_paybehalf = agentadmin_data.get('system_paybehalf') or False
        if self.current_admin_dict.get('role_code') == ROlE_ALL.SYS_OUT_MONEY_USER or self.current_admin_dict.get('role_code') == ROlE_ALL.OUT_MONEY_USER:
            
            order_time, start_time, end_time = self.get_reqorder()
            order_time = {'$gte': start_time, '$lte': end_time}

            selectedcard = WithdrawalCardTable.find_one({"uuid": self.current_admin_dict.get("withdrawalcard_uuid")})
            if selectedcard:
                self.getSelectedCardInfo(selectedcard, order_time)

            _fff = {
                'pay_statu': True
            }
            _fff.update(self.get_filter_dict())
            # 总支付金额
            _actual_amount_total = behalfPayOrderTable.collection().aggregate([
                {"$match": _fff},
                {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
            ])
            actual_amount_total_ll = list(_actual_amount_total)
            actual_amount_tota = 0
            if actual_amount_total_ll:
                actual_amount_tota = actual_amount_total_ll[0].get('actual_amount')

            selectedcard["paid_amount"] = actual_amount_tota
            
        self.context['all_datas'] = datas
        self.context['selectedcard'] = selectedcard
        self.context['system_paybehalf'] = system_paybehalf
        self.context['CallbackState'] = CallbackState
        self.context['CallbankType'] = CallbankType
        self.context['ACCEPTED_STATUS'] = ACCEPTED_STATUS
        self.context['BEHAVIOR_TYPE'] = BEHAVIOR_TYPE

    def get_other_way(self):
        action = request.args.get('action')
        if action == 'refreshOrderList':
            return self.refreshOrderList()

    def post_other_way(self):
        if self.action == 'get_total_info':
            filter_dict = self.get_filter_dict()

            # 总订单金额
            _order_amount_total = behalfPayOrderTable.collection().aggregate([
                {"$match": filter_dict},
                {"$group": {"_id": None, "order_amount": {"$sum": '$order_amount'}}},
            ])
            order_amount_total_ll = list(_order_amount_total)
            order_amount_tota = 0
            if order_amount_total_ll:
                order_amount_tota = order_amount_total_ll[0].get('order_amount')

            # 总订单手续费
            _repay_amount_total = behalfPayOrderTable.collection().aggregate([
                {"$match": filter_dict},
                {"$group": {"_id": None, "repay_amount": {"$sum": '$repay_amount'}}},
            ])
            repay_amount_total_ll = list(_repay_amount_total)
            repay_amount_total = 0
            if repay_amount_total_ll:
                repay_amount_total = repay_amount_total_ll[0].get('repay_amount')

            # 总订单笔数
            total_order_count = behalfPayOrderTable.count(filter_dict) or 0

            # 总支付笔数
            _fff = {
                'pay_statu': True
            }
            _fff.update(filter_dict)
            total_pay_count = behalfPayOrderTable.count(_fff) or 0

            # 总支付金额
            _actual_amount_total = behalfPayOrderTable.collection().aggregate([
                {"$match": _fff},
                {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
            ])
            actual_amount_total_ll = list(_actual_amount_total)
            actual_amount_tota = 0
            if actual_amount_total_ll:
                actual_amount_tota = actual_amount_total_ll[0].get('actual_amount')

            # 总支付手续费
            _repay_amount_total = behalfPayOrderTable.collection().aggregate([
                {"$match": _fff},
                {"$group": {"_id": None, "repay_amount": {"$sum": '$repay_amount'}}},
            ])
            repay_amount_total_ll = list(_repay_amount_total)
            repay_amount_tota = 0
            if repay_amount_total_ll:
                repay_amount_tota = repay_amount_total_ll[0].get('repay_amount')

            # 回调笔数
            fff11 = {}
            fff11.update(filter_dict)
            if 'pay_statu' not in fff11:
                fff11['pay_statu'] = True
            if 'callback_statu' not in fff11:
                fff11['callback_statu'] = CallbackState.SUCCESS
            callback_order_count = behalfPayOrderTable.count(fff11) or 0

            # 回调金额
            _actual_amount_total = behalfPayOrderTable.collection().aggregate([
                {"$match": fff11},
                {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
            ])
            actual_amount_total_ll = list(_actual_amount_total)
            callback_amount_tota = 0
            if actual_amount_total_ll:
                callback_amount_tota = actual_amount_total_ll[0].get('actual_amount')

            # 支付成功未回调或回调失败
            fff22 = {}
            fff22.update(filter_dict)
            if 'pay_statu' not in fff22:
                fff22['pay_statu'] = True
            if 'callback_statu' not in fff22:
                fff22['callback_statu'] = {'$in': [CallbackState.NOT_CALLEDBACK, CallbackState.FAILED]}
            nf_callback_count = behalfPayOrderTable.count(fff22) or 0

            # 回调金额
            _nf_callback_amount_ll = behalfPayOrderTable.collection().aggregate([
                {"$match": fff22},
                {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
            ])
            nf_callback_amount_ll = list(_nf_callback_amount_ll)
            nf_callback_amount = 0
            if nf_callback_amount_ll:
                nf_callback_amount = nf_callback_amount_ll[0].get('actual_amount')

            _data = {
                'order_amount_total': self.format_money(order_amount_tota),
                'actual_amount_total': self.format_money(actual_amount_tota),
                'repay_amount_total': self.format_money(repay_amount_total),
                'pay_repay_amount_total': self.format_money(repay_amount_tota),
                'total_order_count': total_order_count,
                'total_pay_count': total_pay_count,
                'callback_order_count': callback_order_count,
                'callback_amount_total': self.format_money(callback_amount_tota),
                'nf_callback_count': nf_callback_count,
                'nf_callback_amount': self.format_money(nf_callback_amount),
            }
            return self.xtjson.json_result(data=_data)
        if self.action == 'export_order':
            filter_dict = self.get_filter_dict()
            datas = behalfPayOrderTable.find_many(filter_dict)
            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            FOLDER_name = ''
            agentadmin_uuid = ''
            if self.current_admin_roleCode == ROlE_ALL.SUPERADMIN or self.current_admin_roleCode == ROlE_ALL.ADMINISTRATOR:
                FOLDER_name = 'su'
            else:
                if self.is_agentadmin:
                    FOLDER_name = self.current_admin_dict.get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('uuid')
                elif self.current_admin_dict.get('agentadmin_data'):
                    FOLDER_name = self.current_admin_dict.get('agentadmin_data').get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('agentadmin_data').get('uuid')
            if not FOLDER_name:
                return self.xtjson.json_params_error()
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER, FOLDER_name)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            _out_data_dict = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas),
                'out_count': 0,
                'note': '订单数据-' + datetime.datetime.now().strftime('%Y%m%d%H%M%S'),
                'agentadmin_uuid': agentadmin_uuid,
            }
            uuid = ExportDataModel.insert_one(_out_data_dict)
            threading.Thread(target=self.exportData, args=(datas, uuid, export_folder, filename)).start()
            return self.xtjson.json_result(message='数据导出中，请稍后到"导出文件"中查看数据！')
        if self.action == 'get_testOrder_html':
            return self.get_testOrder_html()
        if self.action == 'createOrder':
            return self.createOrder_func()
        if self.action == 'createTestOrder':
            return self.createTestOrder()

    def post_data_other_way(self):
        # 或者支付二维码
        if self.action == 'get_pay_qrcode':
            bank_data = BankTable.find_one({'code': self.data_dict.get('receive_bank_code')}) or {}
            if not bank_data:
                return self.xtjson.json_params_error('该订单收款银行错误！')

            payqrcode_url = self.data_dict.get('payqrcode_url') or ''
            project_static_folder = os.path.join(current_app.static_folder, current_app.config.get('PROJECT_NAME'))
            _state, payQrcode = getBankPayQrcode(
                self.data_uuid,
                self.data_dict.get('order_amount'),
                self.data_dict.get('bank_memo'),
                bank_data,
                payqrcode_url=payqrcode_url,
                project_static_folder=project_static_folder,
                receive_account=self.data_dict.get('receive_account'),
                is_behalfPay=True
            )

            if not _state:
                return self.xtjson.json_params_error(payQrcode)

            return self.xtjson.json_result(data={'generate_qrcode': payQrcode})
        # 获取订单信息
        if self.action == 'get_pay_info':
            bank_data = BankTable.find_one({'code': self.data_dict.get('receive_bank_code')}) or {}
            if not bank_data:
                return self.xtjson.json_params_error('该订单收款银行错误！')

            _data = {
                'order_id': self.data_dict.get('order_id'),
                'receive_bank': bank_data.get('shortName'),
                'receive_account': self.data_dict.get('receive_account'),
                'receive_owner': self.data_dict.get('receive_owner'),
                'order_amount': self.data_dict.get('order_amount'),
                'pay_statu': self.data_dict.get('pay_statu') or False,
                'reject_pay': self.data_dict.get('reject_pay') or False,
                'is_read': self.data_dict.get('is_read') or False,
            }
            if self.data_dict.get('pay_statu'):
                _data['payQrcode'] = '/public/world/img/paysuccess5.png'
            elif self.data_dict.get('reject_pay'):
                _data['payQrcode'] = '/public/world/img/ju.png'
            else:
                payqrcode_url = self.data_dict.get('payqrcode_url') or ''
                project_static_folder = os.path.join(current_app.static_folder, current_app.config.get('PROJECT_NAME'))
                _state, payQrcode = getBankPayQrcode(
                    self.data_uuid,
                    self.data_dict.get('order_amount'),
                    self.data_dict.get('bank_memo'),
                    bank_data,
                    payqrcode_url=payqrcode_url,
                    project_static_folder=project_static_folder,
                    receive_account=self.data_dict.get('receive_account'),
                    is_behalfPay=True
                )
                if not _state:
                    return self.xtjson.json_params_error(payQrcode)
                _data['payQrcode'] = payQrcode
            return self.xtjson.json_result(data=_data)
        # 删除订单
        if self.action == 'delOrder':
            self.MCLS.delete_one({'uuid': self.data_uuid})
            return self.xtjson.json_result()
        # 支付订单
        if self.action == 'payOrder':
            if self.data_dict.get('pay_statu'):
                return self.xtjson.json_params_error('该订单已支付！')
            if self.data_dict.get('reject_pay'):
                return self.xtjson.json_params_error('该订单已处理！')

            merchant_data = MerchantTable.find_one({'merchant_id': self.data_dict.get('merchant_id')})
            if not merchant_data:
                return self.xtjson.json_params_error('商户不存在！')

            order_amount = self.data_dict.get('order_amount')
            repay_amount = self.data_dict.get('repay_amount')
            _amount = order_amount + repay_amount

            if self.current_admin_dict.get("role_code") == ROlE_ALL.SYS_OUT_MONEY_USER or self.current_admin_dict.get("role_code") == ROlE_ALL.OUT_MONEY_USER:
                selectedcard = WithdrawalCardTable.find_one({"uuid": self.current_admin_dict.get("withdrawalcard_uuid")})
                if selectedcard.get("balance_amount") < order_amount:
                    return self.xtjson.json_params_error("余额不足！") 
                ss = {
                    "balance_amount": selectedcard.get("balance_amount") - order_amount
                }
                WithdrawalCardTable.update_one({'uuid': selectedcard.get("uuid")}, {"$set": ss})

            self.data_from['actual_amount'] = order_amount
            self.data_from['force_ispay'] = True
            self.data_from['pay_statu'] = True
            self.data_from['processor_uid'] = self.current_admin_dict.get('uuid')
            self.data_from['pay_time'] = datetime.datetime.now()
            self.MCLS.update_one({'uuid': self.data_uuid},{'$set': self.data_from})

            # 添加订单流程
            bankcard_info_text = ''
            bankcard_info = self.current_admin_dict.get('bankcard_info') or ''
            if bankcard_info.strip():
                bankcard_info_text += '出款信息：' + bankcard_info
            behalfPayOrderProcessTable.insert_one({'order_id': self.data_dict.get('order_id'), 'text': f'付款成功，处理方式：手动，处理人：{self.current_admin_dict.get("account")}，{bankcard_info_text}'})

            # 代付回调
            ptext = '手动回调，回调结果：'
            _state, _res = behalfPayCallbackOrderFunc(self.data_uuid, is_manual=True, admin_uuid=self.current_admin_dict.get('uuid'), note='强制付款')
            if _state:
                ptext += '成功'
            else:
                ptext += '失败'
            behalfPayOrderProcessTable.insert_one({'order_id': self.data_dict.get('order_id'), 'text': ptext})
            return self.xtjson.json_result()
        # 拒绝订单
        if self.action == 'is_reject_pay':
            if self.data_dict.get('pay_statu'):
                return self.xtjson.json_params_error('该订单已支付！')

            order_msg = self.request_data.get('order_msg')
            if not order_msg or not order_msg.strip():
                return self.xtjson.json_params_error('请选择拒绝原因！')

            if self.data_dict.get('reject_pay'):
                return self.xtjson.json_params_error('该订单已处理！')

            merchant_data = MerchantTable.find_one({'merchant_id': self.data_dict.get('merchant_id')})
            if not merchant_data:
                return self.xtjson.json_params_error('商户不存在！')

            self.data_from['reject_pay'] = True
            self.data_from['dealwith_time'] = datetime.datetime.now()
            self.MCLS.update_one({'uuid': self.data_uuid},{'$set': self.data_from})

            # 添加订单流程
            behalfPayOrderProcessTable.insert_one({'order_id': self.data_dict.get('order_id'), 'text': f'拒绝支付，处理方式：手动，处理人：{self.current_admin_dict.get("account")}'})

            ptext = '手动回调，回调结果：'
            _state,_res = behalfPayCallbackOrderFunc(self.data_uuid, is_manual=True, admin_uuid=self.current_admin_dict.get('uuid'), note=order_msg.strip(), msg=order_msg.strip())
            if not _state:
                ptext += '失败'
                behalfPayOrderProcessTable.insert_one({'order_id': self.data_dict.get('order_id'), 'text': ptext})
                return self.xtjson.json_params_error('回调失败！')

            ptext += '成功'
            behalfPayOrderProcessTable.insert_one({'order_id': self.data_dict.get('order_id'), 'text': ptext})
            payBehalf_goback(order_uuid=self.data_uuid, merchant_data=merchant_data)
            return self.xtjson.json_result()
        # 订单信息
        if self.action == 'detailsInfo_html':
            return self.detailsInfo_html()
        # 更新备注
        if self.action == 'updateNote':
            data_value = self.request_data.get('data_value') or ''
            if not data_value:
                return self.xtjson.json_params_error('请收入备注内容!')
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': {'note': data_value.strip()}})
            return self.xtjson.json_result()
        if self.action == 'callbackOrder':
            merchant_data = MerchantTable.find_one({'merchant_id': self.data_dict.get('merchant_id')})
            if not merchant_data:
                return self.xtjson.json_params_error('商户不存在！')

            msg = ''
            note = ''
            cadata = behalfPayCallbackLogTable.find_one({'order_uuid': self.data_uuid}, sort=[['create_time', -1]]) or {}
            if self.data_dict.get('reject_pay'):
                request_text = cadata.get('request_text') or {}
                if isinstance(request_text, dict):
                    msg = request_text.get('msg') or ''
                    note = request_text.get('msg') or ''
            ptext = '手动回调，回调结果：'
            _state, _res = behalfPayCallbackOrderFunc(self.data_uuid, is_manual=True, admin_uuid=self.current_admin_dict.get('uuid'), note=note, msg=msg)
            if not _state:
                ptext += '失败'
                behalfPayOrderProcessTable.insert_one({'order_id': self.data_dict.get('order_id'), 'text': ptext})
                return self.xtjson.json_params_error('回调失败！')
            ptext += '成功'
            behalfPayOrderProcessTable.insert_one({'order_id': self.data_dict.get('order_id'), 'text': ptext})
            if self.data_dict.get('reject_pay'):
                payBehalf_goback(order_uuid=self.data_uuid, merchant_data=merchant_data)
            return self.xtjson.json_result()
        if self.action == 'get_payQrcode':
            bank_data = BankTable.find_one({'code': self.data_dict.get('receive_bank_code')}) or {}
            if not bank_data:
                return self.xtjson.json_params_error('该订单收款银行错误！')
            _data = {
                'order_id': self.data_dict.get('order_id'),
                'receive_bank': bank_data.get('shortName'),
                'receive_account': self.data_dict.get('receive_account'),
                'receive_owner': self.data_dict.get('receive_owner'),
                'order_amount': self.data_dict.get('order_amount'),
                'pay_statu': self.data_dict.get('pay_statu') or False,
                'reject_pay': self.data_dict.get('reject_pay') or False,
            }
            payqrcode_url = self.data_dict.get('payqrcode_url') or ''
            project_static_folder = os.path.join(current_app.static_folder, current_app.config.get('PROJECT_NAME'))
            _state, payQrcode = getBankPayQrcode(
                self.data_uuid,
                self.data_dict.get('order_amount'),
                self.data_dict.get('bank_memo'),
                bank_data,
                payqrcode_url=payqrcode_url,
                project_static_folder=project_static_folder,
                receive_account=self.data_dict.get('receive_account'),
                is_behalfPay=True
            )
            if not _state:
                return self.xtjson.json_params_error(payQrcode)

            _data['payQrcode'] = payQrcode
            return self.xtjson.json_result(data=_data)
        if self.action == 'forcePay':
            if self.data_dict.get('pay_statu'):
                return self.xtjson.json_params_error('该订单已支付！')

            merchant_data = MerchantTable.find_one({'merchant_id': self.data_dict.get('merchant_id')})
            if not merchant_data:
                return self.xtjson.json_params_error('商户不存在！')

            order_amount = self.data_dict.get('order_amount')
            repay_amount = self.data_dict.get('repay_amount')
            _amount = order_amount + repay_amount

            self.data_from['actual_amount'] = order_amount
            self.data_from['force_ispay'] = True
            self.data_from['pay_statu'] = True
            self.data_from['processor_uid'] = self.current_admin_dict.get('uuid')
            self.data_from['pay_time'] = datetime.datetime.now()
            self.MCLS.update_one({'uuid': self.data_uuid},{'$set': self.data_from})

            # 添加订单流程
            bankcard_info_text = ''
            bankcard_info = self.current_admin_dict.get('bankcard_info') or ''
            if bankcard_info.strip():
                bankcard_info_text += '出款信息：' + bankcard_info
            behalfPayOrderProcessTable.insert_one({'order_id': self.data_dict.get('order_id'), 'text': f'付款成功，处理方式：手动，处理人：{self.current_admin_dict.get("account")}，{bankcard_info_text}'})

            # 代付回调
            ptext = '手动回调，回调结果：'
            _state, _res = behalfPayCallbackOrderFunc(self.data_uuid, is_manual=True, admin_uuid=self.current_admin_dict.get('uuid'), note='强制付款')
            if _state:
                ptext += '成功'
            else:
                ptext += '失败'
            behalfPayOrderProcessTable.insert_one({'order_id': self.data_dict.get('order_id'), 'text': ptext})
            return self.xtjson.json_result()
        if self.action == 'update_is_read':
            if not self.data_dict.get('is_read'):
                self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': {'is_read': True}})
            return self.xtjson.json_result()



class behalfPayCallbackLogView(CmsTableViewBase):
    add_url_rules = [['/behalfPay/callbackLog', 'behalfPay_callbackLog']]
    per_page = 30
    MCLS = behalfPayCallbackLogTable
    template = 'cms/behalfPay/callbackLog.html'
    title = '回调记录'

    def get_filter_dict(self):
        fff = {}
        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')
        return fff

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        datas = []
        _dd = {}
        for da in all_datas:
            merchant_uuid = da.get('merchant_uuid')
            merchant_data = _dd.get(merchant_uuid)
            if not merchant_data:
                merchant_data = MerchantTable.find_one({'uuid': merchant_uuid})
                _dd[merchant_uuid] = merchant_data
            da['merchant_data'] = merchant_data

            order_uuid = da.get('order_uuid')
            order_data = da.get(order_uuid)
            if not order_data:
                order_data = behalfPayOrderTable.find_one({'uuid': order_uuid})
            da['order_data'] = order_data
            datas.append(da)
        self.context['all_datas'] = datas



class behalfPayScriptView(CmsTableViewBase):
    add_url_rules = [['/behalfPay/scriptConfig', 'behalfPay_scriptConfig']]
    per_page = 30
    MCLS = behalfPayScriptTable
    template = 'cms/behalfPay/scriptConfig.html'
    title = '代付脚本配置列表'

    def get_data_id(self, staStr=''):
        while True:
            id_str = staStr +str(random.random() * 100000).split('.')[0]
            if self.MCLS.find_one({'$or': [{'desktop_id': id_str}, {'device_id': id_str}]}):
                continue
            return id_str

    def script_form_html(self, data_dict={}):
        if data_dict:
            _action = 'edit_script_data'
        else:
            _action = 'add_script_data'
        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">       
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">名称：</span>
                        <input type="text" class="form-control" id="name" value="{ data_dict.get('name') or '' }" placeholder="名称" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">桌面编号：</span>
                        <input type="text" class="form-control" id="desktop_id" value="{ data_dict.get('desktop_id') or self.get_data_id('D') }" placeholder="桌面编号" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">设备Id：</span>
                        <input type="text" class="form-control" id="device_id" value="{ data_dict.get('device_id') or self.get_data_id('S') }" placeholder="设备Id" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行卡人姓名：</span>
                        <input type="text" class="form-control" id="bankowner" value="{ data_dict.get('bankowner') or '' }" placeholder="银行卡人姓名" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行卡号：</span>
                        <input type="text" class="form-control" id="bankcard_account" value="{ data_dict.get('bankcard_account') or '' }" placeholder="银行卡号" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">金额最小值：</span>
                        <input type="number" class="form-control" id="min_amount" value="{ data_dict.get('min_amount') or '' }" placeholder="金额最小值" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">金额最大值：</span>
                        <input type="number" class="form-control" id="max_amount" value="{ data_dict.get('max_amount') or '' }" placeholder="金额最大值" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">备注：</span>
                        <input type="text" class="form-control" id="note" value="{ data_dict.get('note') or '' }" placeholder="备注" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item" style="display: flex; align-items: center;justify-content: center;">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">IP白名单：</span>
                        <textarea rows="5" class="form-control" id="ips" placeholder="IP白名单，一行一个" aria-label="" style="display: inline-block; width: calc(100% - 180px)">{data_dict.get('ips') or ''}</textarea>
                    </div>                             
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_script_data('{_action}', '{data_dict.get('uuid') if data_dict else ''}')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>
        '''
        return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas') or []
        _datas = []
        for adl in all_datas:
            back_code = adl.get('back_code') or ''
            _bacnk_data = BankTable.find_one({'code': back_code}) or {}
            adl['bank_data'] = _bacnk_data

            _datas.append(adl)
        self.context['all_datas'] = _datas

    def post_other_way(self):
        if self.action == 'add_script_html':
            return self.script_form_html()
        if self.action == 'add_script_data':
            name = self.request_data.get('name')
            note = self.request_data.get('note') or ''
            ips = self.request_data.get('ips') or ''
            desktop_id = self.request_data.get('desktop_id') or ''
            device_id = self.request_data.get('device_id') or ''
            bankowner = self.request_data.get('bankowner') or ''
            bankcard_account = self.request_data.get('bankcard_account') or ''
            min_amount = self.request_data.get('min_amount') or 0
            max_amount = self.request_data.get('max_amount') or 0
            if not name or not bankowner or not bankcard_account or not desktop_id or not device_id:
                return self.xtjson.json_params_error('缺少数据！')

            try:
                min_amount = int(min_amount)
            except:
                return self.xtjson.json_params_error('min_amount: 数据错误！')

            try:
                max_amount = int(max_amount)
            except:
                return self.xtjson.json_params_error('min_amount: 数据错误！')

            uid = shortuuid.uuid()
            _DATA = {
                'uuid': uid,
                'desktop_id': desktop_id,
                'device_id': device_id,
                'name': name.strip(),
                'note': note.strip(),
                'ips': ips.strip(),
                'bankowner': bankowner.strip(),
                'bankcard_account': bankcard_account.strip(),
                'balance_amount': 0,
                'min_amount': min_amount,
                'max_amount': max_amount,
                'statu': True,
                'secret_key': encry_md5(uid)
            }
            self.MCLS.insert_one(_DATA)
            return self.xtjson.json_result()

    def post_data_other_way(self):
        if self.action == 'edit_script_html':
            return self.script_form_html(self.data_dict)
        if self.action == 'edit_script_data':
            name = self.request_data.get('name')
            bankowner = self.request_data.get('bankowner')
            bankcard_account = self.request_data.get('bankcard_account')
            note = self.request_data.get('note') or ''
            ips = self.request_data.get('ips') or ''
            desktop_id = self.request_data.get('desktop_id') or ''
            device_id = self.request_data.get('device_id') or ''
            min_amount = self.request_data.get('min_amount') or 0
            max_amount = self.request_data.get('max_amount') or 0
            if not name or not bankowner or not bankcard_account or not desktop_id or not device_id:
                return self.xtjson.json_params_error('缺少数据！')

            try:
                min_amount = int(min_amount)
            except:
                return self.xtjson.json_params_error('min_amount: 数据错误！')

            try:
                max_amount = int(max_amount)
            except:
                return self.xtjson.json_params_error('min_amount: 数据错误！')

            _DATA = {
                'name': name.strip(),
                'desktop_id': desktop_id.strip(),
                'bankowner': bankowner.strip(),
                'device_id': device_id.strip(),
                'bankcard_account': bankcard_account.strip(),
                'note': note.strip(),
                'ips': ips.strip(),
                'min_amount': min_amount,
                'max_amount': max_amount,
            }
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': _DATA})
            return self.xtjson.json_result()
        if self.action == 'update_script_statu':
            if self.data_dict.get('statu'):
                self.data_from['statu'] = False
            else:
                self.data_from['statu'] = True
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            return self.xtjson.json_result()



class behalfPayTaskView(CmsTableViewBase):
    add_url_rules = [['/behalfPay/taskList', 'behalfPay_taskList']]
    per_page = 30
    MCLS = behalfPayTaskTable
    template = 'cms/behalfPay/taskList.html'
    title = '代付任务列表'

    def get_context(self):
        return {'taskStatus': taskStatus}

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas') or []
        _datas = []
        for adl in all_datas:

            script_id = adl.get('script_id') or ''
            script_data = behalfPayScriptTable.find_one({'uuid': script_id}) or {}
            adl['script_data'] = script_data

            order_id = adl.get('order_id') or ''
            order_data = behalfPayOrderTable.find_one({'order_id': order_id}) or {}
            adl['order_data'] = order_data

            _datas.append(adl)
        self.context['all_datas'] = _datas



class behalfPayConfigView(CmsFormViewBase):
    add_url_rules = [['/behalfPayConfig', 'behalfPayConfig_view']]
    title = '代付配置'
    MCLS = SiteConfigTable
    
    def getConfig(self):
        site_data = {}
        if self.current_admin_dict.get('role_code') in [ROlE_ALL.SUPERADMIN, ROlE_ALL.ADMINISTRATOR]:
            _data = self.MCLS.find_one({})
            ks = [
                'maintain_switch'
            ]
            for k in ks:
                site_data[k] = _data.get(k) or ''
            maintain_bankcodes = _data.get('maintain_bankcodes') or ''
            maintain_bankcodes = maintain_bankcodes.split(',')
            site_data['maintain_bankcodes'] = maintain_bankcodes
            site_data['automatic_maintain_switch'] = _data.get('automatic_maintain_switch')
            site_data['check_aname_switch'] = _data.get('check_aname_switch')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            maintain_bankcodes = self.current_admin_dict.get('maintain_bankcodes') or ''
            maintain_switch = self.current_admin_dict.get('maintain_bankcodes') or False
            site_data['maintain_bankcodes'] = maintain_bankcodes
            site_data['maintain_switch'] = maintain_switch
        elif self.current_admin_dict.get('role_code') == [ROlE_ALL.OUT_MONEY_USER, ROlE_ALL.SYSTEMUSER]:
            maintain_bankcodes = self.current_admin_dict.get('agentadmin_data').get('maintain_bankcodes') or ''
            maintain_switch = self.current_admin_dict.get('agentadmin_data').get('maintain_bankcodes') or False
            site_data['maintain_bankcodes'] = maintain_bankcodes
            site_data['maintain_switch'] = maintain_switch
        else:
            return
        return site_data

    def view_get(self):
        self.context['title'] = self.title
        site_data = self.getConfig()
        if not site_data:
            return abort(404)
        self.context['site_data'] = site_data
        bank_datas = BankTable.find_many({}) or []
        self.context['bank_datas'] = bank_datas
        html =  render_template('cms/behalfPay/bpConfig.html', **self.context)
        return update_language(self.current_admin_dict.get("language"), html)

    def post_other_way(self):
        if self.action == 'getSiteConfig':
            site_data = self.getConfig()
            if not site_data:
                return self.xtjson.json_params_error()
            return self.xtjson.json_result(data=site_data)
        if self.action == 'updateConfig':
            maintain_switch = self.request_data.get('maintain_switch')
            automatic_maintain_switch = self.request_data.get('automatic_maintain_switch')
            check_aname_switch = self.request_data.get('check_aname_switch')
            maintain_bankcodes = self.request_data.get('maintain_bankcodes') or ''

            _maintain_switch = False
            if maintain_switch == '1':
                _maintain_switch = True
            self.data_from['maintain_switch'] = _maintain_switch

            maintain_bankcodes = maintain_bankcodes.strip().strip(',')
            self.data_from['maintain_bankcodes'] = maintain_bankcodes

            # 系统管理员配置功能
            if self.current_admin_dict.get('role_code') in [ROlE_ALL.SUPERADMIN, ROlE_ALL.ADMINISTRATOR]:
                _automatic_maintain_switch = False
                if automatic_maintain_switch == '1':
                    _automatic_maintain_switch = True
                self.data_from['automatic_maintain_switch'] = _automatic_maintain_switch

                _check_aname_switch = False
                if check_aname_switch == '1':
                    _check_aname_switch = True
                self.data_from['check_aname_switch'] = _check_aname_switch

                _site_data = self.MCLS.find_one({})
                if _site_data:
                    self.MCLS.update_one({'uuid': _site_data.get('uuid')}, {'$set': self.data_from})
            elif self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
                CmsUserTable.update_one({'uuid': self.current_admin_dict.get('uuid')}, {'$set': self.data_from})
            elif self.current_admin_dict.get('role_code') == [ROlE_ALL.OUT_MONEY_USER, ROlE_ALL.SYSTEMUSER]:
                CmsUserTable.update_one({'uuid': self.current_admin_dict.get('agentadmin_data').get('uuid')}, {'$set': self.data_from})
            self.add_SystemLog('更新代付配置')
            return self.xtjson.json_result()



class CnBankCardView(CmsTableViewBase):
    add_url_rules = [['/checkName/bnakcards', 'CnBankCardView']]
    per_page = 30
    MCLS = CnBankCardTable
    template = 'cms/behalfPay/cnBankCard.html'
    title = '姓名检测银行卡管理'

    def bankCard_html(self, data_dict={}):
        _action = 'add_bankCard_data'
        if data_dict:
            _action = 'edit_bankCard_data'

        CC = ['ACB']

        bank_html = ''
        for da in BankTable.find_many({}):
            if da.get('code') not in CC:
                continue
            if data_dict:
                bank_html += f'''
                <option value="{da.get('uuid')}" {'selected' if data_dict.get('bank_uid') == da.get('uuid') else ''}>{da.get('shortName')}</option>
                '''
            else:
                bank_html += '''
                <option value="%s">%s</option>
                ''' % (da.get('uuid'), da.get('shortName'))

        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">       
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">名称：</span>
                        <input type="text" class="form-control" id="name" value="{data_dict.get('name') or ''}" placeholder="名称" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行用户名：</span>
                        <input type="text" class="form-control" id="username" value="{data_dict.get('username') or ''}" placeholder="银行用户名" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行密码：</span>
                        <input type="text" class="form-control" id="password" value="{data_dict.get('password') or ''}" placeholder="银行密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行卡账号：</span>
                        <input type="text" class="form-control" id="account" value="{data_dict.get('account') or ''}" placeholder="银行卡账号" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行账户人姓名：</span>
                        <input type="text" class="form-control" id="bank_name" value="{data_dict.get('bank_name') or ''}" placeholder="银行账户人姓名" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行：</span>
                        <select class="form-control" id="bank_uid" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                            <option value="">选择银行</option>
                            {bank_html}
                        </select>
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">备注：</span>
                        <input type="text" class="form-control" id="note" value="{data_dict.get('note') or ''}" placeholder="备注" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                        
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_bank_data('{_action}', '{data_dict.get('uuid') if data_dict else ''}')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>
        '''
        return html

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        datas = []
        _dd = {}
        for da in all_datas:
            bank_uid = da.get('bank_uid')
            bank_data = _dd.get(bank_uid)
            if not bank_data:
                bank_data = BankTable.find_one({'uuid': bank_uid})
                _dd[bank_uid] = bank_data
            da['bank_data'] = bank_data
            datas.append(da)
        self.context['all_datas'] = datas

    def post_other_way(self):
        if self.action == 'add_bankcard_html':
            html = self.bankCard_html()
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        if self.action == 'add_bankCard_data':
            note = request.form.get('note') or ''
            dls = ['name', 'username', 'password', 'account', 'bank_name', 'bank_uid']
            for k in dls:
                _v = request.form.get(k)
                if not _v or not _v.strip():
                    return self.xtjson.json_params_error(f'{k}: 缺少数据！')
                self.data_from[k] = _v.strip()
            if self.MCLS.find_one({'account': self.data_from.get('account')}):
                return self.xtjson.json_params_error('该银行卡账户已存在！')
            self.data_from['note'] = note
            self.data_from['statu'] = True
            self.MCLS.insert_one(self.data_from)
            return self.xtjson.json_result()

    def post_data_other_way(self):
        if self.action == 'edit_bankcard_html':
            html = self.bankCard_html(self.data_dict)
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        if self.action == 'edit_bankCard_data':
            note = request.form.get('note') or ''
            dls = ['name', 'username', 'password', 'account', 'bank_name', 'bank_uid']
            for k in dls:
                _v = request.form.get(k)
                if not _v or not _v.strip():
                    return self.xtjson.json_params_error(f'{k}: 缺少数据！')
                self.data_from[k] = _v.strip()
            _dd = self.MCLS.find_one({'account': self.data_from.get('account')})
            if _dd and _dd.get('uuid') != self.data_uuid:
                return self.xtjson.json_params_error('该银行卡账户已存在！')
            self.data_from['note'] = note
            self.data_from['statu'] = True
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            return self.xtjson.json_result()

class WithdrawalBankCardView(CmsTableViewBase):
    add_url_rules = [['/WithdrawalBankCardList', 'WithdrawalBankCard']]
    per_page = 30
    MCLS = WithdrawalCardTable
    template = 'cms/behalfPay/withdrawBankCard.html'
    title = '出款银行卡'

    def get_filter_dict(self):
        fff = {"is_deleted": {"$ne": True}}

        if self.current_admin_dict.get('role_code') == ROlE_ALL.SYS_OUT_MONEY_USER:
            admin_uuid = CmsUserTable.find_one({"role_code":ROlE_ALL.SUPERADMIN}).get("uuid")
            fff["admin_uuid"] = admin_uuid 
            fff["location"] = LOCATION_TYPE.WITHDRAW 
            fff["is_Disabled"] =  {"$ne": True}
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.OUT_MONEY_USER:
            fff["location"] = LOCATION_TYPE.WITHDRAW 
            fff["admin_uuid"] = self.current_admin_dict.get("agentadmin_uuid")  
            fff["is_Disabled"] =  {"$ne": True}
        elif self.current_admin_dict.get('role_code') in [ROlE_ALL.SUPERADMIN]:
            admin_uuid = CmsUserTable.find_one({"role_code":ROlE_ALL.SUPERADMIN}).get("uuid")
            fff["admin_uuid"] = admin_uuid
            self.search_dict["admin_uuid"] = admin_uuid
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff["admin_uuid"] = self.current_admin_dict.get("uuid")  
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            fff["admin_uuid"] = self.current_admin_dict.get("agentadmin_uuid")  
            fff["payer_uuids"] = self.current_admin_dict.get("uuid") 
            #fff["$or"] =  [{"location": LOCATION_TYPE.WITHDRAW },  { "payer_uuids": self.current_admin_dict.get("uuid") }]
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.ADMINISTRATOR:
            fff["admin_uuid"] = CmsUserTable.find_one({"role_code":ROlE_ALL.SUPERADMIN}).get("uuid")
            fff["payer_uuids"] = self.current_admin_dict.get("uuid")
            #fff["$or"] =  [{"location": LOCATION_TYPE.WITHDRAW },  { "payer_uuids": self.current_admin_dict.get("uuid") }]

        location = request.args.get('location')
        bank_name = request.args.get('bank_name')
        account = request.args.get('account')
        bank_uid = request.args.get('bank_uid')
        note = request.args.get('note')

        if "admin_uuid" in request.args:
            admin_uuid = request.args.get('admin_uuid')
            fff['admin_uuid'] = admin_uuid
            self.search_dict["admin_uuid"] = admin_uuid
            if not admin_uuid:
                fff.pop("admin_uuid")
                self.search_dict.pop("admin_uuid")

        if location:
            fff['location'] = location
        if bank_name:
            fff['bank_name'] = bank_name
        if account:
            fff['account'] = account
        if bank_uid:
            fff['bank_uid'] = bank_uid
        if note:
            fff["note"] = { "$regex": note, "$options": "i" }
        return fff
    
    def dealwith_main_context(self):
        datas = []
        datas1 = []
        datas2 = []
        datas3 = []
        datas4 = []
        owner_datas = []
        bank_datas = BankTable.find_many({})
        admin_uuid = ''

        all_datas = self.context.get('all_datas') or []
        for da in all_datas:
            self.get_deposit_cardinfo(da)
            _dd = BankTable.find_one({'uuid': da.get('bank_uid')})
            da['bank_data'] = _dd
            da['owner'] = CmsUserTable.find_one({'uuid': da["admin_uuid"]})
           
            if self.current_admin_dict.get('role_code') in [ROlE_ALL.SUPERADMIN ,ROlE_ALL.AGENTADMIN, ROlE_ALL.ADMINISTRATOR, ROlE_ALL.SYSTEMUSER]:
                if self.current_admin_dict.get('role_code') == ROlE_ALL.ADMINISTRATOR:
                    admin_uuid = CmsUserTable.find_one({'role_code': ROlE_ALL.SUPERADMIN}).get("uuid")
                elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
                    admin_uuid = self.current_admin_dict.get("agentadmin_uuid")
                else:
                    admin_uuid = self.current_admin_dict["uuid"]
                
                self.reset_statistic_agent(da)
                
                # da["withdrawlogs"] = []
                xx = WithdrawalOrderLogTable.find_one({"former_card_uuid": da.get("uuid"),"admin_uuid": admin_uuid,"order_status": ORDER_STATUS.ORDERED}, sort=[("order_time", -1)])
                if xx:
                    da["new_request"] = True
                da["withdraw_money"] = 0
                if xx:
                    da["withdraw_money"] = xx.get("request_money")
                
                # da["withdrawlogs"] = WithdrawalOrderLogTable.find_many({"former_card_uuid": da.get("uuid"),"admin_uuid": self.current_admin_dict["uuid"] }, limit = 10, sort=[("order_time", -1)])
                
                da["sub_admins"] = []
                for uuid in (da.get("payer_uuids") or []):
                    da["sub_admins"].append(CmsUserTable.find_one({"uuid": uuid}).get("account"))
                
                da["outpayer_account"] = CmsUserTable.find_one({"withdrawalcard_uuid": da.get("uuid")}).get("account")

            if self.current_admin_dict.get('role_code') == ROlE_ALL.OUT_MONEY_USER or self.current_admin_dict.get("role_code") == ROlE_ALL.SYS_OUT_MONEY_USER:
                # da["withdrawlogs"] = WithdrawalOrderLogTable.find_many({"former_card_uuid": da.get("uuid"), "accepted_status": {"$in": ACCEPTED_STATUS.name_arr} }, limit = 10, sort=[("create_time", -1)])
                # da["withdrawlogs"] = []

                if da.get("uuid") == self.current_admin_dict.get("withdrawalcard_uuid"):
                    da["isSelected"] = True

                payers = []
                if self.current_admin_dict.get('role_code') == ROlE_ALL.OUT_MONEY_USER:
                    payers = CmsUserTable.find_many({"agentadmin_uuid": self.current_admin_dict.get("agentadmin_uuid")})
                elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYS_OUT_MONEY_USER:
                    payers = CmsUserTable.find_many({"role_code": ROlE_ALL.SYS_OUT_MONEY_USER})
                
                if da.get("uuid") in [x.get("withdrawalcard_uuid") for x in payers]:
                    if da.get("uuid") != self.current_admin_dict.get("withdrawalcard_uuid"):
                        da["isDisabled"] = True
                if not da.get("isDisabled"):
                    self.get_statistic_info(da)

            
            da['outpayer'] = CmsUserTable.find_one({'withdrawalcard_uuid': da.get("uuid")})
            datas.append(da)
            # if da["withdrawlogs"]:
            # elif da["location"] == LOCATION_TYPE.WITHDRAW:
            #     datas1.append(da)
            # elif da["location"] == LOCATION_TYPE.TRANSIT:
            #     datas2.append(da)
            # elif da["location"] == LOCATION_TYPE.DEPOSIT:
            #     datas3.append(da)
            # elif da["location"] == LOCATION_TYPE.OTHER:
            #     _dd = BankTable.find_one({'uuid': da.get('bank_uid')})
            #     da['bank_data'] = _dd
            #     datas4.append(da)
        
        owner_datas.append(CmsUserTable.find_one({"role_code":ROlE_ALL.SUPERADMIN}))
        agents = CmsUserTable.find_many({"role_code":ROlE_ALL.AGENTADMIN})
        for agent in agents:
            owner_datas.append(agent)

        if self.current_admin_dict.get("is_addcard"):
            self.context['is_addcard'] = True
        if self.current_admin_dict.get("is_clear_deposit"):
            self.context['is_clear_deposit'] = True
        if self.current_admin_dict.get("is_clear_withdraw"):
            self.context['is_clear_withdraw'] = True
        if self.current_admin_dict.get("is_clear_transit"):
            self.context['is_clear_transit'] = True
        if self.current_admin_dict.get("is_clear_other"):
            self.context['is_clear_other'] = True
        self.context['all_datas'] = datas# + datas1 + datas2 + datas3 + datas4
        self.context['bank_datas'] = bank_datas
        self.context['owner_datas'] = owner_datas
        self.context['ACCEPTED_STATUS'] = ACCEPTED_STATUS
        self.context['ORDER_STATUS'] = ORDER_STATUS
        self.context['BEHAVIOR_TYPE'] = BEHAVIOR_TYPE
        self.context['LOCATION_TYPE'] = LOCATION_TYPE
        self.context['admin_uuid'] = admin_uuid

    def check_current_date(self, card_uuid):
        card = WithdrawalCardTable.find_one({"uuid": card_uuid})
        cur_date = datetime.datetime.now().date().strftime("%Y-%m-%d")
        if card.get("current_date") == cur_date:
            return
        
        data_from = {
            "current_date": cur_date,
            "total_request" : 0,
            "total_transfer_in" : 0,
            "total_transfer_out" : 0,
            "total_other_transfer_in" : 0,
            "total_other_transfer_out" : 0,
            "total_issue" : 0,
        }
        WithdrawalCardTable.update_one({"uuid": card_uuid}, {'$set': data_from})

    def reset_statistic_agent(self, dd):
        cur_date = datetime.datetime.now().date().strftime("%Y-%m-%d")
        if dd.get("current_date") == cur_date:
            return
        
        data_from = {
            "current_date": cur_date,
            "total_request" : 0,
            "total_transfer_in" : 0,
            "total_transfer_out" : 0,
            "total_other_transfer_in" : 0,
            "total_other_transfer_out" : 0,
            "total_issue" : 0,
        }

        dd.update(data_from)
        WithdrawalCardTable.update_one({"uuid": dd.get("uuid")}, {'$set': data_from})

    def update_transfer_amount(self, card_uuid, dd):
        self.check_current_date(card_uuid)
        card = WithdrawalCardTable.find_one({"uuid": card_uuid})

        data_from = {}
        if "total_request" in dd:
            data_from["total_request"] = (card.get("total_request") or 0) + (dd.get("total_request") or 0)
        if "total_transfer_in" in dd:
            data_from["total_transfer_in"] = (card.get("total_transfer_in") or 0) + (dd.get("total_transfer_in") or 0)
        if "total_transfer_out" in dd:
            data_from["total_transfer_out"] = (card.get("total_transfer_out") or 0) + (dd.get("total_transfer_out") or 0)
        if "total_other_transfer_in" in dd:
            data_from["total_other_transfer_in"] = (card.get("total_other_transfer_in") or 0) + (dd.get("total_other_transfer_in") or 0)
        if "total_other_transfer_out" in dd:
            data_from["total_other_transfer_out"] = (card.get("total_other_transfer_out") or 0) + (dd.get("total_other_transfer_out") or 0)
        if "total_issue" in dd:
            data_from["total_issue"] = (card.get("total_issue") or 0) + (dd.get("total_issue") or 0),

        WithdrawalCardTable.update_one({"uuid": card_uuid}, {'$set': data_from})

    def get_statistic_info(self, dd):
        # result = list(WithdrawalOrderLogTable.collection().aggregate([
        #     {'$match': 
        #         {"former_card_uuid": dd.get("uuid"), "accepted_status": {"$in": [ACCEPTED_STATUS.NOT_PROCESSED, ACCEPTED_STATUS.ACCEPTED]} }
        #     },
        #     {
        #         '$group': {
        #             '_id': None,
        #             'request_total': {'$sum': f"${'request_money'}"},
        #         }
        #     }
        # ]))
        # dd["request_money"] = result[0].get("received_total") if result else 0
        
        # result = list(WithdrawalOrderLogTable.collection().aggregate([
        #     {'$match': 
        #         {"former_card_uuid": dd.get("uuid"),"behavior": BEHAVIOR_TYPE.TRANSFER_IN ,"accepted_status": {"$in": [ACCEPTED_STATUS.NOT_PROCESSED, ACCEPTED_STATUS.ACCEPTED]} }
        #     },
        #     { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
        # ]))
        # dd["transfer_in_money"] = result[0].get("transfer_total") if result else 0

        # result = list(WithdrawalOrderLogTable.collection().aggregate([
        #     {'$match': 
        #         {"former_card_uuid": dd.get("uuid"),"behavior": BEHAVIOR_TYPE.TRANSFER_OUT ,"accepted_status": {"$in": [ACCEPTED_STATUS.NOT_PROCESSED, ACCEPTED_STATUS.ACCEPTED]} }
        #     },
        #     { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
        # ]))
        # dd["transfer_out_money"] = result[0].get("transfer_total") if result else 0

        # result = list(WithdrawalOrderLogTable.collection().aggregate([
        #     {'$match': 
        #         {"former_card_uuid": dd.get("uuid"),"behavior": BEHAVIOR_TYPE.OTHER_TRANSFER_IN ,"accepted_status": {"$in": [ACCEPTED_STATUS.NOT_PROCESSED, ACCEPTED_STATUS.ACCEPTED]} }
        #     },
        #     { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
        # ]))
        # dd["other_transfer_in_money"] = result[0].get("transfer_total") if result else 0

        # result = list(WithdrawalOrderLogTable.collection().aggregate([
        #     {'$match': 
        #         {"former_card_uuid": dd.get("uuid"),"behavior": BEHAVIOR_TYPE.OTHER_TRANSFER_OUT ,"accepted_status": {"$in": [ACCEPTED_STATUS.NOT_PROCESSED, ACCEPTED_STATUS.ACCEPTED]} }
        #     },
        #     { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
        # ]))
        # dd["other_transfer_out_money"] = result[0].get("transfer_total") if result else 0

        # result = list(WithdrawalOrderLogTable.collection().aggregate([
        #     {'$match': 
        #         {"former_card_uuid": dd.get("uuid"),"behavior": BEHAVIOR_TYPE.ISSUED ,"accepted_status": {"$in": [ACCEPTED_STATUS.NOT_PROCESSED, ACCEPTED_STATUS.ACCEPTED]} }
        #     },
        #     { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
        # ]))
        # dd["issued_money"] = result[0].get("transfer_total") if result else 0
        
        # dd["plus_money"] = dd["transfer_in_money"]+dd.get("other_transfer_in_money")
        # dd["minus_money"] = dd.get("transfer_out_money")+dd.get("other_transfer_out_money")+dd.get("issued_money")

        # dd["error"] = dd.get("start_money") + dd["transfer_in_money"]+dd.get("other_transfer_in_money") - dd.get("transfer_out_money")-dd.get("other_transfer_out_money")-dd.get("issued_money")-dd.get("balance_amount")

        # return
        dd["plus_money"] =(dd.get("total_transfer_in") or 0) + (dd.get("total_other_transfer_in") or 0)
        dd["minus_money"] = (dd.get("total_transfer_out") or 0) + (dd.get("total_other_transfer_out") or 0) + (dd.get("total_issue") or 0)

        dd["error"] = dd["plus_money"] - dd["minus_money"]

        return
    
    def get_deposit_cardinfo(self, dd):
        if dd.get('location') != LOCATION_TYPE.ANOTHER:
            return
        _bankcard = BankCardTable.find_one({'uuid': dd.get('deposit_uuid')})
        dd['account'] = _bankcard['account']
        dd['bank_name'] = _bankcard['name']
        dd['balance_amount'] = _bankcard['balance_amount']
        dd['bank_uid'] = _bankcard.get("bank_uid")
        dd['code'] = _bankcard.get("code")
        dd['balance_amount'] = _bankcard.get("balance_amount")
        dd['start_money'] = _bankcard.get("start_money")
        dd['note'] = _bankcard.get("note")
   
    def update_cardinfo(self, dd, data_from):
        if dd.get('location') != LOCATION_TYPE.ANOTHER:
            WithdrawalCardTable.update_one({ "uuid": dd.get("uuid")},{ "$set": data_from })
            return
        BankCardTable.update_one({'uuid': dd.get('deposit_uuid')},{ "$set": data_from })

    def choose_bankcard_html(self):
        ff = {"location": LOCATION_TYPE.WITHDRAW, "payer_uuid": {"$ne": self.current_admin_dict.get("uuid")}}
        if self.current_admin_dict.get('role_code') == ROlE_ALL.OUT_MONEY_USER:
            ff["admin_uuid"] = self.current_admin_dict["agentadmin_uuid"]
        if self.current_admin_dict.get('role_code') == ROlE_ALL.SYS_OUT_MONEY_USER:
            ff["admin_uuid"] = CmsUserTable.find_one({"role_code": ROlE_ALL.SUPERADMIN}).get("uuid")
        withdrawalcards = WithdrawalCardTable.find_many(ff)
        card_cnt = len(withdrawalcards)
        bankcard_html = ''
        for withdrawalcard in withdrawalcards:
            bank_info = BankTable.find_one({'uuid': withdrawalcard.get('bank_uid')})
            owner_info = CmsUserTable.find_one({"uuid": withdrawalcard.get("admin_uuid")})
            bankcard_html += f'''
            <div class="card mb-3" bankcarduuid="{ withdrawalcard.get("uuid") }" onclick="seleBankCard_func($(this))"  >
                <div class="row g-0" style = "cursor: pointer;">
                    <div class="col-md-4 text-center" style="display: flex; align-items: stretch; justify-content: center; margin: 0px; ">
                        <div style=" flex: 1; background-color: #ebeaea; display: flex; align-items: center; justify-content: center; ">
                            <div>
                                <h1>{bank_info.get("code")}</h1>
                                <h6>(代理账号: {owner_info.get("account")})</h6>
                            </div>
                        </div>
                    </div>
                    <div class="col-md-8">
                        <div class="card-body" style="padding:10px">
                            <h5 class="card-title mb-2">银行卡账号: {withdrawalcard.get("account")} (<span style="font-size:14px">银行名字: {withdrawalcard.get("bank_name")})<span></h5>
                            <p class="card-text mb-1">初始余额: {self.format_money(withdrawalcard.get("start_money"))}, 实时余额: {self.format_money(withdrawalcard.get("balance_amount"))}</p>
                            <p class="card-text"><small class="text-body-secondary">注解: {withdrawalcard.get("note")}</small></p>
                        </div>
                    </div>
                </div>
            </div>
            '''

        html = f'''
            <h4 class="mb-0 text-left" style="margin-left:40px">可用出款卡数：{card_cnt}</h4>
            <div class="formBox m-0" style="padding:0px;">
                <div class="list-group-item" style="display: flex; align-items: center;justify-content: center;">
                    <input type="hidden" id="bank_card_uid">
                    <div class="form-control p-3" style="display: inline-block; width: calc(100% - 40px); height: 400px; overflow-y: scroll; text-align: left;">
                        { bankcard_html }                        
                    </div>
                </div>    
            <div>   
            <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

            <div style="position: relative; text-align: center">
                <span class="btn btn-primary" onclick="post_userbankcard_data('add_card_to_payer')">确定</span>&emsp;
                <span class="btn btn-default" onclick="xtalert.close()">取消</span>
            </div>            
        '''
        return html
    
    def request_money_html(self, data_dict={}, behavior = ''):
        bankcard_opt = ''
        bank_datas = BankTable.find_many({})
        for bdc in bank_datas:
            bankcard_opt += f'''<div class="auto-complete-item" onclick="onclick_insert_code('{bdc.get("code")}')">{ bdc.get("code") }</div>'''

        html = f'''
        <div class="formBox">
            <div style="position: relative; box-sizing: border-box; overflow-y: auto;">    
                <div class="list-group-item">
                    <span style="width: 100px; text-align: right; display: inline-block; position: relative;"></span>
                    <div style = "display: inline-block; width: calc(100% - 120px)">
                        <div class="d-flex justify-content-between px-3" >
                            <span class="ant-tag to_cursor ant-tag-blue-inverse mr-0" onclick="setRequestMoney('10,000,000')">10</span>
                            <span class="ant-tag to_cursor ant-tag-blue-inverse mr-0" onclick="setRequestMoney('30,000,000')">30</span>
                            <span class="ant-tag to_cursor ant-tag-blue-inverse mr-0" onclick="setRequestMoney('50,000,000')">50</span>
                            <span class="ant-tag to_cursor ant-tag-blue-inverse mr-0" onclick="setRequestMoney('70,000,000')">70</span>
                            <span class="ant-tag to_cursor ant-tag-blue-inverse mr-0" onclick="setRequestMoney('100,000,000')">100</span>
                            <span class="ant-tag to_cursor ant-tag-blue-inverse mr-0" onclick="setRequestMoney('150,000,000')">150</span>
                            <span class="ant-tag to_cursor ant-tag-blue-inverse mr-0" onclick="setRequestMoney('200,000,000')">200</span>
                        </div>
                    </div>
                </div>
                <div class="list-group-item">
                    <span class="loglable" style="width: 100px; text-align: right; display: inline-block; position: relative;">请求金额:</span>
                    <input type="text" class="form-control" id="transfer_money" value="" placeholder="请求金额" aria-label="" style="display: inline-block; width: calc(100% - 120px)" onchange="onchange_number_input('transfer_money')">
                </div>    
            '''
        if behavior != 'request_money':
            html +=f'''      
                <div class="list-group-item">
                    <span class="loglable" style="width: 100px; text-align: right; display: inline-block; position: relative;">卡银行卡账号:</span>
                    <input type="text" class="form-control" id="bankcard_account" value="" placeholder="卡银行卡账号" aria-label="" style="display: inline-block; width: calc(100% - 120px)">
                </div>                                        
                 <div class="list-group-item">
                    <span class="loglable" style="width: 100px; text-align: right; display: inline-block; position: relative;">银行类型:</span>
                    <div class="auto-complete" style="display: inline-block; width: calc(100% - 120px)">
                        <input type="text" class="form-control" id="bank_code" value="" placeholder="银行类型" aria-label="" onclick="onclick_bankcode()" oninput="oninput_bankcode()">
                        <div class="auto-complete-items">
                        {bankcard_opt}
                        </div>
                    </div>
                 </div>
                 <div class="list-group-item">
                    <span class="loglable" style="width: 100px; text-align: right; display: inline-block; position: relative;">银行名字:</span>
                    <input type="text" class="form-control" id="bank_name" value="" placeholder="银行名字" aria-label="" style="display: inline-block; width: calc(100% - 120px)">
                </div>   
            '''
        qr_html = ''
        if behavior != 'request_money':
            qr_html = f'''<span class="btn btn-secondary mr-5" onclick="post_qr_code2('{data_dict.get('uuid')}', '{behavior}', '{url_for('admin.WithdrawalBankCard')}')">支付码</span>&emsp;'''
        
        html +=f'''      
                                     
                <div class="list-group-item">
                    <span style="width: 100px; text-align: right; display: inline-block; position: relative;">注解：</span>
                    <input type="text" class="form-control" id="note" value="{self.current_admin_dict.get("last_note") or ''}" placeholder="注解" aria-label="" style="display: inline-block; width: calc(100% - 120px)">
                </div>                                        
            </div>

            <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

            <div style="position: relative; text-align: center">
                {qr_html}
                <span class="btn btn-primary" onclick="post_input_money('{behavior}', '{ data_dict.get('uuid') or ''}')">确定</span>&emsp;
                <span class="btn btn-default" onclick="closeQRButton()">取消</span>
            </div>                                                                                 
        </div>
        '''
        return html

    def bankcard_html(self, data_dict={}):
        _action = 'add_bankcard'
        balance_amount_html = ''
        if data_dict:
            _action = 'edit_bankcard'
            balance_amount_html = f'''
            <div class="list-group-item">
                <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">实时余额：</span>
                <input type="text" class="form-control" id="balance_amount" value="{ self.format_money(data_dict.get('balance_amount') )or '' }" placeholder="初始余额" aria-label="" style="display: inline-block; width: calc(100% - 180px)" onchange="onchange_number_input('balance_amount')">
            </div>          
            '''

        bddl_html = ''
        not_uuids = [x.get("deposit_uuid") for x in WithdrawalCardTable.find_many({"location": LOCATION_TYPE.OTHER})]
        BankCardTable.find_many({})
        for uda in BankCardTable.find_many({"uuid": {"$nin": not_uuids}}):
            selected = 'selected' if uda.get('uuid') == data_dict.get('deposit_uuid') else ''
            bddl_html += f'<option value="{uda.get("uuid")}" {selected} >{uda.get("account")}-{uda.get("username")}</option>'

        selbankcard_html = f'''
        <div class="list-group-item">
            <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">选择卡：</span>
            <select class="form-control" id="deposit_bankcard" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                <option value="">选择卡</option>
                { bddl_html }
            </select>
        </div>         
        '''


        bankcard_opt = ''
        bank_datas = BankTable.find_many({})
        for bdc in bank_datas:
            bankcard_opt += f'''<div class="auto-complete-item" onclick="onclick_insert_code('{bdc.get("code")}')">{ bdc.get("code") }</div>'''

        bank_info = BankTable.find_one({'uuid': data_dict.get('bank_uid')})
        bank_htmll = f'''    
                <div class="list-group-item" style="display: flex; align-items: center;justify-content: center;">
                    <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行类型：</span>
                    <div class="auto-complete" style="display: inline-block; width: calc(100% - 180px);">
                        <input type="text" class="form-control" id="bank_code" value="{bank_info.get("code") or ''}" placeholder="银行类型" aria-label="" onclick="onclick_bankcode()" oninput="oninput_bankcode()" >
                        <div class="auto-complete-items">
                            {bankcard_opt}
                        </div>
                    </div>
                </div>                         
        '''

        html = f'''
            <div class="formBox">
                <div style="height: 24rem; position: relative; box-sizing: border-box; overflow-y: auto;">    
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">位置：</span>
                        <select class="form-control" id="location" aria-label="" style="display: inline-block; width: calc(100% - 180px)" onchange = 'onChangeLocation()'>
                            <option value="">选择位置</option>
                            <option value="{LOCATION_TYPE.TRANSIT}" {'selected' if data_dict.get('location') == LOCATION_TYPE.TRANSIT else ''}>{LOCATION_TYPE.name_dict[LOCATION_TYPE.TRANSIT]}</option>
                            <option value="{LOCATION_TYPE.WITHDRAW}" {'selected' if data_dict.get('location') == LOCATION_TYPE.WITHDRAW else ''}>{LOCATION_TYPE.name_dict[LOCATION_TYPE.WITHDRAW]}</option>
                            <option value="{LOCATION_TYPE.DEPOSIT}" {'selected' if data_dict.get('location') == LOCATION_TYPE.DEPOSIT else ''}>{LOCATION_TYPE.name_dict[LOCATION_TYPE.DEPOSIT]}</option>
                            <option value="{LOCATION_TYPE.OTHER}" {'selected' if data_dict.get('location') == LOCATION_TYPE.OTHER else ''}>{LOCATION_TYPE.name_dict[LOCATION_TYPE.OTHER]}</option>
                        </select>
                    </div> 
        '''
        if data_dict.get('location') == LOCATION_TYPE.ANOTHER:
            html += f'''        
                    <div id = "transit_withdraw" style="display:none"> '''
        else:
            html += f'''
                    <div id = "transit_withdraw">'''
        html += f'''        
                        <div class="list-group-item">
                            <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行卡账户：</span>
                            <input type="text" class="form-control" id="account" value="{ data_dict.get('account') or '' }" placeholder="银行卡账户" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                        </div>
                        <div class="list-group-item">
                            <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行名字：</span>
                            <input type="text" class="form-control" id="bank_name" value="{ data_dict.get('bank_name') or '' }" placeholder="银行名字" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                        </div>     
                        { bank_htmll }                   
                        <div class="list-group-item">
                            <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">初始余额：</span>
                            <input type="text" class="form-control" id="start_money" value="{ self.format_money(data_dict.get('start_money') )or '' }" placeholder="初始余额" aria-label="" style="display: inline-block; width: calc(100% - 180px)" onchange="onchange_number_input('start_money')">
                        </div>    
                        {balance_amount_html}      
                        <div class="list-group-item">
                            <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">注解：</span>
                            <input type="text" class="form-control" id="note" value="{ data_dict.get('note') or '' }" placeholder="注解" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                        </div>     
                    </div>
            '''
        if data_dict.get('location') == LOCATION_TYPE.ANOTHER:
            html += f'''
                    <div id = "other_card"> '''
        else:
            html += f'''
                    <div id = "other_card" style="display:none"> '''
        html += f'''
                    {selbankcard_html}
                    </div>
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_bankcard_data('{_action}', '{ data_dict.get('uuid') if data_dict else '' }')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>
        '''
        return html

    def transfer_in_out_html(self, data_dict={}, behavior = ''):
        # ff = {"location": LOCATION_TYPE.TRANSIT}
        ff = {'uuid': {"$ne": data_dict.get("uuid")}}
        if self.current_admin_dict.get("role_code") in [ROlE_ALL.SUPERADMIN, ROlE_ALL.AGENTADMIN]:
            ff["admin_uuid"] = self.current_admin_dict["uuid"]
        if self.current_admin_dict.get("role_code") == ROlE_ALL.SYS_OUT_MONEY_USER:
            ff["admin_uuid"] = CmsUserTable.find_one({"role_code": ROlE_ALL.SUPERADMIN}).get("uuid")
        if self.current_admin_dict.get("role_code") == ROlE_ALL.OUT_MONEY_USER:
            ff["admin_uuid"] = CmsUserTable.find_one({"uuid": self.current_admin_dict.get("agentadmin_uuid")}).get("uuid")
        
        fff = {"is_deleted": {"$ne": True}, "is_Disabled": {"$ne": True}}
        if self.current_admin_dict.get('role_code') == ROlE_ALL.SYS_OUT_MONEY_USER:
            admin_uuid = CmsUserTable.find_one({"role_code":ROlE_ALL.SUPERADMIN}).get("uuid")
            fff["admin_uuid"] = admin_uuid 
            fff["location"] = LOCATION_TYPE.WITHDRAW 
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.OUT_MONEY_USER:
            fff["location"] = LOCATION_TYPE.WITHDRAW 
            fff["admin_uuid"] = self.current_admin_dict.get("agentadmin_uuid")  
        elif self.current_admin_dict.get('role_code') in [ROlE_ALL.SUPERADMIN]:
            admin_uuid = CmsUserTable.find_one({"role_code":ROlE_ALL.SUPERADMIN}).get("uuid")
            fff["admin_uuid"] = admin_uuid
            self.search_dict["admin_uuid"] = admin_uuid
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff["admin_uuid"] = self.current_admin_dict.get("uuid")  
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            fff["admin_uuid"] = self.current_admin_dict.get("agentadmin_uuid")  
            fff["$or"] =  [{"location": LOCATION_TYPE.WITHDRAW },  { "payer_uuids": self.current_admin_dict.get("uuid") }]
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.ADMINISTRATOR:
            fff["admin_uuid"] = CmsUserTable.find_one({"role_code":ROlE_ALL.SUPERADMIN}).get("uuid")
            fff["$or"] =  [{"location": LOCATION_TYPE.WITHDRAW },  { "payer_uuids": self.current_admin_dict.get("uuid") }]

        ff.update(fff)
        cards = WithdrawalCardTable.find_many(ff, sort=[['location', -1]])
        card_cnt = len(cards)
        bankcard_html = ''
        for card in cards:
            self.get_deposit_cardinfo(card)

            bank_info = BankTable.find_one({'uuid': card.get('bank_uid')})
            bankcard_html += f'''
            <div class="card mb-2" bankcarduuid="{ card.get("uuid") }" onclick="seleBankCard_func($(this))" location="{card.get("location")}" bank_code="{bank_info.get("code")}" bank_name="{card.get("bank_name")}" account="{card.get("account")}" locationName={LOCATION_TYPE.name_dict[card.get("location")]}>
                <div class="row g-0 m-0" style = "cursor: pointer;">
                    <div class="col-3 text-center" style="display: flex; align-items: stretch; justify-content: center; padding: 0px; ">
                        <div style=" flex: 1; background-color: #ebeaea; display: flex; align-items: center; justify-content: center; ">
                            <h5 style="margin:0px">{bank_info.get("code")}({LOCATION_TYPE.name_dict[card.get("location")]})</h5>
                        </div>
                    </div>
                    <div class="col-9">
                        <div class="card-body" style="padding:10px">
                            银行卡账号: {card.get("account")} (名字: {card.get("bank_name")}, 余额: {self.format_money(card.get("balance_amount"))})
                        </div>
                    </div>
                </div>
            </div>
            '''
        cards_html = f'''
        <option value="{LOCATION_TYPE.TRANSIT}"> {LOCATION_TYPE.name_dict[LOCATION_TYPE.TRANSIT]}</option>
        <option value="{LOCATION_TYPE.WITHDRAW}" >{LOCATION_TYPE.name_dict[LOCATION_TYPE.WITHDRAW]}</option>
        <option value="{LOCATION_TYPE.DEPOSIT}" >{LOCATION_TYPE.name_dict[LOCATION_TYPE.DEPOSIT]}</option>
        <option value="{LOCATION_TYPE.OTHER}" >{LOCATION_TYPE.name_dict[LOCATION_TYPE.OTHER]}</option>
        '''
        bankcard_opt = ''
        bank_datas = BankTable.find_many({})
        for bdc in bank_datas:
            bankcard_opt += f'''<div class="auto-complete-item" onclick="onclick_insert_code('{bdc.get("code")}', filter = true)">{ bdc.get("code") }</div>'''

        html = f'''
            <div class="d-flex justify-content-between" style="margin-bottom:10px">
               <div style="margin-left:20px; ">
                    <span style="width: 80px; text-align: right; display: inline-block; position: relative;">转账金额:</span>
                    <input type="text" class="form-control" id="transfer_money" placeholder="转账金额" value="0" aria-label="" style="display: inline-block; width: calc(100% - 100px)" onchange = "onchange_number_input('transfer_money')">
                </div>
                <div style="width:250px" >
                    <span style="width: 40px; text-align: right; display: inline-block; position: relative;">备注:</span>
                    <input type="input" class="form-control" id="note" placeholder="备注" aria-label="" style="display: inline-block; width: calc(100% - 60px)">
                </div>
                <button class="btn btn-primary" style="margin-right:40px;" onclick = "addtransfercardlist()">将卡添加到列表</button>
            </div>
            <div class="d-flex justify-content-between" style="margin-bottom:3px; ">
                <select class="form-control mr-sm-2 mb-2" style="margin-left:40px; width: 200px;" aria-label="" id="location_filter" onchange="on_card_filter_change()">
                    <option value="">选择位置</option>
                   {cards_html}
                </select>
                <div class="auto-complete mr-sm-2 mb-2" >
                    <input type="text" class="form-control" id="bank_code" value="" placeholder="银行类型" aria-label="" onclick="onclick_bankcode()" oninput="on_card_filter_change()" >
                    <div class="auto-complete-items">
                    {bankcard_opt}
                    </div>
                </div>

                <input type="text" class="form-control mr-sm-2 mb-2" id="bankname_filter" placeholder="银行名字" aria-label="" style="display: inline-block; width: 200px; margin-right: 40px;" oninput="on_card_filter_change()" onchange="on_card_filter_change()">
                <input type="text" class="form-control" id="account_filter" placeholder="银行卡账号" aria-label="" style="display: inline-block; width: 200px; margin-right: 40px;" oninput="on_card_filter_change()" onchange="on_card_filter_change()">
            </div>
            <div class="formBox m-0" style="padding:0px;">
                <div class="list-group-item" style="display: flex; align-items: center;justify-content: center; padding-bottom:0px">
                    <input type="hidden" id="bank_card_uid">
                    <div class="form-control p-2" style="display: inline-block; width: calc(100% - 40px); height: 300px; overflow-y: scroll; text-align: left;" id='bankcard_list'>
                        { bankcard_html }                        
                    </div>
                </div>    
            <div>   
             <div class="formBox m-0" style="padding:0px;">
                <div class="list-group-item" style="display: flex; align-items: center;justify-content: center;">
                    <input type="hidden" id="bank_card_uid">
                    <div class="form-control p-2" style="display: inline-block; width: calc(100% - 40px); height: 110px; overflow-y: scroll; text-align: left;" id='transfer_card_list'>
                                                
                    </div>
                </div>    
            <div>  
            <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>
            <div class="d-flex justify-content-between px-5">
                <div>
                    <span>转账卡数：</span>
                    <span id="transfer_card_cnt">0</span>
                </div>
                <div style="position: relative; text-align: center">
                    <span class="btn btn-secondary mr-5" onclick="post_qr_code('{data_dict.get('uuid')}', '{behavior}', '{url_for('admin.WithdrawalBankCard')}')">支付码</span>&emsp;
                    <span class="btn btn-primary" onclick="post_transfer_in_out('null', 'transfer_in_out', '{data_dict.get('uuid')}', '{behavior}')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="closeQRButton()">取消</span>
                </div>            
            </div>            
        '''
        return html

    def select_subadmin_html(self, data_dict = {}):
        sub_admin_html = ''
        total = 0
        checkallstat = 'checked'
        if self.current_admin_dict.get("role_code") == ROlE_ALL.SUPERADMIN:
            total = CmsUserTable.count({"role_code": ROlE_ALL.ADMINISTRATOR})
            for data in CmsUserTable.find_many({"role_code": ROlE_ALL.ADMINISTRATOR}):
                checkstate = "checked" if data.get("uuid") in (data_dict.get("payer_uuids") or []) else ""
                if data.get("uuid") not in (data_dict.get("payer_uuids") or []):
                    checkallstat = ''

                sub_admin_html += f'''
                <tr class="sub-admin-info" id="{data.get("account")}">
                    <td>
                        <input type="checkbox" class="checkbox" onchange="onSelectChange()" id={data.get("uuid")} {checkstate}>
                    </td>
                    <td>{ data.get("account") or '' }</td>
                    <td>{ data.get("username") or '' }</td>
                </tr>
                '''
        elif self.current_admin_dict.get("role_code") == ROlE_ALL.AGENTADMIN:
            total = CmsUserTable.count({"role_code": ROlE_ALL.SYSTEMUSER, "agentadmin_uuid": self.current_admin_dict.get("uuid")})
            for data in CmsUserTable.find_many({"agentadmin_uuid": self.current_admin_dict.get("uuid"), "role_code": ROlE_ALL.SYSTEMUSER}):
                checkstate = "checked" if data.get("uuid") in (data_dict.get("payer_uuids") or []) else ""
                if data.get("uuid") not in (data_dict.get("payer_uuids") or []):
                    checkallstat = ''

                sub_admin_html += f'''
                <tr class="sub-admin-info" id="{data.get("account")}">
                    <td>
                        <input type="checkbox" class="checkbox" onchange="onSelectChange()" id={data.get("uuid")} {checkstate}>
                    </td>
                    <td>{ data.get("account") or '' }</td>
                    <td>{ data.get("username") or '' }</td>
                </tr>
                '''
        html = f'''
        <div style="display: inline-block; width: calc(100% - 40px); height: 400px; overflow-y: scroll; text-align: left;">
            <table class="table table-bordered table-hover text-center" style="background-color: #ffffff;">
                <thead>
                    <tr>
                        <td>
                            <input type="checkbox" id="selectAll" onchange="onSelectAllChange()" {checkallstat}>
                        </td>
                        <th>帐户 [{total or 0 }]</th>
                        <th>用户名</th>
                    </tr>
                </thead>
                <tbody>
                    {sub_admin_html}
                </tbody>
            </table>
        </div>
        <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

        <div style="position: relative; text-align: center; display: flex; align-items: center;justify-content: center;">
            <input class="form-control mr-3" placeholder="请输入帐号"  style="display: inline-block; width: 250px;" oninput="oninput_searchaccount()" id="search_account">
            <span class="btn btn-primary" onclick="post_sub_admins('{data_dict.get("uuid")}')" style="display: inline-block;">确定</span>&emsp;
            <span class="btn btn-default" onclick="xtalert.close()" style="display: inline-block;">取消</span>
        </div>            
        '''

        return html
    
    def post_other_way(self):
        if self.action == 'add_bankcard_html':
            html = self.bankcard_html()
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        if self.action == 'confirmCheck':
            checked_uuid = self.request_data.get("checked_uuid")
            data_from = {
                "withdrawalcard_uuid": checked_uuid,
            }
            CmsUserTable.update_one({ "uuid": self.current_admin_dict.get("uuid")},{ "$set": data_from })
            return self.xtjson.json_result()

        if self.action == 'add_bankcard':
            location = self.request_data.get('location')
            
            if self.current_admin_dict.get("role_code") == ROlE_ALL.SYSTEMUSER:
                admin_uuid = self.current_admin_dict.get("agentadmin_uuid")
                self.data_from.update({"payer_uuids": [self.current_admin_dict.get('uuid')]})
            elif self.current_admin_dict.get("role_code") == ROlE_ALL.ADMINISTRATOR:
                admin_uuid = CmsUserTable.find_one({"role_code": ROlE_ALL.SUPERADMIN}).get("uuid")
                self.data_from.update({"payer_uuids": [self.current_admin_dict.get('uuid')]})
            else:
                admin_uuid = self.current_admin_dict.get('uuid')
            if location == LOCATION_TYPE.ANOTHER:
                deposit_uuid = self.request_data.get('deposit_uuid')
                _dd = {
                    "location":location,
                    "deposit_uuid": deposit_uuid,
                    'admin_uuid': admin_uuid,
               }
                self.MCLS.insert_one(_dd)
                return self.xtjson.json_result()

            bank_code = self.request_data.get('bank_code')
            account = self.request_data.get('account')
            bank_name = self.request_data.get('bank_name')
            note = self.request_data.get('bank_name')
            start_money = float(self.request_data.get('start_money').replace(',','') or '0')
            note = self.request_data.get('note')
            
            balance_amount = start_money
            bankinfo = BankTable.find_one({"code": bank_code})
            if not bankinfo or not account or not bank_name or not account.strip() or not bank_name.strip():
                return self.xtjson.json_params_error('缺少数据！')

            _c = self.MCLS.find_one({"account": account,"bank_name": bank_name,'bank_uid': bankinfo.get("uuid"), "is_deleted": {"$ne": True}})
            if _c:
                return self.xtjson.json_params_error("该卡账户已存在!")

            self.data_from.update({
                'bank_uid': bankinfo.get("uuid"),
                'account': account,
                'bank_name': bank_name,
                'admin_uuid': admin_uuid,
                'location': location,
                'balance_amount': balance_amount,
                'start_money': start_money,
                'note': note,
            })
            self.MCLS.insert_one(self.data_from)
            return self.xtjson.json_result()
        if self.action == 'choose_bankcard_html':
            html = self.choose_bankcard_html()
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        if self.action == "add_card_to_payer":
            bankcard_uid = self.request_data.get('bankcarduuid')
            if not bankcard_uid:
                return self.xtjson.json_params_error('缺少数据！')

            cardinfo = self.MCLS.find_one({"uuid": bankcard_uid})
            if not cardinfo:
                return self.xtjson.json_params_error('缺少数据！')

            cards = self.MCLS.find_many({"payer_uuid": self.current_admin_dict.get('uuid')})
            if not cards:
                data_from = {
                    "withdrawalcard_uuid": bankcard_uid,
                }
                CmsUserTable.update_one({ "uuid": self.current_admin_dict.get("uuid")},{ "$set": data_from })

            card = self.MCLS.find_one({'uuid': bankcard_uid})
            _data = {
                'payer_uuid': card.get("payer_uuid"),
            }
            
            self.MCLS.update_one({'uuid': bankcard_uid}, {'$set': _data})
            return self.xtjson.json_result()
        if self.action == 'clearDeposit':
            fff = self.get_filter_dict()
            fff.update({"location": LOCATION_TYPE.DEPOSIT})
            _cards = self.MCLS.find_many(fff)
            for card in _cards:
                self.MCLS.update_one({"uuid": card.get('uuid')}, {'$set': {
                    "start_money": 0,
                    "balance_amount": 0,
                }})
            return self.xtjson.json_result()
        if self.action == 'clearWidrawal':
            fff = self.get_filter_dict()
            fff.update({"location": LOCATION_TYPE.WITHDRAW})
            _cards = self.MCLS.find_many(fff)
            for card in _cards:
                self.MCLS.update_one({"uuid": card.get('uuid')}, {'$set': {
                    "start_money": 0,
                    "balance_amount": 0,
                }})
            return self.xtjson.json_result()
        if self.action == 'clearTransit':
            fff = self.get_filter_dict()
            fff.update({"location": LOCATION_TYPE.TRANSIT})
            _cards = self.MCLS.find_many(fff)
            for card in _cards:
                self.MCLS.update_one({"uuid": card.get('uuid')}, {'$set': {
                    "start_money": 0,
                    "balance_amount": 0,
                }})
            return self.xtjson.json_result()
        if self.action == 'clearOther':
            fff = self.get_filter_dict()
            fff.update({"location": LOCATION_TYPE.OTHER})
            _cards = self.MCLS.find_many(fff)
            for card in _cards:
                self.MCLS.update_one({"uuid": card.get('uuid')}, {'$set': {
                    "start_money": 0,
                    "balance_amount": 0,
                }})
            return self.xtjson.json_result()
        if self.action == 'transit_statistic':
            fff = self.get_filter_dict()
            fff.update({"location": LOCATION_TYPE.TRANSIT})

            result = list(self.MCLS.collection().aggregate([
                {'$match': fff },
                {
                    '$group': {
                        '_id': None,
                        'balance_amount': {'$sum': f"${'balance_amount'}"},
                    }
                }
            ]))
            total = result[0].get("balance_amount") if result else 0

            html = f'''
            <h5>
                余额总额: {total}
            </h5>
            <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

            <div style="position: relative; text-align: center; display: flex; align-items: center;justify-content: center;">
                <span class="btn btn-default" onclick="xtalert.close()" style="display: inline-block;">取消</span>
            </div>  
            '''
            return self.xtjson.json_result(message = html)

    def post_data_other_way(self):
        if self.action == 'request_money':
            transfer_money = float(self.request_data.get("transfer_money").replace(",", "")) or 0
            note = self.request_data.get("note")

            bank_info = BankTable.find_one({'uuid': self.data_dict.get('bank_uid')})
            former_pre_balance = self.data_dict.get("balance_amount") or 0
            former_cur_balance = former_pre_balance
            self.update_transfer_amount(self.data_dict.get("uuid"), {"total_request": transfer_money})
            WithdrawalOrderLogTable.insert_one({
                "admin_uuid": self.data_dict.get("admin_uuid"),
                "payer_uuid": self.current_admin_dict.get("uuid"),
                
                "former_bankcard_account": self.data_dict.get("account"),
                "former_card_uuid": self.data_dict.get("uuid"),
                "former_bank_code": bank_info.get("code"),
                "former_bank_name": self.data_dict.get("bank_name"),
                "former_location": self.data_dict.get("location"),
                "former_cur_balance": former_cur_balance,
                "former_pre_balance": former_pre_balance,

                "request_money":transfer_money,
                "order_time": datetime.datetime.now(),
                "order_status":ORDER_STATUS.ORDERED,
                'behavior': BEHAVIOR_TYPE.TRANSFER_IN,
                "note": note
            })
            CmsUserTable.update_one({"uuid": self.current_admin_dict.get("uuid")}, {"$set": {"last_note": note}})
            return self.xtjson.json_result()
        if self.action == 'request_money_html':
            behavior = self.request_data.get('behavior')
            html = self.request_money_html(self.data_dict, behavior)
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        if self.action == 'edit_bankcard_html':
            html = self.bankcard_html(self.data_dict)
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        if self.action == 'edit_bankcard':
            location = self.request_data.get('location')
            if location == LOCATION_TYPE.ANOTHER:
                deposit_uuid = self.request_data.get('deposit_uuid')
                _dd = {
                    "deposit_uuid": deposit_uuid,
               }
                self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': _dd})
                return self.xtjson.json_result()

            bank_code = self.request_data.get('bank_code')
            account = self.request_data.get('account')
            bank_name = self.request_data.get('bank_name')
            note = self.request_data.get('bank_name')
            start_money = float(self.request_data.get('start_money').replace(',','') or '0')
            balance_amount = float(self.request_data.get('balance_amount').replace(',','') or '0')
            note = self.request_data.get('note')
            
            bankinfo = BankTable.find_one({"code": bank_code})
            if not bankinfo or not account or not bank_name or not account.strip() or not bank_name.strip():
                return self.xtjson.json_params_error('缺少数据！')

            _c = self.MCLS.find_one({"account": account,"bank_name": bank_name,'bank_uid': bankinfo.get("uuid"), "is_deleted": {"$ne": True}, "uuid":{'$ne': self.data_dict.get("uuid")}})
            if _c:
                return self.xtjson.json_params_error("该卡账户已存在!")

            self.data_from.update({
                'bank_uid': bankinfo.get("uuid"),
                'account': account,
                'bank_name': bank_name,
                'location': location,
                'start_money': start_money,
                'balance_amount': balance_amount,
                'note': note,
            })
            
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            return self.xtjson.json_result()
       
        if self.action == 'del':
            self.data_from.update({
                'is_deleted': True
            })
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            return self.xtjson.json_result()
        if self.action == 'update_status':
            self.data_from.update({
                'is_Disabled': False if self.data_dict.get("is_Disabled") else True
            })
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            return self.xtjson.json_result()
        if self.action == 'update_balance':
            balance_amount = self.data_dict.get("balance_amount")
            return self.xtjson.json_result(data = balance_amount)

        if self.action == 'transfer_in_out_html':
            behavior = self.request_data.get('behavior')
            html = self.transfer_in_out_html(self.data_dict, behavior)
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        if self.action == 'transfer_in_out2':
            behavior = self.request_data.get("behavior") or ''
            transfer_money = float(self.request_data.get("transfer_money").replace(',','')) or 0
            note = self.request_data.get("note") or ''

            latter_card = WithdrawalCardTable.find_one({"uuid": self.request_data.get("bankcarduuid")})

            self.get_deposit_cardinfo(latter_card)
            self.get_deposit_cardinfo(self.data_dict)

            former_cur_balance = former_pre_balance = self.data_dict.get("balance_amount") or 0
            latter_cur_balance = latter_pre_balance = latter_card.get("balance_amount") or 0
    
            former_bank_info = BankTable.find_one({'uuid': self.data_dict.get('bank_uid')})
            latter_bank_info = BankTable.find_one({'uuid': latter_card.get('bank_uid')})

            if behavior == BEHAVIOR_TYPE.TRANSFER_IN:
                latter_cur_balance = latter_cur_balance - transfer_money
                self.update_cardinfo(latter_card, {"balance_amount": latter_cur_balance})
                self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_out": transfer_money})
            
            elif behavior == BEHAVIOR_TYPE.TRANSFER_OUT:
                former_cur_balance = former_pre_balance - transfer_money
                self.update_cardinfo(self.data_dict, {"balance_amount": former_cur_balance})
                self.update_transfer_amount(self.data_dict.get("uuid"), {"total_transfer_out": transfer_money})
            elif behavior == BEHAVIOR_TYPE.ISSUED:
                former_cur_balance = former_pre_balance - transfer_money
                self.update_cardinfo(self.data_dict, {"balance_amount": former_cur_balance})
                self.update_transfer_amount(self.data_dict.get("uuid"), {"total_issue": transfer_money})

            WithdrawalOrderLogTable.insert_one({
                "admin_uuid": self.current_admin_dict.get("uuid"),
                "payer_uuid": self.data_dict.get("payer_uuids"),
                # "payer_uuid": CmsUserTable.find_one({"withdrawalcard_uuid": self.data_dict.get("uuid")}).get("uuid") if self.data_dict.get("location") == LOCATION_TYPE.WITHDRAW else self.data_dict.get("payer_uuids"),

                "former_card_uuid": self.data_dict.get("uuid"),
                "former_bank_code": former_bank_info.get("code"),
                "former_bank_name": self.data_dict.get("bank_name"),
                "former_bankcard_account": self.data_dict.get("account"),
                "former_location": self.data_dict.get("location"),
                "former_cur_balance": former_cur_balance,
                "former_pre_balance": former_pre_balance,
                
                "latter_card_uuid": latter_card.get("uuid"),
                "latter_bank_code": latter_bank_info.get("code"),
                "latter_bank_name": latter_card.get("bank_name"),
                "latter_bankcard_account": latter_card.get("account"),
                "latter_location": latter_card.get("location"),
                "latter_cur_balance": latter_cur_balance,
                "latter_pre_balance": latter_pre_balance,
            
                "transfer_money":transfer_money,
                "request_money":0,
                "order_status":ORDER_STATUS.COMPLETED,
                "accepted_status":ACCEPTED_STATUS.NOT_PROCESSED if self.current_admin_dict.get('role_code') in [ROlE_ALL.SUPERADMIN, ROlE_ALL.AGENTADMIN, ROlE_ALL.ADMINISTRATOR, ROlE_ALL.SYSTEMUSER] else ACCEPTED_STATUS.ACCEPTED,
                # "pay_time": datetime.datetime.now(),
                'behavior': behavior,
                "note": note,
                'operator_agent_uuid': self.current_admin_dict.get('uuid')
            })
            return self.xtjson.json_result()
        if self.action == 'transfer_in_out':
            behavior = self.request_data.get("behavior") or ''
            carddatas = self.request_data.get("carddatas") or ''
            carddatas = json.loads(carddatas)
            for carddata in carddatas:
                transfer_money = float(carddata.get("transfer_money").replace(',','')) or 0
                note = carddata.get("note") or ''
                latter_card = WithdrawalCardTable.find_one({"uuid": carddata.get("bankcarduuid")})

                self.get_deposit_cardinfo(latter_card)
                self.get_deposit_cardinfo(self.data_dict)

                former_cur_balance = former_pre_balance = self.data_dict.get("balance_amount") or 0
                latter_cur_balance = latter_pre_balance = latter_card.get("balance_amount") or 0
        
                former_bank_info = BankTable.find_one({'uuid': self.data_dict.get('bank_uid')})
                latter_bank_info = BankTable.find_one({'uuid': latter_card.get('bank_uid')})

                if behavior == BEHAVIOR_TYPE.TRANSFER_IN:
                    latter_cur_balance = latter_cur_balance - transfer_money
                    self.update_cardinfo(latter_card, {"balance_amount": latter_cur_balance})
                    self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_out": transfer_money})
                
                elif behavior == BEHAVIOR_TYPE.TRANSFER_OUT:
                    former_cur_balance = former_pre_balance - transfer_money
                    self.update_cardinfo(self.data_dict, {"balance_amount": former_cur_balance})
                    self.update_transfer_amount(self.data_dict.get("uuid"), {"total_transfer_out": transfer_money})

                elif behavior == BEHAVIOR_TYPE.ISSUED:
                    former_cur_balance = former_pre_balance - transfer_money
                    self.update_cardinfo(self.data_dict, {"balance_amount": former_cur_balance})
                    self.update_transfer_amount(self.data_dict.get("uuid"), {"total_issue": transfer_money})

                WithdrawalOrderLogTable.insert_one({
                    "admin_uuid": self.current_admin_dict.get("uuid"),
                    "payer_uuid": self.data_dict.get("payer_uuids"),
                    # "payer_uuid": CmsUserTable.find_one({"withdrawalcard_uuid": self.data_dict.get("uuid")}).get("uuid") if self.data_dict.get("location") == LOCATION_TYPE.WITHDRAW else self.data_dict.get("payer_uuids"),

                    "former_card_uuid": self.data_dict.get("uuid"),
                    "former_bank_code": former_bank_info.get("code"),
                    "former_bank_name": self.data_dict.get("bank_name"),
                    "former_bankcard_account": self.data_dict.get("account"),
                    "former_location": self.data_dict.get("location"),
                    "former_cur_balance": former_cur_balance,
                    "former_pre_balance": former_pre_balance,
                    
                    "latter_card_uuid": latter_card.get("uuid"),
                    "latter_bank_code": latter_bank_info.get("code"),
                    "latter_bank_name": latter_card.get("bank_name"),
                    "latter_bankcard_account": latter_card.get("account"),
                    "latter_location": latter_card.get("location"),
                    "latter_cur_balance": latter_cur_balance,
                    "latter_pre_balance": latter_pre_balance,
                
                    "transfer_money":transfer_money,
                    "request_money":0,
                    "order_status":ORDER_STATUS.COMPLETED,
                    "accepted_status":ACCEPTED_STATUS.NOT_PROCESSED if self.current_admin_dict.get('role_code') in [ROlE_ALL.SUPERADMIN, ROlE_ALL.AGENTADMIN, ROlE_ALL.ADMINISTRATOR, ROlE_ALL.SYSTEMUSER] else ACCEPTED_STATUS.ACCEPTED,
                    # "pay_time": datetime.datetime.now(),
                    'behavior': behavior,
                    "note": note,
                    'operator_agent_uuid': self.current_admin_dict.get('uuid')
                })
            return self.xtjson.json_result()

        if self.action == 'other_transfer_in_out':
            transfer_money = float(self.request_data.get("transfer_money").replace(",", "")) or 0
            note = self.request_data.get("note") or ''
            bank_name = self.request_data.get("bank_name")
            bank_code = self.request_data.get("bank_code")
            bankcard_account = self.request_data.get("bankcard_account")
            behavior = self.request_data.get("behavior") or ''

            self.get_deposit_cardinfo(self.data_dict)

            former_pre_balance = self.data_dict.get("balance_amount") or 0
       
            if behavior == BEHAVIOR_TYPE.OTHER_TRANSFER_IN:
                former_cur_balance = former_pre_balance + transfer_money
                self.update_transfer_amount(self.data_dict.get("uuid"), {"total_other_transfer_in": transfer_money})


            elif behavior == BEHAVIOR_TYPE.OTHER_TRANSFER_OUT:
                # if former_pre_balance < transfer_money:
                #     return self.xtjson.json_params_error('卡内余额不足！')
                former_cur_balance = former_pre_balance - transfer_money
                self.update_transfer_amount(self.data_dict.get("uuid"), {"total_other_transfer_out": transfer_money})
            
            elif behavior == BEHAVIOR_TYPE.ISSUED:
                # if former_pre_balance < transfer_money:
                #     return self.xtjson.json_params_error('卡内余额不足！')
                former_cur_balance = former_pre_balance - transfer_money
                self.update_transfer_amount(self.data_dict.get("uuid"), {"total_issue": transfer_money})
            
            else:
                return self.xtjson.json_params_error("错误动作!")
        

            self.update_cardinfo(self.data_dict, {"balance_amount": former_cur_balance})
            
            former_bank_info = BankTable.find_one({'uuid': self.data_dict.get('bank_uid')})

            WithdrawalOrderLogTable.insert_one({
                "admin_uuid": self.current_admin_dict.get("uuid"),
                "payer_uuid": CmsUserTable.find_one({"withdrawalcard_uuid": self.data_dict.get("uuid")}).get("uuid"),

                "former_card_uuid": self.data_dict.get("uuid"),
                "former_bank_code": former_bank_info.get("code"),
                "former_bank_name": self.data_dict.get("bank_name"),
                "former_location": self.data_dict.get("location"),
                "former_bankcard_account": self.data_dict.get("account"),
                "former_cur_balance": former_cur_balance,
                "former_pre_balance": former_pre_balance,
            
                "latter_bankcard_account": bankcard_account,
                "latter_bank_code": bank_code,
                "latter_bank_name": bank_name,

                "transfer_money":transfer_money,
                "request_money":0,
                "order_status":ORDER_STATUS.COMPLETED,
                "accepted_status":ACCEPTED_STATUS.ACCEPTED,#ACCEPTED_STATUS.NOT_PROCESSED if self.current_admin_dict.get('role_code') not in [ROlE_ALL.OUT_MONEY_USER, ROlE_ALL.SYS_OUT_MONEY_USER] else ACCEPTED_STATUS.ACCEPTED,
                "pay_time": datetime.datetime.now(),
                'behavior': behavior,
                "note": note,
                'operator_agent_uuid': self.current_admin_dict.get('uuid')
            })
            return self.xtjson.json_result()
        if self.action == 'get_pay_qrcode':
            self.get_deposit_cardinfo(self.data_dict)
            behavior = self.request_data.get('behavior') or ''
            carddatas = self.request_data.get('carddatas') or ''
            carddatas = json.loads(carddatas)
            
            html = f'''
            <div class="mb-2" style="text-align: center"><button class="btn btn-primary" onclick="post_transfer_in_out('qrcode', 'transfer_in_out', '{self.data_dict.get("uuid")}','{behavior}')">全部已支付</button></div>
            <div class="d-flex justify-content-between" style="overflow:scroll" id="qrboxlist">'''
            for carddata in carddatas:
                bankcarduuid = carddata.get('bankcarduuid') or ''

                if behavior == BEHAVIOR_TYPE.TRANSFER_IN or behavior == BEHAVIOR_TYPE.OTHER_TRANSFER_IN:
                    bank_data = BankTable.find_one({'uuid': self.data_dict.get('bank_uid')}) or {}
                    username = self.data_dict.get("bank_name")
                    cardaccount = self.data_dict.get("account")
                    latter_card = WithdrawalCardTable.find_one({"uuid": bankcarduuid})
                    latter_bankdata = BankTable.find_one({'uuid': latter_card.get('bank_uid')})
                    addinfo = latter_card.get("bank_name") + f' ({latter_bankdata.get("code")})'
                    out_card_code = latter_bankdata.get("code")

                elif behavior == BEHAVIOR_TYPE.TRANSFER_OUT or behavior == BEHAVIOR_TYPE.ISSUED:
                    latter_card = WithdrawalCardTable.find_one({"uuid": bankcarduuid})
                    bank_data = BankTable.find_one({'uuid': latter_card.get("bank_uid")}) or {}
                    username = latter_card.get("bank_name")
                    fomer_bankdata = BankTable.find_one({'uuid': self.data_dict.get('bank_uid')}) or {}
                    addinfo = self.data_dict.get("bank_name") + f' ({fomer_bankdata.get("code")})'
                    cardaccount = latter_card.get("account")
                    out_card_code = fomer_bankdata.get("code")
                else:
                    return self.xtjson.json_params_error('Error！')
                
                addinfo += ' to ' + username + f' {bank_data.get("code")}'
                if out_card_code == 'SHB':
                    addinfo = addinfo[:25]

                transfer_money = carddata.get("transfer_money").replace(",", '') or '0'
                transfer_money = carddata.get("transfer_money").replace(",", '') or '0'
                transfer_money = float(transfer_money)
                if not bank_data:
                    return self.xtjson.json_params_error('该订单收款银行错误！')

                payqrcode_url = self.data_dict.get('payqrcode_url') or ''
                project_static_folder = os.path.join(current_app.static_folder, current_app.config.get('PROJECT_NAME'))
                _state, payQrcode = getWithdrawalBankPayQrcode(
                    self.data_uuid,
                    transfer_money,
                    addinfo,
                    bank_data,
                    payqrcode_url='',
                    project_static_folder=project_static_folder,
                    receive_account=cardaccount,
                )

                if not _state:
                    return self.xtjson.json_params_error(payQrcode)
            
                html += f'''
                <div class="formBox" style="padding: 0px; margin: 0px 10px">
                    <div style="height: 35rem; position: relative; box-sizing: border-box; overflow-y: auto;">
                        <div>
                            <img src="{payQrcode}" alt="" style="display: block; width: 260px; margin: auto; margin-bottom: 30px;">
                            <p style="padding-left: 15%;text-align: left; margin-bottom: 10px; font-size: 14px;">收款银行：{bank_data.get('shortName')}</p>
                            <p style="padding-left: 15%;text-align: left; margin-bottom: 10px; font-size: 14px;">收款卡号：{cardaccount}</p>
                            <p style="padding-left: 15%;text-align: left; margin-bottom: 10px; font-size: 14px;">收款人：{username}</p>
                            <p style="padding-left: 15%;text-align: left; font-size: 14px;">订单金额：{transfer_money}</p>
                        </div>
                        <div class="blank" style="background: #eeeeee; height: 1px; margin: 10px 0 15px;"></div>
                        <div style="position: relative; text-align: center" bankcarduuid='{bankcarduuid}' transfer_money='{transfer_money}'  note='{carddata.get("note")}' paid="" class="qrboxlist">
                            <span class="btn btn-warning" style="color: white;" onclick="post_transfer_in_out(this, 'transfer_in_out2', '{self.data_dict.get("uuid")}','{behavior}')">已支付</span>
                        </div>
                    </div>
                </div>
                '''

            html += '</div>'
            return self.xtjson.json_result(message=html)
        if self.action == 'get_pay_qrcode2':
            self.get_deposit_cardinfo(self.data_dict)

            #bank_data = BankTable.find_one({'uuid': self.data_dict.get('bank_uid')}) or {}
            behavior = self.request_data.get('behavior') or ''
            bankcarduuid = self.request_data.get('bankcarduuid') or ''
            
            bank_name = self.request_data.get('bank_name') or ''
            bankcard_account = self.request_data.get('bankcard_account') or ''
            bank_code = self.request_data.get('bank_code') or ''
            addinfo = ''


            if behavior == BEHAVIOR_TYPE.TRANSFER_IN or behavior == BEHAVIOR_TYPE.OTHER_TRANSFER_IN:
                bank_data = BankTable.find_one({'uuid': self.data_dict.get('bank_uid')}) or {}
                username = self.data_dict.get("bank_name")
                cardaccount = self.data_dict.get("account")
                latter_card = WithdrawalCardTable.find_one({"uuid": bankcarduuid})
                latter_bankdata = BankTable.find_one({'uuid': latter_card.get('bank_uid')})
                addinfo = latter_card.get("bank_name") + f' ({latter_bankdata.get("code")})'
                out_card_code = latter_bankdata.get("code")

            elif behavior == BEHAVIOR_TYPE.TRANSFER_OUT or behavior == BEHAVIOR_TYPE.ISSUED:
                latter_card = WithdrawalCardTable.find_one({"uuid": bankcarduuid})
                bank_data = BankTable.find_one({'uuid': latter_card.get("bank_uid")}) or {}
                username = latter_card.get("bank_name")
                cardaccount = latter_card.get("account")
                fomer_bankdata = BankTable.find_one({'uuid': self.data_dict.get('bank_uid')}) or {}
                addinfo = self.data_dict.get("bank_name") + f' ({fomer_bankdata.get("code")})'
                out_card_code = fomer_bankdata.get("code")
            elif behavior == BEHAVIOR_TYPE.OTHER_TRANSFER_OUT:
                bank_data = BankTable.find_one({'code': bank_code}) or {}
                username = bank_name
                cardaccount = bankcard_account
                fomer_bankdata = BankTable.find_one({'uuid': self.data_dict.get('bank_uid')}) or {}
                addinfo = self.data_dict.get("bank_name") + f' ({fomer_bankdata.get("code")})'
                out_card_code = fomer_bankdata.get("code")
            else:
                return self.xtjson.json_params_error('Error！')
            
            addinfo += ' to ' + username + f' {bank_data.get("code")}'
            if out_card_code == 'SHB':
                addinfo = addinfo[:25]

            transfer_money = self.request_data.get("transfer_money").replace(",", '') or '0'
            transfer_money = float(transfer_money)
            if not bank_data:
                return self.xtjson.json_params_error('该订单收款银行错误！')

            payqrcode_url = self.data_dict.get('payqrcode_url') or ''
            project_static_folder = os.path.join(current_app.static_folder, current_app.config.get('PROJECT_NAME'))

            _state, payQrcode = getWithdrawalBankPayQrcode(
                self.data_uuid,
                transfer_money,
                addinfo,
                bank_data,
                payqrcode_url='',
                project_static_folder=project_static_folder,
                receive_account=cardaccount,
            )

            if not _state:
                return self.xtjson.json_params_error(payQrcode)
           
            html = f'''
            <div class="formBox" style="padding: 0px; margin: 0px 10px">
                <div style="height: 35rem; position: relative; box-sizing: border-box; overflow-y: auto;">
                    <div>
                        <img src="{payQrcode}" alt="" style="display: block; width: 260px; margin: auto; margin-bottom: 30px;">
                        <p style="padding-left: 15%;text-align: left; margin-bottom: 10px; font-size: 14px;">收款银行：{bank_data.get('shortName')}</p>
                        <p style="padding-left: 15%;text-align: left; margin-bottom: 10px; font-size: 14px;">收款卡号：{cardaccount}</p>
                        <p style="padding-left: 15%;text-align: left; margin-bottom: 10px; font-size: 14px;">收款人：{username}</p>
                        <p style="padding-left: 15%;text-align: left; font-size: 14px;">订单金额：{transfer_money}</p>
                    </div>
                    <div class="blank" style="background: #eeeeee; height: 1px; margin: 10px 0 15px;"></div>
                        <div style="position: relative; text-align: center">
                        <span class="btn btn-warning" style="color: white;" onclick="post_transfer_in_out(this, 'transfer_in_out', '{self.data_dict.get("uuid")}','{behavior}')">已支付</span>
                    </div>
                </div>
            </div>
            '''
            return self.xtjson.json_result(message=html)
        if self.action == 'select_subadmin_html':
            html = self.select_subadmin_html(self.data_dict)
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        if self.action == 'set_payer_uuids':
            payer_uuids = json.loads(self.request_data.get("payer_uuids"))
            self.MCLS.update_one({"uuid": self.data_dict.get("uuid")}, {"$set": {
                "payer_uuids": payer_uuids
            }})
            return self.xtjson.json_result()


class TransactionFlowView(CmsTableViewBase):
    add_url_rules = [['/TransactionFlow', 'TransactionFlow']]
    per_page = 30
    MCLS = WithdrawalOrderLogTable
    template = 'cms/behalfPay/transactionFlow.html'
    title = '交易流水'

    def get_filter_dict(self):
        fff = {}
        if self.current_admin_dict.get('role_code') == ROlE_ALL.SUPERADMIN:
            admin_uuid = request.args.get('admin_uuid') or ''
            payer_uuid = request.args.get('payer_uuid') or ''
            if admin_uuid:
                fff["admin_uuid"] = admin_uuid
                self.search_dict['admin_uuid'] = admin_uuid
            if payer_uuid:
                fff["payer_uuid"] = payer_uuid
                self.search_dict['payer_uuid'] = payer_uuid
        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN: 
            admin_uuids = [x.get("uuid") for x in CmsUserTable.find_many({"agentadmin_uuid": self.current_admin_dict.get("uuid") })]
            fff["admin_uuid"] = {"$in": [self.current_admin_dict.get("uuid")]+ admin_uuids} 
            payer_uuid = request.args.get('payer_uuid') or ''
            if payer_uuid:
                fff["payer_uuid"] = payer_uuid
                self.search_dict['payer_uuid'] = payer_uuid
        
        if self.current_admin_dict.get('role_code') == ROlE_ALL.ADMINISTRATOR:
            outpayers = [x.get('uuid') for x in CmsUserTable.find_many({"role_code": ROlE_ALL.SYS_OUT_MONEY_USER })]
            fff['$or'] =  [
                {'payer_uuid': self.current_admin_dict.get("uuid")},  # Check if xxx field is equal to x_value_to_find
                {'payer_uuid': {'$in': outpayers}},  # Check if xxx field is in y_array_to_match
            ]

        if self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            outpayers = [x.get('uuid') for x in CmsUserTable.find_many({"agentadmin_uuid": self.current_admin_dict.get("agentadmin_uuid"), "role_code": ROlE_ALL.OUT_MONEY_USER })]
            fff['$or'] =  [
                {'payer_uuid': self.current_admin_dict.get("uuid")},  # Check if xxx field is equal to x_value_to_find
                {'payer_uuid': {'$in': outpayers}},  # Check if xxx field is in y_array_to_match
            ]
 
        if self.current_admin_dict.get('role_code') == ROlE_ALL.SYS_OUT_MONEY_USER:
            #fff["payer_uuid"] = self.current_admin_dict.get("uuid")
            selectedcard = WithdrawalCardTable.find_one({"uuid": self.current_admin_dict.get("withdrawalcard_uuid")})
            fff["former_card_uuid"] = selectedcard.get("uuid")

        if self.current_admin_dict.get('role_code') == ROlE_ALL.OUT_MONEY_USER :
            selectedcard = WithdrawalCardTable.find_one({"uuid": self.current_admin_dict.get("withdrawalcard_uuid")})
            fff["former_card_uuid"] = selectedcard.get("uuid")
        

        former_bankcard_account = request.args.get('former_bankcard_account') or ''
        if former_bankcard_account:
            fff["former_bankcard_account"] = former_bankcard_account
            self.search_dict['former_bankcard_account'] = former_bankcard_account

        payer_uuid = request.args.get('payer_uuid') or ''
        if payer_uuid:
            fff["payer_uuid"] = payer_uuid
            self.search_dict['payer_uuid'] = payer_uuid

        former_bank_name = request.args.get('former_bank_name') or ''
        if former_bank_name:
            fff["former_bank_name"] = former_bank_name
            self.search_dict['former_bank_name'] = former_bank_name
        
        former_location = request.args.get('former_location') or ''
        if former_location:
            fff["former_location"] = former_location
            self.search_dict['former_location'] = former_location

        former_bank_code = request.args.get('former_bank_code') or ''
        if former_bank_code:
            fff["former_bank_code"] = former_bank_code
            self.search_dict['former_bank_code'] = former_bank_code

        latter_bankcard_account = request.args.get('latter_bankcard_account') or ''
        if latter_bankcard_account:
            fff["latter_bankcard_account"] = latter_bankcard_account
            self.search_dict['latter_bankcard_account'] = latter_bankcard_account

        latter_bank_code = request.args.get('latter_bank_code') or ''
        if latter_bank_code:
            fff["latter_bank_code"] = latter_bank_code
            self.search_dict['latter_bank_code'] = latter_bank_code

        latter_location = request.args.get('latter_location') or ''
        if latter_location:
            fff["latter_location"] = latter_location
            self.search_dict['latter_location'] = latter_location

        accepted_status = request.args.get('accepted_status') or ''
        if accepted_status:
            fff["accepted_status"] = accepted_status
            self.search_dict['accepted_status'] = accepted_status

        behavior = request.args.get('behavior') or ''
        if behavior:
            fff["behavior"] = behavior
            self.search_dict['behavior'] = behavior

        note = request.args.get('note') or ''
        if note:
            fff["note"] = note
            self.search_dict['note'] = note

        transfer_money = request.args.get('transfer_money') or ''
        if transfer_money:
            fff["transfer_money"] = float(transfer_money)

        latter_bank_name = request.args.get('latter_bank_name') or ''
        if latter_bank_name:
            fff["latter_bank_name"] = latter_bank_name
            self.search_dict['latter_bank_name'] = latter_bank_name

        order_status = request.args.get('order_status') or ''
        if order_status:
            fff["order_status"] = order_status
            self.search_dict['order_status'] = order_status

        order_time = request.args.get('order_time')
        if order_time and order_time.strip():
            start_time, end_time = PagingCLS.by_silce(order_time)
        else:
            crrdate = datetime.datetime.now()
            start_time, end_time = datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 0, 0,0), datetime.datetime(crrdate.year, crrdate.month, crrdate.day,23, 59, 59)
        if order_time:
            fff['order_time'] = {'$gte': start_time, '$lte': end_time}
        
        self.search_dict['order_time'] = f'{start_time}|{end_time}'
        return fff

    def dealwith_main_context(self):
        select_cols = session.get(SELECET_COLS_SESSION_KEY)
        if not select_cols:
            select_cols = {
                "admin_check": "true",
                "former_account_check" : "true", 
                "former_name_check" : "true", 
                "former_location_check" : "true", 
                "former_pre_balance_check" : "true", 
                "former_cur_balance_check" : "true", 
                "latter_account_check" : "true", 
                "latter_name_check" : "true", 
                "latter_location_check" : "true", 
                "latter_pre_balance_check" : "true", 
                "latter_cur_balance_check" : "true", 
                "request_money_check" : "true", 
                "transfer_money_check" : "true", 
                "order_time_check" :"true", 
                "withdrw_time_check" :"true", 
                "order_status_check" : "true", 
                "accept_status_check" : "true", 
                "behavior_check" : "true", 
                "operator_check" : "true", 
                "operator_agent_check" : "true", 
                "note_check" :"true", 
            }
        session[SELECET_COLS_SESSION_KEY] = select_cols

        all_datas = self.context.get('all_datas')
        # if self.search_dict.get("payer_uuid"):
        #     outpayer = CmsUserTable.find_one({"uuid": self.search_dict.get("payer_uuid")})
        #     all_datas = [self.MCLS.find_one({"former_card_uuid": outpayer.get("withdrawalcard_uuid")})]
        for da in all_datas:
            da["outpayer"] = [CmsUserTable.find_one({"uuid": x}).get("account") for x in (da.get("payer_uuid") or [])]
            # da["outpayer"] = CmsUserTable.find_one({"withdrawalcard_uuid": da.get("former_card_uuid")}).get("account")
            da["admin"] = CmsUserTable.find_one({"uuid": da.get("admin_uuid")})
            da["operator"] = CmsUserTable.find_one({"uuid": da.get("operator_uuid")}).get("account")
            da["operator_agent"] = CmsUserTable.find_one({"uuid": da.get("operator_agent_uuid")}).get("account")
            payer = CmsUserTable.find_one({"uuid": da.get("payer_uuid")})
            da["accept_order"] = False
            if payer.get('role_code') in [ROlE_ALL.OUT_MONEY_USER, ROlE_ALL.SYS_OUT_MONEY_USER]:
                da["accept_order"] = True
        bank_datas = BankTable.find_many({})
        
        if self.current_admin_dict.get('role_code') in [ ROlE_ALL.SUPERADMIN, ROlE_ALL.ADMINISTRATOR]:
            admins = [self.current_admin_dict]
            admins += CmsUserTable.find_many({"role_code": ROlE_ALL.AGENTADMIN})
            self.context['admins'] = admins
            outpayers = CmsUserTable.find_many({"role_code": ROlE_ALL.SYS_OUT_MONEY_USER})
            outpayers += CmsUserTable.find_many({"role_code": ROlE_ALL.OUT_MONEY_USER})
            self.context['outpayers'] = outpayers
        if self.current_admin_dict.get('role_code') in [ ROlE_ALL.AGENTADMIN, ROlE_ALL.SYSTEMUSER ] :
            outpayers = CmsUserTable.find_many({"role_code": ROlE_ALL.OUT_MONEY_USER, "agentadmin_uuid":self.current_admin_dict.get("uuid")})
            self.context['outpayers'] = outpayers
        if self.current_admin_dict.get("is_clear_log"):
            self.context['is_clear_log'] = True
        if self.current_admin_dict.get("is_test_withdraw_card"):
            self.context['is_test_withdraw_card'] = True

        selectedcard = {}
        if self.current_admin_dict.get('role_code') == ROlE_ALL.SYS_OUT_MONEY_USER or self.current_admin_dict.get('role_code') == ROlE_ALL.OUT_MONEY_USER:
            selectedcard = WithdrawalCardTable.find_one({"uuid": self.current_admin_dict.get("withdrawalcard_uuid")})
                
        self.context['selectedcard'] = selectedcard
        self.context['all_datas'] = all_datas
        self.context['bank_datas'] = bank_datas
        self.context['ORDER_STATUS'] = ORDER_STATUS
        self.context['ACCEPTED_STATUS'] = ACCEPTED_STATUS
        self.context['BEHAVIOR_TYPE'] = BEHAVIOR_TYPE
        self.context['LOCATION_TYPE'] = LOCATION_TYPE
        self.context['COLS'] = select_cols

    def check_current_date(self, card_uuid):
        card = WithdrawalCardTable.find_one({"uuid": card_uuid})
        cur_date = datetime.datetime.now().date().strftime("%Y-%m-%d")
        if card.get("current_date") == cur_date:
            return
        
        data_from = {
            "current_date": cur_date,
            "total_request" : 0,
            "total_transfer_in" : 0,
            "total_transfer_out" : 0,
            "total_other_transfer_in" : 0,
            "total_other_transfer_out" : 0,
            "total_issue" : 0,
            "total_init_money" : card.get("balance_amount"),
        }
        WithdrawalCardTable.update_one({"uuid": card_uuid}, {'$set': data_from})

    def update_transfer_amount(self, card_uuid, dd):
        self.check_current_date(card_uuid)
        card = WithdrawalCardTable.find_one({"uuid": card_uuid})

        data_from = {}
        if "total_request" in dd:
            data_from["total_request"] = (card.get("total_request") or 0) + (dd.get("total_request") or 0)
        if "total_transfer_in" in dd:
            data_from["total_transfer_in"] = (card.get("total_transfer_in") or 0) + (dd.get("total_transfer_in") or 0)
        if "total_transfer_out" in dd:
            data_from["total_transfer_out"] = (card.get("total_transfer_out") or 0) + (dd.get("total_transfer_out") or 0)
        if "total_other_transfer_in" in dd:
            data_from["total_other_transfer_in"] = (card.get("total_other_transfer_in") or 0) + (dd.get("total_other_transfer_in") or 0)
        if "total_other_transfer_out" in dd:
            data_from["total_other_transfer_out"] = (card.get("total_other_transfer_out") or 0) + (dd.get("total_other_transfer_out") or 0)
        if "total_issue" in dd:
            data_from["total_issue"] = (card.get("total_issue") or 0) + (dd.get("total_issue") or 0)

        WithdrawalCardTable.update_one({"uuid": card_uuid}, {'$set': data_from})

    def retransfer_html(self, data_dict={}):
        fff = {"is_deleted": {"$ne": True}, "is_Disabled": {"$ne": True}}
        fff["admin_uuid"] = self.current_admin_dict["uuid"]
        fff['uuid'] = {"$ne": data_dict.get("former_card_uuid")}

        if self.current_admin_dict.get('role_code') == ROlE_ALL.SYS_OUT_MONEY_USER:
            admin_uuid = CmsUserTable.find_one({"role_code":ROlE_ALL.SUPERADMIN}).get("uuid")
            fff["admin_uuid"] = admin_uuid 
            fff["location"] = LOCATION_TYPE.WITHDRAW 
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.OUT_MONEY_USER:
            fff["location"] = LOCATION_TYPE.WITHDRAW 
            fff["admin_uuid"] = self.current_admin_dict.get("agentadmin_uuid")  
        elif self.current_admin_dict.get('role_code') in [ROlE_ALL.SUPERADMIN]:
            admin_uuid = CmsUserTable.find_one({"role_code":ROlE_ALL.SUPERADMIN}).get("uuid")
            fff["admin_uuid"] = admin_uuid
            self.search_dict["admin_uuid"] = admin_uuid
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff["admin_uuid"] = self.current_admin_dict.get("uuid")  
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            fff["admin_uuid"] = self.current_admin_dict.get("agentadmin_uuid")  
            fff["payer_uuids"] =  self.current_admin_dict.get("uuid")
            # fff["$or"] =  [{"location": LOCATION_TYPE.WITHDRAW },  { "payer_uuids": self.current_admin_dict.get("uuid") }]
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.ADMINISTRATOR:
            fff["admin_uuid"] = CmsUserTable.find_one({"role_code":ROlE_ALL.SUPERADMIN}).get("uuid")
            fff["payer_uuids"] = self.current_admin_dict.get("uuid")
            # fff["$or"] =  [{"location": LOCATION_TYPE.WITHDRAW },  { "payer_uuids": self.current_admin_dict.get("uuid") }]


        withdrawalcards = WithdrawalCardTable.find_many(fff)
        card_cnt = len(withdrawalcards)
        bankcard_html = ''
        
        for withdrawalcard in withdrawalcards:
            self.get_deposit_cardinfo(withdrawalcard)
            if withdrawalcard.get("uuid") in [self.data_dict.get("former_card_uuid"), self.data_dict.get("latter_card_uuid")]:
                continue
            bank_info = BankTable.find_one({'uuid': withdrawalcard.get('bank_uid')})
            bankcard_html += f'''
            <div class="card mb-3" bankcarduuid="{ withdrawalcard.get("uuid") }" onclick="seleBankCard_func($(this))" location="{withdrawalcard.get("location")}" bank_code="{bank_info.get("code")}" bank_name="{withdrawalcard.get("bank_name")}" account="{withdrawalcard.get("account")}" >
                <div class="row g-0 m-0" style = "cursor: pointer;">
                    <div class="col-3 text-center" style="display: flex; align-items: stretch; justify-content: center; padding: 0px; ">
                        <div style=" flex: 1; background-color: #ebeaea; display: flex; align-items: center; justify-content: center; ">
                            <h5 style="margin:0px">{bank_info.get("code")}({LOCATION_TYPE.name_dict[withdrawalcard.get("location")]})</h5>
                        </div>
                    </div>
                    <div class="col-9">
                        <div class="card-body" style="padding:10px">
                            银行卡账号: {withdrawalcard.get("account")} (名字: {withdrawalcard.get("bank_name")}, 余额: {self.format_money(withdrawalcard.get("balance_amount"))})
                        </div>
                    </div>
                </div>
            </div>
            '''
        cards_html = f'''
        <option value="{LOCATION_TYPE.TRANSIT}"> {LOCATION_TYPE.name_dict[LOCATION_TYPE.TRANSIT]}</option>
        <option value="{LOCATION_TYPE.WITHDRAW}" >{LOCATION_TYPE.name_dict[LOCATION_TYPE.WITHDRAW]}</option>
        <option value="{LOCATION_TYPE.DEPOSIT}" >{LOCATION_TYPE.name_dict[LOCATION_TYPE.DEPOSIT]}</option>
        <option value="{LOCATION_TYPE.OTHER}" >{LOCATION_TYPE.name_dict[LOCATION_TYPE.OTHER]}</option>
        '''

        bankcard_opt = ''
        bank_datas = BankTable.find_many({})
        for bdc in bank_datas:
            bankcard_opt += f'''<div class="auto-complete-item" onclick="onclick_insert_code('{bdc.get("code")}', filter = true)">{ bdc.get("code") }</div>'''

        html = f'''
            <div class="d-flex justify-content-between mb-2">
                <h5 class="mb-0 text-left py-2" style="margin-left:40px" >可用卡数: <span class="text-danger">{card_cnt}<span></h5>
        '''
        if data_dict.get("request_money") != 0:
            html += f'''
                <h5 class="mb-0 text-left py-2">订购的钱: <span class="text-danger">{self.format_money(data_dict.get("request_money"))}<span></h5>
            '''
        html += f'''
                <div style="margin-right:33px">
                    <span style="width: 80px; text-align: right; display: inline-block; position: relative;">转账金额:</span>
                    <input type="text" class="form-control" id="transfer_money" placeholder="转账金额" value="{self.format_money(data_dict.get("transfer_money"))}" aria-label="" style="display: inline-block; width: calc(100% - 100px)" onchange="onchange_number_input('transfer_money')">
                </div>
            </div>
            <div class="d-flex justify-content-between" style="margin-bottom:3px">
                <select class="form-control mr-sm-2 mb-2" style="margin-left:40px; width: 200px;" aria-label="" id="location_filter" onchange="on_card_filter_change()">
                    <option value="">选择位置</option>
                   {cards_html}
                </select>
                <div class="auto-complete mr-sm-2 mb-2" >
                    <input type="text" class="form-control" id="bank_code" value="" placeholder="银行类型" aria-label="" onclick="onclick_bankcode()" oninput="on_card_filter_change()" >
                    <div class="auto-complete-items">
                    {bankcard_opt}
                    </div>
                </div>

                <input type="text" class="form-control mr-sm-2 mb-2" id="bankname_filter" placeholder="银行名字" aria-label="" style="display: inline-block; width: 200px; margin-right: 40px;" oninput="on_card_filter_change()" onchange="on_card_filter_change()">
                <input type="text" class="form-control" id="account_filter" placeholder="银行卡账号" aria-label="" style="display: inline-block; width: 200px; margin-right: 40px;" oninput="on_card_filter_change()" onchange="on_card_filter_change()">
            </div>
            <div class="formBox m-0" style="padding:0px;">
                <div class="list-group-item" style="display: flex; align-items: center;justify-content: center;">
                    <input type="hidden" id="bank_card_uid">
                    <div class="form-control p-3" style="display: inline-block; width: calc(100% - 40px); height: 400px; overflow-y: scroll; text-align: left;" id='bankcard_list'>
                        { bankcard_html }                        
                    </div>
                </div>    
            <div>   
            <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

            <div style="position: relative; text-align: center">
                <span class="btn btn-secondary mr-5" onclick="post_qr_code('{data_dict.get('uuid')}', '{data_dict.get("behavior")}' , 'retransfer_in_out', '{url_for('admin.TransactionFlow')}' )">支付码</span>&emsp;
                <span class="btn btn-primary" onclick="post_transfer_in('retransfer_in_out', '{data_dict.get('uuid')}')">确定</span>&emsp;
                <span class="btn btn-default" onclick="closeQRButton()">取消</span>
            </div>            
        '''
        return html

    def test_card_html(self, data_dict={}):
        fff = {"is_deleted": {"$ne": True}, "is_Disabled": {"$ne": True}}
        
        if self.current_admin_dict.get('role_code') == ROlE_ALL.SYS_OUT_MONEY_USER:
            admin_uuid = CmsUserTable.find_one({"role_code":ROlE_ALL.SUPERADMIN}).get("uuid")
            fff["admin_uuid"] = admin_uuid 
            fff["location"] = LOCATION_TYPE.TRANSIT 
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.OUT_MONEY_USER:
            fff["location"] = LOCATION_TYPE.TRANSIT 
            fff["admin_uuid"] = self.current_admin_dict.get("agentadmin_uuid")  

        withdrawalcards = WithdrawalCardTable.find_many(fff)
        card_cnt = len(withdrawalcards)
        bankcard_html = ''
        
        for withdrawalcard in withdrawalcards:
            self.get_deposit_cardinfo(withdrawalcard)
            bank_info = BankTable.find_one({'uuid': withdrawalcard.get('bank_uid')})
            bankcard_html += f'''
            <div class="card mb-3" bankcarduuid="{ withdrawalcard.get("uuid") }" onclick="seleBankCard_func($(this))" location="{withdrawalcard.get("location")}" bank_code="{bank_info.get("code")}" bank_name="{withdrawalcard.get("bank_name")}" account="{withdrawalcard.get("account")}" >
                <div class="row g-0 m-0" style = "cursor: pointer;">
                    <div class="col-3 text-center" style="display: flex; align-items: stretch; justify-content: center; padding: 0px; ">
                        <div style=" flex: 1; background-color: #ebeaea; display: flex; align-items: center; justify-content: center; ">
                            <h5 style="margin:0px">{bank_info.get("code")}({LOCATION_TYPE.name_dict[withdrawalcard.get("location")]})</h5>
                        </div>
                    </div>
                    <div class="col-9">
                        <div class="card-body" style="padding:10px">
                            银行卡账号: {withdrawalcard.get("account")} (名字: {withdrawalcard.get("bank_name")}, 余额: {self.format_money(withdrawalcard.get("balance_amount"))})
                        </div>
                    </div>
                </div>
            </div>
            '''
        cards_html = f'''
        <option value="{LOCATION_TYPE.TRANSIT}"> {LOCATION_TYPE.name_dict[LOCATION_TYPE.TRANSIT]}</option>
        <option value="{LOCATION_TYPE.WITHDRAW}" >{LOCATION_TYPE.name_dict[LOCATION_TYPE.WITHDRAW]}</option>
        <option value="{LOCATION_TYPE.DEPOSIT}" >{LOCATION_TYPE.name_dict[LOCATION_TYPE.DEPOSIT]}</option>
        <option value="{LOCATION_TYPE.OTHER}" >{LOCATION_TYPE.name_dict[LOCATION_TYPE.OTHER]}</option>
        '''

        bankcard_opt = ''
        bank_datas = BankTable.find_many({})
        for bdc in bank_datas:
            bankcard_opt += f'''<div class="auto-complete-item" onclick="onclick_insert_code('{bdc.get("code")}', filter = true)">{ bdc.get("code") }</div>'''

        html = f'''
            <div class="d-flex justify-content-between mb-2">
                <h5 class="mb-0 text-left py-2" style="margin-left:40px" >可用卡数: <span class="text-danger">{card_cnt}<span></h5>
        '''
        html += f'''
                <div class="d-flex justify-content-end" style="margin-right:40px">
                    <div>
                        <span style="width: 80px; text-align: right; display: inline-block; position: relative;">转账金额:</span>
                        <input type="text" class="form-control" id="transfer_money" value = "0" placeholder="转账金额" aria-label="" style="display: inline-block; width: 150px" onchange="onchange_number_input('transfer_money')">
                    </div>
                    <div class="ml-1">
                        <input type="text" class="form-control" id="note" placeholder="注解" aria-label="" style="display: inline-block; width: 150px">
                    </div>
                    <div class="ml-1">
                        <select class="form-control" id="behavior_type" style="width: 150px">
                            <option value="{BEHAVIOR_TYPE.TRANSFER_OUT}" >{BEHAVIOR_TYPE.name_dict[BEHAVIOR_TYPE.TRANSFER_OUT]}</option>
                            <option value="{BEHAVIOR_TYPE.TRANSFER_IN}"> {BEHAVIOR_TYPE.name_dict[BEHAVIOR_TYPE.TRANSFER_IN]}</option>
                        </select>
                    </div>
                </div>
            </div>
            <div class="d-flex justify-content-between" style="margin-bottom:3px">
                <div class="auto-complete mr-sm-2 mb-2" style="margin-left:40px; width: 200px;">
                    <input type="text" class="form-control" id="bank_code" value="" placeholder="银行类型" aria-label="" onclick="onclick_bankcode()" oninput="on_card_filter_change()" >
                    <div class="auto-complete-items">
                    {bankcard_opt}
                    </div>
                </div>

                <input type="text" class="form-control mr-sm-2 mb-2" id="bankname_filter" placeholder="银行名字" aria-label="" style="display: inline-block; width: 200px; margin-right: 40px;" oninput="on_card_filter_change()" onchange="on_card_filter_change()">
                <input type="text" class="form-control" id="account_filter" placeholder="银行卡账号" aria-label="" style="display: inline-block; width: 200px; margin-right: 40px;" oninput="on_card_filter_change()" onchange="on_card_filter_change()">
            </div>
            <div class="formBox m-0" style="padding:0px;">
                <div class="list-group-item" style="display: flex; align-items: center;justify-content: center;">
                    <input type="hidden" id="bank_card_uid">
                    <div class="form-control p-3" style="display: inline-block; width: calc(100% - 40px); height: 400px; overflow-y: scroll; text-align: left;" id='bankcard_list'>
                        { bankcard_html }                        
                    </div>
                </div>    
            <div>   
            <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

            <div style="position: relative; text-align: center">
                <span class="btn btn-secondary mr-5" onclick="test_post_qr_code()">支付码</span>&emsp;
                <span class="btn btn-primary" onclick="test_transfer_in_out()">确定</span>&emsp;
                <span class="btn btn-default" onclick="closeQRButton()">取消</span>
            </div>            
        '''
        return html

    def transfer_in_html(self, data_dict={}):
        ff = {"is_deleted": {"$ne": True}, "is_Disabled": {"$ne": True}}
        if self.current_admin_dict.get("role_code") in [ROlE_ALL.SUPERADMIN, ROlE_ALL.AGENTADMIN]:
            ff["admin_uuid"] = self.current_admin_dict["uuid"]
        elif self.current_admin_dict.get("role_code") in [ROlE_ALL.ADMINISTRATOR, ROlE_ALL.SYSTEMUSER]:
            ff["payer_uuids"] = {'$elemMatch': {'$eq': self.current_admin_dict["uuid"]}}
        else:
            return f'Error!'
        ff['uuid'] = {'$ne': data_dict.get('former_card_uuid')}

        cards = WithdrawalCardTable.find_many(ff, sort=[['location', -1]])
        card_cnt = len(cards)
        bankcard_html = ''
        for card in cards:
            self.get_deposit_cardinfo(card)

            bank_info = BankTable.find_one({'uuid': card.get('bank_uid')})
            bankcard_html += f'''
            <div class="card mb-2" bankcarduuid="{ card.get("uuid") }" onclick="seleBankCard_func($(this))" location="{card.get("location")}" bank_code="{bank_info.get("code")}" bank_name="{card.get("bank_name")}" account="{card.get("account")}" locationName={LOCATION_TYPE.name_dict[card.get("location")]}>
                <div class="row g-0 m-0" style = "cursor: pointer;">
                    <div class="col-3 text-center" style="display: flex; align-items: stretch; justify-content: center; padding: 0px; ">
                        <div style=" flex: 1; background-color: #ebeaea; display: flex; align-items: center; justify-content: center; ">
                            <h5 style="margin:0px">{bank_info.get("code")}({LOCATION_TYPE.name_dict[card.get("location")]})</h5>
                        </div>
                    </div>
                    <div class="col-9">
                        <div class="card-body" style="padding:10px">
                            银行卡账号: {card.get("account")} (名字: {card.get("bank_name")}, 余额: {self.format_money(card.get("balance_amount"))})
                        </div>
                    </div>
                </div>
            </div>
            '''
        cards_html = f'''
        <option value="{LOCATION_TYPE.TRANSIT}"> {LOCATION_TYPE.name_dict[LOCATION_TYPE.TRANSIT]}</option>
        <option value="{LOCATION_TYPE.WITHDRAW}" >{LOCATION_TYPE.name_dict[LOCATION_TYPE.WITHDRAW]}</option>
        <option value="{LOCATION_TYPE.DEPOSIT}" >{LOCATION_TYPE.name_dict[LOCATION_TYPE.DEPOSIT]}</option>
        <option value="{LOCATION_TYPE.OTHER}" >{LOCATION_TYPE.name_dict[LOCATION_TYPE.OTHER]}</option>
        '''
        bankcard_opt = ''
        bank_datas = BankTable.find_many({})
        for bdc in bank_datas:
            bankcard_opt += f'''<div class="auto-complete-item" onclick="onclick_insert_code('{bdc.get("code")}', filter = true)">{ bdc.get("code") }</div>'''

        html = f'''
            <h5 class="mb-0 text-left py-2 ml-5">订购的钱: <span class="text-danger" id="requested_amount">{self.format_money(data_dict.get("request_money"))}<span></h5>
            <div class="d-flex justify-content-between" style="margin-bottom:10px">
               <div style="margin-left:20px; ">
                    <span style="width: 80px; text-align: right; display: inline-block; position: relative;">转账金额:</span>
                    <input type="text" class="form-control" id="transfer_money" placeholder="转账金额" aria-label="" value = {self.format_money(data_dict.get("request_money"))} style="display: inline-block; width: calc(100% - 100px)" onchange = "onchange_number_input('transfer_money')">
                </div>
                <div style="width:250px" >
                    <span style="width: 40px; text-align: right; display: inline-block; position: relative;">备注:</span>
                    <input type="input" class="form-control" id="note" placeholder="备注" aria-label="" style="display: inline-block; width: calc(100% - 60px)">
                </div>
                <button class="btn btn-primary" style="margin-right:40px;" onclick = "addtransfercardlist()">将卡添加到列表</button>
            </div>
            <div class="d-flex justify-content-between" style="margin-bottom:3px; ">
                <select class="form-control mr-sm-2 mb-2" style="margin-left:40px; width: 200px;" aria-label="" id="location_filter" onchange="on_card_filter_change()">
                    <option value="">选择位置</option>
                   {cards_html}
                </select>
                <div class="auto-complete mr-sm-2 mb-2" >
                    <input type="text" class="form-control" id="bank_code" value="" placeholder="银行类型" aria-label="" onclick="onclick_bankcode()" oninput="on_card_filter_change()" >
                    <div class="auto-complete-items">
                    {bankcard_opt}
                    </div>
                </div>

                <input type="text" class="form-control mr-sm-2 mb-2" id="bankname_filter" placeholder="银行名字" aria-label="" style="display: inline-block; width: 200px; margin-right: 40px;" oninput="on_card_filter_change()" onchange="on_card_filter_change()">
                <input type="text" class="form-control" id="account_filter" placeholder="银行卡账号" aria-label="" style="display: inline-block; width: 200px; margin-right: 40px;" oninput="on_card_filter_change()" onchange="on_card_filter_change()">
            </div>
            <div class="formBox m-0" style="padding:0px;">
                <div class="list-group-item" style="display: flex; align-items: center;justify-content: center; padding-bottom:0px">
                    <input type="hidden" id="bank_card_uid">
                    <div class="form-control p-2" style="display: inline-block; width: calc(100% - 40px); height: 300px; overflow-y: scroll; text-align: left;" id='bankcard_list'>
                        { bankcard_html }                        
                    </div>
                </div>    
            <div>   
             <div class="formBox m-0" style="padding:0px;">
                <div class="list-group-item" style="display: flex; align-items: center;justify-content: center;">
                    <input type="hidden" id="bank_card_uid">
                    <div class="form-control p-2" style="display: inline-block; width: calc(100% - 40px); height: 110px; overflow-y: scroll; text-align: left;" id='transfer_card_list'>
                                                
                    </div>
                </div>    
            <div>  
            <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>
            <div class="d-flex justify-content-between px-5">
                <div>
                    <span>转账卡数：</span>
                    <span id="transfer_card_cnt">0</span>
                </div>
                <div style="position: relative; text-align: center">
                    <span class="btn btn-secondary mr-5" onclick="post_multi_qr_code('{data_dict.get('uuid')}', '{BEHAVIOR_TYPE.TRANSFER_IN}')">支付码</span>&emsp;
                    <span class="btn btn-primary" onclick="post_transfer_in_out('null', 'transfer_in_out', '{data_dict.get('uuid')}', '{BEHAVIOR_TYPE.TRANSFER_IN}')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="closeQRButton()">取消</span>
                </div>            
            </div>            
        '''
        return html


    def update_cardinfo(self, dd, data_from):
        if dd.get('location') != LOCATION_TYPE.OTHER:
            WithdrawalCardTable.update_one({ "uuid": dd.get("uuid")},{ "$set": data_from })
            return
        BankCardTable.update_one({'uuid': dd.get('deposit_uuid')},{ "$set": data_from })

    def get_deposit_cardinfo(self, dd):
        if dd.get('location') != LOCATION_TYPE.ANOTHER:
            return
        _bankcard = BankCardTable.find_one({'uuid': dd.get('deposit_uuid')})
        dd['account'] = _bankcard['account']
        dd['bank_name'] = _bankcard['name']
        dd['balance_amount'] = _bankcard['balance_amount']
        dd['bank_uid'] = _bankcard.get("bank_uid")
        dd['code'] = _bankcard.get("code")
        dd['balance_amount'] = _bankcard.get("balance_amount")
        dd['start_money'] = _bankcard.get("start_money")
        dd['note'] = _bankcard.get("note")

    def post_other_way(self):
        # 统计
        if self.action == 'get_total_info_html':
            cardinfo = {}
            filter_dict = self.get_filter_dict()

            result = list(WithdrawalOrderLogTable.collection().aggregate([
                {'$match':filter_dict },
                { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'request_money'}"}}}
            ]))
            cardinfo["request_money"] = result[0].get("transfer_total") if result else 0
            
            filter_dict["behavior"] = BEHAVIOR_TYPE.TRANSFER_IN
            result = list(WithdrawalOrderLogTable.collection().aggregate([
                {'$match':filter_dict },
                { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
            ]))
            cardinfo["transfer_in_money"] = result[0].get("transfer_total") if result else 0

            filter_dict["behavior"] = BEHAVIOR_TYPE.TRANSFER_OUT
            result = list(WithdrawalOrderLogTable.collection().aggregate([
                {'$match':filter_dict },
                { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
            ]))
            cardinfo["transfer_out_money"] = result[0].get("transfer_total") if result else 0

            filter_dict["behavior"] = BEHAVIOR_TYPE.OTHER_TRANSFER_IN
            result = list(WithdrawalOrderLogTable.collection().aggregate([
                {'$match':filter_dict },
                { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
            ]))
            cardinfo["other_transfer_in_money"] = result[0].get("transfer_total") if result else 0

            filter_dict["behavior"] = BEHAVIOR_TYPE.OTHER_TRANSFER_OUT
            result = list(WithdrawalOrderLogTable.collection().aggregate([
                {'$match':filter_dict },
                { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
            ]))
            cardinfo["other_transfer_out_money"] = result[0].get("transfer_total") if result else 0

            filter_dict["behavior"] = BEHAVIOR_TYPE.ISSUED
            result = list(WithdrawalOrderLogTable.collection().aggregate([
                {'$match':filter_dict },
                { '$group': { '_id': None, 'transfer_total': {'$sum': f"${'transfer_money'}"}}}
            ]))
            cardinfo["issued_money"] = result[0].get("transfer_total") if result else 0

            html = f"""
            <div class="formBox">
                <div style="position: relative; box-sizing: border-box; overflow-y: auto;">
                    <p style="margin-bottom: 20px; font-size: 18px; padding-left: 30px; text-align: left;">
                        <b>订购金额: </b> {cardinfo.get("request_money") or 0}
                    </p>
                    <p style="margin-bottom: 20px; font-size: 18px; padding-left: 30px; text-align: left;">
                        <b>转入金额: </b> {cardinfo.get("transfer_in_money") or 0 }
                    </p>
                    <p style="margin-bottom: 20px; font-size: 18px; padding-left: 30px; text-align: left;">
                        <b>转出金额:</b>{cardinfo.get("transfer_out_money") or 0 }
                    </p>
                    <p style="margin-bottom: 20px; font-size: 18px; padding-left: 30px; text-align: left;">
                        <b>其他转入金额	: </b> {cardinfo.get("other_transfer_in_money") or 0 }
                    </p>
                    <p style="margin-bottom: 20px; font-size: 18px; padding-left: 30px; text-align: left;">
                        <b>其他转出金额: </b> {cardinfo.get("other_transfer_out_money") or 0 }
                    </p>
                    <p style="margin-bottom: 20px; font-size: 18px; padding-left: 30px; text-align: left;">
                        <b>下发金额: </b> {cardinfo.get("issued_money") or 0 }
                    </p>
                </div>
            </div>"""
            return self.xtjson.json_result(message = update_language(self.current_admin_dict.get("language"), html))
        if self.action == 'get_selectCols_html':
            select_cols = session.get(SELECET_COLS_SESSION_KEY)
            if not select_cols:
                select_cols = {
                    "admin_check": "true",
                    "former_account_check" : "true", 
                    "former_name_check" : "true", 
                    "former_location_check" : "true", 
                    "former_pre_balance_check" : "true", 
                    "former_cur_balance_check" : "true", 
                    "latter_account_check" : "true", 
                    "latter_name_check" : "true", 
                    "latter_location_check" : "true", 
                    "latter_pre_balance_check" : "true", 
                    "latter_cur_balance_check" : "true", 
                    "request_money_check" : "true", 
                    "transfer_money_check" : "true", 
                    "order_time_check" :"true", 
                    "withdrw_time_check" :"true", 
                    "order_status_check" : "true", 
                    "accept_status_check" : "true", 
                    "behavior_check" : "true", 
                    "operator_check" : "true", 
                    "operator_check_agent" : "true", 
                    "note_check" :"true", 
                }
            session[SELECET_COLS_SESSION_KEY] = select_cols

            html = f'''
            <div class="d-flex flex-wrap align-content-start p-2">
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("admin_check") == 'true' else ''} id="admin_check"> 管理员
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("former_account_check") == 'true' else ''} id="former_account_check"> 卡银行卡账号
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("former_name_check") == 'true' else ''} id="former_name_check"> 银行名字
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("former_location_check") == 'true' else ''} id="former_location_check"> 位置
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("former_pre_balance_check") == 'true' else ''} id="former_pre_balance_check"> 先前余额
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("former_cur_balance_check") == 'true' else ''} id="former_cur_balance_check"> 当前余额
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("latter_account_check") == 'true' else ''} id="latter_account_check"> 卡银行卡账号
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("latter_name_check") == 'true' else ''} id="latter_name_check"> 银行名字
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("latter_location_check") == 'true' else ''} id="latter_location_check"> 位置
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("latter_pre_balance_check") == 'true' else ''} id="latter_pre_balance_check"> 先前余额
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("latter_cur_balance_check") == 'true' else ''} id="latter_cur_balance_check"> 当前余额
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("request_money_check") == 'true' else ''} id="request_money_check"> 要求的钱
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("transfer_money_check") == 'true' else ''} id="transfer_money_check"> 转移金额
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("order_time_check") == 'true' else ''} id="order_time_check"> 订单时间
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("withdrw_time_check") == 'true' else ''} id="withdrw_time_check"> 付款时间
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("order_status_check") == 'true' else ''} id="order_status_check"> 订单状态
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("accept_status_check") == 'true' else ''} id="accept_status_check"> 接受状态
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("behavior_check") == 'true' else ''} id="behavior_check"> 行为
                    </label>
                </div>
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("operator_check") == 'true' else ''} id="operator_check"> 接受员
                    </label>
                </div>                
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("operator_agent_check") == 'true' else ''} id="operator_agent_check"> 操作员
                    </label>
                </div>                
                <div class="form-group form-check p-2 m-2">
                    <label class="form-check-label">
                    <input class="form-check-input" type="checkbox" {'checked' if select_cols.get("note_check") == 'true' else ''} id="note_check"> 注解
                    </label>
                </div>
            </div>
            <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

            <div style="position: relative; text-align: center">
                <span class="btn btn-primary" onclick="post_select_cols()">确定</span>&emsp;
                <span class="btn btn-default" onclick="xtalert.close()">取消</span>
            </div> 
            '''            
            return self.xtjson.json_result(message = update_language(self.current_admin_dict.get("language"), html))
        if self.action == 'showcols':
            data = {}
            data["admin_check"] = self.request_data.get("admin_check")
            data["former_account_check"] = self.request_data.get("former_account_check")
            data["former_name_check"] = self.request_data.get("former_name_check")
            data["former_location_check"] = self.request_data.get("former_location_check")
            data["former_pre_balance_check"] = self.request_data.get("former_pre_balance_check")
            data["former_cur_balance_check"] = self.request_data.get("former_cur_balance_check")
            data["latter_account_check"] = self.request_data.get("latter_account_check")
            data["latter_name_check"] = self.request_data.get("latter_name_check")
            data["latter_location_check"] = self.request_data.get("latter_location_check")
            data["latter_pre_balance_check"] = self.request_data.get("latter_pre_balance_check")
            data["latter_cur_balance_check"] = self.request_data.get("latter_cur_balance_check")
            data["request_money_check"] = self.request_data.get("request_money_check")
            data["transfer_money_check"] = self.request_data.get("transfer_money_check")
            data["withdrw_time_check"] = self.request_data.get("withdrw_time_check")
            data["order_time_check"] = self.request_data.get("order_time_check")
            data["order_status_check"] = self.request_data.get("order_status_check")
            data["accept_status_check"] = self.request_data.get("accept_status_check")
            data["behavior_check"] = self.request_data.get("behavior_check")
            data["operator_check"] = self.request_data.get("operator_check")
            data["operator_agent_check"] = self.request_data.get("operator_agent_check")
            data["note_check"] = self.request_data.get("note_check")

            session[SELECET_COLS_SESSION_KEY] = data

            return self.xtjson.json_result()
        if self.action == 'clearlog':
            fff = self.get_filter_dict()
            self.MCLS.delete_many(fff)
            
            return self.xtjson.json_result()
        if self.action == 'test_card_html':
            html = self.test_card_html()
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        if self.action == 'test_tranfer_in_out':
            behavior = self.request_data.get("behavior_type") or ''
            note = self.request_data.get("note") or ''
            transfer_money = float(self.request_data.get("transfer_money").replace(',','')) or 0

            former_card = WithdrawalCardTable.find_one({"uuid": self.current_admin_dict.get("withdrawalcard_uuid")})
            latter_card = WithdrawalCardTable.find_one({"uuid": self.request_data.get("bankcarduuid")})

            self.get_deposit_cardinfo(latter_card)
            self.get_deposit_cardinfo(former_card)

            former_cur_balance = former_pre_balance = former_card.get("balance_amount") or 0
            latter_cur_balance = latter_pre_balance = latter_card.get("balance_amount") or 0
    
            former_bank_info = BankTable.find_one({'uuid': former_card.get('bank_uid')})
            latter_bank_info = BankTable.find_one({'uuid': latter_card.get('bank_uid')})

            if behavior == BEHAVIOR_TYPE.TRANSFER_IN:
                latter_cur_balance = latter_cur_balance - transfer_money
                self.update_cardinfo(latter_card, {"balance_amount": latter_cur_balance})
                self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_out": transfer_money})
            
            elif behavior == BEHAVIOR_TYPE.TRANSFER_OUT:
                former_cur_balance = former_pre_balance - transfer_money
                self.update_cardinfo(former_card, {"balance_amount": former_cur_balance})
                self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_out": transfer_money})
            elif behavior == BEHAVIOR_TYPE.ISSUED:
                return self.xtjson.json_params_error("Behavior Type error!")

            WithdrawalOrderLogTable.insert_one({
                "admin_uuid": self.current_admin_dict.get("uuid"),
                "payer_uuid": former_card.get("payer_uuids"),
                # "payer_uuid": CmsUserTable.find_one({"withdrawalcard_uuid": former_card.get("uuid")}).get("uuid") if former_card.get("location") == LOCATION_TYPE.WITHDRAW else former_card.get("payer_uuids"),

                "former_card_uuid": former_card.get("uuid"),
                "former_bank_code": former_bank_info.get("code"),
                "former_bank_name": former_card.get("bank_name"),
                "former_bankcard_account": former_card.get("account"),
                "former_location": former_card.get("location"),
                "former_cur_balance": former_cur_balance,
                "former_pre_balance": former_pre_balance,
                
                "latter_card_uuid": latter_card.get("uuid"),
                "latter_bank_code": latter_bank_info.get("code"),
                "latter_bank_name": latter_card.get("bank_name"),
                "latter_bankcard_account": latter_card.get("account"),
                "latter_location": latter_card.get("location"),
                "latter_cur_balance": latter_cur_balance,
                "latter_pre_balance": latter_pre_balance,
            
                "transfer_money":transfer_money,
                "request_money":0,
                "order_status":ORDER_STATUS.COMPLETED,
                "accepted_status":ACCEPTED_STATUS.NOT_PROCESSED ,
                # "pay_time": datetime.datetime.now(),
                'behavior': behavior,
                "note": note,
                'operator_agent_uuid': self.current_admin_dict.get('uuid')
            })
            return self.xtjson.json_result()
        if self.action == 'test_pay_qrcode':
            behavior = self.request_data.get('behavior')
            bankcarduuid = self.request_data.get('bankcarduuid')
            former_card = WithdrawalCardTable.find_one({"uuid": self.current_admin_dict.get("withdrawalcard_uuid")})
            
            addinfo = ''
            if behavior == BEHAVIOR_TYPE.TRANSFER_IN:
                bank_data = BankTable.find_one({'uuid': former_card.get('bank_uid')}) or {}
                username = former_card.get("bank_name")
                cardaccount = former_card.get("account")
                latter_card = WithdrawalCardTable.find_one({"uuid": bankcarduuid})
                latter_bankdata = BankTable.find_one({'uuid': latter_card.get('bank_uid')})
                addinfo = latter_card.get("bank_name") + f' ({latter_bankdata.get("code")})'
                out_card_code = latter_bankdata.get("code")

            else:
                latter_card = WithdrawalCardTable.find_one({"uuid": bankcarduuid})
                bank_data = BankTable.find_one({'uuid': latter_card.get("bank_uid")}) or {}
                username = latter_card.get("bank_name")
                cardaccount = latter_card.get("account")
                
                former_bank_data = BankTable.find_one({'uuid': former_card.get('bank_uid')}) or {}
                addinfo = former_card.get("bank_name") + f' {former_bank_data.get("code")}'
                out_card_code = former_bank_data.get("code")
            
            addinfo += ' to ' + username + f' {bank_data.get("code")}'
            if out_card_code == 'SHB':
                addinfo = addinfo[:25]
            transfer_money = self.request_data.get("transfer_money").replace(",", '') or '0'
            transfer_money = float(transfer_money)

            if not bank_data:
                return self.xtjson.json_params_error('该订单收款银行错误！')

            _data = {
                'receive_bank': bank_data.get('shortName'),
                'former_bankcard_account': cardaccount,
                'order_amount': transfer_money,
                "receive_owner": username,

                'uuid': shortuuid.uuid(),
                'accepted_status': ACCEPTED_STATUS.NOT_PROCESSED,
            }

            project_static_folder = os.path.join(current_app.static_folder, current_app.config.get('PROJECT_NAME'))

            _state, payQrcode = getWithdrawalBankPayQrcode(
                self.data_uuid,
                transfer_money,
                addinfo,
                bank_data,
                payqrcode_url='',
                project_static_folder=project_static_folder,
                receive_account=cardaccount,
            )

            if not _state:
                return self.xtjson.json_params_error(payQrcode)

            _data['payQrcode'] = payQrcode

            return self.xtjson.json_result(data=_data)
        
    def post_data_other_way(self):
        if self.action == 'withdrawSuccess':
            return self.xtjson.json_result()
        if self.action == 'retransfer_in_out':
            request_money = self.data_dict.get('request_money')
            old_transfer_money= self.data_dict.get("transfer_money")
            accepted_status= self.data_dict.get("accepted_status")
            behavior = self.data_dict.get("behavior") or ''
            
            former_card = WithdrawalCardTable.find_one({"uuid": self.data_dict.get("former_card_uuid")})
            old_latter_card = WithdrawalCardTable.find_one({"uuid": self.data_dict.get("latter_card_uuid")})

            former_cur_balance = former_card.get("balance_amount") or 0
            old_latter_cur_balance = old_latter_card.get("balance_amount") or 0

            if behavior == BEHAVIOR_TYPE.TRANSFER_IN:
                if accepted_status == ACCEPTED_STATUS.NOT_PROCESSED:
                    old_latter_cur_balance = old_latter_cur_balance + old_transfer_money
                    self.update_transfer_amount(old_latter_card.get("uuid"), {"total_transfer_out": -old_transfer_money})
                    self.update_cardinfo(old_latter_card, {"balance_amount": old_latter_cur_balance})
                elif accepted_status == ACCEPTED_STATUS.NOT_ACCEPTED:
                    old_latter_cur_balance = old_latter_cur_balance + old_transfer_money
                    self.update_transfer_amount(old_latter_card.get("uuid"), {"total_transfer_out": -old_transfer_money})
                    self.update_cardinfo(old_latter_card, {"balance_amount": old_latter_cur_balance})
                elif accepted_status == ACCEPTED_STATUS.ACCEPTED:
                    old_latter_cur_balance = old_latter_cur_balance + old_transfer_money
                    former_cur_balance = former_cur_balance - old_transfer_money
                    self.update_transfer_amount(old_latter_card.get("uuid"), {"total_transfer_out": -old_transfer_money})
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_in":  -old_transfer_money})
                    self.update_cardinfo(former_card, {"balance_amount": former_cur_balance})
                    self.update_cardinfo(old_latter_card, {"balance_amount": old_latter_cur_balance})
                    
            elif behavior == BEHAVIOR_TYPE.TRANSFER_OUT:
                if accepted_status == ACCEPTED_STATUS.NOT_PROCESSED:
                    former_cur_balance = former_cur_balance + old_transfer_money
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_out": -old_transfer_money})
                    self.update_cardinfo(former_card, {"balance_amount": former_cur_balance})
                elif accepted_status == ACCEPTED_STATUS.NOT_ACCEPTED:
                    former_cur_balance = former_cur_balance + old_transfer_money
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_out": -old_transfer_money})
                    self.update_cardinfo(former_card, {"balance_amount": former_cur_balance})
                elif accepted_status == ACCEPTED_STATUS.ACCEPTED:
                    former_cur_balance = former_cur_balance + old_transfer_money
                    latter_cur_balance = old_latter_cur_balance - old_transfer_money
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_out": -old_transfer_money})
                    self.update_transfer_amount(old_latter_card.get("uuid"), {"total_transfer_in":  -old_transfer_money})
                    self.update_cardinfo(former_card, {"balance_amount": former_cur_balance})
                    self.update_cardinfo(old_latter_card, {"balance_amount": old_latter_cur_balance})
                    
            elif behavior == BEHAVIOR_TYPE.OTHER_TRANSFER_IN:
                former_cur_balance = former_cur_balance - old_transfer_money
                self.update_transfer_amount(former_card.get("uuid"), {"total_other_transfer_in":  -old_transfer_money})
                self.update_cardinfo(former_card, {"balance_amount": former_cur_balance})

            elif behavior == BEHAVIOR_TYPE.OTHER_TRANSFER_OUT:
                former_cur_balance = former_cur_balance + old_transfer_money
                self.update_transfer_amount(former_card.get("uuid"), {"total_other_transfer_out":  -old_transfer_money})
                self.update_cardinfo(former_card, {"balance_amount": former_cur_balance})

            else:
                self.xtjson.json_params_error("Behavior Error!")
            
            ##############################################################################
            transfer_money = float(self.request_data.get("transfer_money").replace(',','')) or 0
            latter_card = WithdrawalCardTable.find_one({"uuid": self.request_data.get("bankcarduuid")})
            latter_bank_info = BankTable.find_one({'uuid': latter_card.get('bank_uid')})
            
            self.get_deposit_cardinfo(latter_card)

            former_pre_balance = former_cur_balance
            latter_cur_balance = latter_pre_balance = latter_card.get("balance_amount") or 0

            """ Transfer function """
            if behavior == BEHAVIOR_TYPE.TRANSFER_IN:
                if accepted_status == ACCEPTED_STATUS.NOT_PROCESSED:
                    latter_cur_balance = latter_cur_balance - transfer_money
                    self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_out": transfer_money})
                    self.update_cardinfo(latter_card, {"balance_amount": latter_cur_balance})
                elif accepted_status == ACCEPTED_STATUS.NOT_ACCEPTED:
                    latter_cur_balance = latter_cur_balance - transfer_money
                    self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_out": transfer_money})
                    self.update_cardinfo(latter_card, {"balance_amount": latter_cur_balance})
                elif accepted_status == ACCEPTED_STATUS.ACCEPTED:
                    latter_cur_balance = latter_cur_balance - transfer_money
                    former_cur_balance = former_cur_balance + transfer_money
                    self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_out": transfer_money})
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_in":  transfer_money})
                    self.update_cardinfo(latter_card, {"balance_amount": latter_cur_balance})
                    self.update_cardinfo(former_card, {"balance_amount": former_cur_balance})

            elif behavior == BEHAVIOR_TYPE.TRANSFER_OUT:
                if accepted_status == ACCEPTED_STATUS.NOT_PROCESSED:
                    former_cur_balance = former_cur_balance - transfer_money
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_out": transfer_money})
                    self.update_cardinfo(former_card, {"balance_amount": former_cur_balance})
                elif accepted_status == ACCEPTED_STATUS.NOT_ACCEPTED:
                    former_cur_balance = former_cur_balance - transfer_money
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_out": transfer_money})
                    self.update_cardinfo(former_card, {"balance_amount": former_cur_balance})
                elif accepted_status == ACCEPTED_STATUS.ACCEPTED:
                    former_cur_balance = former_cur_balance - transfer_money
                    latter_cur_balance = latter_cur_balance + transfer_money
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_out": transfer_money})
                    self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_in":  transfer_money})
                    self.update_cardinfo(latter_card, {"balance_amount": latter_cur_balance})
                    self.update_cardinfo(former_card, {"balance_amount": former_cur_balance})
            elif behavior == BEHAVIOR_TYPE.OTHER_TRANSFER_IN:
                former_cur_balance = former_cur_balance + old_transfer_money
                self.update_transfer_amount(former_card.get("uuid"), {"total_other_transfer_in":  old_transfer_money})
                self.update_cardinfo(former_card, {"balance_amount": former_cur_balance})

            elif behavior == BEHAVIOR_TYPE.OTHER_TRANSFER_OUT:
                former_cur_balance = former_cur_balance - old_transfer_money
                self.update_transfer_amount(former_card.get("uuid"), {"total_other_transfer_out":  old_transfer_money})
                self.update_cardinfo(former_card, {"balance_amount": former_cur_balance})
            else:
                self.xtjson.json.params_error("Behavior Error!")

            data_from = {
                "former_cur_balance": former_cur_balance,
                "former_pre_balance": former_pre_balance,
                
                "latter_card_uuid": latter_card.get("uuid"),
                "latter_bank_code": latter_bank_info.get("code"),
                "latter_bank_name": latter_card.get("bank_name"),
                "latter_location": latter_card.get("location"),
                "latter_bankcard_account": latter_card.get("account"),
                "latter_cur_balance": latter_cur_balance,
                "latter_pre_balance": latter_pre_balance,
            
                "transfer_money":transfer_money,
                "order_status":ORDER_STATUS.INCOMPLETED if transfer_money < request_money else ORDER_STATUS.COMPLETED,
                "pay_time": datetime.datetime.now(),
                'operator_agent_uuid': self.current_admin_dict.get('uuid')
         
            }
            if self.data_dict.get("order_time"):
                data_from["order_time"] = datetime.datetime.now()
            if self.data_dict.get("pay_time"):
                data_from["pay_time"] = datetime.datetime.now()

            WithdrawalOrderLogTable.update_one({"uuid": self.data_dict.get("uuid")}, {"$set": data_from})
            return self.xtjson.json_result()

        if self.action == 'transfer_in_out':
            data_uuid = self.data_uuid
            request_money = self.data_dict.get("request_money")
            behavior = self.request_data.get("behavior") or ''
            carddatas = self.request_data.get("carddatas") or ''
            carddatas = json.loads(carddatas)
            
            for carddata in carddatas:
                transfer_money = float(carddata.get("transfer_money").replace(',','')) or 0
                note = carddata.get("note") or ''
                latter_card = WithdrawalCardTable.find_one({"uuid": carddata.get("bankcarduuid")})
                former_card = WithdrawalCardTable.find_one({"uuid": self.data_dict.get("former_card_uuid")})

                self.get_deposit_cardinfo(latter_card)
                self.get_deposit_cardinfo(former_card)

                former_cur_balance = former_pre_balance = former_card.get("balance_amount") or 0
                latter_cur_balance = latter_pre_balance = latter_card.get("balance_amount") or 0
        
                former_bank_info = BankTable.find_one({'uuid': former_card.get('bank_uid')})
                latter_bank_info = BankTable.find_one({'uuid': latter_card.get('bank_uid')})

                if behavior == BEHAVIOR_TYPE.TRANSFER_IN:
                    latter_cur_balance = latter_cur_balance - transfer_money
                    self.update_cardinfo(latter_card, {"balance_amount": latter_cur_balance})
                    self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_out": transfer_money})

                data_from = self.data_dict
                data_from.pop("uuid")
                data_from.pop("_id")
                data_from.update({
                    "former_pre_balance": former_pre_balance,
                    "former_cur_balance": former_cur_balance,
                    "latter_card_uuid": latter_card.get("uuid"),
                    "latter_bank_code": latter_bank_info.get("code"),
                    "latter_bank_name": latter_card.get("bank_name"),
                    "latter_location": latter_card.get("location"),
                    "latter_bankcard_account": latter_card.get("account"),
                    "latter_cur_balance": latter_cur_balance,
                    "latter_pre_balance": latter_pre_balance,

                    "transfer_money":transfer_money,
                    "request_money": request_money,
                    "order_status":ORDER_STATUS.COMPLETED,
                    "accepted_status":ACCEPTED_STATUS.NOT_PROCESSED if self.current_admin_dict.get('role_code') in [ROlE_ALL.SUPERADMIN, ROlE_ALL.AGENTADMIN, ROlE_ALL.ADMINISTRATOR, ROlE_ALL.SYSTEMUSER] else ACCEPTED_STATUS.ACCEPTED,
                    "pay_time": datetime.datetime.now(),
                    'behavior': behavior,
                    "note": note,
                    'operator_agent_uuid': self.current_admin_dict.get('uuid')
                })
                request_money = request_money - transfer_money
                WithdrawalOrderLogTable.insert_one(data_from)

            if request_money > 0:
                WithdrawalOrderLogTable.update_one({"uuid": data_uuid}, {'$set':{ 
                    "request_money": request_money}})
            else:
                WithdrawalOrderLogTable.delete_one({"uuid": data_uuid})

            return self.xtjson.json_result()
        if self.action == 'transfer_in_out2':
            data_uuid = self.data_uuid
            request_money = self.data_dict.get("request_money")
            behavior = self.request_data.get("behavior") or ''
            transfer_money = float(self.request_data.get("transfer_money").replace(',','')) or 0
            note = self.request_data.get("note") or ''

            latter_card = WithdrawalCardTable.find_one({"uuid": self.request_data.get("bankcarduuid")})
            former_card = WithdrawalCardTable.find_one({"uuid": self.data_dict.get("former_card_uuid")})

            self.get_deposit_cardinfo(latter_card)
            self.get_deposit_cardinfo(former_card)

            former_cur_balance = former_pre_balance = former_card.get("balance_amount") or 0
            latter_cur_balance = latter_pre_balance = latter_card.get("balance_amount") or 0
    
            former_bank_info = BankTable.find_one({'uuid': former_card.get('bank_uid')})
            latter_bank_info = BankTable.find_one({'uuid': latter_card.get('bank_uid')})

            if behavior == BEHAVIOR_TYPE.TRANSFER_IN:
                latter_cur_balance = latter_cur_balance - transfer_money
                self.update_cardinfo(latter_card, {"balance_amount": latter_cur_balance})
                self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_out": transfer_money})
            
            data_from = self.data_dict
            data_from.pop("uuid")
            data_from.pop("_id")
            data_from.update({
                "former_pre_balance": former_pre_balance,
                "former_cur_balance": former_cur_balance,
                "latter_card_uuid": latter_card.get("uuid"),
                "latter_bank_code": latter_bank_info.get("code"),
                "latter_bank_name": latter_card.get("bank_name"),
                "latter_location": latter_card.get("location"),
                "latter_bankcard_account": latter_card.get("account"),
                "latter_cur_balance": latter_cur_balance,
                "latter_pre_balance": latter_pre_balance,

                "transfer_money":transfer_money,
                "order_status":ORDER_STATUS.COMPLETED,
                "accepted_status":ACCEPTED_STATUS.NOT_PROCESSED if self.current_admin_dict.get('role_code') in [ROlE_ALL.SUPERADMIN, ROlE_ALL.AGENTADMIN, ROlE_ALL.ADMINISTRATOR, ROlE_ALL.SYSTEMUSER] else ACCEPTED_STATUS.ACCEPTED,
                "pay_time": datetime.datetime.now(),
                'behavior': behavior,
                "note": note,
                'operator_agent_uuid': self.current_admin_dict.get('uuid')
            })

            WithdrawalOrderLogTable.insert_one(data_from)
            if request_money > transfer_money:
                WithdrawalOrderLogTable.update_one({'uuid': data_uuid}, {'$set':{
                    "request_money": request_money - transfer_money
                }})
            else:
                WithdrawalOrderLogTable.delete_one({"uuid": data_uuid})

            return self.xtjson.json_result()
        if self.action == 'retransfer_html':
            html = self.retransfer_html(self.data_dict)
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        if self.action == 'transfer_in_html':
            html = self.transfer_in_html(self.data_dict)
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        if self.action == 'accept_confirm':
            accepted_status = self.request_data.get("accepted_status")
            behavior = self.data_dict.get("behavior")
            accepted_status_old = self.data_dict.get("accepted_status")
            transfer_money= self.data_dict.get("transfer_money")
            former_card = WithdrawalCardTable.find_one({"uuid": self.data_dict.get("former_card_uuid")})
            latter_card = WithdrawalCardTable.find_one({"uuid": self.data_dict.get("latter_card_uuid")})
            if (not former_card) or (not latter_card):
                return self.xtjson.json_params_error('找不到卡!')
            former_cur_balance = former_pre_balance = former_card.get("balance_amount") or 0
            latter_cur_balance = latter_pre_balance = latter_card.get("balance_amount") or 0

            """ Initialize all of cards """
            if behavior == BEHAVIOR_TYPE.TRANSFER_IN:
                if accepted_status_old == ACCEPTED_STATUS.NOT_PROCESSED:
                    latter_cur_balance = latter_pre_balance + transfer_money
                    self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_out": -transfer_money})
                elif accepted_status_old == ACCEPTED_STATUS.NOT_ACCEPTED:
                    latter_cur_balance = latter_pre_balance + transfer_money
                    self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_out": -transfer_money})
                elif accepted_status_old == ACCEPTED_STATUS.ACCEPTED:
                    latter_cur_balance = latter_pre_balance + transfer_money
                    former_cur_balance = former_pre_balance - transfer_money
                    self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_out": -transfer_money})
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_in":  -transfer_money})
                    
            elif behavior == BEHAVIOR_TYPE.TRANSFER_OUT:
                if accepted_status_old == ACCEPTED_STATUS.NOT_PROCESSED:
                    former_cur_balance = former_pre_balance + transfer_money
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_out": -transfer_money})
                elif accepted_status_old == ACCEPTED_STATUS.NOT_ACCEPTED:
                    former_cur_balance = former_pre_balance + transfer_money
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_out": -transfer_money})
                elif accepted_status_old == ACCEPTED_STATUS.ACCEPTED:
                    former_cur_balance = former_pre_balance + transfer_money
                    latter_cur_balance = latter_pre_balance - transfer_money
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_out": -transfer_money})
                    self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_in":  -transfer_money})
            else:
                self.xtjson.json_params_error("Behavior Error!")

            """ Transfer function """
            if behavior == BEHAVIOR_TYPE.TRANSFER_IN:
                if accepted_status == ACCEPTED_STATUS.NOT_ACCEPTED:
                    latter_cur_balance = latter_cur_balance - transfer_money
                    self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_out": transfer_money})
                elif accepted_status == ACCEPTED_STATUS.ACCEPTED:
                    latter_cur_balance = latter_cur_balance - transfer_money
                    former_cur_balance = former_cur_balance + transfer_money
                    self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_out": transfer_money})
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_in":  transfer_money})

            elif behavior == BEHAVIOR_TYPE.TRANSFER_OUT:
                if accepted_status == ACCEPTED_STATUS.NOT_ACCEPTED:
                    former_cur_balance = former_cur_balance - transfer_money
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_out": transfer_money})
                elif accepted_status == ACCEPTED_STATUS.ACCEPTED:
                    former_cur_balance = former_cur_balance - transfer_money
                    latter_cur_balance = latter_cur_balance + transfer_money
                    self.update_transfer_amount(former_card.get("uuid"), {"total_transfer_out": transfer_money})
                    self.update_transfer_amount(latter_card.get("uuid"), {"total_transfer_in":  transfer_money})
            else:
                self.xtjson.json.params_error("Behavior Error!")

            """ complete transaction log """
            self.update_cardinfo(former_card,    {"balance_amount": former_cur_balance})
            self.update_cardinfo(latter_card   , {"balance_amount": latter_cur_balance})

            self.MCLS.update_one({"uuid": self.data_dict.get("uuid")}, {"$set": {
                "former_cur_balance": former_cur_balance,
                "former_pre_balance": former_pre_balance,
                
                "latter_cur_balance": latter_cur_balance,
                "latter_pre_balance": latter_pre_balance,
            
                "accepted_status":accepted_status,
                "pay_time": datetime.datetime.now(),
                'behavior': behavior,
                'operator_uuid': self.current_admin_dict.get("uuid"),
                "accepted_status":accepted_status,
                'operator_uuid': self.current_admin_dict.get("uuid"),
            }})
            return self.xtjson.json_result()
        
        if self.action == 'get_pay_qrcode':
            behavior = self.request_data.get('behavior')
            bankcarduuid = self.request_data.get('bankcarduuid')
            addinfo = ''
            if behavior == BEHAVIOR_TYPE.TRANSFER_IN:
                bank_data = BankTable.find_one({'code': self.data_dict.get('former_bank_code')}) or {}
                username = self.data_dict.get("former_bank_name")
                cardaccount = self.data_dict.get("former_bankcard_account")
                latter_card = WithdrawalCardTable.find_one({"uuid": bankcarduuid})
                latter_bankdata = BankTable.find_one({'uuid': latter_card.get('bank_uid')})
                addinfo = latter_card.get("bank_name") + f' ({latter_bankdata.get("code")})'
                out_card_code = latter_bankdata.get("code")

            else:
                latter_card = WithdrawalCardTable.find_one({"uuid": bankcarduuid})
                bank_data = BankTable.find_one({'uuid': latter_card.get("bank_uid")}) or {}
                username = latter_card.get("bank_name")
                cardaccount = latter_card.get("account")
                addinfo = self.data_dict.get("former_bank_name") + f' {self.data_dict.get("former_bank_code")}'
                out_card_code = self.data_dict.get("former_bank_code")
            
            addinfo += ' to ' + username + f' {bank_data.get("code")}'
            if out_card_code == 'SHB':
                addinfo = addinfo[:25]
            transfer_money = self.request_data.get("transfer_money").replace(",", '') or '0'
            transfer_money = float(transfer_money)

            if not bank_data:
                return self.xtjson.json_params_error('该订单收款银行错误！')

            _data = {
                'receive_bank': bank_data.get('shortName'),
                'former_bankcard_account': cardaccount,
                'order_amount': transfer_money,
                "receive_owner": username,

                'uuid': self.data_dict.get('uuid'),
                'accepted_status': self.data_dict.get('accepted_status') or '',
            }

            # if not self.data_dict.get( "payer_uuid"):
            #     return self.xtjson.json_params_error("未定义付款人!")
            # _data["receive_owner"]  = CmsUserTable.find_one({"uuid": self.data_dict.get("payer_uuid")}).get("account") or ''

            payqrcode_url = self.data_dict.get('payqrcode_url') or ''
            project_static_folder = os.path.join(current_app.static_folder, current_app.config.get('PROJECT_NAME'))

            _state, payQrcode = getWithdrawalBankPayQrcode(
                self.data_uuid,
                transfer_money,
                addinfo,
                bank_data,
                payqrcode_url='',
                project_static_folder=project_static_folder,
                receive_account=cardaccount,
            )

            if not _state:
                return self.xtjson.json_params_error(payQrcode)

            _data['payQrcode'] = payQrcode

            return self.xtjson.json_result(data=_data)
        if self.action == 'get_multi_qrcode':
            former_card = WithdrawalCardTable.find_one({"uuid": self.data_dict.get("former_card_uuid")})
            self.get_deposit_cardinfo(former_card)
            behavior = self.request_data.get('behavior') or ''
            carddatas = self.request_data.get('carddatas') or ''
            carddatas = json.loads(carddatas)
            
            html = f'''
            <div class="mb-2" style="text-align: center"><button class="btn btn-primary" onclick="post_transfer_in_out('qrcode', 'transfer_in_out', '{self.data_dict.get("uuid")}','{behavior}')">全部已支付</button></div>
            <div class="d-flex justify-content-between" style="overflow:scroll" id="qrboxlist">'''
            for carddata in carddatas:
                bankcarduuid = carddata.get('bankcarduuid') or ''
                
                if behavior == BEHAVIOR_TYPE.TRANSFER_IN or behavior == BEHAVIOR_TYPE.OTHER_TRANSFER_IN:
                    bank_data = BankTable.find_one({'uuid': former_card.get('bank_uid')}) or {}
                    username = former_card.get("bank_name")
                    cardaccount = former_card.get("account")
                    latter_card = WithdrawalCardTable.find_one({"uuid": bankcarduuid})
                    latter_bankdata = BankTable.find_one({'uuid': latter_card.get('bank_uid')})
                    addinfo = latter_card.get("bank_name") + f' ({latter_bankdata.get("code")})'
                    out_card_code = latter_bankdata.get("code")
                else:
                    return self.xtjson.json_params_error("Error transfer type!")

                addinfo += ' to ' + username + f' {bank_data.get("code")}'
                if out_card_code == 'SHB':
                    addinfo = addinfo[:25]
                transfer_money = carddata.get("transfer_money").replace(",", '') or '0'
                transfer_money = carddata.get("transfer_money").replace(",", '') or '0'
                transfer_money = float(transfer_money)
                if not bank_data:
                    return self.xtjson.json_params_error('该订单收款银行错误！')

                payqrcode_url = former_card.get('payqrcode_url') or ''
                project_static_folder = os.path.join(current_app.static_folder, current_app.config.get('PROJECT_NAME'))
                
                _state, payQrcode = getWithdrawalBankPayQrcode(
                    self.data_uuid,
                    transfer_money,
                    addinfo,
                    bank_data,
                    payqrcode_url='',
                    project_static_folder=project_static_folder,
                    receive_account=cardaccount,
                )

                if not _state:
                    return self.xtjson.json_params_error(payQrcode)
            
                html += f'''
                <div class="formBox" style="padding: 0px; margin: 0px 10px">
                    <div style="height: 35rem; position: relative; box-sizing: border-box; overflow-y: auto;">
                        <div>
                            <img src="{payQrcode}" alt="" style="display: block; width: 260px; margin: auto; margin-bottom: 30px;">
                            <p style="padding-left: 15%;text-align: left; margin-bottom: 10px; font-size: 14px;">收款银行：{bank_data.get('shortName')}</p>
                            <p style="padding-left: 15%;text-align: left; margin-bottom: 10px; font-size: 14px;">收款卡号：{cardaccount}</p>
                            <p style="padding-left: 15%;text-align: left; margin-bottom: 10px; font-size: 14px;">收款人：{username}</p>
                            <p style="padding-left: 15%;text-align: left; font-size: 14px;">订单金额：{transfer_money}</p>
                        </div>
                        <div class="blank" style="background: #eeeeee; height: 1px; margin: 10px 0 15px;"></div>
                        <div style="position: relative; text-align: center" bankcarduuid='{bankcarduuid}' transfer_money='{transfer_money}'  note='{carddata.get("note")}' paid="" class="qrboxlist">
                            <span class="btn btn-warning" style="color: white;" onclick="post_transfer_in_out(this, 'transfer_in_out2', '{self.data_dict.get("uuid")}','{behavior}')">已支付</span>
                        </div>
                    </div>
                </div>
                '''

            html += '</div>'
            return self.xtjson.json_result(message=html)
        if self.action == 'show_qrcode':
            addinfo = ''
            if self.data_dict.get("order_status") != ORDER_STATUS.COMPLETED:
                return self.xtjson.json_params_error('订单未处理！')

            behavior = self.data_dict.get("behavior")
            if behavior == BEHAVIOR_TYPE.TRANSFER_IN or behavior == BEHAVIOR_TYPE.OTHER_TRANSFER_IN:
                bank_data = BankTable.find_one({'code': self.data_dict.get('former_bank_code')}) or {}
                username = self.data_dict.get("former_bank_name")
                cardaccount = self.data_dict.get("former_bankcard_account")
                
                latter_card = WithdrawalCardTable.find_one({"uuid": self.data_dict.get("latter_card_uuid")})
                latter_bankdata = BankTable.find_one({'uuid': latter_card.get('bank_uid')})
                addinfo = latter_card.get("bank_name") + f' ({latter_bankdata.get("code")})'
                out_card_code = latter_bankdata.get("code")

            else:
                latter_card = WithdrawalCardTable.find_one({"uuid": self.data_dict.get("latter_card_uuid")})
                bank_data = BankTable.find_one({'uuid': latter_card.get("bank_uid")}) or {}
                username = latter_card.get("bank_name")
                cardaccount = latter_card.get("account")
                addinfo = self.data_dict.get("former_bank_name") + f' {self.data_dict.get("former_bank_code")}'
                out_card_code = self.data_dict.get("former_bank_code")
            
            addinfo += ' to ' + username + f' {bank_data.get("code")}'
            if out_card_code == 'SHB':
                addinfo = addinfo[:25]
            transfer_money = self.data_dict.get("transfer_money")

            if not bank_data:
                return self.xtjson.json_params_error('该订单收款银行错误！')

            # _data = {
            #     'receive_bank': bank_data.get('shortName'),
            #     'former_bankcard_account': cardaccount,
            #     'order_amount': transfer_money,
            #     "receive_owner": username,

            #     'uuid': self.data_dict.get('uuid'),
            #     'accepted_status': self.data_dict.get('accepted_status') or '',
            # }

            # if not self.data_dict.get( "payer_uuid"):
            #     return self.xtjson.json_params_error("未定义付款人!")
            # _data["receive_owner"]  = CmsUserTable.find_one({"uuid": self.data_dict.get("payer_uuid")}).get("account") or ''

            payqrcode_url = self.data_dict.get('payqrcode_url') or ''
            project_static_folder = os.path.join(current_app.static_folder, current_app.config.get('PROJECT_NAME'))

            _state, payQrcode = getWithdrawalBankPayQrcode(
                self.data_uuid,
                transfer_money,
                addinfo,
                bank_data,
                payqrcode_url=payqrcode_url,
                project_static_folder=project_static_folder,
                receive_account=cardaccount,
            )

            if not _state:
                return self.xtjson.json_params_error(payQrcode)

            html = f'''
            <div class="formBox" style="padding: 0px;">
                <div style="height: 24rem; position: relative; box-sizing: border-box; overflow-y: auto;">
                    <div class='mb-2'>
                        <img src="{payQrcode}" alt="" style="display: block; width: 260px; margin: auto; margin-bottom: 10px;">
                        <a href='{payQrcode}' target="_blank">QR Code</a>
                    </div>
                    <span class="btn btn-warning" style="color: white;" onclick="closeQRButton()">取消</span>
                </div>
            </div>
            '''

            return self.xtjson.json_result(message=html)
