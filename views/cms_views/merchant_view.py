import os
import string
import requests
import datetime, time, json
import random
import threading
import shortuuid
from decimal import Decimal
from openpyxl import Workbook
from flask import render_template, abort, request, url_for, current_app
from .cms_base import CmsTableViewBase, CmsFormViewBase
from models.pay_table import MerchantTable, MerchantTunnleTable, TunnelTable, CollectionOrderTable, BankCardTable, RechargeMoneyTable,  ReduceMoneyTable, MerchantLogTable, BankTable, MerchantBillStatementTable, callbackLogTable, WithdrawTable, unknownIncomeTable, AgentadminBillLogTable, MerchantBankCardTable
from models.cms_user import CmsUserTable
from modules.google_module.google_verify import GooleVerifyCls
from common_utils.utils_funcs import encry_md5, getMonthDateSilce, getDayDateSilce
from constants import TUNNLE_METHOD, PAY_METHOD, WITHDRAW_STATUS, CallbackState, CallbankType, BILL_STATEMEN_TYPES, ROlE_ALL, ExportStatu, ASSETS_FOLDER, EXPORT_FOLDER, MERCHANT_ROLES, BANK_CODE
from modules.view_helpres.view_func import CallbackPayOrderFunc, getBankPayQrcode, payIncome_addto
from common_utils.utils_funcs import PagingCLS, update_language
from models.site_table import ExportDataModel
from modules.view_helpres.view_func import MerchantUpdateAmout_func



class MerchantListView(CmsTableViewBase):
    add_url_rules = [['/merchantList', 'merchant_view']]
    per_page = 20
    MCLS = MerchantTable
    template = 'cms/merchant/merchant_list.html'
    title = '商户列表'

    def getOrderId(self, startText, mdl):
        while True:
            fff = list('ABCDEFGHJKLMNPQRSTUVWXYZABCDEFGHJKLMNPQRSTUVWXYZ')
            text = ''
            text += startText
            crr_Date = datetime.datetime.now()
            crr_day = crr_Date.day
            text += fff[crr_day]
            text += str(mdl)
            for i in range(7):
                ssd = list(string.digits)
                random.shuffle(ssd)
                text += random.choice(ssd)
            if CollectionOrderTable.find_one({'order_id': text}):
                continue
            return text

    def get_filter_dict(self):
        fff = {
            'is_review': True,
            'role_code': MERCHANT_ROLES.MERCHANT,
        }
        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')

        sort_type = request.args.get('sort_type')
        if sort_type and sort_type.strip():
            if len(sort_type.rsplit('_', 1)) == 2:
                filed, vv = sort_type.rsplit('_', 1)
                if vv == '0':
                    self.sort = [[filed, -1]]
                if vv == '1':
                    self.sort = [[filed, 1]]

        agentadmin_account = request.args.get('agentadmin_account') or ''
        if agentadmin_account and agentadmin_account.strip():
            agentadmin_data = CmsUserTable.find_one({'role_code': ROlE_ALL.AGENTADMIN, 'account': agentadmin_account.strip()}) or {}
            if agentadmin_data:
                fff['agentadmin_uuid'] = agentadmin_data.get('uuid')
            self.search_dict['agentadmin_account'] = agentadmin_account

        return fff

    def dealwith_main_context(self):
        search_res = self.context.get('search_res') or {}
        sort_type = request.args.get('sort_type')
        if sort_type and sort_type.strip():
            search_res['sort_type'] = sort_type
            self.context['search_res'] =search_res

        all_datas = self.context.get('all_datas') or []
        _datas = []
        agentadmin_datas = {}

        crr_time = datetime.datetime.now()
        start_time = datetime.datetime(crr_time.year, crr_time.month, crr_time.day, 0, 0, 0)
        end_time = datetime.datetime(crr_time.year, crr_time.month, crr_time.day, 23, 59, 59)
        for adl in all_datas:
            agentadmin_uuid = adl.get('agentadmin_uuid') or ''
            agentadmin_data = agentadmin_datas.get(agentadmin_uuid) or {}
            if not agentadmin_data:
                agentadmin_data = CmsUserTable.find_one({'uuid': agentadmin_uuid}) or {}
                agentadmin_datas[agentadmin_uuid] = agentadmin_data
            adl['agentadmin_data'] = agentadmin_data

            # ds_success_money = '0'
            # ds_success_rate = '0%'
            # orderdatas = CollectionOrderTable.find_many({'order_time': {'$gte': start_time, '$lte': end_time}, 'merchant_id': adl.get('merchant_id'), 'agentadmin_uuid': adl.get('agentadmin_uuid')})
            # if orderdatas:
            #     success_count = 0
            #     failed_count = 0
            #     total_money = 0
            #     for odata in orderdatas:
            #         if odata.get('pay_statu'):
            #             success_count += 1
            #             total_money += odata.get('actual_amount') or 0
            #         else:
            #             failed_count += 1
            #     if success_count == 0:
            #         ds_success_rate = 0
            #     else:
            #         ds_success_rate = round(success_count / (success_count + failed_count) * 100, 3)
            #
            #     ds_success_money = str(total_money)
            #     ds_success_rate = str(ds_success_rate) + '%'
            # adl['ds_success_money'] = float(ds_success_money)
            # adl['ds_success_rate'] = ds_success_rate
            _datas.append(adl)

        # sort_type = request.args.get('sort_type')
        # if sort_type and sort_type == 'ds_success_money':
        #     _datas = sorted(_datas, key=lambda x: x['ds_success_money'], reverse=True)
            
        self.context['all_datas'] = _datas
        self.context['req_host'] = str(request.host_url).strip('/')

    def cz_html(self):
        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">       
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">商户Id：</span>
                        <span style="display: inline-block;width: calc(100% - 180px);text-align: left;font-size: 16px;vertical-align: -1px;">{ self.data_dict.get('merchant_id') }</span>
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">商户名称：</span>
                        <span style="display: inline-block;width: calc(100% - 180px);text-align: left;font-size: 16px;vertical-align: -1px;">{ self.data_dict.get('merchant_name') }</span>
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">充值金额：</span>
                        <input type="number" class="form-control" id="cz_amount" value="" placeholder="充值金额" aria-label="" style="display: inline-block; width: calc(100% - 180px)" onchange="monitorRechargeMoney()">
                    </div>                    
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">内充利率%：</span>
                        <input type="number" class="form-control" id="recharge_money_rate" value="{ float(Decimal(str(self.data_dict.get('recharge_money_rate') or 0)) * Decimal('100')) if self.data_dict.get('recharge_money_rate') else 0 }" placeholder="内充利率%" aria-label="" style="display: inline-block; width: calc(100% - 180px)" onchange="monitorRechargeMoney()">
                    </div>     
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">手续费：</span>
                        <span style="display: inline-block;width: calc(100% - 180px);text-align: left;font-size: 16px;vertical-align: -1px;"><span id="repay_amount_text">0</span>元</span>
                    </div>                                   
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">备注：</span>
                        <input type="text" class="form-control" id="note" value="" placeholder="备注" aria-label="" style="display: inline-block; width: calc(100% - 180px)" onchange="monitorRechargeMoney()">
                    </div>
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_recharge_money('{self.data_dict.get("uuid")}')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>
        '''
        return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))

    def reduce_html(self):
        html = f'''
            <div class="formBox">
                <div style="height: 15rem; position: relative; box-sizing: border-box; overflow-y: auto;">       
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">商户Id：</span>
                        <span style="display: inline-block;width: calc(100% - 180px);text-align: left;font-size: 16px;vertical-align: -1px;">{ self.data_dict.get('merchant_id') }</span>
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">商户名称：</span>
                        <span style="display: inline-block;width: calc(100% - 180px);text-align: left;font-size: 16px;vertical-align: -1px;">{ self.data_dict.get('merchant_name') }</span>
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">减金额：</span>
                        <input type="text" class="form-control" id="reduce_amount" value="" placeholder="减金额" aria-label="" style="display: inline-block; width: calc(100% - 180px)" onchange="onchange_number_input('reduce_amount')">
                    </div>                    
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">备注：</span>
                        <input type="text" class="form-control" id="note" value="" placeholder="备注" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_reduce_money('{self.data_dict.get("uuid")}')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>
        '''
        return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))

    def get_data_id(self):
        while True:
            id_str = '16' +str(random.random() * 100000).split('.')[0]
            if self.MCLS.find_one({'account': id_str}):
                continue
            return id_str

    def form_merchant_html(self, data_dict={}):
        _action = 'add_merchant_data'
        if data_dict:
            _action = 'edit_merchant_data'

        bddl_html = ''
        for uda in CmsUserTable.find_many({'role_code': ROlE_ALL.AGENTADMIN}):
            bddl_html += f'<option value="{uda.get("uuid")}">{uda.get("account")}-{uda.get("username")}</option>'

        select_dl_html = ''
        if not data_dict and (self.current_admin_user.is_superadmin or self.current_admin_user.is_administrator):
            select_dl_html += f'''
                <div class="list-group-item">
                    <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">选择绑定代理：</span>
                    <select class="form-control" id="agentadmin_uuid" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                        <option value="">选择绑定代理</option>
                        {bddl_html}
                    </select>
                </div>                   
            '''

        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">
                    {select_dl_html}       
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">商户名称：</span>
                        <input type="text" class="form-control" id="merchant_name" value="{ data_dict.get('merchant_name') or ''}" placeholder="商户名称" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">账户：</span>
                        <input type="text" class="form-control" id="account" value="{ data_dict.get('account') or ''}" placeholder="账户" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span {'' if data_dict else 'class="loglable"'} style="width: 120px; text-align: right; display: inline-block; position: relative;">密码：</span>
                        <input type="text" class="form-control" id="password" placeholder="密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span {'' if data_dict else 'class="loglable"'} style="width: 120px; text-align: right; display: inline-block; position: relative;">确认密码：</span>
                        <input type="text" class="form-control" id="confirm_password" placeholder="密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                  
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">代付默认利率%：</span>
                        <input type="number" class="form-control" id="payment_rate" value="{ round(data_dict.get('payment_rate') *100, 5) if data_dict.get('payment_rate') else '' }" placeholder="代付默认利率%" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                        
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">下发默认利率%：</span>
                        <input type="number" class="form-control" id="issued_money_rate" value="{ round(data_dict.get('issued_money_rate') *100, 5) if data_dict.get('issued_money_rate') else '' }" placeholder="下发默认利率%" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">内充默认利率%：</span>
                        <input type="number" class="form-control" id="recharge_money_rate" value="{ round(data_dict.get('recharge_money_rate') *100, 5) if data_dict.get('recharge_money_rate') else '' }" placeholder="内充默认利率%" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">代付金额最小值：</span>
                        <input type="number" class="form-control" id="paybehalf_min_money" value="{ int(data_dict.get('paybehalf_min_money') or 10000) }" placeholder="代付金额最小值" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">代付金额最大值：</span>
                        <input type="number" class="form-control" id="paybehalf_max_money" value="{ int(data_dict.get('paybehalf_max_money') or 300000000) }" placeholder="代付金额最大值" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">备注：</span>
                        <input type="text" class="form-control" id="note" value="{data_dict.get('note') or ''}" placeholder="备注" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>     
                    <div class="list-group-item" style="display: flex; justify-content: center;">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">代付IP白名单：</span>
                        <textarea class="form-control" id="ip_whitelist" rows="5" placeholder="IP白名单/一行一个" style="display: inline-block; width: calc(100% - 180px)">{data_dict.get('ip_whitelist') or ''}</textarea>                
                    </div>       
                    <div class="list-group-item" style="display: flex;align-items: center; justify-content:center; margin-top: 0;margin-bottom: 0; padding-top: 0;padding-bottom: 0;">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">代收功能开关：</span>
                        <input type="hidden" alt="" aria-label="" value="{ '1' if data_dict.get('collect_money_switch') else '0' }" id="collect_money_switch">
                        <div style="display: inline-block; width: calc(100% - 180px); text-align: left;">
                            <i class="iconfont { 'icon-kaiguan4' if data_dict.get('collect_money_switch') else 'icon-kaiguanguan' } pointer" style="font-size: 40px;" onclick="switch_func($(this))"></i>                        
                        </div>
                    </div>                                      
                    <div class="list-group-item" style="display: flex;align-items: center; justify-content:center; margin-top: 0;margin-bottom: 0; padding-top: 0;padding-bottom: 0;">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">代付功能开关：</span>
                        <input type="hidden" alt="" aria-label="" value="{ '1' if data_dict.get('paybehalf_switch') else '0' }" id="paybehalf_switch">
                        <div style="display: inline-block; width: calc(100% - 180px); text-align: left;">
                            <i class="iconfont { 'icon-kaiguan4' if data_dict.get('paybehalf_switch') else 'icon-kaiguanguan' } pointer" style="font-size: 40px;" onclick="switch_func($(this))"></i>                        
                        </div>
                    </div>                                      
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_merchant_data('{_action}', '{data_dict.get('uuid') if data_dict else ''}')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>
        '''
        return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))

    def add_sub_html(self):
        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">账户名称：</span>
                        <input type="text" class="form-control" id="account_name" value="" placeholder="商户名称" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">账户：</span>
                        <input type="text" class="form-control" id="account" value="" placeholder="账户" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">密码：</span>
                        <input type="text" class="form-control" id="password" placeholder="密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">确认密码：</span>
                        <input type="text" class="form-control" id="confirm_password" placeholder="密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                  
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">备注：</span>
                        <input type="text" class="form-control" id="note" value="" placeholder="备注" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                                    
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_sub_data('{ self.data_uuid }')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>
        '''
        return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))

    def post_other_way(self):
        if self.action == 'add_merchant_html':
            return self.form_merchant_html()
        if self.action == 'add_merchant_data':
            account = self.request_data.get('account')
            merchant_name = self.request_data.get('merchant_name')
            password = self.request_data.get('password')
            note = self.request_data.get('note')
            issued_money_rate = self.request_data.get('issued_money_rate')
            payment_rate = self.request_data.get('payment_rate')
            recharge_money_rate = self.request_data.get('recharge_money_rate')
            ip_whitelist = self.request_data.get('ip_whitelist') or ''
            paybehalf_switch = self.request_data.get('paybehalf_switch') or ''
            collect_money_switch = self.request_data.get('collect_money_switch') or ''
            
            if not account or not merchant_name or not password or not account.strip() or not merchant_name.strip() or not password.strip():
                return self.xtjson.json_params_error('缺少数据！')

            if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
                agentadmin_uuid = self.current_admin_dict.get('uuid')
            elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
                agentadmin_uuid = self.current_admin_dict.get('agentadmin_uuid')
            elif self.current_admin_user.is_superadmin or self.current_admin_user.is_administrator:
                agentadmin_uuid = self.request_data.get('agentadmin_uuid')
                if not agentadmin_uuid:
                    return self.xtjson.json_params_error('请选择绑定的代理！')
            else:
                return self.xtjson.json_params_error('添加失败！')

            m_data = self.MCLS.find_one({'account': account.strip()})
            if m_data:
                return self.xtjson.json_params_error(f'账户：{account},已存在！')

            m_data = self.MCLS.find_one({'merchant_name': merchant_name.strip()})
            if m_data:
                return self.xtjson.json_params_error(f'商户名称：{merchant_name},已存在！')

            if issued_money_rate:
                try:
                    issued_money_rate = float(issued_money_rate) or 0
                except:
                    return self.xtjson.json_params_error('issued_money_rate: 参数错误！')
            else:
                issued_money_rate = 0

            if payment_rate:
                try:
                    payment_rate = float(payment_rate) or 0
                except:
                    return self.xtjson.json_params_error('issued_money_rate: 参数错误！')
            else:
                payment_rate = 0

            if recharge_money_rate:
                try:
                    recharge_money_rate = float(recharge_money_rate) or 0
                except:
                    return self.xtjson.json_params_error('recharge_money_rate: 参数错误！')
            else:
                recharge_money_rate = 0

            paybehalf_max_money = self.request_data.get('paybehalf_max_money') or 300000000
            paybehalf_min_money = self.request_data.get('paybehalf_min_money') or 10000
            try:
                paybehalf_max_money = int(paybehalf_max_money)
            except:
                return self.xtjson.json_params_error('代付金额最大值：错误！')
            try:
                paybehalf_min_money = int(paybehalf_min_money)
            except:
                return self.xtjson.json_params_error('代付金额最小值：错误！')

            uuid = shortuuid.uuid()
            secret_key = encry_md5(uuid)

            _paybehalf_switch = False
            if paybehalf_switch == '1':
                _payment_rate = True
            self.data_from['paybehalf_switch'] = _paybehalf_switch

            _collect_money_switch = False
            if collect_money_switch == '1':
                _collect_money_switch = True
            self.data_from['collect_money_switch'] = _collect_money_switch

            _data = {
                'uuid': uuid,
                'secret_key': secret_key,
                'merchant_id': self.get_data_id(),
                'account': account.strip(),
                'merchant_name': merchant_name.strip(),
                'ip_whitelist': ip_whitelist.strip(),
                'password': self.MCLS.encry_password(password.strip()),
                'note': note or '',
                'balance_amount': 0,
                'agentadmin_uuid': agentadmin_uuid,
                'statu': True,
                'is_activate': False,
                'is_review': True,
                'issued_money_rate': issued_money_rate / 100,
                'payment_rate': payment_rate / 100,
                'recharge_money_rate': recharge_money_rate / 100,
                'collect_money_switch': True,
                'paybehalf_switch': True,
                'paybehalf_max_money': paybehalf_max_money,
                'paybehalf_min_money': paybehalf_min_money,
                'role_code': MERCHANT_ROLES.MERCHANT,
            }
            _data.update(self.data_from)
            merchant_uuid = self.MCLS.insert_one(_data)

            for tt in TunnelTable.find_many({}):
                _data = {
                    'tunnle_id': tt.get('uuid'),
                    'statu': False,
                    'single_amount_max': 0,
                    'single_amount_min': 0,
                    'merchant_uuid': merchant_uuid,
                    'rate': 0,
                    'tunnle_method': TUNNLE_METHOD.collection,
                }
                MerchantTunnleTable.insert_one(_data)
            return self.xtjson.json_result()

    def post_data_other_way(self):
        if self.action == 'edit_merchant_html':
            return self.form_merchant_html(self.data_dict)
        if self.action == 'edit_merchant_data':
            account = self.request_data.get('account')
            merchant_name = self.request_data.get('merchant_name')
            password = self.request_data.get('password')
            note = self.request_data.get('note')
            issued_money_rate = self.request_data.get('issued_money_rate')
            ip_whitelist = self.request_data.get('ip_whitelist') or ''
            payment_rate = self.request_data.get('payment_rate')
            recharge_money_rate = self.request_data.get('recharge_money_rate')
            paybehalf_switch = self.request_data.get('paybehalf_switch') or ''
            collect_money_switch = self.request_data.get('collect_money_switch') or ''
            if not account or not merchant_name or not account.strip() or not merchant_name.strip():
                return self.xtjson.json_params_error('缺少数据！')

            m_data = self.MCLS.find_one({'account': account.strip()})
            if m_data and m_data.get('uuid') != self.data_uuid:
                return self.xtjson.json_params_error(f'账户：{account},已存在！')

            m_data = self.MCLS.find_one({'merchant_name': merchant_name.strip()})
            if m_data and m_data.get('uuid') != self.data_uuid:
                return self.xtjson.json_params_error(f'商户名称：{merchant_name},已存在！')

            if issued_money_rate:
                try:
                    issued_money_rate = float(issued_money_rate) or 0
                except:
                    return self.xtjson.json_params_error('issued_money_rate: 参数错误！')
            else:
                issued_money_rate = 0

            if payment_rate:
                try:
                    payment_rate = float(payment_rate) or 0
                except:
                    return self.xtjson.json_params_error('issued_money_rate: 参数错误！')
            else:
                payment_rate = 0

            if recharge_money_rate:
                try:
                    recharge_money_rate = float(recharge_money_rate) or 0
                except:
                    return self.xtjson.json_params_error('recharge_money_rate: 参数错误！')
            else:
                recharge_money_rate = 0

            paybehalf_max_money = self.request_data.get('paybehalf_max_money') or 300000000
            paybehalf_min_money = self.request_data.get('paybehalf_min_money') or 10000
            try:
                paybehalf_max_money = int(paybehalf_max_money)
            except:
                return self.xtjson.json_params_error('代付金额最大值：错误！')
            try:
                paybehalf_min_money = int(paybehalf_min_money)
            except:
                return self.xtjson.json_params_error('代付金额最小值：错误！')

            _paybehalf_switch = False
            if paybehalf_switch == '1':
                _paybehalf_switch = True
            self.data_from['paybehalf_switch'] = _paybehalf_switch

            _collect_money_switch = False
            if collect_money_switch == '1':
                _collect_money_switch = True
            self.data_from['collect_money_switch'] = _collect_money_switch

            _data = {
                'account': account.strip(),
                'merchant_name': merchant_name.strip(),
                'ip_whitelist': ip_whitelist.strip(),
                'note': note or '',
                'issued_money_rate': issued_money_rate / 100,
                'payment_rate': payment_rate / 100,
                'recharge_money_rate': recharge_money_rate / 100,
                'paybehalf_max_money': paybehalf_max_money,
                'paybehalf_min_money': paybehalf_min_money,
            }
            if password:
                _data['password'] = self.MCLS.encry_password(password.strip())
            _data.update(self.data_from)
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': _data})
            return self.xtjson.json_result()
        if self.action == 'getGoogleQrcode':
            if not self.data_uuid:
                return self.xtjson.json_params_error()

            user_dict = self.MCLS.find_one({'uuid': self.data_uuid})
            if not user_dict:
                return self.xtjson.json_params_error()

            google_cls = GooleVerifyCls(pwd=self.data_uuid, account=user_dict.get('account'), s_label='pay2wold')
            generate_qrcode = google_cls.secret_generate_qrcode()
            return self.xtjson.json_result(data={'generate_qrcode': generate_qrcode})
        if self.action == 'update_statu':
            if self.data_dict.get('statu'):
                self.data_from['statu'] = False
            else:
                self.data_from['statu'] = True
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            return self.xtjson.json_result()
        if self.action == 'cz_html':
            return self.cz_html()
        if self.action == 'reduce_html':
            return self.reduce_html()
        if self.action == 'rechargeMoney':
            note = self.request_data.get('note') or ''
            cz_amount = self.request_data.get('cz_amount')
            recharge_money_rate = self.request_data.get('recharge_money_rate')

            if recharge_money_rate:
                try:
                    recharge_money_rate = float(recharge_money_rate) or 0
                except:
                    return self.xtjson.json_params_error('recharge_money_rate: 参数错误！')
            else:
                recharge_money_rate = 0

            if cz_amount:
                try:
                    cz_amount = float(cz_amount) or 0
                except:
                    return self.xtjson.json_params_error('cz_amount: 参数错误！')
            else:
                cz_amount = 0

            if cz_amount <= 0:
                return self.xtjson.json_params_error('充值金额要大于0！')

            repay_amount = 0
            if recharge_money_rate:
                repay_amount = round(float(cz_amount * recharge_money_rate / 100), 2)

            zs_cz_money = cz_amount - repay_amount
            _balance_amount = self.data_dict.get('balance_amount') or 0

            _state, _balance_amount = MerchantUpdateAmout_func(zs_cz_money, self.data_uuid)
            if not _state:
                return self.xtjson.json_params_error('金额变动失败！')

            _RechargeMoney = {
                'operate_user_uuid': self.current_admin_dict.get('uuid'),
                'merchant_id': self.data_dict.get('merchant_id'),
                'amount': cz_amount,
                'actual_amount': zs_cz_money,
                'repay_amount': repay_amount,
                'note': note.strip(),
                'agentadmin_uuid': self.data_dict.get('agentadmin_uuid'),
            }
            RechargeMoneyTable.insert_one(_RechargeMoney)

            _mbs_data = {
                'merchant_uuid': self.data_uuid,
                'amount': int(zs_cz_money),
                'balance_amount': _balance_amount,
                'note': note,
                'repay_amount': repay_amount,
                'bill_type': BILL_STATEMEN_TYPES.RECHARGE,
                'agentadmin_uuid': self.data_dict.get('agentadmin_uuid'),
                'order_id': self.getOrderId('C', self.data_dict.get('merchant_id')[-3:]),
            }
            MerchantBillStatementTable.insert_one(_mbs_data)

            return self.xtjson.json_result()
        if self.action == 'reduceMoney':
            note = self.request_data.get('note') or ''
            reduce_amount = self.request_data.get('reduce_amount')

            if not reduce_amount:
                return self.xtjson.json_params_error()
            try:
                aumot = int(reduce_amount.replace(',', '').replace(' ',''))
            except:
                return self.xtjson.json_params_error()
            if aumot <= 0:
                return self.xtjson.json_params_error()
            balance_amount = self.data_dict.get('balance_amount')
            if balance_amount < aumot:
                return self.xtjson.json_params_error('余额不足！')

            _state, _balance_amount = MerchantUpdateAmout_func(aumot, self.data_uuid, is_add=False)
            if not _state:
                return self.xtjson.json_params_error('金额变动失败，'+str(_balance_amount))

            _ReduceMoney = {
                'operate_user_uuid': self.current_admin_dict.get('uuid'),
                'merchant_id': self.data_dict.get('merchant_id'),
                'amount': aumot,
                'note': note.strip(),
                'agentadmin_uuid': self.data_dict.get('agentadmin_uuid'),
            }
            ReduceMoneyTable.insert_one(_ReduceMoney)

            _mbs_data = {
                'merchant_uuid': self.data_uuid,
                'amount': int('-'+str(int(aumot))),
                'balance_amount': _balance_amount,
                'note': note,
                'repay_amount': 0,
                'bill_type': BILL_STATEMEN_TYPES.REDUCE,
                'agentadmin_uuid': self.data_dict.get('agentadmin_uuid'),
                'order_id': self.getOrderId('J', self.data_dict.get('merchant_id')[-3:]),
            }
            MerchantBillStatementTable.insert_one(_mbs_data)
            return self.xtjson.json_result()
        if self.action == 'del':
            MerchantBillStatementTable.delete_many({'merchant_uuid': self.data_uuid})
            CollectionOrderTable.delete_many({'merchant_id': self.data_dict.get('merchant_id')})
            RechargeMoneyTable.delete_many({'merchant_id': self.data_dict.get('merchant_id')})
            callbackLogTable.delete_many({'merchant_uuid': self.data_uuid})
            WithdrawTable.delete_many({'merchant_uuid': self.data_uuid})
            MerchantTunnleTable.delete_many({'merchant_uuid': self.data_uuid})
            MerchantBankCardTable.delete_many({'merchant_uuid': self.data_uuid})
            MerchantLogTable.delete_many({'merchant_uuid': self.data_uuid})
            self.MCLS.delete_one({'uuid': self.data_uuid})
            return self.xtjson.json_result()
        if self.action == 'add_sub_html':
            return self.add_sub_html()
        if self.action == 'add_sub_data':
            account = self.request_data.get('account')
            account_name = self.request_data.get('account_name')
            password = self.request_data.get('password')
            note = self.request_data.get('note') or ''
            if not account or not account_name or not password:
                return self.xtjson.json_params_error('缺少数据！')
            if self.MCLS.find_one({'account':account.strip(), 'upper_mid': self.data_uuid}):
                return self.xtjson.json_params_error('该账户已存在！')
            if self.MCLS.find_one({'account_name':account_name.strip(), 'upper_mid': self.data_uuid}):
                return self.xtjson.json_params_error('该账户名称已存在！')
            if not password.isalnum():
                return self.xtjson.json_params_error('密码格式错误！')
            if len(password.strip()) < 5 or len(password.strip()) > 15:
                return self.xtjson.json_params_error('密码长度在5~15位之间！')
            _data = {
                'account': account,
                'account_name': account_name,
                'password': self.MCLS.encry_password(password),
                'note': note.strip(),
                'role_code': MERCHANT_ROLES.SUBMERCHANT,
                'upper_mid': self.data_uuid,
                'statu': True
            }
            self.MCLS.insert_one(_data)
            return self.xtjson.json_result()



class SubMerchantView(CmsTableViewBase):
    add_url_rules = [['/subMerchantList/<string:muid>', 'SubMerchantView']]
    per_page = 20
    MCLS = MerchantTable
    template = 'cms/merchant/subMerchantList.html'
    title = '商户子账户列表'

    def edit_submerchant_html(self, data_dict):
        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">账户名称：</span>
                        <input type="text" class="form-control" id="account_name" value="{ data_dict.get('account_name') or '' }" placeholder="商户名称" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">账户：</span>
                        <input type="text" class="form-control" id="account" value="{ data_dict.get('account') or '' }" placeholder="账户" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">密码：</span>
                        <input type="text" class="form-control" id="password" value="" placeholder="密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">确认密码：</span>
                        <input type="text" class="form-control" id="confirm_password" value="{ data_dict.get('confirm_password') or '' }" placeholder="密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">备注：</span>
                        <input type="text" class="form-control" id="note" value="" placeholder="备注" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                                    
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_sub_data('{ self.data_uuid }')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>
        '''
        return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))

    def get_filter_dict(self):
        muid = self.kwargs.get('muid') or ''
        if not muid:
            return abort(404)
        fff = {
            'upper_mid': muid,
            'role_code': MERCHANT_ROLES.SUBMERCHANT,
        }
        return fff

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas') or []
        _datas = []
        m_datas = {}
        for adl in all_datas:
            upper_mid = adl.get('upper_mid') or ''
            upper_data = adl.get(upper_mid) or {}
            if not upper_data:
                upper_data = MerchantTable.find_one({'uuid': upper_mid}) or {}
                m_datas[upper_mid] = upper_data
            adl['upper_data'] = upper_data
            _datas.append(adl)
        self.context['all_datas'] = _datas

    def post_data_other_way(self):
        if self.action == 'edit_submerchant_html':
            return self.edit_submerchant_html(self.data_dict)
        if self.action == 'edit_sub_data':
            account = self.request_data.get('account')
            account_name = self.request_data.get('account_name')
            password = self.request_data.get('password')
            note = self.request_data.get('note') or ''
            if not account or not account_name:
                return self.xtjson.json_params_error('缺少数据！')
            if password:
                if not password.isalnum():
                    return self.xtjson.json_params_error('密码格式错误！')
                if len(password.strip()) < 5 or len(password.strip()) > 15:
                    return self.xtjson.json_params_error('密码长度在5~15位之间！')
            _dd = self.MCLS.find_one({'account': account})
            if _dd and _dd.get('uuid') != self.data_uuid:
                return self.xtjson.json_params_error('该账户已存在！')
            _dd = self.MCLS.find_one({'account_name': account_name})
            if _dd and _dd.get('uuid') != self.data_uuid:
                return self.xtjson.json_params_error('该账户名已存在！')
            _data = {
                'account': account,
                'account_name': account_name,
                'note': note.strip(),
            }
            if password:
                _data['password'] = self.MCLS.encry_password(password)
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': _data})
            return self.xtjson.json_result()
        if self.action == 'del':
            self.MCLS.delete_one({'uuid': self.data_uuid})
            return self.xtjson.json_result()
        if self.action == 'getGoogleQrcode':
            if not self.data_uuid:
                return self.xtjson.json_params_error()

            user_dict = self.MCLS.find_one({'uuid': self.data_uuid})
            if not user_dict:
                return self.xtjson.json_params_error()

            google_cls = GooleVerifyCls(pwd=self.data_uuid, account=user_dict.get('account'), s_label='pay2wold')
            generate_qrcode = google_cls.secret_generate_qrcode()
            return self.xtjson.json_result(data={'generate_qrcode': generate_qrcode})



class MerchantInfoView(CmsFormViewBase):
    add_url_rules = [['/merchant/Info/<string:merchantId>', 'merchant_info']]
    MCLS = MerchantTable
    template = 'cms/merchant/merchantInfo.html'
    title = '商户详细信息'

    def view_get(self, merchantId):
        merchant_data = self.MCLS.find_one({'merchant_id': merchantId})
        if not merchant_data:
            return abort(404)

        merchanttunnle_datas = MerchantTunnleTable.find_many({'merchant_uuid': merchant_data.get('uuid'), 'tunnle_method': TUNNLE_METHOD.collection})
        if not merchanttunnle_datas:
            return abort(404)

        mdatas = []
        for da in merchanttunnle_datas:
            tunnle_data = TunnelTable.find_one({'uuid': da.get('tunnle_id')})
            da['tunnle_data'] = tunnle_data
            rate = da.get('rate') or 0
            if rate:
                rate = float(Decimal(str(rate or 0)) * Decimal('100'))
                da['rate'] = rate
            mdatas.append(da)

        mlogs = MerchantLogTable.find_many({'merchant_uuid': merchant_data.get('uuid')}, limit=6, sort=[['create_time', -1]])

        MerchantBankCard_datas = MerchantBankCardTable.find_many({'merchant_uuid': merchant_data.get('uuid')})
        mbcard_datas = []
        for md in MerchantBankCard_datas:
            cdata = BankTable.find_one({'uuid': md.get('bank_uid')})
            md['bank_data'] = cdata
            mbcard_datas.append(md)

        # 日订单量
        time_start, time_end = getDayDateSilce()
        orderDayCount = CollectionOrderTable.count({'order_time': {'$gte':time_start, '$lte': time_end}, 'merchant_id': merchantId})
        self.context['orderDayCount'] = orderDayCount or 0
        # 总订单量
        allOrderCount = CollectionOrderTable.count({'merchant_id': merchantId})
        self.context['allOrderCount'] = allOrderCount or 0

        # 日交易额
        ddsls = CollectionOrderTable.collection().aggregate([
            {"$match": {'order_time': {'$gte':time_start, '$lte': time_end}, 'pay_statu': True, 'merchant_id': merchantId}},
            {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
        ])
        day_actual_amount = 0
        if ddsls:
            ddsls = list(ddsls)
        if ddsls:
            day_actual_amount = list(ddsls)[0].get('actual_amount') or 0
        self.context['day_actual_amount'] = day_actual_amount
        # 月交易额
        ytime_start, ytime_end = getMonthDateSilce()
        ddsls = CollectionOrderTable.collection().aggregate([
            {"$match": {'order_time': {'$gte':ytime_start, '$lte': ytime_end}, 'pay_statu': True, 'merchant_id': merchantId}},
            {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
        ])
        month_actual_amount = 0
        if ddsls:
            ddsls = list(ddsls)
        if ddsls:
            month_actual_amount = list(ddsls)[0].get('actual_amount') or 0
        self.context['month_actual_amount'] = month_actual_amount

        self.context['mlogs'] = mlogs
        self.context['mdatas'] = mdatas
        self.context['mbcard_datas'] = mbcard_datas
        self.context['merchant_data'] = merchant_data
        
        html =  render_template(self.template, **self.context)
        return update_language(self.current_admin_dict.get("language"), html)

    def post_other_way(self):
        if self.action == 'update_tunble_statu':
            mdata = MerchantTunnleTable.find_one({'uuid': self.data_uuid})
            if not mdata:
                return self.xtjson.json_params_error('数据不存在!')
            if mdata.get('statu'):
                self.data_from['statu'] = False
            else:
                tdata = TunnelTable.find_one({'uuid': mdata.get('tunnle_id')})
                if not tdata:
                    return self.xtjson.json_params_error('通道数据不存在！')
                if not tdata.get('tunnel_statu'):
                    return self.xtjson.json_params_error(f'{tdata.get("tunnel_name")}: 主通道未开通！')
                self.data_from['statu'] = True

            MerchantTunnleTable.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            return self.xtjson.json_result()
        if self.action == 'update_tunnleInfo':
            rate = self.request_data.get('rate')
            single_amount_max = self.request_data.get('single_amount_max')
            single_amount_min = self.request_data.get('single_amount_min')
            mdata = MerchantTunnleTable.find_one({'uuid': self.data_uuid})
            if not mdata:
                return self.xtjson.json_params_error('数据不存在！')

            if rate:
                try:
                    rate = float(rate) or 0
                except:
                    return self.xtjson.json_params_error('issued_money_rate: 参数错误！')
            else:
                rate = 0

            if single_amount_max:
                try:
                    single_amount_max = int(single_amount_max) or 0
                except:
                    return self.xtjson.json_params_error('issued_money_rate: 参数错误！')
            else:
                single_amount_max = 0

            if single_amount_min:
                try:
                    single_amount_min = int(single_amount_min) or 0
                except:
                    return self.xtjson.json_params_error('issued_money_rate: 参数错误！')
            else:
                single_amount_min = 0
            self.data_from['rate'] = float(round(rate / 100, 5))
            self.data_from['single_amount_max'] = single_amount_max or 0
            self.data_from['single_amount_min'] = single_amount_min or 0
            MerchantTunnleTable.update_one({'uuid': mdata.get('uuid')}, {'$set': self.data_from})
            return self.xtjson.json_result()



class OrderListView(CmsTableViewBase):
    add_url_rules = [['/collectionOrder', 'collectionOrder']]
    per_page = 30
    MCLS = CollectionOrderTable
    template = 'cms/merchant/collectionOrder.html'
    title = '订单列表'

    def exportData(self, datas, log_uuid, export_folder, filename):
        ''' 导出数据 '''
        export_data = ExportDataModel.find_one({'uuid': log_uuid})
        if not export_data:
            return
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1
            header = ['订单号', '商户名称', '商户订单号', '收款账号', '收款人', '支付模式', '订单金额', '手续费',  '支付金额', '订单时间', '支付状态', '回调状态', '回调类型', '回调时间']
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])

            merchant_data_dict = {}
            bankcard_data_dict = {}
            for data in datas:
                row += 1

                merchant_id = data.get('merchant_id') or ''
                bankcard_id = data.get('bankcard_id') or ''
                merchant_data = merchant_data_dict.get(merchant_id)
                if not merchant_data:
                    merchant_data = MerchantTable.find_one({'merchant_id': merchant_id}) or {}
                    merchant_data_dict[merchant_id] = merchant_data
                bankcard_data = bankcard_data_dict.get(bankcard_id) or {}
                if not bankcard_data:
                    bankcard_data = BankCardTable.find_one({'uuid': bankcard_id}) or {}
                    bankcard_data_dict[bankcard_id] = bankcard_data

                wa.cell(row=row, column=1, value=str(data.get('order_id') or ''))
                wa.cell(row=row, column=2, value=str(merchant_data.get('merchant_name') or ''))
                wa.cell(row=row, column=3, value=str(data.get('merchant_order_id') or ''))
                wa.cell(row=row, column=4, value=str(bankcard_data.get('account') or ''))
                wa.cell(row=row, column=5, value=str(bankcard_data.get('account_username') or ''))
                wa.cell(row=row, column=6, value=str(PAY_METHOD.name_dict.get(data.get('pay_method')) or ''))
                wa.cell(row=row, column=7, value=self.format_money(str(data.get('order_amount') or '0')))
                wa.cell(row=row, column=8, value=self.format_money(str(data.get('repay_amount') or '0')))
                wa.cell(row=row, column=9, value=self.format_money(str(data.get('actual_amount') or '0')))
                wa.cell(row=row, column=10, value=data.get('order_time').strftime('%Y-%m-%d %H:%M:%S'))
                wa.cell(row=row, column=11, value='已支付' if data.get('pay_statu') else '未支付')
                call_text = ''
                if data.get('callback_statu') == CallbackState.SUCCESS:
                    call_text = '回调成功'
                elif data.get('callback_statu') == CallbackState.FAILED:
                    call_text = '回调失败'
                else:
                    call_text = '未回调'
                wa.cell(row=row, column=12, value=call_text)
                wa.cell(row=row, column=13, value=CallbankType.name_dict.get(data.get('callbank_type')) or '' if data.get('callback_statu') == CallbackState.SUCCESS else '')
                wa.cell(row=row, column=14, value=data.get('callback_time').strftime('%Y-%m-%d %H:%M:%S') if data.get('callback_statu') == CallbackState.SUCCESS else '')

                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    # xhr get请求
    def is_xhr(self):
        X_Requested_With = request.headers.get('X-Requested-With')
        if not X_Requested_With or X_Requested_With.lower() != 'xmlhttprequest':
            return
        return True
    
    def get_other_way(self):
        if not self.is_xhr():
            return
        action = request.args.get('action')
        if action == 'refreshOrderList':
            filter_dict = {}
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if not statu:
                return self.xtjson.json_params_error(res)
            filter_dict.update(res[0])
            order_time, start_time, end_time = self.get_reqorder()
            filter_dict.update({'order_time': {'$gte': start_time, '$lte': end_time}})

            if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
                filter_dict['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
            elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
                filter_dict['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')

            searchType = request.args.get('searchType')
            searchValue = request.args.get('searchValue')
            if searchType and searchValue and searchValue.strip() and searchType.strip():
                if searchType == 'bankcard_account':
                    bankcard_data = BankCardTable.find_one({'account': searchValue.strip()})
                    if bankcard_data:
                        filter_dict['bankcard_id'] = bankcard_data.get('uuid')
                    else:
                        filter_dict['bankcard_id'] = searchValue.strip()
                else:
                    dd_fields = self.MCLS.fields()
                    if hasattr(self.MCLS, 'field_search'):
                        field_search = getattr(self.MCLS, 'field_search')()
                        for db_field in field_search:
                            if db_field == searchType:
                                col_value = searchValue.strip().replace(',', '')
                                field_cls = dd_fields.get(db_field)
                                if not field_cls:
                                    break
                                statu, res = field_cls.search_validate(col_value)
                                if statu:
                                    filter_dict[db_field] = res
                                break

            page = request.args.get('page', 1, int)
            skip = (page - 1) * self.per_page
            all_datas = self.MCLS.find_many(filter_dict, limit=self.per_page, skip=skip, sort=self.sort)
            html = '''
            <tr style="background-color: #f7f7f7;">
                <td>订单号</td>
                <td>商户名称</td>
                <td>商户订单号</td>
                <td>收款账号</td>
                <td>收款人</td>                
                <td>支付模式</td>
                <td>订单金额</td>
                <td>手续费</td>
                <td>支付金额</td>
                <td>订单时间</td>
                <td>支付状态</td>
                <td>回调状态</td>
                <td>回调类型</td>                
                <td>回调时间</td>
                <td width="260">操作</td>
            </tr>            
            '''
            for data in all_datas:
                merchant_data = MerchantTable.find_one({'merchant_id': data.get('merchant_id')})
                bankcard_data = BankCardTable.find_one({'uuid': data.get('bankcard_id')})
                html += f'''
                    <tr>
                        <td>{ data.get('order_id') }</td>
                        <td>{ merchant_data.get('merchant_name') or '' }</td>
                        <td>{ data.get('merchant_order_id') }</td>
                        <td>{ bankcard_data.get('account') or '' }</td>
                        <td>{ bankcard_data.get('account_username') or '' }</td>
                        <td>{ PAY_METHOD.name_dict.get(data.get('pay_method')) }</td>
                        <td>{ self.format_money(data.get('order_amount') or 0) }</td>
                        <td>{ self.format_money(data.get('repay_amount') or 0) }</td>
                        <td>{ self.format_money(data.get('actual_amount') or 0) }</td>
                        <td>{ self.format_time_func(data.get('order_time'), '%H:%M:%S') }</td>
                '''
                if data.get('pay_statu'):
                    html += f'''
                    <td><span class="ant-tag ant-tag-green">已支付</span></td>
                    '''
                else:
                    html += f'''
                    <td><span class="ant-tag ant-tag-red">未支付</span></td>                
                    '''
                if data.get('callback_statu') == CallbackState.SUCCESS:
                    html += f'''
                    <td><span class="ant-tag ant-tag-green">已回调</span></td>
                    '''
                elif data.get('callback_statu') == CallbackState.FAILED:
                    html += f'''
                    <td><span class="ant-tag ant-tag-red">回调失败</span></td>
                    '''
                else:
                    html += f'''
                    <td><span class="ant-tag ant-tag-cyan">未回调</span></td>
                    '''

                ttt = '<span class="ant-tag to_cursor ant-tag-blue mr-0" onclick="post_form_html({\'action\': \'detailsInfo_html\', \'data_uuid\': \'%s\'}, \'订单详细信息\', 800)">详细</span>' % data.get('uuid')
                html += f'''
                <td>{ CallbankType.name_dict.get(data.get('callbank_type')) or '' }</td>                
                <td>{ self.format_time_func(data.get('callback_time'), '%H:%M:%S') or '' }</td>
                <td>
                    <span class="ant-tag to_cursor ant-tag-blue mr-0" onclick="post_update_statu('forceIsPay', '{ data.get('uuid') }', '确定强制入款？')">强制入款</span>
                    { ttt }
                    <span class="ant-tag to_cursor mr-0 dropdown-toggle" data-toggle="dropdown" aria-expanded="false">更多</span>
                    <div class="dropdown-menu">
                        <span class="dropdown-item" onclick="getPayQrcode('{ data.get('uuid') }')">支付码</span>                    
                        <span class="dropdown-item" onclick="post_update_statu('callbackOrder', '{ data.get('uuid') }', '确认回调该订单？')">订单回调</span>
                        <a class="dropdown-item" href="{ url_for('admin.callbackLog', order_id=data.get('order_id')) }" target="_blank">回调记录</a>
                        <a class="dropdown-item" href="{ url_for('admin.BankCardBillLog', bc_id=data.get('bankcard_id')) }" target="_blank">爬虫明细</a>                        
                        <span class="dropdown-item" onclick="post_update_statu('delOrder', '{ data.get('uuid') }', '确定删除该订单？删除后不可恢复')">删除</span>
                    </div>
                </td>         
            </div>   
                '''
            return self.xtjson.json_result(data={'table_html': html, 'is_data': True if all_datas else False})

        return self.xtjson.json_params_error()

    def get_reqorder(self):
        order_time = request.args.get('order_time')
        if order_time and order_time.strip():
            start_time, end_time = PagingCLS.by_silce(order_time)
        else:
            crrdate = datetime.datetime.now()
            start_time, end_time = datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 0, 0,0), datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 23,59, 59)
            order_time = start_time.strftime('%Y-%m-%d %H:%M:%S') + '|' + end_time.strftime('%Y-%m-%d %H:%M:%S')

        return order_time, start_time, end_time

    def detailsInfo_html(self):
        html = ''
        html += f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto; text-align: left; overflow-x: hidden; box-sizing: border-box;">               
                    <ul class="layui-timeline">
                        <li class="layui-timeline-item">
                            <i class="layui-icon layui-timeline-axis"></i>
                            <div class="layui-timeline-content layui-text">
                                <div class="layui-timeline-title">{ self.format_time_func(self.data_dict.get('order_time')) }，创建订单</div>
                            </div>
                        </li>
    
        '''

        if self.data_dict.get('pay_statu'):
            html += f'''
                        <li class="layui-timeline-item">
                            <i class="layui-icon layui-timeline-axis"></i>
                            <div class="layui-timeline-content layui-text">
                                <div class="layui-timeline-title">{self.format_time_func(self.data_dict.get('pay_time'))}，支付成功</div>
                            </div>
                        </li>            
            '''     
        calllogs = callbackLogTable.find_many({'order_uuid': self.data_uuid}, sort=[['create_time', 1]])
        if calllogs:
            if not self.data_dict.get('pay_statu'):
                html += f'''
                            <li class="layui-timeline-item">
                                <i class="layui-icon layui-timeline-axis"></i>
                                <div class="layui-timeline-content layui-text">
                                    <div class="layui-timeline-title">{self.format_time_func(self.data_dict.get('pay_time'))}，支付失败</div>
                                </div>
                            </li>            
                '''
        for cladata in calllogs:
            text = ''
            if cladata.get('callbank_type') == CallbankType.MANUAL:
                text += '手动回调，'
                udata = CmsUserTable.find_one({'uuid': cladata.get('admin_uuid')})
                if udata:
                    text += f"操作人：{udata.get('account')}，{ '备注：'+cladata.get('note')+',' if cladata.get('note') else '' }"
            else:
                text += '自动回调，'
            if cladata.get('statu'):
                html += f'''
                    <li class="layui-timeline-item">
                        <i class="layui-icon layui-timeline-axis"></i>
                        <div class="layui-timeline-content layui-text">
                            <div class="layui-timeline-title">{ self.format_time_func(cladata.get('create_time')) }，{ text } 回调成功！</div>
                        </div>
                    </li>            
                '''
                break
            else:
                html += f'''
                    <li class="layui-timeline-item">
                        <i class="layui-icon layui-timeline-axis"></i>
                        <div class="layui-timeline-content layui-text">
                            <div class="layui-timeline-title">{ self.format_time_func(cladata.get('create_time')) }，{ text } 回调失败！</div>
                        </div>
                    </li>            
                '''

        if not self.data_dict.get('pay_statu'):
            order_time = self.data_dict.get('order_time')
            if order_time+datetime.timedelta(minutes=15) < datetime.datetime.now():
                html += f'''
                    <li class="layui-timeline-item">
                        <i class="layui-icon layui-timeline-axis"></i>
                        <div class="layui-timeline-content layui-text">
                            <div class="layui-timeline-title">{ self.format_time_func(order_time+datetime.timedelta(minutes=15)) }，超时未支付，取消订单！</div>
                        </div>
                    </li>            
                '''
        if self.data_dict.get('note'):
            html += f'''
                <li class="layui-timeline-item">
                    <i class="layui-icon layui-timeline-axis"></i>
                    <div class="layui-timeline-content layui-text">
                        <div class="layui-timeline-title">备注：{ self.data_dict.get('note') or '' }</div>
                    </div>
                </li>            
            '''
        html += '''
                    </ul>        
                </div>
            </div>        
        '''
        return self.xtjson.json_result(message=html)

    def get_filter_dict(self):
        order_time, start_time, end_time = self.get_reqorder()
        dd = {}
        order_time = {'$gte': start_time, '$lte': end_time}

        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            dd['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            dd['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')

        sort_type = request.args.get('sort_type')
        if sort_type and sort_type.strip():
            if len(sort_type.rsplit('_', 1)) == 2:
                filed, vv = sort_type.rsplit('_', 1)
                if vv == '0':
                    self.sort = [[filed, -1]]
                if vv == '1':
                    self.sort = [[filed, 1]]

        searchType = request.args.get('searchType')
        searchValue = request.args.get('searchValue')
        if searchType and searchValue and searchValue.strip() and searchType.strip():
            if searchType == 'bankcard_account':
                bankcard_data = BankCardTable.find_one({'account': searchValue.strip()})
                if bankcard_data:
                    dd['bankcard_id'] = bankcard_data.get('uuid')
                else:
                    dd['bankcard_id'] = searchValue.strip()
            elif searchType == 'merchant_name':
                merchant_data = MerchantTable.find_one({'merchant_name': searchValue}) or {}
                dd['merchant_id'] = merchant_data.get('merchant_id')
                self.search_dict['merchant_name'] = searchValue
            else:
                dd_fields = self.MCLS.fields()
                if hasattr(self.MCLS, 'field_search'):
                    field_search = getattr(self.MCLS, 'field_search')()
                    for db_field in field_search:
                        if db_field == searchType:
                            col_value = searchValue.strip().replace(',','')
                            field_cls = dd_fields.get(db_field)
                            if not field_cls:
                                break
                            statu, res = field_cls.search_validate(col_value)
                            if statu:
                                dd[db_field] = res
                            break

        fields = self.MCLS.fields()
        statu, res = self.search_func(fields)
        if statu:
            dd.update(res[0])

        if 'callback_time' not in dd and not request.args.get('order_time'):
            dd['order_time'] = order_time
        if request.args.get('order_time'):
            dd['order_time'] = order_time
        pay_method = request.args.get('pay_method')
        if pay_method:
            dd['pay_method'] = pay_method

        bank_uid = request.args.get('bank_uid')
        if bank_uid:
            bank_code = BankTable.find_one({"uuid":bank_uid}).get("code")
            dd['bank_code'] = bank_code
            self.search_dict['bank_uid'] = bank_uid

        is_search= False
        if 'callbank_type' in str(request.url):
            is_search = True
        self.search_dict['is_search'] = is_search
        return dd

    def get_context(self):
            _back_datas = BankTable.find_many({})
            back_datas = []
            for d in _back_datas:
                if d.get('code') in BANK_CODE:
                    back_datas.append(d)
            res = {
                'back_datas': back_datas
            }
            return res
    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        datas = []
        merchant_datas = {}
        bankcard_datas = {}
        for da in all_datas:
            merchant_data = merchant_datas.get(da.get('merchant_id'))
            if not merchant_data:
                merchant_data = MerchantTable.find_one({'merchant_id': da.get('merchant_id')})
                merchant_datas[da.get('merchant_id')] = merchant_data
            da['merchant_data'] = merchant_data

            bankcard_data = bankcard_datas.get(da.get('bankcard_id'))
            if not bankcard_data:
                bankcard_data = BankCardTable.find_one({'uuid': da.get('bankcard_id')})
                bankcard_datas[da.get('bankcard_id')] = bankcard_data
            da['bankcard_data'] = bankcard_data
            datas.append(da)

        self.context['all_datas'] = datas
        self.context['PAY_METHOD'] = PAY_METHOD
        self.context['CallbackState'] = CallbackState
        self.context['CallbankType'] = CallbankType
        self.context['back_datas'] = self.get_context()["back_datas"]

        order_time, start_time, end_time = self.get_reqorder()
        search_res = self.context.get('search_res') or {}
        search_res['order_time'] = order_time

        searchType = request.args.get('searchType')
        searchValue = request.args.get('searchValue')
        if searchType:
            search_res['searchType'] = searchType
        if searchValue:
            search_res['searchValue'] = searchValue
        self.context['search_res'] = search_res

    def post_data_other_way(self):
        # 删除订单
        # if self.action == 'delOrder':
        #     callbackLogTable.delete_many({'order_uuid': self.data_uuid})
        #     self.MCLS.delete_one({'uuid': self.data_uuid})
        #     return self.xtjson.json_result()
        # 强制入款
        if self.action == 'forceIsPay':
            if self.data_dict.get('pay_statu'):
                return self.xtjson.json_params_error('该订单已支付！ 不可重复操作')

            if self.data_dict.get('force_ispay'):
                return self.xtjson.json_params_error('该订单已处理，不可重复操作！')

            merchant_data = MerchantTable.find_one({'merchant_id': self.data_dict.get('merchant_id')})
            if not merchant_data:
                return self.xtjson.json_params_error('商户不存在！')

            order_amount = self.data_dict.get('order_amount')

            self.data_from['actual_amount'] = order_amount
            self.data_from['force_ispay'] = True
            self.data_from['pay_statu'] = True
            self.data_from['pay_time'] = datetime.datetime.now()
            self.MCLS.update_one({'uuid': self.data_uuid},{'$set': self.data_from})

            _state, _res = CallbackPayOrderFunc(self.data_uuid, is_manual=True, admin_uuid=self.current_admin_dict.get('uuid'), note='强制入款')
            if _state:
                payIncome_addto(order_uuid=self.data_uuid, merchant_data=merchant_data)
                return self.xtjson.json_result()
            return self.xtjson.json_params_error('回调失败')
        # 订单回调
        if self.action == 'callbackOrder':
            order_amount = self.data_dict.get('order_amount') or 0
            actual_amount = self.data_dict.get('actual_amount') or 0
            if actual_amount and order_amount != actual_amount:
                return self.xtjson.json_params_error('实际支付金额与订单金额不一致，禁止回调！')

            merchant_data = MerchantTable.find_one({'merchant_id': self.data_dict.get('merchant_id')})
            if not merchant_data:
                return self.xtjson.json_params_error('商户不存在！')

            if self.data_dict.get('callback_statu') == CallbackState.SUCCESS:
                return self.xtjson.json_params_error('该订单已回调')
            _state, res = CallbackPayOrderFunc(self.data_uuid, is_manual=True,admin_uuid=self.current_admin_dict.get('uuid'), note='订单回调')
            if not _state:
                return self.xtjson.json_params_error('回调失败！')

            payIncome_addto(order_uuid=self.data_uuid, merchant_data=merchant_data)
            return self.xtjson.json_result()
        # 订单信息
        if self.action == 'detailsInfo_html':
            return self.detailsInfo_html()
        # 支付二维码
        if self.action == 'get_pay_qrcode':
            bankcard_data = BankCardTable.find_one({'uuid': self.data_dict.get('bankcard_id')})
            if not bankcard_data:
                return self.xtjson.json_params_error('该订单暂未绑定收款银行卡！')
            bank_data = BankTable.find_one({'uuid': bankcard_data.get('bank_uid')}) or {}
            if not bank_data:
                return self.xtjson.json_params_error('数据错误！')

            payqrcode_url = self.data_dict.get('payqrcode_url') or ''
            project_static_folder = os.path.join(current_app.static_folder, current_app.config.get('PROJECT_NAME'))
            _state, payQrcode = getBankPayQrcode(
                self.data_uuid,
                self.data_dict.get('order_amount'),
                self.data_dict.get('bank_memo'),
                bank_data,
                payqrcode_url=payqrcode_url,
                project_static_folder=project_static_folder,
                receive_account=bankcard_data.get('account'),
            )
            if not _state:
                return self.xtjson.json_params_error(payQrcode)

            return self.xtjson.json_result(data={'generate_qrcode': payQrcode})
        # 更新备注
        if self.action == 'updateNote':
            data_value = self.request_data.get('data_value') or ''
            if not data_value:
                return self.xtjson.json_params_error('请收入备注内容!')
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': {'note': data_value.strip()}})
            return self.xtjson.json_result()

    def post_other_way(self):
        if self.action == 'totalData':
            filter_dict = self.get_filter_dict()
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if statu:
                filter_dict.update(res[0])

            # 订单金额统计
            fff_d1 = {}
            fff_d1.update(filter_dict)
            _order_amount_total_ll = CollectionOrderTable.collection().aggregate([
                {"$match": filter_dict},
                {"$group": {"_id": None, "order_amount": {"$sum": '$order_amount'}}},
            ])
            order_amount_total_ll = list(_order_amount_total_ll)
            order_amount_tota = 0
            if order_amount_total_ll:
                order_amount_tota = order_amount_total_ll[0].get('order_amount')

            # 实付金额统计
            fff1 = {}
            fff1.update(filter_dict)
            fff1.update({
                'pay_statu': True
            })
            _actual_amount_total_ll = CollectionOrderTable.collection().aggregate([
                {"$match": fff1},
                {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
            ])
            actual_amount_total_ll = list(_actual_amount_total_ll)
            actual_amount_tota = 0
            if actual_amount_total_ll:
                actual_amount_tota = actual_amount_total_ll[0].get('actual_amount')
            actual_amount_count = CollectionOrderTable.count(fff1)

            # 实付金额总手续费
            _repay_amount_total_ll = CollectionOrderTable.collection().aggregate([
                {"$match": fff1},
                {"$group": {"_id": None, "repay_amount": {"$sum": '$repay_amount'}}},
            ])
            repay_amount_total_ll = list(_repay_amount_total_ll)
            repay_amount_total = 0
            if repay_amount_total_ll:
                repay_amount_total = repay_amount_total_ll[0].get('repay_amount')

            # 回调金额统计
            fff2 = {}
            fff2.update(filter_dict)
            fff2.update({
                'pay_statu': True,
                'callback_statu': CallbackState.SUCCESS,
            })
            _callback_amount_total_ll = CollectionOrderTable.collection().aggregate([
                {"$match": fff2},
                {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
            ])
            callback_amount_total_ll = list(_callback_amount_total_ll)
            callback_amount_tota = 0
            if callback_amount_total_ll:
                callback_amount_tota = callback_amount_total_ll[0].get('actual_amount')
            callback_amount_count = CollectionOrderTable.count(fff2)

            # 掉单金额统计
            fff3 = {}
            fff3.update(filter_dict)
            fff3.update({
                'is_lose': True,
            })
            _lose_amount_total_ll = CollectionOrderTable.collection().aggregate([
                {"$match": fff3},
                {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
            ])
            lose_amount_total_ll = list(_lose_amount_total_ll)
            lose_amount_tota = 0
            if lose_amount_total_ll:
                lose_amount_tota = lose_amount_total_ll[0].get('actual_amount')
            lose_amount_count = CollectionOrderTable.count(fff3)

            html = f'''
                <div class="formBox">
                    <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto; text-align: left; overflow-x: hidden; font-size: 14px; line-height: 25px; color: #333333;">
                        <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">订单金额：</span>{ self.format_money(order_amount_tota) }</p>
                        <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">实付金额：</span>{ self.format_money(actual_amount_tota) }</p>
                        <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">实付笔数：</span>{ actual_amount_count }</p>
                        <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">实付手续费：</span>{ self.format_money(repay_amount_total) }</p>                        
                        <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">回调金额：</span>{ self.format_money(callback_amount_tota) }</p>
                        <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">回调笔数：</span>{ callback_amount_count }</p>
                        <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">掉单金额：</span>{ self.format_money(lose_amount_tota) }</p>
                        <p style="margin-bottom: 1rem;"><span style="width: 150px; display: inline-block; text-align: right;">掉单笔数：</span>{ lose_amount_count }</p>
                    </div>
                </div>
            '''
            return self.xtjson.json_result(message=html)
        if self.action == 'export_order':
            filter_dict = self.get_filter_dict()
            datas = CollectionOrderTable.find_many(filter_dict)
            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            FOLDER_name = ''
            agentadmin_uuid = ''
            if self.current_admin_roleCode == ROlE_ALL.SUPERADMIN or self.current_admin_roleCode == ROlE_ALL.ADMINISTRATOR:
                FOLDER_name = 'su'
            else:
                if self.is_agentadmin:
                    FOLDER_name = self.current_admin_dict.get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('uuid')
                elif self.current_admin_dict.get('agentadmin_data'):
                    FOLDER_name = self.current_admin_dict.get('agentadmin_data').get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('agentadmin_data').get('uuid')
            if not FOLDER_name:
                return self.xtjson.json_params_error()
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER, FOLDER_name)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            _out_data_dict = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas),
                'out_count': 0,
                'note': '订单数据-' + datetime.datetime.now().strftime('%Y%m%d%H%M%S'),
                'agentadmin_uuid': agentadmin_uuid,
            }
            uuid = ExportDataModel.insert_one(_out_data_dict)
            threading.Thread(target=self.exportData, args=(datas, uuid, export_folder, filename)).start()
            return self.xtjson.json_result(message='数据导出中，请稍后到"导出文件"中查看数据！')



class MerchantReviewView(CmsTableViewBase):
    add_url_rules = [['/merchantReview', 'merchantReview']]
    per_page = 30
    MCLS = MerchantTable
    template = 'cms/merchant/merchantReview.html'
    title = '审核列表'

    def get_filter_dict(self):
        fff = {
            'is_review': False,
        }
        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')
        return fff

    def post_data_other_way(self):
        if self.action == 'update_is_review':
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': {'is_review': True}})
            return self.xtjson.json_result()



class RechargeMoneyView(CmsTableViewBase):
    add_url_rules = [['/rechargeMoney', 'rechargeMoney']]
    per_page = 30
    MCLS = RechargeMoneyTable
    template = 'cms/merchant/rechargeMoney.html'
    title = '内充列表'

    def get_reqorder(self):
        create_time = request.args.get('create_time')
        start_time =  end_time = ''
        
        if create_time and create_time.strip():
            start_time, end_time = PagingCLS.by_silce(create_time)
        else:
            crrdate = datetime.datetime.now()
            start_time, end_time = datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 0, 0,0), datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 23,59, 59)
            create_time = start_time.strftime('%Y-%m-%d %H:%M:%S') + '|' + end_time.strftime('%Y-%m-%d %H:%M:%S')

        return create_time, start_time, end_time
    
    def get_filter_dict(self):
        fff = {}
        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')

        operate_user = request.args.get('operate_user_uuid')
        if operate_user and operate_user.strip():
            fff['operate_user_uuid'] = {"$in": [x.get("uuid") for x in CmsUserTable.find_many({"account":operate_user})]}
            self.search_dict['operate_user_uuid'] = operate_user

        merchant_name = request.args.get('merchant_name')
        if merchant_name and merchant_name.strip():
            fff['merchant_id'] = {"$in": [x.get("merchant_id") for x in MerchantTable.find_many({"merchant_name":merchant_name})]}
            self.search_dict['merchant_name'] = merchant_name

        merchant_id = request.args.get('merchant_id')
        if merchant_id and merchant_id.strip():
            fff['merchant_id'] = merchant_id
            self.search_dict['merchant_id'] = merchant_id

        note = request.args.get('note')
        if note and note.strip():
            fff['note'] = {"$regex": note}

        create_time, start_time, end_time = self.get_reqorder()
        if request.args.get("create_time"):
            fff.update({'create_time': {'$gte': start_time, '$lte': end_time}})
        else :
            self.search_dict['create_time'] = create_time

        return fff

    def  dealwith_main_context(self):
        all_datas = self.context.get('all_datas') or []
        datas = []
        for dd in all_datas:
            udata = CmsUserTable.find_one({'uuid': dd.get('operate_user_uuid')})
            dd['user_data'] = udata
            merchant_data = MerchantTable.find_one({'merchant_id': dd.get('merchant_id')})
            dd['merchant_data'] = merchant_data
            datas.append(dd)
        self.context['all_datas'] = datas

    def post_other_way(self):
        if self.action == 'get_statistic_html':
            _total_data = RechargeMoneyTable.collection().aggregate([
                {"$match": self.get_filter_dict()},
                {"$group": {"_id": None, 
                    "amount": {"$sum": '$amount'},
                    "repay_amount": {"$sum": '$repay_amount'}
                }},
            ])
            total_data = list(_total_data)
            amount = 0
            repay_amount = 0
            if total_data:
                amount = total_data[0].get('amount')
                repay_amount = total_data[0].get('repay_amount')

            html = f'''
            <div class="formBox">
                <div style="height: 10rem; position: relative; box-sizing: border-box; overflow-y: auto;">      
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">内充统计：</span>
                        <span style="display: inline-block;width: calc(100% - 180px);text-align: left;font-size: 16px;vertical-align: -1px;">{ amount }</span>
                    </div>


                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">手续费统计：</span>
                        <span style="display: inline-block;width: calc(100% - 180px);text-align: left;font-size: 16px;vertical-align: -1px;">{ repay_amount }</span>
                    </div>
                </div>
                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>
                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="xtalert.close()">确定</span>
                </div>                                                                                 
            </div> 
            '''

            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
class ReduceMoneyView(CmsTableViewBase):
    add_url_rules = [['/reduceMoney', 'reduceMoney']]
    per_page = 30
    MCLS = ReduceMoneyTable
    template = 'cms/merchant/reduceMoney.html'
    title = '内充列表'

    def get_reqorder(self):
        create_time = request.args.get('create_time')
        start_time =  end_time = ''
        
        if create_time and create_time.strip():
            start_time, end_time = PagingCLS.by_silce(create_time)
        else:
            crrdate = datetime.datetime.now()
            start_time, end_time = datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 0, 0,0), datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 23,59, 59)
            create_time = start_time.strftime('%Y-%m-%d %H:%M:%S') + '|' + end_time.strftime('%Y-%m-%d %H:%M:%S')

        return create_time, start_time, end_time
    
    def get_filter_dict(self):
        fff = {}
        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')

        operate_user = request.args.get('operate_user_uuid')
        if operate_user and operate_user.strip():
            fff['operate_user_uuid'] = {"$in": [x.get("uuid") for x in CmsUserTable.find_many({"account":operate_user})]}
            self.search_dict['operate_user_uuid'] = operate_user

        merchant_name = request.args.get('merchant_name')
        if merchant_name and merchant_name.strip():
            fff['merchant_id'] = {"$in": [x.get("merchant_id") for x in MerchantTable.find_many({"merchant_name":merchant_name})]}
            self.search_dict['merchant_name'] = merchant_name

        merchant_id = request.args.get('merchant_id')
        if merchant_id and merchant_id.strip():
            fff['merchant_id'] = merchant_id
            self.search_dict['merchant_id'] = merchant_id

        note = request.args.get('note')
        if note and note.strip():
            fff['note'] = {"$regex": note}

        create_time, start_time, end_time = self.get_reqorder()
        if request.args.get("create_time"):
            fff.update({'create_time': {'$gte': start_time, '$lte': end_time}})
        else :
            self.search_dict['create_time'] = create_time

        return fff

    def  dealwith_main_context(self):
        all_datas = self.context.get('all_datas') or []
        datas = []
        for dd in all_datas:
            udata = CmsUserTable.find_one({'uuid': dd.get('operate_user_uuid')})
            dd['user_data'] = udata
            merchant_data = MerchantTable.find_one({'merchant_id': dd.get('merchant_id')})
            dd['merchant_data'] = merchant_data
            datas.append(dd)
        self.context['all_datas'] = datas

    def post_other_way(self):
        if self.action == 'get_statistic_html':
            _total_data = ReduceMoneyTable.collection().aggregate([
                {"$match": self.get_filter_dict()},
                {"$group": {"_id": None, 
                    "amount": {"$sum": '$amount'},
                }},
            ])
            total_data = list(_total_data)
            amount = 0
            repay_amount = 0
            if total_data:
                amount = total_data[0].get('amount')
                repay_amount = total_data[0].get('repay_amount')

            html = f'''
            <div class="formBox">
                <div style="height: 10rem; position: relative; box-sizing: border-box; overflow-y: auto;">      
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">减少统计：</span>
                        <span style="display: inline-block;width: calc(100% - 180px);text-align: left;font-size: 16px;vertical-align: -1px;">- { amount }</span>
                    </div>
                </div>
                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>
                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="xtalert.close()">确定</span>
                </div>                                                                                 
            </div> 
            '''

            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))


class LoseOrderListView(CmsTableViewBase):
    add_url_rules = [['/loseOrderList', 'loseOrderList']]
    per_page = 30
    MCLS = CollectionOrderTable
    template = 'cms/merchant/loseOrderList.html'
    title = '掉单列表'

    def exportData(self, datas, log_uuid, export_folder, filename):
        ''' 导出数据 '''
        export_data = ExportDataModel.find_one({'uuid': log_uuid})
        if not export_data:
            return
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1
            header = ['订单号','商户名称','商户订单号','收款卡号','订单金额','支付金额','来源','原因','状态','处理状态','备注','备注码','订单时间','处理时间']
            
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])

            merchant_data_dict = {}
            bankcard_data_dict = {}
            for data in datas:
                row += 1

                merchant_id = data.get('merchant_id') or ''
                bankcard_id = data.get('bankcard_id') or ''
                merchant_data = merchant_data_dict.get(merchant_id)
                if not merchant_data:
                    merchant_data = MerchantTable.find_one({'merchant_id': merchant_id}) or {}
                    merchant_data_dict[merchant_id] = merchant_data
                bankcard_data = bankcard_data_dict.get(bankcard_id) or {}
                if not bankcard_data:
                    bankcard_data = BankCardTable.find_one({'uuid': bankcard_id}) or {}
                    bankcard_data_dict[bankcard_id] = bankcard_data

                wa.cell(row=row, column=1, value=str(data.get('order_id') or ''))
                wa.cell(row=row, column=2, value=str(merchant_data.get('merchant_name') or ''))
                wa.cell(row=row, column=3, value=str(data.get('merchant_order_id') or ''))
                wa.cell(row=row, column=4, value=str(bankcard_data.get('account') or ''))
                wa.cell(row=row, column=5, value=self.format_money(data.get('order_amount')))
                wa.cell(row=row, column=6, value=self.format_money(data.get('actual_amount' )) )
                wa.cell(row=row, column=7, value=PAY_METHOD.name_dict.get(data.get('pay_method')))
                wa.cell(row=row, column=8, value=str(data.get('lose_reason') or ''))
                wa.cell(row=row, column=9, value='已支付' if data.get('pay_statu') else '未支付')
                wa.cell(row=row, column=10, value='已处理' if data.get('proc_statu') else '未处理')
                wa.cell(row=row, column=11, value=data.get('lose_note') or '')
                wa.cell(row=row, column=12, value=data.get('bank_memo') if self.xs_func(data.get('order_time')) else '')
                wa.cell(row=row, column=13, value=data.get('order_time').strftime('%Y-%m-%d %H:%M:%S') ) if data.get('order_time') else ''
                wa.cell(row=row, column=14, value=data.get('proc_time').strftime('%Y-%m-%d %H:%M:%S')) if data.get('proc_time') else ''

                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def get_reqorder(self):
        order_time = request.args.get('order_time')
        if order_time and order_time.strip():
            start_time, end_time = PagingCLS.by_silce(order_time)
        else:
            crrdate = datetime.datetime.now()
            start_time, end_time = datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 0, 0,0), datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 23,59, 59)
            order_time = start_time.strftime('%Y-%m-%d %H:%M:%S') + '|' + end_time.strftime('%Y-%m-%d %H:%M:%S')

        proc_time = request.args.get('proc_time')
        proc_start_time = proc_end_time = ''
        if proc_time and proc_time.strip():
            proc_start_time, proc_end_time = PagingCLS.by_silce(proc_time)

        # else:
        #     crrdate = datetime.datetime.now()
        #     proc_start_time, proc_end_time = datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 0, 0,0), datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 23,59, 59)
        #     proc_time = proc_start_time.strftime('%Y-%m-%d %H:%M:%S') + '|' + proc_end_time.strftime('%Y-%m-%d %H:%M:%S')

        return order_time, start_time, end_time, proc_time, proc_start_time, proc_end_time

    def get_filter_dict(self):
        order_time, start_time, end_time, proc_time, proc_start_time, proc_end_time = self.get_reqorder()
        fff = {'is_lose': True}

        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')

        if request.args.get("proc_time"):
            fff.update({'proc_time': {'$gte': proc_start_time, '$lte': proc_end_time}})
        
        if request.args.get("order_time"):
            fff.update({'order_time': {'$gte': start_time, '$lte': end_time}})

        merchant_name = request.args.get('merchant_name')
        if merchant_name and merchant_name.strip():
            merchant_data = MerchantTable.find_one({'merchant_name': merchant_name.strip()}) or {}
            fff['merchant_id'] = merchant_data.get('merchant_id')
            self.search_dict['merchant_name'] = merchant_name
        
        lose_reason = request.args.get('lose_reason')
        if lose_reason and lose_reason.strip():
            fff['lose_reason'] = lose_reason
            self.search_dict['lose_reason'] = lose_reason

        return fff

    def xs_func(self, order_time):
        xs = datetime.datetime.now() + datetime.timedelta(minutes=15)
        if order_time <= xs:
            return True
        return

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        datas = []
        for da in all_datas:
            merchant_data = MerchantTable.find_one({'merchant_id': da.get('merchant_id')})
            da['merchant_data'] = merchant_data
            bankcard_data = BankCardTable.find_one({'uuid': da.get('bankcard_id')})
            da['bankcard_data'] = bankcard_data
            datas.append(da)
        self.context['all_datas'] = datas
        self.context['PAY_METHOD'] = PAY_METHOD
        self.context['xs_func'] = self.xs_func

        order_time, start_time, end_time, proc_time, proc_start_time, proc_end_time = self.get_reqorder()
        search_res = self.context.get('search_res') or {}
        search_res['order_time'] = order_time
        search_res['proc_time'] = proc_time or ''
        self.context['search_res'] = search_res

    def orderInfo_html(self):
        bankcard_data = BankCardTable.find_one({'uuid': self.data_dict.get('bankcard_id')}) or {}
        merchant_data = MerchantTable.find_one({'merchant_id': self.data_dict.get('merchant_id')}) or {}
        agentadmin_data = CmsUserTable.find_one({'uuid': self.data_dict.get('agentadmin_uuid')}) or {}
        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto; text-align: left; overflow-x: hidden; font-size: 14px; line-height: 25px; color: #333333;">
                    <p><span style="width: 150px; display: inline-block; text-align: right;">订单号：</span>{ self.data_dict.get('order_id') or '' }</p>
                    <p><span style="width: 150px; display: inline-block; text-align: right;">商户订单号：</span>{ self.data_dict.get('merchant_order_id') or '' }</p>
                    <p><span style="width: 150px; display: inline-block; text-align: right;">商户账户：</span>{ merchant_data.get('account') or '' }</p>
                    <p><span style="width: 150px; display: inline-block; text-align: right;">所属代理账户：</span>{ agentadmin_data.get('account') or '' }</p>
                    <p><span style="width: 150px; display: inline-block; text-align: right;">收款卡号：</span>{ bankcard_data.get('account') or '' }</p>
                    <p><span style="width: 150px; display: inline-block; text-align: right;">收款人姓名：</span>{ bankcard_data.get('account_username') or '0' }</p>
                    <p><span style="width: 150px; display: inline-block; text-align: right;">订单金额：</span>{ self.format_money(self.data_dict.get('order_amount') or '0') }</p>
                    <p><span style="width: 150px; display: inline-block; text-align: right;">手续费：</span>{ self.format_money(self.data_dict.get('repay_amount') or '0') }</p>
                    <p><span style="width: 150px; display: inline-block; text-align: right;">实际支付金额：</span>{ self.format_money(self.data_dict.get('actual_amount') or '') }</p>
                    <p><span style="width: 150px; display: inline-block; text-align: right;">支付时间：</span>{ self.format_time_func(self.data_dict.get('pay_time') or '') }</p>
                </div>
            </div>
        '''
        return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))

    def post_data_other_way(self):
        # if self.action == 'delOrder':
        #     callbackLogTable.delete_many({'order_uuid': self.data_uuid})
        #     self.MCLS.delete_one({'uuid': self.data_uuid})
        #     return self.xtjson.json_result()
        if self.action == 'forceIsPay':
            if self.data_dict.get('pay_statu'):
                return self.xtjson.json_params_error('该订单已支付！ 不可强制入款')

            if self.data_dict.get('force_ispay'):
                return self.xtjson.json_params_error('该订单已强制入款，不可重复操作！')

            merchant_data = MerchantTable.find_one({'merchant_id': self.data_dict.get('merchant_id')})
            if not merchant_data:
                return self.xtjson.json_params_error('商户不存在！')

            order_amount = self.data_dict.get('order_amount')
            repay_amount = self.data_dict.get('repay_amount')
            _amount = order_amount - repay_amount

            self.data_from['actual_amount'] = order_amount
            self.data_from['force_ispay'] = True
            self.data_from['pay_statu'] = True
            self.data_from['pay_time'] = datetime.datetime.now()
            self.MCLS.update_one({'uuid': self.data_uuid},{'$set': self.data_from})

            _state, _res = CallbackPayOrderFunc(self.data_uuid, is_manual=True, admin_uuid=self.current_admin_dict.get('uuid'))
            if not _state:
                return self.xtjson.json_params_error('回调失败！')
            payIncome_addto(order_uuid=self.data_uuid, merchant_data=merchant_data)
            return self.xtjson.json_result()
        if self.action == 'callbackOrder':
            order_amount = self.data_dict.get('order_amount') or 0
            actual_amount = self.data_dict.get('actual_amount') or 0
            if actual_amount and order_amount != actual_amount:
                return self.xtjson.json_params_error('实际支付金额与订单金额不一致，禁止回调！')

            merchant_data = MerchantTable.find_one({'merchant_id': self.data_dict.get('merchant_id')})
            if not merchant_data:
                return self.xtjson.json_params_error('商户不存在！')

            _state, res = CallbackPayOrderFunc(self.data_uuid, is_manual=True, admin_uuid=self.current_admin_dict.get('uuid'))
            if _state:
                payIncome_addto(order_uuid=self.data_uuid, merchant_data=merchant_data)
                return self.xtjson.json_result()
            return self.xtjson.json_params_error(res)
        if self.action == 'updateNote':
            if not self.data_value:
                return self.xtjson.json_params_error()
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': {
                'lose_note': self.data_value, 
                'proc_time': datetime.datetime.now(),
                'proc_statu': True,
                }})
            return self.xtjson.json_result()
        if self.action == 'orderInfo_html':
            return self.orderInfo_html()

    def createTestOrder(self):
        bankcc = [
            {
                "bankCode": "NAB",
                "bankAccount": "622226319100001",
                "bankOwner": "NGUYEN VU PHUONG",
            },
            {
                "bankCode": "VAB",
                "bankAccount": "00195363",
                "bankOwner": "NGO TAN PHUONG TAY",
            },
            {
                "bankCode": "SEAB",
                "bankAccount": "000009731380",
                "bankOwner": "DINH THI QUYNH NHU",
            },
            {
                "bankCode": "VAB",
                "bankAccount": "00340337",
                "bankOwner": "NGUYEN HONG SON",
            },
        ]
        for i in range(5):
            bank_data = random.choice(bankcc)
            bankCode = bank_data.get('bankCode')
            bankAccount = bank_data.get('bankAccount')
            bankOwner = bank_data.get('bankOwner')

            merchant_data = MerchantTable.find_one({})
            if not merchant_data:
                return self.xtjson.json_params_error('无商户数据！')
            mchId = MerchantTable.find_one({}).get('merchant_id')
            mchOrderId = 'Y'+str(int(time.time()*1000))
            amount = random.choice(list(range(1, 100)))

            rate = merchant_data.get('payment_rate') or 0
            repay_amount = round(amount * rate, 2)

            balance_amount = merchant_data.get('balance_amount') or 0
            if balance_amount < (repay_amount + amount):
                return self.xtjson.json_params_error('商户余额不足！', code=408)

            ouuid = shortuuid.uuid()
            _order_daat = {
                'merchant_id': mchId,
                'merchant_order_id': mchOrderId,
                'order_id': ouuid,
                'order_amount': amount,
                'actual_amount': amount,
                'repay_amount': repay_amount,
                'order_time': datetime.datetime.now(),
                'pay_statu': False,
                'reject_pay': False,
                'is_lose': True,
                'callback_statu': CallbackState.NOT_CALLEDBACK,
                'callback_url': 'https://baidu.com',
                'bank_memo': ouuid,
                'out_money_userid': '',
                "lose_reason":"",
                'ip': '127.0.0.1',
                'agentadmin_uuid': merchant_data.get('agentadmin_uuid'),
            }
            # print('_order_daat:', _order_daat)
            # print('*'*50)
            self.MCLS.insert_one(_order_daat)

        return self.xtjson.json_result()
    
    def post_other_way(self):
        if self.action == 'createTestOrder':
            self.createTestOrder()
            return self.xtjson.json_result()
        if self.action == 'get_total_info':
            filter_dict = self.get_filter_dict()
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if statu:
                filter_dict.update(res[0])

            # 总笔数
            number_count = self.MCLS.count(filter_dict) or 0
            # 总订单金额
            amount_total_ll = self.MCLS.collection().aggregate([
                {"$match": filter_dict},
                {"$group": {"_id": None, "actual_amount": {"$sum": '$actual_amount'}}},
            ])
            amount_total_l = list(amount_total_ll)
            amount_total = 0
            if amount_total_l:
                amount_total = amount_total_l[0].get('actual_amount')

            filter_dict['proc_statu'] = True
            processed_cnt = self.MCLS.count(filter_dict) or 0
            filter_dict['proc_statu'] = {"$ne": True}
            unprocessed_cnt = self.MCLS.count(filter_dict) or 0

            _data = {
                'amount_total': amount_total,
                'processed_cnt': processed_cnt,
                'unprocessed_cnt': unprocessed_cnt,
                'number_count': self.format_money(number_count),
            }
            return self.xtjson.json_result(data=_data)
        if self.action == 'export_order':
            filter_dict = self.get_filter_dict()
            
            sort_query = 'create_time'
            print(filter_dict)
            if 'order_time' in filter_dict:
                sort_query = 'order_time'
            elif 'proc_time' in filter_dict:
                sort_query = 'proc_time'
            datas = CollectionOrderTable.find_many(filter_dict, sort=[[sort_query, -1]])

            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            FOLDER_name = ''
            agentadmin_uuid = ''
            if self.current_admin_roleCode == ROlE_ALL.SUPERADMIN or self.current_admin_roleCode == ROlE_ALL.ADMINISTRATOR:
                FOLDER_name = 'su'
            else:
                if self.is_agentadmin:
                    FOLDER_name = self.current_admin_dict.get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('uuid')
                elif self.current_admin_dict.get('agentadmin_data'):
                    FOLDER_name = self.current_admin_dict.get('agentadmin_data').get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('agentadmin_data').get('uuid')
            if not FOLDER_name:
                return self.xtjson.json_params_error()
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER, FOLDER_name)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            _out_data_dict = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas),
                'out_count': 0,
                'note': '掉单列表-' + datetime.datetime.now().strftime('%Y%m%d%H%M%S'),
                'agentadmin_uuid': agentadmin_uuid,
            }
            uuid = ExportDataModel.insert_one(_out_data_dict)
            threading.Thread(target=self.exportData, args=(datas, uuid, export_folder, filename)).start()
            return self.xtjson.json_result(message='数据导出中，请稍后到"导出文件"中查看数据！')


class WithdrawListView(CmsTableViewBase):
    add_url_rules = [['/withdrawList', 'withdrawList']]
    per_page = 30
    MCLS = WithdrawTable
    template = 'cms/merchant/withdrawList.html'
    title = '提现申请'

    def exportData(self, datas, log_uuid, export_folder, filename):
        ''' 导出数据 '''
        export_data = ExportDataModel.find_one({'uuid': log_uuid})
        if not export_data:
            return
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1
            header = ['商户名称', '代理账户', '商家ID', '收款银行', '收款卡号', '收款人姓名', '提现金额',  '手续费', '实际扣除金额', '处理状态', '处理人', '处理时间', '申请时间']
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])

            for data in datas:
                row += 1
                merchant_data = MerchantTable.find_one({'uuid': data.get('merchant_uuid')})
                bankcard_data = MerchantBankCardTable.find_one({'uuid': data.get('bankcard_uuid')}) or {}
                bank_data = BankTable.find_one({'uuid': bankcard_data.get('bank_uid')}) or {}
                adminuser = CmsUserTable.find_one({'uuid': data.get('admin_uuid')})
                agentadmin_data = CmsUserTable.find_one({'uuid': merchant_data.get('agentadmin_uuid')}) or {}

                wa.cell(row=row, column=1, value=str(merchant_data.get('merchant_name') or ''))
                wa.cell(row=row, column=2, value=str(agentadmin_data.get('account') or ''))
                wa.cell(row=row, column=3, value=str(merchant_data.get('merchant_id') or ''))
                wa.cell(row=row, column=4, value=str(bank_data.get('shortName') or bank_data.get('code') or data.get("payee_bank") or ''))
                wa.cell(row=row, column=5, value=str(bankcard_data.get('account') or data.get('payee_bankcard') or ''))

                wa.cell(row=row, column=6, value=str(bankcard_data.get('account_username') or data.get('payee_username') or ''))
                wa.cell(row=row, column=7, value=self.format_money(str(data.get('amount') or '0')))
                wa.cell(row=row, column=8, value=self.format_money(str(data.get('repay_amount') or '0')))
                wa.cell(row=row, column=9, value=self.format_money(str(data.get('actual_amount') or '0')))
                wa.cell(row=row, column=10, value=str(WITHDRAW_STATUS.name_dict.get(data.get("statu"))))
                wa.cell(row=row, column=11, value=str(adminuser.get('account') or ''))

                wa.cell(row=row, column=12, value=str(self.format_time_func(data.get("dealwith_time")) or ''))
                wa.cell(row=row, column=13, value=str(self.format_time_func(data.get("create_time")) or ''))

                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return


    def get_filter_dict(self):
        fff = {}
        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')
        
        merchant_name_id = request.args.get('merchant_name_id')
        if merchant_name_id and merchant_name_id.strip():
            merchant_data = MerchantTable.find_one({'$or': [{'merchant_name': merchant_name_id.strip()}, {'merchant_id': merchant_name_id.strip()}]}) or {}
            fff['merchant_uuid'] = merchant_data.get('uuid') or ''
            self.search_dict['merchant_name_id'] = merchant_name_id

        bankcard_account = request.args.get('bankcard_account')
        if bankcard_account and bankcard_account.strip():
            bankcard_data = MerchantBankCardTable.find_one({'account': bankcard_account.strip()}) or {}
            fff['bankcard_uuid'] = bankcard_data.get('uuid') or ''
            self.search_dict['bankcard_account'] = bankcard_account

        agentadmin_account = request.args.get('agentadmin_account')
        if agentadmin_account and agentadmin_account.strip():
            agentadmin_data = CmsUserTable.find_one({'account': agentadmin_account.strip()})
            fff['agentadmin_uuid'] = agentadmin_data.get('uuid')
            self.search_dict['agentadmin_account'] = agentadmin_account

        return fff

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        datas = []
        for da in all_datas:
            merchant_data = MerchantTable.find_one({'uuid': da.get('merchant_uuid')})
            da['merchant_data'] = merchant_data
            bankcard_data = MerchantBankCardTable.find_one({'uuid': da.get('bankcard_uuid')}) or {}
            da['bankcard_data'] = bankcard_data
            bank_data = BankTable.find_one({'uuid': bankcard_data.get('bank_uid')}) or {}
            da['bank_data'] = bank_data
            adminuser = CmsUserTable.find_one({'uuid': da.get('admin_uuid')})
            da['adminuser'] = adminuser
            agentadmin_data = CmsUserTable.find_one({'uuid': merchant_data.get('agentadmin_uuid')}) or {}
            da['agentadmin_data'] = agentadmin_data
            datas.append(da)
        self.context['all_datas'] = datas
        self.context['WITHDRAW_STATUS'] = WITHDRAW_STATUS

    def post_other_way(self):
        # 统计
        if self.action == 'get_total_info':
            filter_dict = self.get_filter_dict()
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if statu:
                filter_dict.update(res[0])

            # 下发总笔数
            number_count = self.MCLS.count(filter_dict) or 0

            # 下发总金额
            amount_total_ll = self.MCLS.collection().aggregate([
                {"$match": filter_dict},
                {"$group": {"_id": None, "amount": {"$sum": '$amount'}}},
            ])
            amount_total_l = list(amount_total_ll)
            amount_total = 0
            if amount_total_l:
                amount_total = amount_total_l[0].get('amount')

            _data = {
                'amount_total': amount_total,
                'number_count': self.format_money(number_count),
            }
            return self.xtjson.json_result(data=_data)
        if self.action == 'export_order':
            filter_dict = self.get_filter_dict()
            datas = WithdrawTable.find_many(filter_dict)

            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            FOLDER_name = ''
            agentadmin_uuid = ''
            if self.current_admin_roleCode == ROlE_ALL.SUPERADMIN or self.current_admin_roleCode == ROlE_ALL.ADMINISTRATOR:
                FOLDER_name = 'su'
            else:
                if self.is_agentadmin:
                    FOLDER_name = self.current_admin_dict.get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('uuid')
                elif self.current_admin_dict.get('agentadmin_data'):
                    FOLDER_name = self.current_admin_dict.get('agentadmin_data').get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('agentadmin_data').get('uuid')
            if not FOLDER_name:
                return self.xtjson.json_params_error()
            
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER, FOLDER_name)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            _out_data_dict = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas),
                'out_count': 0,
                'note': '提现申请-' + datetime.datetime.now().strftime('%Y%m%d%H%M%S'),
                'agentadmin_uuid': agentadmin_uuid,
            }
            uuid = ExportDataModel.insert_one(_out_data_dict)
            threading.Thread(target=self.exportData, args=(datas, uuid, export_folder, filename)).start()
            return self.xtjson.json_result(message='数据导出中，请稍后到"导出文件"中查看数据！')

    def post_data_other_way(self):
        if self.action == 'withdrawSuccess':
            merchant_data = MerchantTable.find_one({'uuid': self.data_dict.get('merchant_uuid')})
            if not merchant_data:
                return self.xtjson.json_params_error()

            _data = {
                'statu': WITHDRAW_STATUS.success,
                'admin_uuid': self.current_admin_dict.get('uuid'),
                'dealwith_time': datetime.datetime.now(),
            }
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': _data})
            return self.xtjson.json_result()
        if self.action == 'withdrawReject':
            merchant_data = MerchantTable.find_one({'uuid': self.data_dict.get('merchant_uuid')})
            if not merchant_data:
                return self.xtjson.json_params_error()

            amount = self.data_dict.get('amount') or 0
            repay_amount = self.data_dict.get('repay_amount') or 0
            _aa = amount + repay_amount
            _state, _balance_amount = MerchantUpdateAmout_func(_aa, self.data_dict.get('merchant_uuid'))
            if not _state:
                return self.xtjson.json_params_error('处理失败！')

            _data = {
                'statu': WITHDRAW_STATUS.reject,
                'admin_uuid': self.current_admin_dict.get('uuid'),
                'dealwith_time': datetime.datetime.now(),
            }
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': _data})
            
            _mbs_data = {
                'merchant_uuid': self.data_dict.get('merchant_uuid'),
                'amount': amount,
                'repay_amount': repay_amount,
                'balance_amount': _balance_amount,
                'note': '',
                'bill_type': BILL_STATEMEN_TYPES.GOBACK_SETTLEMENT,
                'agentadmin_uuid': self.data_dict.get('agentadmin_uuid'),
            }
            MerchantBillStatementTable.insert_one(_mbs_data)
            return self.xtjson.json_result()
        # 获取订单信息
        if self.action == 'get_pay_info':
            bank_data = BankTable.find_one({'code': self.data_dict.get('payee_bank')}) or {}
            if not bank_data:
                return self.xtjson.json_params_error('该订单收款银行错误！')

            _data = {
                'receive_bank': bank_data.get('shortName'),
                'receive_account': self.data_dict.get('payee_bankcard'),
                'receive_owner': self.data_dict.get('payee_username'),
                'order_amount': self.data_dict.get('amount'),
                'statu': self.data_dict.get('statu'),
            }
            if self.data_dict.get('statu') == WITHDRAW_STATUS.success:
                _data['payQrcode'] = '/public/world/img/paysuccess5.png'
            elif self.data_dict.get('statu') == WITHDRAW_STATUS.reject:
                _data['payQrcode'] = '/public/world/img/ju.png'
            else:
                payqrcode_url = self.data_dict.get('payqrcode_url') or ''
                project_static_folder = os.path.join(current_app.static_folder, current_app.config.get('PROJECT_NAME'))
                _state, payQrcode = getBankPayQrcode(
                    self.data_uuid,
                    self.data_dict.get('amount'),
                    '',
                    bank_data,
                    payqrcode_url=payqrcode_url,
                    project_static_folder=project_static_folder,
                    receive_account=self.data_dict.get('payee_bankcard'),
                )
                if not _state:
                    return self.xtjson.json_params_error(payQrcode)

                _data['payQrcode'] = payQrcode
            return self.xtjson.json_result(data=_data)



class CallbackLogView(CmsTableViewBase):
    add_url_rules = [['/callbackLog', 'callbackLog']]
    per_page = 30
    MCLS = callbackLogTable
    template = 'cms/merchant/callbackLog.html'
    title = '回调记录'

    def get_filter_dict(self):
        fff = {}
        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')

        create_time = request.args.get('create_time')
        if create_time:
            start_time, end_time = PagingCLS.by_silce(create_time)
            fff.update({'create_time': {'$gte': start_time, '$lte': end_time}})
            self.search_dict['create_time'] = create_time

        order_id = request.args.get('order_id')
        if order_id and order_id.strip():
            fff['order_id'] = order_id
            self.search_dict['order_id'] = order_id
        
        request_text = request.args.get('request_text')
        if request_text and request_text.strip():
            self.search_dict['request_text'] = request_text
        
        statu = request.args.get('statu')
        if statu and statu.strip():
            fff['statu'] = True if statu == '1' else {"$ne": True}
            self.search_dict['statu'] = statu

        return fff

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        datas = []
        _dd = {}
        for da in all_datas:
            merchant_uuid = da.get('merchant_uuid')
            merchant_data = _dd.get(merchant_uuid)
            if not merchant_data:
                merchant_data = MerchantTable.find_one({'uuid': merchant_uuid})
                _dd[merchant_uuid] = merchant_data
            da['merchant_data'] = merchant_data

            order_uuid = da.get('order_uuid')
            order_data = da.get(order_uuid)
            if not order_data:
                order_data = CollectionOrderTable.find_one({'uuid': order_uuid})
            da['order_data'] = order_data
            req_str = json.dumps(da.get("request_text") or '').replace(" ", "").replace('"', "'")
            print( json.dumps(da.get("request_text") or '').replace(" ", ""))
            if 'request_text' in self.search_dict:
                req_txt = self.search_dict['request_text'].replace(" ", "").replace('"', "'")
                print(req_txt)
                print(req_str)
                if req_txt in req_str:
                    datas.append(da)
            else:
                datas.append(da)
        self.context['all_datas'] = datas

    def createTestOrder(self):
        for i in range(5):
            m_uuids = [x.get("uuid") for x in MerchantTable.find_many()]
            m_uuid = m_uuids[random.randint(0, len(m_uuids)-1)]
            
            o_uuids = [x.get("uuid") for x in CollectionOrderTable.find_many()]
            o_uuid = o_uuids[random.randint(0, len(o_uuids)-1)]

            order_ids = [x.get("order_id") for x in CollectionOrderTable.find_many()]
            order_id = order_ids[random.randint(0, len(order_ids)-1)]

            _order_daat = {
                'order_uuid': o_uuid,
                'order_id': order_id,
                'merchant_uuid': m_uuid,
                'affidavit_text': str(shortuuid.uuid()).replace("A", ' '),
                'callback_url': 'https://baidu.com',
                'callbank_type': 'manual',
                'admin_uuid': self.current_admin_dict.get("uuid"),
                'response_code': 500,
                'statu': False,
                'note': "HTTPSConnectionPool(host='baidu.com', port=443): Read timed out.",
                'response_text': '',
                'request_text': {
                    "mchId": str(random.randint(100000,200000)),
                    "mchOrderId": shortuuid.uuid(),
                    "amount": random.randint(100,200),
                    "payAmount": random.randint(100,200),
                    "isPaid": 1,
                    "payMethod": 7,
                    "sign": "0a627c0e775fdeeb149eb6c7d6591148",                
                },
            }
            # print('_order_daat:', _order_daat)
            # print('*'*50)
            self.MCLS.insert_one(_order_daat)

        return self.xtjson.json_result()

    def post_other_way(self):
        if self.action == 'createTestOrder':
            self.createTestOrder()
            return self.xtjson.json_result()


class unknownIncomeView(CmsTableViewBase):
    add_url_rules = [['/unknownIncome', 'unknownIncome']]
    per_page = 30
    MCLS = unknownIncomeTable
    template = 'cms/merchant/unknownIncome.html'
    title = '不明收入'
    
    def exportData(self, datas, log_uuid, export_folder, filename):
        ''' 导出数据 '''
        export_data = ExportDataModel.find_one({'uuid': log_uuid})
        if not export_data:
            return
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1
            header = ['收款银行卡号','原订单ID','金额','记录时间','代理商','处理时间','处理状态','处理人','订单备注','备注']
            
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])

            for data in datas:
                row += 1
                agenetdata = CmsUserTable.find_one({"uuid": data.get('agentadmin_uuid')})
                userdata = CmsUserTable.find_one({"uuid": data.get('admin_uid')})

                wa.cell(row=row, column=1, value=str(data.get('receive_bankacrd_account') or ''))
                wa.cell(row=row, column=2, value=str(data.get('bank_bill_id') or ''))
                wa.cell(row=row, column=3, value=str(self.format_money(data.get('amount')) or '0'))
                wa.cell(row=row, column=4, value=str(self.format_time_func(data.get('create_time')) or ''))
                wa.cell(row=row, column=5, value=self.format_money(data.get('order_amount')))
                wa.cell(row=row, column=6, value=f'''{ agenetdata.get('account') or '' }-{agenetdata.get('username') or '' }''' )
                wa.cell(row=row, column=7, value=str(self.format_time_func(data.get('proc_time')) or ''))
                wa.cell(row=row, column=8, value='已处理' if data.get('state') else '未处理')
                wa.cell(row=row, column=9, value=userdata.get('account')) or ''
                wa.cell(row=row, column=10, value=str(data.get('order_note') or ''))
                wa.cell(row=row, column=11, value=str(data.get('note') or ''))

                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def add_data_html(self, data_dict={}):
        _action = 'add_data'
        if data_dict:
            _action = 'edit_data'

        bddl_html = ''
        for uda in CmsUserTable.find_many({'role_code': ROlE_ALL.AGENTADMIN}):
            bddl_html += f'<option value="{uda.get("uuid")}">{uda.get("account")}-{uda.get("username")}</option>'

        select_dl_html = ''
        if not data_dict and (self.current_admin_user.is_superadmin or self.current_admin_user.is_administrator):
            select_dl_html += f'''
                <div class="list-group-item" style="display: none;">
                    <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">选择绑定代理：</span>
                    <select class="form-control" id="agentadmin_uuid" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                        <option value="">选择绑定代理</option>
                        {bddl_html}
                    </select>
                </div>
            '''
        ttt = '''
            <div class="list-group-item">
                <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">收款时间：</span>
                <input type="text" class="form-control singleTime" id="receive_time" value="{data_dict.get('receive_time') or ''}" placeholder="收款时间" onmouseenter="$.single_time('.singleTime')" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
            </div>        
        '''
        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">    
                    {select_dl_html}   
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">收款银行卡号：</span>
                        <input type="text" class="form-control" id="receive_bankacrd_account" value="{data_dict.get('receive_bankacrd_account') or ''}" placeholder="收款银行卡号" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">金额：</span>
                        <input type="number" class="form-control" id="amount" value="{data_dict.get('amount') or ''}" placeholder="金额" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                                                                   
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">备注：</span>
                        <input type="text" class="form-control" id="note" value="{data_dict.get('note') or ''}" placeholder="备注" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_form_data('{_action}', '{data_dict.get('uuid') if data_dict else ''}')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>
        '''
        return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))

    def get_reqorder(self):
        create_time = request.args.get('create_time')
        if create_time and create_time.strip():
            start_time, end_time = PagingCLS.by_silce(create_time)
        else:
            crrdate = datetime.datetime.now()
            start_time, end_time = datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 0, 0,0), datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 23,59, 59)
            create_time = start_time.strftime('%Y-%m-%d %H:%M:%S') + '|' + end_time.strftime('%Y-%m-%d %H:%M:%S')

        proc_time = request.args.get('proc_time')
        proc_start_time = proc_end_time = ''

        if proc_time and proc_time.strip():
            proc_start_time, proc_end_time = PagingCLS.by_silce(proc_time)
        # else:
        #     crrdate = datetime.datetime.now()
        #     proc_start_time, proc_end_time = datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 0, 0,0), datetime.datetime(crrdate.year, crrdate.month, crrdate.day, 23,59, 59)
        #     proc_time = proc_start_time.strftime('%Y-%m-%d %H:%M:%S') + '|' + proc_end_time.strftime('%Y-%m-%d %H:%M:%S')

        return create_time, start_time, end_time, proc_time, proc_start_time, proc_end_time

    def get_filter_dict(self):
        fff = {}
        create_time, start_time, end_time, proc_time, proc_start_time, proc_end_time = self.get_reqorder()

        if request.args.get("proc_time"):
            fff.update({'proc_time': {'$gte': proc_start_time, '$lte': proc_end_time}})
        if request.args.get("create_time"):
            fff.update({'create_time': {'$gte': start_time, '$lte': end_time}})

        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')
        return fff

    def createTestOrder(self):
        for i in range(5):
            amount = random.choice(list(range(10, 1000)))
            agentadmin_uuid = ''
            if self.current_admin_dict.get("role_code") == ROlE_ALL.AGENTADMIN:
                agentadmin_uuid = self.current_admin_dict.get("uuid")
            else :
                agents = [x.get("uuid") for x in CmsUserTable.find_many({"role_code": ROlE_ALL.AGENTADMIN})]
                agentadmin_uuid = agents[random.randint(0, len(agents)-1)]
            
            bankcards = [x.get("account") for x in BankCardTable.find_many()]
            bankcard = bankcards[random.randint(0, len(bankcards)-1)]
            _order_daat = {
                'amount': amount,
                'admin_uid': self.current_admin_dict.get("uuid"),
                'agentadmin_uuid': agentadmin_uuid,
                'bank_bill_id': str(shortuuid.uuid())[:8],
                'receive_bankacrd_account': bankcard,
                'state': False,
                'order_note': str(shortuuid.uuid())[:5],
            }
            # print('_order_daat:', _order_daat)
            # print('*'*50)
            self.MCLS.insert_one(_order_daat)

        return self.xtjson.json_result()

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        udicts = {}
        datas = []
        for da in all_datas:
            udata = udicts.get(da.get('admin_uid')) or {}
            if not udata and da.get('admin_uid'):
                udata = CmsUserTable.find_one({'uuid': da.get('admin_uid')}) or {}
                udicts[da.get('admin_uid')] = udata
            da['user_data'] = udata

            udata1 = udicts.get(da.get('agentadmin_uuid')) or {}
            if not udata1 and da.get('agentadmin_uuid'):
                udata1 = CmsUserTable.find_one({'uuid': da.get('agentadmin_uuid')}) or {}
                udicts[da.get('agentadmin_uuid')] = udata1
            da['agentadmin_data'] = udata1

            datas.append(da)

        create_time, start_time, end_time , proc_time, proc_start_time, proc_end_time= self.get_reqorder()
        search_res = self.context.get('search_res') or {}
        search_res['proc_time'] = proc_time or ''
        search_res['create_time'] = create_time
        self.context['search_res'] = search_res
        self.context['all_datas'] = datas
        self.context['PAY_METHOD'] = PAY_METHOD
        self.context['CallbackState'] = CallbackState

    def post_other_way(self):
        if self.action == 'add_data_html':
            return self.add_data_html()
        if self.action == 'add_data':
            receive_bankacrd_account = self.request_data.get('receive_bankacrd_account')
            amount = self.request_data.get('amount')
            # receive_time = self.request_data.get('receive_time')
            note = self.request_data.get('note')
            if not receive_bankacrd_account or not amount:
                return self.xtjson.json_params_error('缺少数据！')

            try:
                amount = int(amount)
            except:
                return self.xtjson.json_params_error('金额输入错误！')
            if amount <= 0:
                return self.xtjson.json_params_error('金额输入错误！')

            # try:
            #     receive_time = datetime.datetime.strptime(receive_time, '%Y-%m-%d %H:%M:%S')
            # except:
            #     return self.xtjson.json_params_error('收款时间错误！')

            if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
                agentadmin_uuid = self.current_admin_dict.get('uuid')
            elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
                agentadmin_uuid = self.current_admin_dict.get('agentadmin_uuid')
            elif self.current_admin_user.is_superadmin or self.current_admin_user.is_administrator:
                agentadmin_uuid = self.request_data.get('agentadmin_uuid')
                if not agentadmin_uuid:
                    return self.xtjson.json_params_error('请选择绑定的代理！')
            else:
                return self.xtjson.json_params_error('添加失败！')

            _data = {
                'receive_bankacrd_account': receive_bankacrd_account.strip(),
                # 'receive_time': receive_time,
                'amount': amount,
                'note': note,
                'agentadmin_uuid': agentadmin_uuid,
            }
            self.MCLS.insert_one(_data)
            return self.xtjson.json_result()
        if self.action == 'del_all':
            return self.xtjson.json_params_error()
        if self.action == 'get_total_info':
            filter_dict = self.get_filter_dict()
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if statu:
                filter_dict.update(res[0])

            # 总笔数
            number_count = self.MCLS.count(filter_dict) or 0
            # 总订单金额
            amount_total_ll = self.MCLS.collection().aggregate([
                {"$match": filter_dict},
                {"$group": {"_id": None, "amount": {"$sum": '$amount'}}},
            ])
            amount_total_l = list(amount_total_ll)
            amount_total = 0
            if amount_total_l:
                amount_total = amount_total_l[0].get('amount')

            filter_dict['state'] = True
            processed_cnt = self.MCLS.count(filter_dict) or 0
            filter_dict['state'] = {"$ne": True}
            unprocessed_cnt = self.MCLS.count(filter_dict) or 0


            _data = {
                'amount_total': amount_total,
                'processed_cnt': processed_cnt,
                'unprocessed_cnt': unprocessed_cnt,
                'number_count': self.format_money(number_count),
            }
            return self.xtjson.json_result(data=_data)
        if self.action == 'createTestOrder':
            self.createTestOrder()
            return self.xtjson.json_result()
        if self.action == 'export_order':
            filter_dict = self.get_filter_dict()
            
            sort_query = 'create_time'
            if 'create_time' in filter_dict:
                sort_query = 'create_time'
            elif 'proc_time' in filter_dict:
                sort_query = 'proc_time'
            datas = self.MCLS.find_many(filter_dict, sort=[[sort_query, -1]])

            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            FOLDER_name = ''
            agentadmin_uuid = ''
            if self.current_admin_roleCode == ROlE_ALL.SUPERADMIN or self.current_admin_roleCode == ROlE_ALL.ADMINISTRATOR:
                FOLDER_name = 'su'
            else:
                if self.is_agentadmin:
                    FOLDER_name = self.current_admin_dict.get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('uuid')
                elif self.current_admin_dict.get('agentadmin_data'):
                    FOLDER_name = self.current_admin_dict.get('agentadmin_data').get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('agentadmin_data').get('uuid')
            if not FOLDER_name:
                return self.xtjson.json_params_error()
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER, FOLDER_name)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            _out_data_dict = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas),
                'out_count': 0,
                'note': '掉单列表-' + datetime.datetime.now().strftime('%Y%m%d%H%M%S'),
                'agentadmin_uuid': agentadmin_uuid,
            }
            uuid = ExportDataModel.insert_one(_out_data_dict)
            threading.Thread(target=self.exportData, args=(datas, uuid, export_folder, filename)).start()
            return self.xtjson.json_result(message='数据导出中，请稍后到"导出文件"中查看数据！')

    def post_data_other_way(self):
        if self.action == 'edit_data_html':
            return self.add_data_html(self.data_dict)
        if self.action == 'edit_data':
            receive_bankacrd_account = self.request_data.get('receive_bankacrd_account')
            amount = self.request_data.get('amount')
            # receive_time = self.request_data.get('receive_time')
            note = self.request_data.get('note')
            if not receive_bankacrd_account or not amount:
                return self.xtjson.json_params_error('缺少数据！')

            try:
                amount = int(amount)
            except:
                return self.xtjson.json_params_error('金额输入错误！')
            if amount <= 0:
                return self.xtjson.json_params_error('金额输入错误！')

            # try:
            #     receive_time = datetime.datetime.strptime(receive_time, '%Y-%m-%d %H:%M:%S')
            # except:
            #     return self.xtjson.json_params_error('收款时间错误！')

            _data = {
                'receive_bankacrd_account': receive_bankacrd_account.strip(),
                # 'receive_time': receive_time,
                'amount': amount,
                'note': note,
            }
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': _data})
            return self.xtjson.json_result()
        if self.action == 'updateState':
            if self.data_dict.get('state'):
                return self.xtjson.json_params_error('该订单已处理！不可重复操作！')
            self.data_from['state'] = True
            self.data_from['proc_time'] = datetime.datetime.now()
            self.data_from['admin_uid'] = self.current_admin_dict.get('uuid')
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            return self.xtjson.json_result()
        if self.action == 'updateNote':
            data_value = self.request_data.get('data_value') or ''

            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': {
                'note': data_value.strip(),
                'proc_time': datetime.datetime.now(),
                'state': True
            }})
            return self.xtjson.json_result()
        if self.action == 'del':
            self.MCLS.delete_one({"uuid": self.data_dict.get("uuid")})
            return self.xtjson.json_result()
class MerchantFundsDetailView(CmsTableViewBase):
    add_url_rules = [['/merchant/fundsDetail/<string:muid>', 'merchant_fundsDetail']]
    per_page = 30
    MCLS = MerchantBillStatementTable
    template = 'cms/merchant/fundsDetail.html'
    title = '资金流水'

    def exportData(self, datas, log_uuid, export_folder, filename):
        ''' 导出数据 '''
        export_data = ExportDataModel.find_one({'uuid': log_uuid})
        if not export_data:
            return
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1
            header = ['商户ID', '商户名称', '交易订单号', '交易商户订单号','交易类型', '交易金额', '手续费', '当前余额', '账单时间']
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])

            merchant_data_dict = {}
            for data in datas:
                row += 1
                merchant_uuid = data.get('merchant_uuid') or ''
                merchant_data = merchant_data_dict.get(merchant_uuid)
                if not merchant_data:
                    merchant_data = MerchantTable.find_one({'uuid': merchant_uuid}) or {}
                    merchant_data_dict[merchant_uuid] = merchant_data

                wa.cell(row=row, column=1, value=str(merchant_data.get('amount') or ''))
                wa.cell(row=row, column=2, value=str(merchant_data.get('merchant_name') or ''))
                wa.cell(row=row, column=3, value=str(data.get('order_id') or ''))
                wa.cell(row=row, column=4, value=str(data.get('merchant_order_id') or ''))
                wa.cell(row=row, column=5, value=str(BILL_STATEMEN_TYPES.name_dict.get(data.get('bill_type')) or ''))
                wa.cell(row=row, column=6, value=str(data.get('amount') or ''))
                wa.cell(row=row, column=7, value=str(data.get('repay_amount') or ''))
                wa.cell(row=row, column=8, value=str(data.get('balance_amount') or ''))
                wa.cell(row=row, column=9, value=data.get('create_time').strftime('%Y-%m-%d %H:%M:%S'))

                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def get_filter_dict(self):
        fff = {}
        muid = self.kwargs.get('muid')
        if muid:
            fff['merchant_uuid'] = muid
        
        note = request.args.get('note')
        if note:
            fff['note'] = {'$regex': note}

        return fff

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        datas = []
        for da in all_datas:
            merchant_data = MerchantTable.find_one({'uuid': da.get('merchant_uuid')})
            da['merchant_data'] = merchant_data
            datas.append(da)
        self.context['all_datas'] = datas
        self.context['BILL_STATEMEN_TYPES'] = BILL_STATEMEN_TYPES

    def post_other_way(self):
        if self.action == 'get_total_info':
            filter_dict = {}
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if not statu:
                return self.xtjson.json_params_error(res)
            filter_dict.update(res[0])
            muid = self.kwargs.get('muid')
            if muid:
                filter_dict['merchant_uuid'] = muid

            fff1 = {}
            fff1.update(filter_dict)
            fff1.update({
                'bill_type': BILL_STATEMEN_TYPES.INCOME_ORDER,
            })
            # 代收成功笔数
            ds_success_count = MerchantBillStatementTable.count(fff1) or 0
            # 代收成功总金额
            _amount_total_ll = MerchantBillStatementTable.collection().aggregate([
                {"$match": fff1},
                {"$group": {"_id": None, "amount": {"$sum": '$amount'}}},
            ])
            ds_success_amount_total_ll = list(_amount_total_ll)
            ds_success_amount_total = 0
            if ds_success_amount_total_ll:
                ds_success_amount_total = ds_success_amount_total_ll[0].get('amount')

            # 代收总手续费
            _repay_amount_total_ll = MerchantBillStatementTable.collection().aggregate([
                {"$match": fff1},
                {"$group": {"_id": None, "repay_amount": {"$sum": '$repay_amount'}}},
            ])
            repay_amount_total_ll = list(_repay_amount_total_ll)
            ds_repay_amount_total = 0
            if repay_amount_total_ll:
                ds_repay_amount_total = repay_amount_total_ll[0].get('repay_amount')
            
            fff2 = {}
            fff2.update(filter_dict)
            fff2.update({
                'bill_type': BILL_STATEMEN_TYPES.PAY_BEHALF,
            })
            # 代付成功笔数
            df_success_count = MerchantBillStatementTable.count(fff1) or 0
            # 代付成功总金额
            _amount_total_ll = MerchantBillStatementTable.collection().aggregate([
                {"$match": fff2},
                {"$group": {"_id": None, "amount": {"$sum": '$amount'}}},
            ])
            df_success_amount_total_ll = list(_amount_total_ll)
            df_success_amount_total = 0
            if df_success_amount_total_ll:
                df_success_amount_total = df_success_amount_total_ll[0].get('amount')
            # 代付总手续费
            _repay_amount_total_ll = MerchantBillStatementTable.collection().aggregate([
                {"$match": fff2},
                {"$group": {"_id": None, "repay_amount": {"$sum": '$repay_amount'}}},
            ])
            repay_amount_total_ll = list(_repay_amount_total_ll)
            df_repay_amount_total = 0
            if repay_amount_total_ll:
                df_repay_amount_total = repay_amount_total_ll[0].get('repay_amount')

            fff3 = {}
            fff3.update(filter_dict)
            fff3.update({
                'bill_type': BILL_STATEMEN_TYPES.SETTLEMENT,
            })
            # 结算成功总手续费
            js_success_ll = MerchantBillStatementTable.collection().aggregate([
                {"$match": fff3},
                {"$group": {"_id": None, "amount": {"$sum": '$amount'}}},
            ])
            js_success_ll = list(js_success_ll)
            js_success_amount = 0
            if js_success_ll:
                js_success_amount = js_success_ll[0].get('amount')

            fff3.update({
                'bill_type': BILL_STATEMEN_TYPES.GOBACK_SETTLEMENT,
            })
            # 结算成功总金额
            js_success_th_ll = MerchantBillStatementTable.collection().aggregate([
                {"$match": fff3},
                {"$group": {"_id": None, "amount": {"$sum": '$amount'}}},
            ])
            js_success_th_ll = list(js_success_th_ll)
            js_success_amount_th = 0
            if js_success_th_ll:
                js_success_amount_th = js_success_th_ll[0].get('amount')
            js_successAmount = abs(js_success_amount) - abs(js_success_amount_th)

            # 内充和减金额差值
            fff3 = {}
            fff3.update(filter_dict)
            fff3.update({
                'bill_type': BILL_STATEMEN_TYPES.RECHARGE,
            })
            nc_amount_ll = MerchantBillStatementTable.collection().aggregate([
                {"$match": fff3},
                {"$group": {"_id": None, "amount": {"$sum": '$amount'}}},
            ])
            nc_amount_ll = list(nc_amount_ll)
            nc_amount = 0
            if nc_amount_ll:
                nc_amount = nc_amount_ll[0].get('amount')
            fff3.update({
                'bill_type': BILL_STATEMEN_TYPES.REDUCE,
            })
            jje_amount_ll = MerchantBillStatementTable.collection().aggregate([
                {"$match": fff3},
                {"$group": {"_id": None, "amount": {"$sum": '$amount'}}},
            ])
            jje_amount_ll = list(jje_amount_ll)
            jje_amount = 0
            if jje_amount_ll:
                jje_amount = jje_amount_ll[0].get('amount')

            nc_jje_vv = nc_amount - abs(jje_amount)

            _data = {
                'ds_success_count': self.format_money(ds_success_count),
                'ds_success_amount_total': self.format_money(ds_success_amount_total),
                'ds_repay_amount_total': self.format_money(ds_repay_amount_total),
                'df_success_count': self.format_money(df_success_count),
                'df_success_amount_total': self.format_money(df_success_amount_total),
                'df_repay_amount_total': self.format_money(df_repay_amount_total),
                'js_successAmount': self.format_money(js_successAmount),
                'nc_jje_vv': self.format_money(nc_jje_vv),
            }
            return self.xtjson.json_result(data=_data)
        if self.action == 'exportBill':
            filter_dict = {}
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if not statu:
                return self.xtjson.json_params_error(res)
            filter_dict.update(res[0])
            muid = self.kwargs.get('muid')
            if muid:
                filter_dict['merchant_uuid'] = muid

            datas = self.MCLS.find_many(filter_dict)
            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            FOLDER_name = ''
            agentadmin_uuid = ''
            if self.current_admin_roleCode == ROlE_ALL.SUPERADMIN or self.current_admin_roleCode == ROlE_ALL.ADMINISTRATOR:
                FOLDER_name = 'su'
            else:
                if self.is_agentadmin:
                    FOLDER_name = self.current_admin_dict.get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('uuid')
                elif self.current_admin_dict.get('agentadmin_data'):
                    FOLDER_name = self.current_admin_dict.get('agentadmin_data').get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('agentadmin_data').get('uuid')
            if not FOLDER_name:
                return self.xtjson.json_params_error()
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER, FOLDER_name)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            _out_data_dict = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas),
                'out_count': 0,
                'note': '商户资金流水-' + datetime.datetime.now().strftime('%Y%m%d%H%M%S'),
                'agentadmin_uuid': agentadmin_uuid,
            }
            uuid = ExportDataModel.insert_one(_out_data_dict)
            threading.Thread(target=self.exportData, args=(datas, uuid, export_folder, filename)).start()
            return self.xtjson.json_result(message='数据导出中，请稍后到"导出文件"中查看数据！')



class AgentadminBillLogView(CmsTableViewBase):
    add_url_rules = [['/agentadmin/bills/<string:auid>', 'AgentadminBillLogView']]
    per_page = 30
    MCLS = AgentadminBillLogTable
    template = 'cms/userManage/agentadminBill.html'
    title = '代理资金流水'

    def get_filter_dict(self):
        fff = {}
        auid = self.kwargs.get('auid')
        if auid:
            fff['agentadmin_uuid'] = auid
            
        merchant_name_id = request.args.get('merchant_name_id')
        if merchant_name_id and merchant_name_id.strip():
            merchant_data = MerchantTable.find_one({'$or': [{'merchant_name': merchant_name_id.strip()}, {'merchant_id': merchant_name_id.strip()}]}) or {}
            fff['merchant_id'] = merchant_data.get('merchant_id') or ''
            self.search_dict['merchant_name_id'] = merchant_name_id
            
        return fff

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        datas = []
        merchant_data_dict = {}
        agentadmin_data_dict = {}
        for da in all_datas:
            agentadmin_uuid = da.get('agentadmin_uuid') or ''
            agentadmin_data = agentadmin_data_dict.get(agentadmin_uuid)
            if not agentadmin_data:
                agentadmin_data = CmsUserTable.find_one({'uuid': agentadmin_uuid})
                agentadmin_data_dict[agentadmin_uuid] = agentadmin_data
            da['agentadmin_data'] = agentadmin_data
            merchant_data = merchant_data_dict.get(da.get('merchant_id'))
            if not merchant_data:
                merchant_data = MerchantTable.find_one({'merchant_id': da.get('merchant_id')}) or {}
                merchant_data_dict[da.get('merchant_id')] = merchant_data
            da['merchant_data'] = merchant_data
            datas.append(da)
        self.context['all_datas'] = datas
        self.context['BILL_STATEMEN_TYPES'] = BILL_STATEMEN_TYPES


