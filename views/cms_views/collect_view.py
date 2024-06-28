''' 收款 '''
import json
import threading, os, random, string, datetime, time
import shortuuid
from openpyxl import Workbook
from flask import abort, current_app, request
from .cms_base import CmsTableViewBase
from models.cms_user import CmsUserTable
from models.pay_table import BankCardTable, BankTable, MerchantTable, BankScriptLogTable, VpnTable, BankCardBillTable, ScriptLogTable
from constants import BANK_CODE, ROlE_ALL, ASSETS_FOLDER, EXPORT_FOLDER, ExportStatu, unusualTypes, BankCardType, BankBillTypes, PAY_METHOD, OTP_BANK_CODE, SC_OTP_BANK_CODE
import ctypes
# from modules.bank_module.ACB import BANK_ACB
# from modules.bank_module.TPB import BANK_TPB
# from modules.bank_module.SEAB import BANK_SEAB
from modules.bank_module.ICB import BANK_ICB
# from modules.bank_module.MB import BANK_MB
# from modules.bank_module.NAB import BANK_NAB
# from modules.bank_module.MSB import BNAK_MSB
from modules.bank_module.VAB import BANK_VAB
# from modules.bank_module.BAB import BANK_BAB
# from modules.bank_module.VIETBANK import BANK_VIETBANK
# from modules.bank_module.VPB import BANK_VPB
from modules.bank_module.VCB import BANK_VCB
from modules.bank_module.TCB import BANK_TCB
from modules.view_helpres.bank_func import ACB_script_func, SEAB_script_func, TPB_script_func, ICB_script_func, MB_script_func, balanceData_func, NAB_script_func, VAB_script_func, BAB_script_func, VIETBANK_script_func, VPB_script_func, VCB_script_func
from models.site_table import ExportDataModel
from common_utils.lqredis import SiteRedis
from modules.view_helpres.view_func import get_vpnurl
from common_utils.utils_funcs import update_language
ACB = os.path.join(os.path.dirname(__file__), 'modules\\bank_module', 'ACB.so')
TPB = os.path.join(os.path.dirname(__file__), 'modules\\bank_module', 'TPB.so')
SEAB = os.path.join(os.path.dirname(__file__), 'modules\\bank_module', 'SEAB.so')
MB = os.path.join(os.path.dirname(__file__), 'modules\\bank_module', 'MB.so')
NAB = os.path.join(os.path.dirname(__file__), 'modules\\bank_module', 'NAB.so')
MSB = os.path.join(os.path.dirname(__file__), 'modules\\bank_module', 'MSB.so')
BAB = os.path.join(os.path.dirname(__file__), 'modules\\bank_module', 'BAB.so')
VIETBANK = os.path.join(os.path.dirname(__file__), 'modules\\bank_module', 'VIETBANK.so')
VPB = os.path.join(os.path.dirname(__file__), 'modules\\bank_module', 'VPB.so')

class BankCardView(CmsTableViewBase):
    add_url_rules = [['/bankCard', 'bankCard_view']]
    title = '银行卡管理'
    MCLS = BankCardTable
    per_page = 20
    template = 'cms/bankCard/bankCard.html'
    sort = [['statu', -1], ['start_time', -1]]

    def aout_xjlog(self):
        if not self.data_dict:
            return self.xtjson.json_params_error()
        datas = ScriptLogTable.find_many({'note': {'$regex': self.data_dict.get('account')}}, sort=[['create_time', -1]], limit=30)
        if not datas:
            return self.xtjson.json_params_error('暂无数据！')
        html = ''
        html += f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto; text-align: left; overflow-x: hidden; box-sizing: border-box;">               
                    <ul class="layui-timeline">
        '''
        for data in datas:
            html += f'''
                <li class="layui-timeline-item">
                    <i class="layui-icon layui-timeline-axis"></i>
                    <div class="layui-timeline-content layui-text">
                        <div class="layui-timeline-title">{self.format_time_func(data.get('create_time'))}，{data.get('note') or ''}</div>
                    </div>
                </li>            
            '''
        html += '''
                    </ul>        
                </div>
            </div>        
        '''
        return self.xtjson.json_result(message=html)

    def check_bankcard_abnormal(self, bankcard_data):
        '''
        检测银行卡是否异常
        '''
        update_balance_amount_time = bankcard_data.get('update_balance_amount_time')
        update_newbalance_amount_time = bankcard_data.get('update_newbalance_amount_time')
        crr_time = datetime.datetime.now() - datetime.timedelta(minutes=5)
        if not update_balance_amount_time or update_balance_amount_time < crr_time:
            return True, '余额更新异常！'
        if not update_newbalance_amount_time:
            return True, '余额更新异常！'
        # bankcard_bill_data = BankCardBillTable.find_one({'bankacrd_uuid': bankcard_data.get('uuid')}, sort=[['create_time', -1]])
        # update_newbalance_amount_time = update_newbalance_amount_time + datetime.timedelta(minutes=5)
        # if bankcard_bill_data.get('create_time') > update_newbalance_amount_time:
        #     return True, '余额更新异常！'
        return False, ''

    def get_filter_dict(self):
        bankcard_type = request.args.get('bankcard_type')
        fff = {
#            'method_type': PAY_METHOD.VNBANKQR
        }
        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
            if not bankcard_type:
                fff['bankcard_type'] = BankCardType.AGENTADMIN_CARD
                self.search_dict['bankcard_type'] = BankCardType.AGENTADMIN_CARD
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            if self.current_admin_dict.get('agentadmin_data').get('is_syscard'):
                fff['$or'] = [{'bankcard_type': BankCardType.SYSTEM_CARD}, {'agentadmin_uuid': self.current_admin_dict.get('agentadmin_uuid')}]
            else:
                fff['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.ADMINISTRATOR:
            if not bankcard_type:
                fff['bankcard_type'] = BankCardType.SYSTEM_CARD
                self.search_dict['bankcard_type'] = BankCardType.SYSTEM_CARD

        agentadmin_account = request.args.get('agentadmin_account') or ''
        if agentadmin_account and agentadmin_account.strip():
            agentadmin_data = CmsUserTable.find_one({'role_code': ROlE_ALL.AGENTADMIN, 'account': agentadmin_account.strip()}) or {}
            if agentadmin_data:
                fff['agentadmin_uuid'] = agentadmin_data.get('uuid')
            self.search_dict['agentadmin_account'] = agentadmin_account
        return fff

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas') or []
        _datas = []
        agentadmin_datas = {}
        vpn_datas = {}
        for adl in all_datas:
            bank_uid = adl.get('bank_uid') or ''
            _bacnk_data = BankTable.find_one({'uuid': bank_uid}) or {}
            adl['back_data'] = _bacnk_data

            agentadmin_uuid = adl.get('agentadmin_uuid') or ''
            if agentadmin_uuid:
                agentadmin_data = agentadmin_datas.get(agentadmin_uuid) or {}
                if not agentadmin_data:
                    agentadmin_data = CmsUserTable.find_one({'uuid': agentadmin_uuid}) or {}
                    agentadmin_datas[agentadmin_uuid] = agentadmin_data
                adl['agentadmin_data'] = agentadmin_data

            vpn_uuid = adl.get('vpn_uuid') or ''
            if vpn_uuid:
                vpn_data = vpn_datas.get(vpn_uuid) or {}
                if not vpn_data:
                    vpn_data = VpnTable.find_one({'uuid': vpn_uuid}) or {}
                    vpn_datas[vpn_uuid] = vpn_data
                adl['vpn_data'] = vpn_data
            else:
                adl['vpn_data'] = {}

            # # 检测更新是否异常
            is_abnormal2, is_abnormal_text = self.check_bankcard_abnormal(adl)
            adl['is_abnormal2'] = is_abnormal2
            _datas.append(adl)
        self.context['all_datas'] = _datas
        self.context['BankCardType'] = BankCardType

    def get_context(self):
        _back_datas = BankTable.find_many({})
        back_datas = []
        for d in _back_datas:
            if d.get('code') in BANK_CODE:
                back_datas.append(d)
        res = {
            'back_datas': back_datas
        }
        return res

    def bankCard_html(self, data_dict={}):
        _action = 'add_bankCard_data'
        if data_dict:
            _action = 'edit_bankCard_data'

        bank_html = ''
        for da in BankTable.find_many({}):
            if da.get('code') not in BANK_CODE:
                continue
            if data_dict:
                bank_html += f'''
                <option value="{da.get('uuid')}" { 'selected' if data_dict.get('bank_uid') == da.get('uuid') else '' }>{da.get('shortName')}</option>
                '''
            else:
                bank_html += '''
                <option value="%s">%s</option>
                ''' % (da.get('uuid'), da.get('shortName'))

        select_vpn_html = ''
        if data_dict:
            if data_dict.get('bankcard_type') == BankCardType.SYSTEM_CARD:
                vpn_datas = VpnTable.find_many({'statu': True})
                for vp in vpn_datas:
                    if data_dict:
                        select_vpn_html += f'''
                        <option value="{vp.get('uuid')}" { 'selected' if data_dict.get('vpn_uuid') == vp.get('uuid') else '' }>{vp.get('name')}</option>
                        '''
                    else:
                        select_vpn_html += '''
                        <option value="%s">%s</option>
                        ''' % (vp.get('uuid'), vp.get('name'))
            if data_dict.get('bankcard_type') == BankCardType.AGENTADMIN_CARD:
                vpn_datas = VpnTable.find_many({'statu': True, 'agentadmin_uuid': data_dict.get('agentadmin_uuid')})
                for vp in vpn_datas:
                    if data_dict:
                        select_vpn_html += f'''
                        <option value="{vp.get('uuid')}" { 'selected' if data_dict.get('vpn_uuid') == vp.get('uuid') else '' }>{vp.get('name')}</option>
                        '''
                    else:
                        select_vpn_html += '''
                        <option value="%s">%s</option>
                        ''' % (vp.get('uuid'), vp.get('name'))
        else:
            if not self.current_admin_user.is_superadmin or not self.current_admin_user.is_administrator:
                crr_agentadmin_uuid = ''
                if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
                    crr_agentadmin_uuid = self.current_admin_dict.get('uuid')
                elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
                    crr_agentadmin_uuid = self.current_admin_dict.get('agentadmin_uuid')

                vpn_datas = VpnTable.find_many({'statu': True, 'agentadmin_uuid': crr_agentadmin_uuid})
                for vp in vpn_datas:
                    if data_dict:
                        select_vpn_html += f'''
                        <option value="{vp.get('uuid')}" { 'selected' if data_dict.get('vpn_uuid') == vp.get('uuid') else '' }>{vp.get('name')}</option>
                        '''
                    else:
                        select_vpn_html += '''
                        <option value="%s">%s</option>
                        ''' % (vp.get('uuid'), vp.get('name'))

        mdatas = []
        merchant_html = ''
        if data_dict:
            mdatas = MerchantTable.find_many({'agentadmin_uuid': data_dict.get('agentadmin_uuid')})
            for mdata in mdatas:
                if data_dict:
                    merchant_html += f'''
                    <option value="{mdata.get('uuid')}" {'selected' if data_dict.get('merchant_uid') == mdata.get('uuid') else ''}>{mdata.get('merchant_id')}  ({mdata.get('merchant_name')})</option>
                    '''
                else:
                    merchant_html += '''
                    <option value="%s">%s  (%s)</option>
                    ''' % (mdata.get('uuid'), mdata.get('merchant_id'), mdata.get('merchant_name'))
        else:
            if not self.current_admin_user.is_superadmin or not self.current_admin_user.is_administrator:
                if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
                    mdatas = MerchantTable.find_many({'agentadmin_uuid': self.current_admin_dict.get('uuid')})
                if self.current_admin_dict.get('role_code') == ROlE_ALL.SUPERADMIN:
                    mdatas = MerchantTable.find_many({'agentadmin_uuid': self.current_admin_dict.get('agentadmin_uuid')})
                for mdata in mdatas:
                    if data_dict:
                        merchant_html += f'''
                        <option value="{mdata.get('uuid')}" { 'selected' if data_dict.get('merchant_uid') == mdata.get('uuid') else '' }>{ mdata.get('merchant_id') }  ({ mdata.get('merchant_name')})</option>
                        '''
                    else:
                        merchant_html += '''
                        <option value="%s">%s  (%s)</option>
                        ''' % (mdata.get('uuid'), mdata.get('merchant_id'), mdata.get('merchant_name'))

        bddl_html = ''
        for uda in CmsUserTable.find_many({'role_code': ROlE_ALL.AGENTADMIN}):
            bddl_html += f'<option value="{uda.get("uuid")}">{uda.get("account")}-{uda.get("username")}</option>'

        bankcard_opt = ''
        for bdc in BankCardType.name_arr:
            bankcard_opt += f'<option value="{bdc}">{ BankCardType.name_dict.get(bdc) }</option>'

        select_dl_html = ''
        bankcard_type_html = ''
        if not data_dict and (self.current_admin_user.is_superadmin or self.current_admin_user.is_administrator):
            select_dl_html += f'''
                <div class="list-group-item" style="display: none;">
                    <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">选择绑定代理：</span>
                    <select class="form-control" id="agentadmin_uuid" aria-label="" style="display: inline-block; width: calc(100% - 180px)" onchange="select_agentadmin()">
                        <option value="">选择绑定代理</option>
                        {bddl_html}
                    </select>
                </div>                   
            '''

            bankcard_type_html += f'''
                <div class="list-group-item">
                    <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">选择卡类型：</span>
                    <select class="form-control" id="bankcard_type" aria-label="" style="display: inline-block; width: calc(100% - 180px)" onchange="select_bankcard_type()">
                        <option value="">选择卡类型</option>
                        { bankcard_opt }
                    </select>
                </div>                    
            '''

        show_merchant_uid_html = False
        merchant_uid_html = f'''
            <div class="list-group-item">
                <span style="width: 120px; text-align: right; display: inline-block; position: relative;">绑定商户：</span>
                <select class="form-control" id="merchant_uid" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    <option value="">选择商户账户</option>
                    { merchant_html }
                </select>
            </div>
        '''
        if not data_dict:
            show_merchant_uid_html = True
        elif data_dict.get('bankcard_type') == BankCardType.AGENTADMIN_CARD:
            show_merchant_uid_html = True

        html = f'''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">       
                    {bankcard_type_html}
                    {select_dl_html}
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">名称：</span>
                        <input type="text" class="form-control" id="name" value="{ data_dict.get('name') or '' }" placeholder="名称" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行用户名：</span>
                        <input type="text" class="form-control" id="username" value="{ data_dict.get('username') or '' }" placeholder="银行用户名" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行密码：</span>
                        <input type="text" class="form-control" id="password" value="{ data_dict.get('password') or '' }" placeholder="银行密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行卡账号：</span>
                        <input type="text" class="form-control" id="account" value="{ data_dict.get('account') or '' }" placeholder="银行卡账号" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行账户人姓名：</span>
                        <input type="text" class="form-control" id="account_username" value="{ data_dict.get('account_username') or '' }" placeholder="银行账户人姓名" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">银行：</span>
                        <select class="form-control" id="bank_uid" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                            <option value="">选择银行</option>
                            { bank_html }
                        </select>
                    </div>
                    { merchant_uid_html if show_merchant_uid_html else '' }
                    <div class="list-group-item">
                        <span class="loglable" style="width: 120px; text-align: right; display: inline-block; position: relative;">异常自动下架时间/分钟：</span>
                        <input type="text" class="form-control" id="auto_removal_time" value="{ data_dict.get('auto_removal_time') or '' }" placeholder="异常自动下架时间/分钟" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                    
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">限制余额最大额度：</span>
                        <input type="number" class="form-control" id="stint_money" value="{ data_dict.get('stint_money') or '' }" placeholder="限制余额最大额度" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">收款最小金额：</span>
                        <input type="number" class="form-control" id="collection_money_min" value="{ data_dict.get('collection_money_min') or '' }" placeholder="收款最小金额" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">收款最大金额：</span>
                        <input type="number" class="form-control" id="collection_money_max" value="{ data_dict.get('collection_money_max') or '' }" placeholder="收款最大金额" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">日收款上限：</span>
                        <input type="number" class="form-control" id="day_collection_money_limit" value="{ data_dict.get('day_collection_money_limit') or '' }" placeholder="日收款上限" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">日收款笔数上限：</span>
                        <input type="number" class="form-control" id="day_collection_pencount_limit" value="{ data_dict.get('day_collection_pencount_limit') or '' }" placeholder="日收款笔数上限" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                    
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">初始资金:</span>
                        <input type="number" class="form-control" id="start_money" value="{ data_dict.get('start_money') or '' }" placeholder="初始资金" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">选择VPN：</span>
                        <select class="form-control" id="vpn_uuid" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                            <option value="">选择VPN</option>
                            { select_vpn_html }
                        </select>
                    </div>                    
                            
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">备注：</span>
                        <input type="text" class="form-control" id="note" value="{ data_dict.get('note') or '' }" placeholder="备注" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    </div>                        
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_bank_data('{_action}', '{ data_dict.get('uuid') if data_dict else '' }')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>
        '''
        return html

    def add_bankCard_data(self):
        name = self.request_data.get('name')
        note = self.request_data.get('note')
        username = self.request_data.get('username')
        password = self.request_data.get('password')
        account = self.request_data.get('account')
        bank_uid = self.request_data.get('bank_uid')
        vpn_uuid = self.request_data.get('vpn_uuid')
        stint_money = self.request_data.get('stint_money')
        start_money = self.request_data.get('start_money')
        merchant_uid = self.request_data.get('merchant_uid')
        account_username = self.request_data.get('account_username')
        collection_money_min = self.request_data.get('collection_money_min')
        collection_money_max = self.request_data.get('collection_money_max')
        day_collection_money_limit = self.request_data.get('day_collection_money_limit')
        day_collection_pencount_limit = self.request_data.get('day_collection_pencount_limit')
        bankcard_type = self.request_data.get('bankcard_type') or ''
        agentadmin_uuid = self.request_data.get('agentadmin_uuid') or ''
        auto_removal_time = self.request_data.get('auto_removal_time') or ''
        if not name or not bank_uid or not username or not password or not account_username or not account:
            return self.xtjson.json_params_error()
        _back_data = BankTable.find_one({'uuid': bank_uid})
        if not _back_data:
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('银行数据不存在！')

        if self.MCLS.find_one({'name': name.strip()}):
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('当前名称已存在！')

        if self.MCLS.find_one({'account': account.strip()}):
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('当前银行卡已存在！')

        if collection_money_min:
            try:
                collection_money_min = int(collection_money_min)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_min：参数错误！')
        else:
            collection_money_min = 0

        if collection_money_max:
            try:
                collection_money_max = int(collection_money_max)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_max：参数错误！')
        else:
            collection_money_max = 0

        if day_collection_money_limit:
            try:
                day_collection_money_limit = int(day_collection_money_limit)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_max：参数错误！')
        else:
            day_collection_money_limit = 0

        if stint_money:
            try:
                stint_money = int(stint_money)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('stint_money：参数错误！')
        else:
            stint_money = 0

        if auto_removal_time:
            try:
                auto_removal_time = int(auto_removal_time)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('auto_removal_time：参数错误！')
            if auto_removal_time < 3:
                return self.xtjson.json_params_error('auto_removal_time：不可小于3分钟！')
            if auto_removal_time > 1440:
                return self.xtjson.json_params_error('auto_removal_time：不可大于1440分钟！')
        else:
            auto_removal_time = 0


        if day_collection_pencount_limit:
            try:
                day_collection_pencount_limit = int(day_collection_pencount_limit)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_max：参数错误！')
        else:
            day_collection_pencount_limit = 0

        if not account.strip().isdigit():
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('银行卡账户输入错误！')

        if not account.strip().isdigit():
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('银行卡账户输入错误！')

        bank_data = BankTable.find_one({'uuid': bank_uid}) or {}
        if not bank_data:
            return self.xtjson.json_params_error('银行信息不存在！')

        if self.current_admin_user.is_superadmin or self.current_admin_user.is_administrator:
            if not bankcard_type:
                return self.xtjson.json_params_error('缺少卡类型！')
            if bankcard_type.strip() == BankCardType.AGENTADMIN_CARD:
                if not agentadmin_uuid:
                    return self.xtjson.json_params_error('缺少绑定的代理！')
        else:
            bankcard_type = BankCardType.AGENTADMIN_CARD
            if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
                agentadmin_uuid = self.current_admin_dict.get('uuid')
            elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
                agentadmin_uuid = self.current_admin_dict.get('agentadmin_uuid')
        _dd = {
            'name': name,
            'note': note,
            'account': account,
            'username': username.strip(),
            'password': password.strip(),
            'bank_uid': bank_uid,
            'balance_amount': start_money,
            'total_balance': 0,
            'statu': False,
            'script_statu': False,
            'stint_money': stint_money,
            'start_money': start_money,
            'vpn_uuid': vpn_uuid or '',
            'account_username': account_username,
            'collection_money_min': collection_money_min or 0,
            'collection_money_max': collection_money_max or 0,
            'day_collection_money_limit': day_collection_money_limit or 0,
            'day_collection_pencount_limit': day_collection_pencount_limit or 0,
            'bankcard_type': bankcard_type,
            'auto_removal_time': auto_removal_time,
            'turn_auto_removal_time': datetime.datetime.now() if auto_removal_time else '',
            'agentadmin_uuid': agentadmin_uuid,
            'method_type': PAY_METHOD.VNBANKQR,
        }
        if bankcard_type == BankCardType.AGENTADMIN_CARD:
            _dd['merchant_uid'] = merchant_uid or ''
        if bank_data.get('code') in OTP_BANK_CODE:
            _dd['is_otp_login'] = True
            _dd['login_lock'] = True

        self.MCLS.insert_one(_dd)
        self.add_SystemLog('添加银行卡信息')
        return self.xtjson.json_result()

    def post_other_way(self):
        # 添加银行卡html
        if self.action == 'add_bankCard_html':
            html = self.bankCard_html()
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        # 添加银行卡
        if self.action == 'add_bankCard_data':
            return self.add_bankCard_data()
        # 获取VPN列表
        if self.action == 'getVpnList':
            agentadmin_uuid = self.request_data.get('agentadmin_uuid')
            if not agentadmin_uuid:
                return self.xtjson.json_params_error()
            vpn_datas = VpnTable.find_many({'statu': True, 'agentadmin_uuid': agentadmin_uuid}) or []
            datas = []
            for v in vpn_datas:
                datas.append({
                    'uuid': v.get('uuid'),
                    'name': v.get('name'),
                })

            mdatas = []
            _mdatas = MerchantTable.find_many({'agentadmin_uuid': agentadmin_uuid}) or []
            for data in _mdatas:
                mdatas.append({
                    'uuid': data.get('uuid'),
                    'name': '%s (%s)'%(data.get('account'), data.get('merchant_name')),
                })
            return self.xtjson.json_result(data={'datalist': datas, 'mdatas': mdatas})
        if self.action == 'getAllVpn':
            vpn_datas = VpnTable.find_many({'statu': True}) or []
            datas = []
            for v in vpn_datas:
                datas.append({
                    'uuid': v.get('uuid'),
                    'name': v.get('name'),
                })
            return self.xtjson.json_result(data={'datalist': datas})

    def post_data_other_way(self):
        # 编辑银行卡信息HTML
        if self.action == 'edit_bankCard_html':
            html = self.bankCard_html(self.data_dict)
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        # 编辑银行卡信息
        if self.action == 'edit_bankCard_data':
            name = self.request_data.get('name')
            note = self.request_data.get('note')
            account = self.request_data.get('account')
            username = self.request_data.get('username')
            password = self.request_data.get('password')
            bank_uid = self.request_data.get('bank_uid')
            vpn_uuid = self.request_data.get('vpn_uuid')
            stint_money = self.request_data.get('stint_money')
            start_money = self.request_data.get('start_money')
            merchant_uid = self.request_data.get('merchant_uid')
            account_username = self.request_data.get('account_username')
            collection_money_min = self.request_data.get('collection_money_min')
            collection_money_max = self.request_data.get('collection_money_max')
            day_collection_money_limit = self.request_data.get('day_collection_money_limit')
            day_collection_pencount_limit = self.request_data.get('day_collection_pencount_limit')
            day_collection_pencount_limit = self.request_data.get('day_collection_pencount_limit')
            auto_removal_time = self.request_data.get('auto_removal_time') or ''
            if not name or not bank_uid or not account or not username or not password or not account_username:
                return self.xtjson.json_params_error()
            _back_data = BankTable.find_one({'uuid': bank_uid})
            if not _back_data:
                self.add_SystemLog('修改银行卡信息', code=400)
                return self.xtjson.json_params_error('银行数据不存在！')

            _ddd = self.MCLS.find_one({'name': name.strip()})
            if _ddd and _ddd.get('uuid') != self.data_uuid:
                self.add_SystemLog('修改银行卡信息', code=400)
                return self.xtjson.json_params_error('当前名称已存在！')

            _ddd = self.MCLS.find_one({'account': account.strip()})
            if _ddd and _ddd.get('uuid') != self.data_uuid:
                self.add_SystemLog('修改银行卡信息', code=400)
                return self.xtjson.json_params_error('当前银行卡已存在！')

            if collection_money_min:
                try:
                    collection_money_min = int(collection_money_min)
                except:
                    self.add_SystemLog('修改银行卡信息', code=400)
                    return self.xtjson.json_params_error('collection_money_min：参数错误！')

            if collection_money_max:
                try:
                    collection_money_max = int(collection_money_max)
                except:
                    return self.xtjson.json_params_error('collection_money_max：参数错误！')

            if stint_money:
                try:
                    stint_money = int(stint_money)
                except:
                    return self.xtjson.json_params_error('stint_money：参数错误！')
            else:
                stint_money = 0

            if day_collection_money_limit:
                try:
                    day_collection_money_limit = int(day_collection_money_limit)
                except:
                    return self.xtjson.json_params_error('day_collection_money_limit：参数错误！')

            if day_collection_pencount_limit:
                try:
                    day_collection_pencount_limit = int(day_collection_pencount_limit)
                except:
                    return self.xtjson.json_params_error('day_collection_pencount_limit：参数错误！')

            if not account.strip().isdigit():
                self.add_SystemLog('修改银行卡信息', code=400)
                return self.xtjson.json_params_error('银行卡账户输入错误！')

            if auto_removal_time:
                try:
                    auto_removal_time = int(auto_removal_time)
                except:
                    self.add_SystemLog('添加银行卡信息', code=400)
                    return self.xtjson.json_params_error('auto_removal_time：参数错误！')
                if auto_removal_time < 3:
                    return self.xtjson.json_params_error('auto_removal_time：不可小于3分钟！')
                if auto_removal_time > 1440:
                    return self.xtjson.json_params_error('auto_removal_time：不可大于1440分钟！')
            else:
                auto_removal_time = 0

            _dd = {
                'name': name,
                'note': note,
                'account': account,
                'username': username,
                'password': password,
                'bank_uid': bank_uid,
                'vpn_uuid': vpn_uuid or '',
                'stint_money': stint_money,
                'start_money': start_money,
                'account_username': account_username,
                'auto_removal_time': auto_removal_time,
                'collection_money_min': collection_money_min or 0,
                'collection_money_max': collection_money_max or 0,
                'day_collection_money_limit': day_collection_money_limit or 0,
                'day_collection_pencount_limit': day_collection_pencount_limit or 0,
            }
            if not auto_removal_time:
                _dd['turn_auto_removal_time'] = ''
            if not self.data_dict.get('auto_removal_time') and auto_removal_time:
                _dd['turn_auto_removal_time'] = datetime.datetime.now()
            if stint_money and self.data_dict.get('balance_amount') >= stint_money and self.data_dict.get('balance_amount') > 0:
                _dd['statu'] = False
            if self.data_dict.get('bankcard_type') == BankCardType.AGENTADMIN_CARD:
                _dd['merchant_uid'] = merchant_uid or ''
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': _dd})
            self.add_SystemLog('修改银行卡信息')
            return self.xtjson.json_result()
        # 更新银行卡状态
        if self.action == 'update_bankcard_statu':
            if self.data_dict.get('statu'):
                self.data_from['statu'] = False
            else:
                self.data_from['statu'] = True
                self.data_from['is_abnormal'] = False
                self.data_from['start_time'] = datetime.datetime.now()
                self.data_from['turn_auto_removal_time'] = datetime.datetime.now()
                stint_money = self.data_dict.get('stint_money') or 0
                if stint_money and self.data_dict.get('balance_amount') >= stint_money and self.data_dict.get('balance_amount') > 0:
                    return self.xtjson.json_params_error('余额已达最大值，请先转出！')
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            self.add_SystemLog('修改银行卡开启状态')
            return self.xtjson.json_result()
        # 更新银行卡脚本状态
        if self.action == 'update_script_statu':
            if self.data_dict.get('script_statu'):
                self.data_from['script_statu'] = False
            else:
                self.data_from['script_statu'] = True
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            self.add_SystemLog(f'修改银行卡开启状态：{ self.data_dict.get("account") or "" }')
            return self.xtjson.json_result()
        # 更新银行卡余额
        if self.action == 'updateBalance':
            res = balanceData_func(self.data_dict)
            if not res:
                return self.xtjson.json_params_error('更新失败！')
            self.add_SystemLog('更新银行卡余额', code=200)
            return self.xtjson.json_result()
        # 测试连接
        if self.action == 'testCard':
            bank_data = BankTable.find_one({'uuid': self.data_dict.get('bank_uid')}) or {}
            if not bank_data:
                self.add_SystemLog(note='测试银行卡脚本连接', code=400)
                return self.xtjson.json_params_error()
            vpn_url = get_vpnurl(bank_data, vpn_uuid=self.data_dict.get('vpn_uuid'), bankcard_uuid=self.data_uuid)
            bcode = bank_data.get('code')
            if bcode == 'ACB':
                bank_cls = ACB.BANK_ACB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url)
                result = bank_cls.login_acb()
            elif bcode == 'TPB':
                bank_cls = TPB.BANK_TPB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url)
                result = bank_cls.doLogin()
            elif bcode == 'SEAB':
                bank_cls = SEAB.BANK_SEAB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url)
                result = bank_cls.doLogin()
            elif bcode == 'ICB':
                bank_cls = BANK_ICB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url)
                result = bank_cls.doLogin()
            elif bcode == 'MB':
                bank_cls = MB.BANK_MB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url)
                result = bank_cls.doLogin()
            elif bcode == 'NAB':
                bank_cls = NAB.BANK_NAB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'))
                result = bank_cls.loginConnect()
            elif bcode == 'MSB':
                bank_cls = MSB.BNAK_MSB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url)
                result = bank_cls.doLogin()
            elif bcode == 'VAB':
                bank_cls = BANK_VAB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url)
                result = bank_cls.doLogin()
            elif bcode == 'BAB':
                bank_cls = BAB.BANK_BAB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url)
                result = bank_cls.doLogin()
            elif bcode == 'VIETBANK':
                bank_cls = VIETBANK.BANK_VIETBANK(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url)
                result = bank_cls.doLogin()
            elif bcode == 'VPB':
                bank_cls = VPB.BANK_VPB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url, login_lock=self.data_dict.get('login_lock') or False)
                result = bank_cls.doLogin()
            elif bcode == 'VCB':
                bank_cls = BANK_VCB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url, login_lock=self.data_dict.get('login_lock') or False)
                result = bank_cls.doLogin()
            elif bcode == 'TCB':
                bank_cls = BANK_TCB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url, login_lock=self.data_dict.get('login_lock') or False)
                result = bank_cls.doLogin()
            else:
                self.add_SystemLog('测试银行卡脚本连接', code=400)
                return self.xtjson.json_params_error('连接失败！')
            self.add_SystemLog('测试银行卡脚本连接', code=400)
            if bcode == 'VPB' and result.get('status') != 'success':
                fffk = 'vpn_' + vpn_url.strip()
                SiteRedis.set(fffk, '1', expire=60 * 25)
                return self.xtjson.json_params_error(result.get('message'))
            elif bcode == 'VCB' and not result:
                fffk = 'vpn_' + vpn_url.strip()
                SiteRedis.set(fffk, '1', expire=60 * 25)
                return self.xtjson.json_params_error('登录失败！')
            else:
                if not result:
                    fffk = 'vpn_' + vpn_url.strip()
                    SiteRedis.set(fffk, '1', expire=60 * 25)
                    return self.xtjson.json_params_error('连接失败！')

            self.add_SystemLog('测试银行卡脚本连接')
            return self.xtjson.json_result(message='连接成功！')
        # 脚本运行日志
        if self.action == 'scriptRunLog':
            html = '''<div class="formBox" style="height: 453px; overflow-y: scroll; text-align: left; line-height: 35px;">'''
            datas = BankScriptLogTable.find_many({'bankcrad_uid': self.data_uuid}, sort=[['create_time', -1]], limit=20)
            if not datas:
                html += f'''
                <div style="position: relative; display: block; margin-bottom: 20px; border-left: 3px solid #cfedf0; background-color: #f7f7f7; padding: 20px">无数据!</div>                
                '''
                return self.xtjson.json_result(message=html)
            for dd in datas:
                tt = (dd.get('log_text') or '').replace('\n', '<br>')
                html += f'''
                <div style="position: relative; display: block; margin-bottom: 20px; border-left: 3px solid #cfedf0; background-color: #f7f7f7; padding: 15px 10px 10px 10px; font-size: 13px;">
                    <p style="margin-bottom: 0px; line-height: 30px;">执行类型：{ '手动执行' if dd.get('is_is_manual') else '自动执行' }</p>
                    <p style="margin-bottom: 0px; line-height: 30px;">获取时间：{ self.format_time_func(dd.get('create_time')) }</p>
                    <p style="margin-bottom: 0px; line-height: 30px;">{tt}</p>
                </div>      
                '''
            html += '</div>'
            return self.xtjson.json_result(message=html)
        # 运行脚本
        if self.action == 'runScript':
            bank_uid = self.data_dict.get('bank_uid')
            if not bank_uid:
                return self.xtjson.json_params_error('运行失败！')
            bank_data = BankTable.find_one({'uuid': bank_uid})
            if not bank_data:
                return self.xtjson.json_params_error('银行数据不存在！')
            bank_code = bank_data.get('code')
            if bank_code == 'ACB':
                threading.Thread(target=ACB_script_func, args=(self.data_uuid,True)).start()
                return self.xtjson.json_result(message='请求成功！后台处理中...')
            elif bank_code == 'TPB':
                threading.Thread(target=TPB_script_func, args=(self.data_uuid,True)).start()
                return self.xtjson.json_result(message='请求成功！后台处理中...')
            elif bank_code == 'SEAB':
                threading.Thread(target=SEAB_script_func, args=(self.data_uuid,True)).start()
                return self.xtjson.json_result(message='请求成功！后台处理中...')
            elif bank_code == 'ICB':
                threading.Thread(target=ICB_script_func, args=(self.data_uuid,True)).start()
                return self.xtjson.json_result(message='请求成功！后台处理中...')
            elif bank_code == 'MB':
                threading.Thread(target=MB_script_func, args=(self.data_uuid,True)).start()
                return self.xtjson.json_result(message='请求成功！后台处理中...')
            elif bank_code == 'NAB':
                threading.Thread(target=NAB_script_func, args=(self.data_uuid,True)).start()
                return self.xtjson.json_result(message='请求成功！后台处理中...')
            elif bank_code == 'MSB':
                threading.Thread(target=MB_script_func, args=(self.data_uuid,True)).start()
                return self.xtjson.json_result(message='请求成功！后台处理中...')
            elif bank_code == 'VAB':
                threading.Thread(target=VAB_script_func, args=(self.data_uuid,True)).start()
                return self.xtjson.json_result(message='请求成功！后台处理中...')
            elif bank_code == 'BAB':
                threading.Thread(target=BAB_script_func, args=(self.data_uuid,True)).start()
                return self.xtjson.json_result(message='请求成功！后台处理中...')
            elif bank_code == 'VIETBANK':
                threading.Thread(target=VIETBANK_script_func, args=(self.data_uuid,True)).start()
                return self.xtjson.json_result(message='请求成功！后台处理中...')
            elif bank_code == 'VPB':
                threading.Thread(target=VPB_script_func, args=(self.data_uuid,True)).start()
                return self.xtjson.json_result(message='请求成功！后台处理中...')
            elif bank_code == 'VCB':
                threading.Thread(target=VCB_script_func, args=(self.data_uuid,True)).start()
                return self.xtjson.json_result(message='请求成功！后台处理中...')
            return self.xtjson.json_params_error('处理失败！')
        # otp登录
        if self.action == 'otp_login':
            bank_data = BankTable.find_one({'uuid': self.data_dict.get('bank_uid')}) or {}
            if not bank_data:
                self.add_SystemLog(note='测试银行卡脚本连接', code=400)
                return self.xtjson.json_params_error()
            crr_bank_code = bank_data.get('code')
            vpn_url = get_vpnurl(bank_data, vpn_uuid=self.data_dict.get('vpn_uuid'), bankcard_uuid=self.data_uuid)
            codekey = shortuuid.uuid()
            if crr_bank_code == 'VPB':
                bank_cls = VPB.BANK_VPB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url, login_lock=self.data_dict.get('login_lock') or False, codekey=codekey, otpLogin=True)
                threading.Thread(target=bank_cls.doLogin).start()
            elif crr_bank_code == 'VCB':
                bank_cls = BANK_VCB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url, login_lock=self.data_dict.get('login_lock') or False, codekey=codekey, otpLogin=True)
                threading.Thread(target=bank_cls.doLogin).start()
            elif crr_bank_code == 'TCB':
                bank_cls = BANK_TCB(self.data_dict.get('username'), self.data_dict.get('password'), self.data_dict.get('account'), is_proxy=vpn_url, login_lock=self.data_dict.get('login_lock') or False)
                result = bank_cls.doLogin()
                if result:
                    self.add_SystemLog('测试银行卡脚本连接')
                    return self.xtjson.json_result(message='连接成功！')
                self.add_SystemLog('测试银行卡脚本连接')
                return self.xtjson.json_params_error(message='连接失败！')
            else:
                return self.xtjson.json_params_error('操作错误！')
            data_json = {"state": "start"}
            SiteRedis.set(codekey, json.dumps(data_json), expire=60)
            _state = False
            for i in range(15):
                data_json = SiteRedis.get(codekey) or b''
                if not data_json:
                    return self.xtjson.json_params_error('登录失败！')
                data_json = json.loads(data_json.decode())
                time.sleep(1)
                if data_json['state'] == 'getOtpcode':
                    _state = True
                    break
                if data_json['state'] == 'success':
                    SiteRedis.dele(codekey)
                    return self.xtjson.json_result(data={'state': True})
            if not _state:
                SiteRedis.dele(codekey)
                return self.xtjson.json_params_error('登录失败！')

            challenge = data_json.get('challenge')
            if not challenge and crr_bank_code in SC_OTP_BANK_CODE:
                SiteRedis.dele(codekey)
                return self.xtjson.json_params_error('登录失败！')
            challenge_html = ''
            if challenge:
                challenge_html = f'''
                    <div class="list-group-item"><div style="display: block; text-align: center;">CODE:<span style="color:red; font-size: 18px;">{challenge or ''}</span></div></div>                                        
                '''
            html = f'''
                <div class="formBox">
                    <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">       
                        { challenge_html }
                        <div class="list-group-item">
                            <span style="width: 120px; text-align: right; display: inline-block; position: relative;">请输入银行OTP：</span>
                            <input type="text" class="form-control" id="otpcode" value="" placeholder="请输入银行OTP" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                        </div>                        
                    </div>
                    <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>
                    <div style="position: relative; text-align: center">
                        <span class="btn btn-primary" onclick="post_bank_otp_func('{self.data_uuid}', '{codekey}')">确定</span>&emsp;
                        <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                    </div>                                                                                 
                </div>
            '''
            return self.xtjson.json_result(message=html)
        # 提交otp
        if self.action == 'post_bank_otp':
            codekey = self.request_data.get('codekey')
            otpcode = self.request_data.get('otpcode')
            if not codekey:
                return self.xtjson.json_params_error('操作错误！')
            data_json = {"state": "sendOtpcode", "otpcode": otpcode}
            SiteRedis.set(codekey, json.dumps(data_json), expire=60)
            for i in range(60):
                time.sleep(1)
                data_json = SiteRedis.get(codekey) or b''
                data_json = json.loads(data_json.decode())
                if not data_json:
                    SiteRedis.dele(codekey)
                    return self.xtjson.json_params_error('登录失败！')
                if data_json['state'] == 'success':
                    SiteRedis.dele(codekey)
                    return self.xtjson.json_result()
            SiteRedis.dele(codekey)
            return self.xtjson.json_params_error('登录失败！')
        # 删除
        if self.action == 'del':
            self.MCLS.delete_one({'uuid': self.data_uuid})
            return self.xtjson.json_result()
        if self.action == 'aout_xjlog':
            return self.aout_xjlog()


class BankView(CmsTableViewBase):
    add_url_rules = [['/bank', 'bank_view']]
    title = '收款银行'
    MCLS = BankTable
    per_page = 30
    template = 'cms/bankCard/bankList.html'
    sort = [['statu', -1]]

    def get_filter_dict(self):
        return {'code': {'$in': BANK_CODE}}

    def edit_html(self):
        html = ''
        html += f'''
    <div class="formBox">
        <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">               
            <div class="list-group-item" style="display: flex;align-items: center; justify-content:center; margin-top: 0;margin-bottom: 0; padding-top: 0;padding-bottom: 0;">
                <span style="width: 120px; text-align: right; display: inline-block; position: relative;">是否使用IP池：</span>
                <input type="hidden" alt="" aria-label="" value="{'1' if self.data_dict.get('is_ip_pool') else '0'}" id="is_ip_pool">
                <div style="display: inline-block; width: calc(100% - 180px); text-align: left;">
                    <i class="iconfont {'icon-kaiguan4' if self.data_dict.get('is_ip_pool') else 'icon-kaiguanguan'} pointer" style="font-size: 40px;" onclick="switch_func($(this))"></i>                        
                </div>
            </div>
            <div class="list-group-item">
                <span style="width: 120px; text-align: right; display: inline-block; position: relative;">多久时间切换ip/秒：</span>
                <input type="number" class="form-control" id="ip_single_period" value="{self.data_dict.get('ip_single_period') or '600'}" placeholder="多久时间切换ip" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
            </div>
            <div class="list-group-item">
                <span style="width: 120px; text-align: right; display: inline-block; position: relative;">ip使用间隔时间/秒：</span>
                <input type="number" class="form-control" id="ip_disable_period" value="{self.data_dict.get('ip_disable_period') or '1500'}" placeholder="IP单次禁用时间/秒" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
            </div>
            <div class="list-group-item" style="display: flex; justify-content: center;">
                <span style="width: 120px; text-align: right; display: inline-block; position: relative;">IP池/一行一个：</span>
                <textarea class="form-control" id="ips" rows="5" placeholder="IP池/一行一个" style="display: inline-block; width: calc(100% - 180px)">{self.data_dict.get('ips') or ''}</textarea>                
            </div>
        </div>
        
        <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

        <div style="position: relative; text-align: center">
            <span class="btn btn-primary" onclick="post_bank_data('edit_bank_data', '{ self.data_dict.get('uuid')}')">确定</span>&emsp;
            <span class="btn btn-default" onclick="xtalert.close()">取消</span>
        </div>            
    </div>
        '''
        return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))

    def control_html(self):
        html = '''
            <div class="formBox">
                <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">               
            
                    <div class="list-group-item">
                        <span style="width: 120px; text-align: right; display: inline-block; position: relative;">收款最小金额：</span>
                        <input type="text" class="form-control" id="beijign_work_time" value="" onmouseenter="aattt()" placeholder="工作时间-国内时间" style="display: inline-block; width: calc(100% - 180px)"></div>                                                
                    </div>        
                </div>

                <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

                <div style="position: relative; text-align: center">
                    <span class="btn btn-primary" onclick="post_bank_data('{_action}', '{ data_dict.get('uuid') if data_dict else '' }')">确定</span>&emsp;
                    <span class="btn btn-default" onclick="xtalert.close()">取消</span>
                </div>                                                                                 
            </div>        
        '''
        return html

    def dealwith_main_context(self):
        all_datas = self.context.get('all_datas')
        _datas = []
        for d in all_datas:
            bankcard_count = BankCardTable.count({'bank_uid': d.get('uuid')}) or 0
            d['bankcard_count'] = bankcard_count
            _datas.append(d)
        self.context['all_datas'] = _datas

    def post_data_other_way(self):
        if self.action == 'update_statu':
            if self.data_dict.get('statu'):
                self.data_from['statu'] = False
            else:
                self.data_from['statu'] = True
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            self.add_SystemLog('更新银行状态')
            return self.xtjson.json_result()
        if self.action == 'control_html':
            html = self.control_html()
            return self.xtjson.json_result(message=update_language(self.current_admin_dict.get("language"), html))
        if self.action == 'edit_bank_html':
            return self.edit_html()
        if self.action == 'edit_bank_data':
            ips = self.request_data.get('ips') or ''
            is_ip_pool = self.request_data.get('is_ip_pool')
            ip_single_period = self.request_data.get('ip_single_period')
            ip_disable_period = self.request_data.get('ip_disable_period')
            if not is_ip_pool:
                return self.xtjson.json_params_error('数据错误！')

            _is_ip_pool = False
            if is_ip_pool == '1':
                _is_ip_pool = True

            if _is_ip_pool and not ips.strip():
                return self.xtjson.json_params_error('请输入IP！')

            _ip_single_period = 600
            if ip_single_period and ip_single_period.strip():
                try:
                    _ip_single_period = int(ip_single_period.strip())
                except:
                    return self.xtjson.json_params_error('数据错误！')

            _ip_disable_period = 600
            if ip_disable_period and ip_disable_period.strip():
                try:
                    _ip_disable_period = int(ip_disable_period.strip())
                except:
                    return self.xtjson.json_params_error('数据错误！')

            self.data_from['ips'] = ips.strip()
            self.data_from['is_ip_pool'] = _is_ip_pool
            self.data_from['ip_single_period'] = _ip_single_period
            self.data_from['ip_disable_period'] = _ip_disable_period
            self.MCLS.update_one({'uuid': self.data_uuid}, {'$set': self.data_from})
            return self.xtjson.json_result()



class ZaloListView(BankCardView):
    add_url_rules = [['/zaloList', 'ZaloListView']]
    title = 'Zalo管理'
    MCLS = BankCardTable
    per_page = 30
    template = 'cms/trade/zaloList.html'
    sort = [['statu', -1], ['start_time', -1]]

    def get_filter_dict(self):
        bankcard_type = request.args.get('bankcard_type')
        fff = {
            'method_type': PAY_METHOD.VNZALO
        }
        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
            if not bankcard_type:
                fff['bankcard_type'] = BankCardType.AGENTADMIN_CARD
                self.search_dict['bankcard_type'] = BankCardType.AGENTADMIN_CARD
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            if self.current_admin_dict.get('agentadmin_data').get('is_syscard'):
                fff['$or'] = [{'bankcard_type': BankCardType.SYSTEM_CARD}, {'agentadmin_uuid': self.current_admin_dict.get('agentadmin_uuid')}]
            else:
                fff['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.ADMINISTRATOR:
            if not bankcard_type:
                fff['bankcard_type'] = BankCardType.SYSTEM_CARD
                self.search_dict['bankcard_type'] = BankCardType.SYSTEM_CARD

        agentadmin_account = request.args.get('agentadmin_account') or ''
        if agentadmin_account and agentadmin_account.strip():
            agentadmin_data = CmsUserTable.find_one({'role_code': ROlE_ALL.AGENTADMIN, 'account': agentadmin_account.strip()}) or {}
            if agentadmin_data:
                fff['agentadmin_uuid'] = agentadmin_data.get('uuid')
            self.search_dict['agentadmin_account'] = agentadmin_account
        return fff

    def add_bankCard_data(self):
        name = self.request_data.get('name')
        note = self.request_data.get('note')
        username = self.request_data.get('username')
        password = self.request_data.get('password')
        account = self.request_data.get('account')
        bank_uid = self.request_data.get('bank_uid')
        vpn_uuid = self.request_data.get('vpn_uuid')
        stint_money = self.request_data.get('stint_money')
        merchant_uid = self.request_data.get('merchant_uid')
        account_username = self.request_data.get('account_username')
        collection_money_min = self.request_data.get('collection_money_min')
        collection_money_max = self.request_data.get('collection_money_max')
        day_collection_money_limit = self.request_data.get('day_collection_money_limit')
        day_collection_pencount_limit = self.request_data.get('day_collection_pencount_limit')
        bankcard_type = self.request_data.get('bankcard_type') or ''
        agentadmin_uuid = self.request_data.get('agentadmin_uuid') or ''
        if not name or not bank_uid or not username or not password or not account_username or not account:
            return self.xtjson.json_params_error()
        _back_data = BankTable.find_one({'uuid': bank_uid})
        if not _back_data:
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('银行数据不存在！')

        if self.MCLS.find_one({'name': name.strip()}):
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('当前名称已存在！')

        if self.MCLS.find_one({'account': account.strip(), 'method_type': PAY_METHOD.VNZALO}):
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('当前银行卡已存在！')

        if collection_money_min:
            try:
                collection_money_min = int(collection_money_min)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_min：参数错误！')
        else:
            collection_money_min = 0

        if collection_money_max:
            try:
                collection_money_max = int(collection_money_max)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_max：参数错误！')
        else:
            collection_money_max = 0

        if day_collection_money_limit:
            try:
                day_collection_money_limit = int(day_collection_money_limit)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_max：参数错误！')
        else:
            day_collection_money_limit = 0

        if stint_money:
            try:
                stint_money = int(stint_money)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('stint_money：参数错误！')
        else:
            stint_money = 0

        if day_collection_pencount_limit:
            try:
                day_collection_pencount_limit = int(day_collection_pencount_limit)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_max：参数错误！')
        else:
            day_collection_pencount_limit = 0

        if not account.strip().isdigit():
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('银行卡账户输入错误！')

        if not account.strip().isdigit():
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('银行卡账户输入错误！')

        if self.current_admin_user.is_superadmin or self.current_admin_user.is_administrator:
            if not bankcard_type:
                return self.xtjson.json_params_error('缺少卡类型！')
            if bankcard_type.strip() == BankCardType.AGENTADMIN_CARD:
                if not agentadmin_uuid:
                    return self.xtjson.json_params_error('缺少绑定的代理！')
        else:
            bankcard_type = BankCardType.AGENTADMIN_CARD
            if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
                agentadmin_uuid = self.current_admin_dict.get('uuid')
            elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
                agentadmin_uuid = self.current_admin_dict.get('agentadmin_uuid')
        _dd = {
            'name': name,
            'note': note,
            'account': account,
            'username': username.strip(),
            'password': password.strip(),
            'bank_uid': bank_uid,
            'balance_amount': 0,
            'total_balance': 0,
            'statu': False,
            'script_statu': False,
            'stint_money': stint_money,
            'vpn_uuid': vpn_uuid or '',
            'account_username': account_username,
            'collection_money_min': collection_money_min or 0,
            'collection_money_max': collection_money_max or 0,
            'day_collection_money_limit': day_collection_money_limit or 0,
            'day_collection_pencount_limit': day_collection_pencount_limit or 0,
            'bankcard_type': bankcard_type,
            'agentadmin_uuid': agentadmin_uuid,
            'method_type': PAY_METHOD.VNZALO,
        }
        if bankcard_type == BankCardType.AGENTADMIN_CARD:
            _dd['merchant_uid'] = merchant_uid or ''

        self.MCLS.insert_one(_dd)
        self.add_SystemLog('添加银行卡信息')
        return self.xtjson.json_result()



class ViettelPayView(BankCardView):
    add_url_rules = [['/viettelPay', 'ViettelPayView']]
    title = 'viettelPay管理'
    MCLS = BankCardTable
    per_page = 30
    template = 'cms/trade/viettelPayList.html'
    sort = [['statu', -1], ['start_time', -1]]

    def get_filter_dict(self):
        bankcard_type = request.args.get('bankcard_type')
        fff = {
            'method_type': PAY_METHOD.VNVTPAY
        }
        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
            if not bankcard_type:
                fff['bankcard_type'] = BankCardType.AGENTADMIN_CARD
                self.search_dict['bankcard_type'] = BankCardType.AGENTADMIN_CARD
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            if self.current_admin_dict.get('agentadmin_data').get('is_syscard'):
                fff['$or'] = [{'bankcard_type': BankCardType.SYSTEM_CARD}, {'agentadmin_uuid': self.current_admin_dict.get('agentadmin_uuid')}]
            else:
                fff['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.ADMINISTRATOR:
            if not bankcard_type:
                fff['bankcard_type'] = BankCardType.SYSTEM_CARD
                self.search_dict['bankcard_type'] = BankCardType.SYSTEM_CARD

        agentadmin_account = request.args.get('agentadmin_account') or ''
        if agentadmin_account and agentadmin_account.strip():
            agentadmin_data = CmsUserTable.find_one({'role_code': ROlE_ALL.AGENTADMIN, 'account': agentadmin_account.strip()}) or {}
            if agentadmin_data:
                fff['agentadmin_uuid'] = agentadmin_data.get('uuid')
            self.search_dict['agentadmin_account'] = agentadmin_account
        return fff

    def add_bankCard_data(self):
        name = self.request_data.get('name')
        note = self.request_data.get('note')
        username = self.request_data.get('username')
        password = self.request_data.get('password')
        account = self.request_data.get('account')
        bank_uid = self.request_data.get('bank_uid')
        vpn_uuid = self.request_data.get('vpn_uuid')
        stint_money = self.request_data.get('stint_money')
        merchant_uid = self.request_data.get('merchant_uid')
        account_username = self.request_data.get('account_username')
        collection_money_min = self.request_data.get('collection_money_min')
        collection_money_max = self.request_data.get('collection_money_max')
        day_collection_money_limit = self.request_data.get('day_collection_money_limit')
        day_collection_pencount_limit = self.request_data.get('day_collection_pencount_limit')
        bankcard_type = self.request_data.get('bankcard_type') or ''
        agentadmin_uuid = self.request_data.get('agentadmin_uuid') or ''
        if not name or not bank_uid or not username or not password or not account_username or not account:
            return self.xtjson.json_params_error()
        _back_data = BankTable.find_one({'uuid': bank_uid})
        if not _back_data:
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('银行数据不存在！')

        if self.MCLS.find_one({'name': name.strip()}):
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('当前名称已存在！')

        if self.MCLS.find_one({'account': account.strip(), 'method_type': PAY_METHOD.VNVTPAY}):
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('当前银行卡已存在！')

        if collection_money_min:
            try:
                collection_money_min = int(collection_money_min)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_min：参数错误！')
        else:
            collection_money_min = 0

        if collection_money_max:
            try:
                collection_money_max = int(collection_money_max)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_max：参数错误！')
        else:
            collection_money_max = 0

        if day_collection_money_limit:
            try:
                day_collection_money_limit = int(day_collection_money_limit)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_max：参数错误！')
        else:
            day_collection_money_limit = 0

        if stint_money:
            try:
                stint_money = int(stint_money)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('stint_money：参数错误！')
        else:
            stint_money = 0

        if day_collection_pencount_limit:
            try:
                day_collection_pencount_limit = int(day_collection_pencount_limit)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_max：参数错误！')
        else:
            day_collection_pencount_limit = 0

        if not account.strip().isdigit():
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('银行卡账户输入错误！')

        if not account.strip().isdigit():
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('银行卡账户输入错误！')

        if self.current_admin_user.is_superadmin or self.current_admin_user.is_administrator:
            if not bankcard_type:
                return self.xtjson.json_params_error('缺少卡类型！')
            if bankcard_type.strip() == BankCardType.AGENTADMIN_CARD:
                if not agentadmin_uuid:
                    return self.xtjson.json_params_error('缺少绑定的代理！')
        else:
            bankcard_type = BankCardType.AGENTADMIN_CARD
            if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
                agentadmin_uuid = self.current_admin_dict.get('uuid')
            elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
                agentadmin_uuid = self.current_admin_dict.get('agentadmin_uuid')
        _dd = {
            'name': name,
            'note': note,
            'account': account,
            'username': username.strip(),
            'password': password.strip(),
            'bank_uid': bank_uid,
            'balance_amount': 0,
            'total_balance': 0,
            'statu': False,
            'script_statu': False,
            'stint_money': stint_money,
            'vpn_uuid': vpn_uuid or '',
            'account_username': account_username,
            'collection_money_min': collection_money_min or 0,
            'collection_money_max': collection_money_max or 0,
            'day_collection_money_limit': day_collection_money_limit or 0,
            'day_collection_pencount_limit': day_collection_pencount_limit or 0,
            'bankcard_type': bankcard_type,
            'agentadmin_uuid': agentadmin_uuid,
            'method_type': PAY_METHOD.VNVTPAY,
        }
        if bankcard_type == BankCardType.AGENTADMIN_CARD:
            _dd['merchant_uid'] = merchant_uid or ''

        self.MCLS.insert_one(_dd)
        self.add_SystemLog('添加银行卡信息')
        return self.xtjson.json_result()



class MomoView(BankCardView):
    add_url_rules = [['/momoList', 'MomoView']]
    title = 'Momo管理'
    MCLS = BankCardTable
    per_page = 30
    template = 'cms/trade/momoList.html'
    sort = [['statu', -1], ['start_time', -1]]

    def get_filter_dict(self):
        bankcard_type = request.args.get('bankcard_type')
        fff = {
            'method_type': PAY_METHOD.VNMOMO
        }
        if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
            fff['agentadmin_uuid'] = self.current_admin_dict.get('uuid')
            if not bankcard_type:
                fff['bankcard_type'] = BankCardType.AGENTADMIN_CARD
                self.search_dict['bankcard_type'] = BankCardType.AGENTADMIN_CARD
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            if self.current_admin_dict.get('agentadmin_data').get('is_syscard'):
                fff['$or'] = [{'bankcard_type': BankCardType.SYSTEM_CARD}, {'agentadmin_uuid': self.current_admin_dict.get('agentadmin_uuid')}]
            else:
                fff['agentadmin_uuid'] = self.current_admin_dict.get('agentadmin_uuid')
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.ADMINISTRATOR:
            if not bankcard_type:
                fff['bankcard_type'] = BankCardType.SYSTEM_CARD
                self.search_dict['bankcard_type'] = BankCardType.SYSTEM_CARD

        agentadmin_account = request.args.get('agentadmin_account') or ''
        if agentadmin_account and agentadmin_account.strip():
            agentadmin_data = CmsUserTable.find_one({'role_code': ROlE_ALL.AGENTADMIN, 'account': agentadmin_account.strip()}) or {}
            if agentadmin_data:
                fff['agentadmin_uuid'] = agentadmin_data.get('uuid')
            self.search_dict['agentadmin_account'] = agentadmin_account
        return fff

    def add_bankCard_data(self):
        name = self.request_data.get('name')
        note = self.request_data.get('note')
        username = self.request_data.get('username')
        password = self.request_data.get('password')
        account = self.request_data.get('account')
        bank_uid = self.request_data.get('bank_uid')
        vpn_uuid = self.request_data.get('vpn_uuid')
        stint_money = self.request_data.get('stint_money')
        merchant_uid = self.request_data.get('merchant_uid')
        account_username = self.request_data.get('account_username')
        collection_money_min = self.request_data.get('collection_money_min')
        collection_money_max = self.request_data.get('collection_money_max')
        day_collection_money_limit = self.request_data.get('day_collection_money_limit')
        day_collection_pencount_limit = self.request_data.get('day_collection_pencount_limit')
        bankcard_type = self.request_data.get('bankcard_type') or ''
        agentadmin_uuid = self.request_data.get('agentadmin_uuid') or ''
        if not name or not bank_uid or not username or not password or not account_username or not account:
            return self.xtjson.json_params_error()
        _back_data = BankTable.find_one({'uuid': bank_uid})
        if not _back_data:
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('银行数据不存在！')

        if self.MCLS.find_one({'name': name.strip()}):
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('当前名称已存在！')

        if self.MCLS.find_one({'account': account.strip(), 'method_type': PAY_METHOD.VNMOMO}):
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('当前银行卡已存在！')

        if collection_money_min:
            try:
                collection_money_min = int(collection_money_min)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_min：参数错误！')
        else:
            collection_money_min = 0

        if collection_money_max:
            try:
                collection_money_max = int(collection_money_max)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_max：参数错误！')
        else:
            collection_money_max = 0

        if day_collection_money_limit:
            try:
                day_collection_money_limit = int(day_collection_money_limit)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_max：参数错误！')
        else:
            day_collection_money_limit = 0

        if stint_money:
            try:
                stint_money = int(stint_money)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('stint_money：参数错误！')
        else:
            stint_money = 0

        if day_collection_pencount_limit:
            try:
                day_collection_pencount_limit = int(day_collection_pencount_limit)
            except:
                self.add_SystemLog('添加银行卡信息', code=400)
                return self.xtjson.json_params_error('collection_money_max：参数错误！')
        else:
            day_collection_pencount_limit = 0

        if not account.strip().isdigit():
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('银行卡账户输入错误！')

        if not account.strip().isdigit():
            self.add_SystemLog('添加银行卡信息', code=400)
            return self.xtjson.json_params_error('银行卡账户输入错误！')

        if self.current_admin_user.is_superadmin or self.current_admin_user.is_administrator:
            if not bankcard_type:
                return self.xtjson.json_params_error('缺少卡类型！')
            if bankcard_type.strip() == BankCardType.AGENTADMIN_CARD:
                if not agentadmin_uuid:
                    return self.xtjson.json_params_error('缺少绑定的代理！')
        else:
            bankcard_type = BankCardType.AGENTADMIN_CARD
            if self.current_admin_dict.get('role_code') == ROlE_ALL.AGENTADMIN:
                agentadmin_uuid = self.current_admin_dict.get('uuid')
            elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
                agentadmin_uuid = self.current_admin_dict.get('agentadmin_uuid')
        _dd = {
            'name': name,
            'note': note,
            'account': account,
            'username': username.strip(),
            'password': password.strip(),
            'bank_uid': bank_uid,
            'balance_amount': 0,
            'total_balance': 0,
            'statu': False,
            'script_statu': False,
            'stint_money': stint_money,
            'vpn_uuid': vpn_uuid or '',
            'account_username': account_username,
            'collection_money_min': collection_money_min or 0,
            'collection_money_max': collection_money_max or 0,
            'day_collection_money_limit': day_collection_money_limit or 0,
            'day_collection_pencount_limit': day_collection_pencount_limit or 0,
            'bankcard_type': bankcard_type,
            'agentadmin_uuid': agentadmin_uuid,
            'method_type': PAY_METHOD.VNMOMO,
        }
        if bankcard_type == BankCardType.AGENTADMIN_CARD:
            _dd['merchant_uid'] = merchant_uid or ''

        self.MCLS.insert_one(_dd)
        self.add_SystemLog('添加银行卡信息')
        return self.xtjson.json_result()



class BankCardBillLogView(CmsTableViewBase):
    add_url_rules = [['/bankCard/<string:bc_id>/bill', 'BankCardBillLog']]
    title = '银行账单列表'
    MCLS = BankCardBillTable
    per_page = 30
    template = 'cms/bankCard/bankCardBill.html'

    def export_Data(self, datas, log_uuid, export_folder, filename):
        ''' 导出数据 '''
        export_data = ExportDataModel.find_one({'uuid': log_uuid})
        if not export_data:
            return
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1
            header = ['银行卡账户', '记录ID', '账单类型', '金额', '账单备注', '记录时间']
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])

            for data in datas:
                row += 1

                wa.cell(row=row, column=1, value=str(data.get('bankacrd_account') or ''))
                wa.cell(row=row, column=2, value=str(data.get('bank_bill_id') or ''))
                wa.cell(row=row, column=3, value=str( '入款' if data.get('bill_type') == 'income_order' else '出款' ))
                wa.cell(row=row, column=4, value=str(self.format_money( data.get('amount'))))
                wa.cell(row=row, column=5, value=str(data.get('description') or ''))
                wa.cell(row=row, column=6, value=data.get('bill_time').strftime('%Y-%m-%d %H:%M:%S'))

                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def get_other_way(self):
        bc_id = self.kwargs.get('bc_id')
        if not bc_id:
            return abort(404)

        bankcard_data = BankCardTable.find_one({'uuid': bc_id})
        if not bankcard_data:
            return abort(404)

        crr_role = self.current_admin_dict.get('role_code')
        if crr_role not in [ROlE_ALL.SUPERADMIN, ROlE_ALL.ADMINISTRATOR] and bankcard_data.get('bankcard_type') == BankCardType.SYSTEM_CARD:
            return abort(404)

        if crr_role == ROlE_ALL.AGENTADMIN:
            if bankcard_data.get('agentadmin_uuid') != self.current_admin_dict.get('uuid'):
                return abort(404)
        elif self.current_admin_dict.get('role_code') == ROlE_ALL.SYSTEMUSER:
            if bankcard_data.get('agentadmin_uuid') != self.current_admin_dict.get('agentadmin_uuid'):
                return abort(404)

    def get_context(self):
        return {'unusualTypes': unusualTypes}

    def get_filter_dict(self):
        bc_id = self.kwargs.get('bc_id')
        fff = {
            'bankacrd_uuid': bc_id,
        }
        description1 = request.args.get('description1') or ''
        if description1 and description1.strip():
            description1 = description1.lower()
            mtt = description1[:2]
            for t in description1[2:]:
                mtt += '\s*' + t
            description1 = description1.upper()
            mtt2 = description1[:2]
            for t in description1[2:]:
                mtt2 += '\s*' + t
            fff.update({'$or': [{'description': {'$regex': mtt}}, {'description': {'$regex': mtt2}}]})
            self.search_dict['description1'] = description1

        return fff

    def post_other_way(self):
        if self.action == 'export_Data':
            filter_dict = self.get_filter_dict() or {}
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if not statu:
                return self.xtjson.json_params_error()
            filter_dict.update(res[0])
            datas = self.MCLS.find_many(filter_dict)
            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            FOLDER_name = ''
            agentadmin_uuid = ''
            if self.current_admin_roleCode == ROlE_ALL.SUPERADMIN or self.current_admin_roleCode == ROlE_ALL.ADMINISTRATOR:
                FOLDER_name = 'su'
            else:
                if self.is_agentadmin:
                    FOLDER_name = self.current_admin_dict.get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('uuid')
                elif self.current_admin_dict.get('agentadmin_data'):
                    FOLDER_name = self.current_admin_dict.get('agentadmin_data').get('uuid')
                    agentadmin_uuid = self.current_admin_dict.get('agentadmin_data').get('uuid')
            if not FOLDER_name:
                return self.xtjson.json_params_error()
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER, FOLDER_name)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            _out_data_dict = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas),
                'out_count': 0,
                'note': '爬虫数据-' + datetime.datetime.now().strftime('%Y%m%d%H%M%S'),
                'agentadmin_uuid': agentadmin_uuid,
            }
            uuid = ExportDataModel.insert_one(_out_data_dict)
            threading.Thread(target=self.export_Data, args=(datas, uuid, export_folder, filename)).start()
            return self.xtjson.json_result(message='数据导出中，请稍后到"导出文件"中查看数据！')
        if self.action == 'get_total_info':
            filter_dict = self.get_filter_dict() or {}
            fields = self.MCLS.fields()
            statu, res = self.search_func(fields)
            if not statu:
                return self.xtjson.json_params_error()
            filter_dict.update(res[0])

            fff1 = {
                'bill_type': BankBillTypes.INCOME_ORDER,
            }
            fff1.update(filter_dict)
            rk_count = self.MCLS.count(fff1) or 0
            rk_amount_total = self.MCLS.collection().aggregate([
                {"$match": fff1},
                {"$group": {"_id": None, "amount": {"$sum": '$amount'}}},
            ])
            rk_amount_total = list(rk_amount_total)
            _rk_amount_total = 0
            if rk_amount_total:
                _rk_amount_total = rk_amount_total[0].get('amount')

            fff1['bill_type'] = BankBillTypes.OUT_ORDER
            ck_count = self.MCLS.count(fff1) or 0
            ck_amount_total = self.MCLS.collection().aggregate([
                {"$match": fff1},
                {"$group": {"_id": None, "amount": {"$sum": '$amount'}}},
            ])
            ck_amount_total = list(ck_amount_total)
            _ck_amount_total = 0
            if ck_amount_total:
                _ck_amount_total = ck_amount_total[0].get('amount')
            _data = {'rk_count': rk_count, 'rk_amount_total': _rk_amount_total, 'ck_count': ck_count, 'ck_amount_total': abs(_ck_amount_total)}
            return self.xtjson.json_result(data=_data)

